import os
import tempfile
import sys
import re
import json
import math
import colorsys
import itertools
import shutil
import subprocess
import unicodedata
import webbrowser
from datetime import datetime, timedelta
import numpy as np
import pandas as pd
import openpyxl

# Garante que print() nunca derrube o programa por causa de um caractere
# fora da tabela de código do console (acentos, "→" etc.). Testado e
# confirmado com um .exe de diagnóstico: no modo --noconsole do
# PyInstaller, sys.stdout começa como None (nesse caso o print() do próprio
# Python já vira um no-op sozinho, sem erro) — só que alguma importação
# mais abaixo (o suspeito é a cadeia pywebview/pythonnet, usada pro
# WebView2) REATRIBUI sys.stdout pra um stream real com codificação cp1252
# no meio do caminho, e QUALQUER print() com acento depois disso derruba o
# programa inteiro antes mesmo da janela abrir. Um reconfigure() feito uma
# vez só aqui em cima não sobrevive a essa troca; por isso a proteção é no
# print() em si, que lê sys.stdout de novo a cada chamada.
_print_original = print


def _print_seguro(*args, **kwargs):
    try:
        _print_original(*args, **kwargs)
    except UnicodeEncodeError:
        destino = kwargs.get("file") or sys.stdout
        if destino is None:
            return
        texto = kwargs.get("sep", " ").join(str(a) for a in args) + kwargs.get("end", "\n")
        try:
            destino.buffer.write(
                texto.encode(getattr(destino, "encoding", None) or "utf-8", errors="replace")
            )
        except Exception:
            pass
    except AttributeError:
        pass  # sys.stdout é None (modo --noconsole sem stream nenhum) — descarta


print = _print_seguro

# tkinter e pywebview só existem/fazem sentido no modo desktop (janela nativa
# + diálogos de erro). No servidor web (Linux, sem Tcl/Tk nem backend de
# WebView instalado) essas duas libs não estão disponíveis — e não fazem
# falta lá, porque todo uso delas está atrás de checagens `if tk is not
# None` / dentro de `abrir_interface_filtros` (só chamada em modo desktop,
# ver `if __name__ == "__main__"` no fim do arquivo). Sem esse fallback, o
# simples `import` deste arquivo pelo servidor web já quebraria de cara.
try:
    import tkinter as tk
    from tkinter import messagebox
except Exception:
    tk = None
    messagebox = None

try:
    import webview  # painel de filtros (HTML/CSS/JS) — pip install pywebview
except Exception:
    webview = None

# Página do mapa mental (botão "MAPA MENTAL", ao lado de GERENCIAL/DASHBOARD
# no topo do painel) — mesmo componente HTML/CSS/JS já usado no projeto
# "Controle de Prazos", adaptado para a árvore Secretaria/Órgão > Eixo >
# Objeto > Ação e para a paleta de FASE já usada no resto deste painel (ver
# mapa_mental_html.py, montar_html_mapa_mental).
from mapa_mental_html import montar_html_mapa_mental

# Ligada pela variável de ambiente PAC_WEB_MODE=1, definida pelo servidor
# web (servidor_web.py). Controla os pontos do código que só fazem sentido
# numa máquina desktop com o Windows na frente do usuário (abrir o PDF
# gerado direto no leitor padrão do sistema via os.startfile) — no servidor,
# quem abre o PDF é o navegador de quem está acessando, não o processo
# Python rodando no host remoto.
MODO_WEB = os.environ.get("PAC_WEB_MODE") == "1"

from reportlab.lib import colors, pagesizes
from reportlab.lib.units import mm
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.graphics.shapes import Drawing, String, Circle, Line, Path, Rect, Wedge
from reportlab.graphics.charts.piecharts import Pie
from reportlab.graphics.charts.barcharts import VerticalBarChart
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import (
    Flowable,
    KeepTogether,
    PageBreak,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)
from reportlab.pdfgen import canvas as reportlab_canvas
from reportlab.graphics.barcode.qr import QrCodeWidget
from reportlab.graphics import renderPDF

# =====================================================
# 1. CAMINHOS, CONFIGURAÇÕES E TIMESTAMPS
# =====================================================

# Pasta onde os arquivos de apoio devem estar (planilha, imagens PAC.png/
# GOVERNO.PNG e fontes DIN): sempre a pasta onde o programa está salvo, seja
# rodando como script .py ou como .exe empacotado (PyInstaller) — assim
# funciona igual em qualquer computador da rede, independente de onde o
# usuário deu duplo clique.
if getattr(sys, "frozen", False):
    PASTA_BASE = os.path.dirname(sys.executable)
else:
    PASTA_BASE = os.path.dirname(os.path.abspath(__file__))

def caminho_recurso(nome_arquivo):
    return os.path.join(PASTA_BASE, nome_arquivo)


def _erro_fatal_inicializacao(mensagem):
    # Erro que impede o programa de continuar (planilha ilegível, coluna
    # essencial faltando etc.), disparado ainda durante o carregamento dos
    # dados — antes de existir janela ou servidor. Sempre imprime no
    # console/log; no modo desktop, também mostra um alerta nativo (só
    # possível quando tkinter está disponível, ver import defensivo no topo
    # do arquivo). No modo web isso vira um erro de inicialização visível no
    # log do servidor, em vez de uma janela travada esperando clique em
    # máquina nenhuma.
    print(f"[PAC - Relatório Gerencial] ERRO FATAL: {mensagem}", file=sys.stderr)
    if tk is not None and not MODO_WEB:
        _raiz_erro = tk.Tk()
        _raiz_erro.withdraw()
        messagebox.showerror("PAC - Relatório Gerencial", mensagem)
    sys.exit(1)


arquivo_excel = caminho_recurso("PANORAMA - PAC ORIGINAL - PAC SELEÇÕES - 2026.xlsx")

# A data/hora de "atualizado em" vem das PROPRIEDADES internas do próprio
# arquivo .xlsx (o campo "Modificado em" que o Excel grava sozinho ao
# salvar) — não da data de modificação do arquivo no sistema operacional
# (os.path.getmtime). Essa segunda opção muda sozinha sempre que o arquivo
# é só COPIADO pra outro lugar, sem ninguém ter mexido nos dados: é
# exatamente o que acontece a cada deploy do servidor web (o Git faz um
# checkout novo do repositório a cada publicação, e todo arquivo "nasce" com
# a data/hora desse checkout, não a da última edição real) e também numa
# migração de máquina (ver MIGRACAO.md). A metadata interna do Excel viaja
# junto com o CONTEÚDO do arquivo e não muda em nenhuma dessas duas
# situações — por isso é a fonte confiável aqui, nos dois modos (desktop e
# web).
try:
    _propriedades_excel = openpyxl.load_workbook(arquivo_excel, read_only=True).properties
    _modificado_utc = _propriedades_excel.modified or _propriedades_excel.created
    if _modificado_utc is None:
        raise ValueError("planilha sem data de modificação nas propriedades internas")
    # O Excel grava esse campo sempre em UTC (sufixo "Z" no XML interno da
    # planilha, docProps/core.xml) — o openpyxl devolve um datetime "naive"
    # (sem fuso), mas o valor em si É UTC. Sem esse ajuste de -3h, o horário
    # mostrado sairia adiantado (Bahia/Brasília é UTC-3 o ano todo desde o
    # fim do horário de verão no Brasil em 2019, então um deslocamento fixo
    # já resolve, sem depender de biblioteca de fuso horário).
    ultima_atualizacao = _modificado_utc - timedelta(hours=3)
except Exception:
    # Reserva: se por algum motivo a planilha não tiver essa metadata (ou o
    # arquivo nem existir ainda — esse erro específico só vai aparecer de
    # verdade mais abaixo, onde já existe tratamento amigável pra isso),
    # cai de volta pro comportamento antigo.
    ultima_atualizacao = datetime.fromtimestamp(os.path.getmtime(arquivo_excel))
ultima_atualizacao_txt = ultima_atualizacao.strftime(
    "Fonte: CASA CIVIL / CGAPE - Planilha Panorama - Atualizado em: %d/%m/%Y às %Hhs%Mmin"
)
MESES_PT = {
    1: "janeiro", 2: "fevereiro", 3: "março", 4: "abril",
    5: "maio", 6: "junho", 7: "julho", 8: "agosto",
    9: "setembro", 10: "outubro", 11: "novembro", 12: "dezembro",
}
data_capa_txt = f"{ultima_atualizacao.day} de {MESES_PT[ultima_atualizacao.month]} de {ultima_atualizacao.year}"
timestamp = ultima_atualizacao.strftime("%Y.%m.%d_%Hhs%Mm")

nome_arquivo_pdf = f"RELATORIO GERENCIAL PAC_{timestamp}.pdf"
PASTA_DOWNLOADS_PADRAO = os.path.join(os.path.expanduser("~"), "Downloads")

# Geometria da Página - A3 PAISAGEM
PAGINA = (841.8897637795277 * 16 / 9, 841.8897637795277)  # 16:9 — mesma altura da A3 paisagem anterior, largura recalculada
MARGEM_ESQ = 30
MARGEM_DIR = 30
MARGEM_SUP = 58  
MARGEM_INF = 42

LARGURA_UTIL = PAGINA[0] - (MARGEM_ESQ + MARGEM_DIR)
ALTURA_UTIL = PAGINA[1] - (MARGEM_SUP + MARGEM_INF)

# Padding interno do Frame que o SimpleDocTemplate cria (valor padrão do
# ReportLab, de cada lado). Não é margem de página: é um recuo a mais,
# aplicado dentro dela, e é sobre a largura JÁ descontada dele que as
# tabelas centralizadas se posicionam.
PADDING_FRAME_DOCUMENTO = 6
LARGURA_PADRAO = LARGURA_UTIL
FATOR_REDUCAO_RESUMO = 0.60
LARGURA_RESUMO = LARGURA_PADRAO * FATOR_REDUCAO_RESUMO

col_eixo = "EIXO"
col_tipo = "TIPO"
col_fase = "FASE"
col_objeto = "OBJETO"
col_descricao = "DESCRICAO"
col_municipio = "MUNICIPIO"
col_status = "STATUS"
col_invest = "INVESTIMENTO TOTAL"
col_fonte = "FONTE DE RECURSO"
col_fonte_financiamento = "FONTE DE FINANCIAMENTO"
col_orgao = "SECRETARIA/ ORGAO"
col_executor = "ORGAO EXECUTOR"
col_tc = "Nº DO TERMO DE COMP. (TC) / CT FINANCIAMENTO"
col_prazo = "PRAZO DE CONCLUSAO DA FASE"
col_avanco = "AVANCO DA OBRA (%)"
col_vigencia = "VIGENCIA"
col_prazo_atual = "PREVISAO DE CONCLUSAO ATUAL"
col_clausula_suspensiva = "CLAUSULA SUSPENSIVA"
col_item = "ITEM"
col_financiamento = "FINANCIAMENTO"

# Colunas exclusivas da Ficha Cadastral (janela de detalhe de uma única
# ação) — além das que o relatório já usa. Acesso sempre defensivo (via
# .get() com "" de reserva) nas funções que leem essas colunas: se algum
# nome aqui não bater 100% com o cabeçalho real da planilha, o campo
# aparece em branco na ficha em vez de travar o programa.
col_pendencia = "PENDENCIAS / TAREFA"
col_providencias = "PROVIDENCIAS (DATAS)"
col_prazo_pendencia = "PRAZO DA PENDENCIA / TAREFA"
col_proximos_passos = "PROXIMOS PASSOS"
col_motivo_clausula_suspensiva = "MOTIVO DA CLAUSULA SUSPENSIVA"
col_link_monitora = "LINK MONITORA"
col_link_localizacao = "LINK LOCALIZACAO"
col_aviso_licitacao = "AVISO DE LICITACAO"
col_abertura_licitacao = "ABERTURA DE LICITACAO"
col_emissao_os = "EMISSAO DE O.S."
col_apoiado = "APOIADO"
col_contrapartida = "CONTRAPARTIDA"
col_complementar = "COMPLEMENTAR"
col_valor_contratado = "VALOR CONTRATADO"
col_financiamento_ajustado = "FINANCIAMENTO AJUSTADO"
col_apoiado_ajustado = "APOIADO AJUSTADO"
col_contrapartida_ajustado = "CONTRAPARTIDA AJUSTADO"
col_complementar_ajustado = "COMPLEMENTAR AJUSTADO"

# Só remove registros com STATUS em branco. Todos os demais status
# (INAUGURADA, COMPLEMENTADA E ENVIADA, etc.) aparecem normalmente no
# relatório — nenhum é excluído automaticamente pelo nome.
status_excluir = [""]

# Tenta registrar a fonte Calibri (fonte padrão do relatório). Procura os
# arquivos .ttf na mesma pasta do script/.exe; se não encontrar, usa
# Helvetica/Helvetica-Bold automaticamente (únicas fontes garantidas no
# ReportLab), sem interromper a geração do relatório.
try:
    pdfmetrics.registerFont(TTFont("Calibri", caminho_recurso("calibri.ttf")))
    pdfmetrics.registerFont(TTFont("Calibri-Bold", caminho_recurso("calibrib.ttf")))
    FONTE_PADRAO = "Calibri"
    FONTE_PADRAO_NEGRITO = "Calibri-Bold"
except Exception:
    FONTE_PADRAO = "Helvetica"
    FONTE_PADRAO_NEGRITO = "Helvetica-Bold"

# Tenta registrar a fonte DIN (bastante usada para destacar valores numéricos
# em painéis/dashboards). Procura os arquivos .ttf na mesma pasta do script;
# se não encontrar, usa a fonte padrão do relatório (Calibri/Helvetica)
# automaticamente, sem interromper a geração do relatório.
try:
    pdfmetrics.registerFont(TTFont("DIN", caminho_recurso("DIN.ttf")))
    pdfmetrics.registerFont(TTFont("DIN-Bold", caminho_recurso("DIN-Bold.ttf")))
    FONTE_DIN = "DIN-Bold"
    FONTE_DIN_REGULAR = "DIN"
except Exception:
    FONTE_DIN = FONTE_PADRAO_NEGRITO
    FONTE_DIN_REGULAR = FONTE_PADRAO

# Tenta registrar a fonte Bahnschrift — usada apenas nos valores dos cards de
# resumo (QTD TOTAL/INVESTIMENTO) e nos rótulos de dados dos gráficos do
# Painel Geral. Procura "bahnschrift.ttf" na mesma pasta do script/.exe; se
# não encontrar, cai para a fonte padrão em negrito, sem travar o relatório.
try:
    pdfmetrics.registerFont(TTFont("Bahnschrift", caminho_recurso("bahnschrift.ttf")))
    FONTE_BAHNSCHRIFT = "Bahnschrift"
except Exception:
    FONTE_BAHNSCHRIFT = FONTE_PADRAO_NEGRITO

# =====================================================
# IDENTIDADE VISUAL — mesma paleta de destaque do dashboard (teal + gold).
# Definida logo aqui, antes de qualquer ParagraphStyle, porque vários
# estilos de cabeçalho de tabela (resumos, detalhamento analítico) usam
# essas cores pra ficarem visualmente unificados com os títulos dos
# gráficos do Painel Geral.
# --- Saturação das cores dos gráficos ---
# Um único botão de ajuste para toda a paleta de gráficos (PDF e painel).
# As cores-base continuam escritas no código como sempre foram; o que sai
# nos gráficos é a versão dessaturada delas, o que evita ter que reescrever
# dezenas de códigos hexadecimais toda vez que a paleta precisa ficar mais
# suave — e garante que o PDF e o painel nunca fiquem com tons diferentes,
# já que os dois passam pela mesma função.
#
# 1.0 = cor original; abaixo disso, mais lavada. A luminosidade é
# preservada, então o contraste com o texto por cima não muda.
FATOR_SATURACAO_GRAFICOS = 0.82

def dessaturar_hex(hex_base, fator=None):
    fator = FATOR_SATURACAO_GRAFICOS if fator is None else fator
    limpo = str(hex_base).lstrip("#")
    vermelho, verde, azul = (int(limpo[i:i + 2], 16) / 255.0 for i in (0, 2, 4))
    matiz, luminosidade, saturacao = colorsys.rgb_to_hls(vermelho, verde, azul)
    vermelho, verde, azul = colorsys.hls_to_rgb(
        matiz, luminosidade, max(0.0, min(1.0, saturacao * fator))
    )
    return "#{:02X}{:02X}{:02X}".format(
        round(vermelho * 255), round(verde * 255), round(azul * 255)
    )

def cor_grafico(hex_base):
    # Versão ReportLab da cor já dessaturada.
    return colors.HexColor(dessaturar_hex(hex_base))

COR_MARCA_TEAL = colors.HexColor("#72B4AE")        # teal puro do dashboard
COR_MARCA_TEAL_ESCURO = colors.HexColor("#2E6F68")  # versão escura, p/ texto e fundos
COR_MARCA_TEAL_CLARA = colors.HexColor("#E3EFED")   # tom clarinho, p/ fundos com texto escuro
COR_MARCA_GOLD = colors.HexColor("#E0AB45")         # gold puro do dashboard
COR_MARCA_GOLD_ESCURO = colors.HexColor("#B8863A")  # versão escura, p/ texto

# =====================================================
# 2. CONFIGURAÇÃO DE ESTILOS (REPORTLAB)
# =====================================================

styles = getSampleStyleSheet()


cell_nowrap = ParagraphStyle(
    "nowrap",
    fontName=FONTE_PADRAO,
    fontSize=12,
    leading=14,
    alignment=TA_LEFT,
)
cell_center = ParagraphStyle("center", parent=cell_nowrap, alignment=TA_CENTER)
cell_municipio = ParagraphStyle("municipio", parent=cell_nowrap)
valor_style = ParagraphStyle("valor", parent=cell_nowrap, alignment=TA_RIGHT)
# Cabeçalho padrão das colunas da tabela analítica (OBJETO, FASE, STATUS,
# AVANÇO, MUNICÍPIOS, INVESTIMENTO, FONTE, PRAZO): fonte maior para melhorar
# a legibilidade.
header_center = ParagraphStyle(
    "header_center",
    parent=cell_nowrap,
    alignment=TA_CENTER,
    fontName=FONTE_PADRAO_NEGRITO,
    fontSize=13,
    leading=15,
    textColor=COR_MARCA_TEAL_ESCURO,
)
# Cabeçalho exclusivo da coluna "TERMO DE COMPROMISSO/ FINANCIAMENTO":
# mantido no tamanho original (10), já que o texto é longo e a coluna é
# estreita — aumentar aqui faria o cabeçalho quebrar/estourar a célula.
header_center_termo = ParagraphStyle(
    "header_center_termo",
    parent=cell_nowrap,
    alignment=TA_CENTER,
    fontName=FONTE_PADRAO_NEGRITO,
    fontSize=10,
    leading=12,
    textColor=COR_MARCA_TEAL_ESCURO,
)

titulo_style = ParagraphStyle(
    "titulo", parent=styles["Heading1"], alignment=1, fontName=FONTE_PADRAO_NEGRITO,
    fontSize=22, leading=25,
)
# Nota de orientação de leitura exibida no Painel Geral, explicando que os
# gráficos mostram o investimento acompanhado da quantidade de ações entre
# parênteses.
orientacao_leitura_style = ParagraphStyle(
    "orientacao_leitura",
    fontName=FONTE_PADRAO,
    fontSize=12,
    leading=15,
    alignment=TA_LEFT,
    textColor=colors.HexColor("#555555"),
)
texto_padrao_style = ParagraphStyle(
    "texto_padrao",
    fontName=FONTE_PADRAO,
    fontSize=16,
    leading=20,
    alignment=TA_LEFT,
)
detalhamento_texto_style = ParagraphStyle(
    "detalhamento_texto",
    fontName=FONTE_PADRAO,
    fontSize=17,
    leading=20,
    alignment=TA_LEFT,
)

gestao_style = ParagraphStyle(
    "gestao", parent=styles["Heading1"], textColor=colors.darkred, keepWithNext=True,
    fontName=FONTE_PADRAO_NEGRITO,
    spaceAfter=10,
)
eixo_style = ParagraphStyle(
    "eixo", parent=styles["Heading3"], textColor=colors.black, keepWithNext=True,
    fontName=FONTE_PADRAO_NEGRITO,
    fontSize=18,
    leading=20,
    spaceAfter=5,
)

obs_titulo = ParagraphStyle(
    "obs_titulo", fontName=FONTE_PADRAO_NEGRITO, fontSize=34, leading=38, alignment=TA_LEFT
)
obs_texto = ParagraphStyle(
    "obs_texto", fontName=FONTE_PADRAO, fontSize=20, leading=25, alignment=TA_LEFT
)
# Estilos da página de Metodologia do Índice de Desempenho — mais
# compactos que os de obs_titulo/obs_texto (Observação), já que essa
# página tem bem mais texto (7 critérios explicados, mais tabela de
# pesos) e precisa caber inteira numa página só.
metodologia_titulo = ParagraphStyle(
    "metodologia_titulo", fontName=FONTE_PADRAO_NEGRITO, fontSize=30, leading=34, alignment=TA_LEFT
)
metodologia_texto = ParagraphStyle(
    "metodologia_texto", fontName=FONTE_PADRAO, fontSize=14, leading=18, alignment=TA_LEFT
)
metodologia_tabela_header = ParagraphStyle(
    "metodologia_tabela_header", fontName=FONTE_PADRAO_NEGRITO, fontSize=13, leading=16, alignment=TA_LEFT
)
# Mesmo cabeçalho, centralizado — usado nas colunas Conta e Resultado da
# tabela do exemplo, que têm o conteúdo centralizado.
metodologia_tabela_header_centro = ParagraphStyle(
    "metodologia_tabela_header_centro", parent=metodologia_tabela_header, alignment=TA_CENTER
)
# Faixa de nota escrita embaixo de cada medidor de exemplo, na página de
# Metodologia ("0 a 24", "25 a 49"...). Centralizada e menor que o corpo do
# texto: é rótulo de legenda, não leitura corrida.
metodologia_exemplo_faixa = ParagraphStyle(
    "metodologia_exemplo_faixa", fontName=FONTE_PADRAO, fontSize=11, leading=14,
    alignment=TA_CENTER, textColor=colors.HexColor("#555555"),
)
# Tabela do exemplo passo a passo: descrição da etapa à esquerda, a conta
# no meio e o resultado parcial à direita.
metodologia_exemplo_desc = ParagraphStyle(
    "metodologia_exemplo_desc", fontName=FONTE_PADRAO, fontSize=12, leading=15, alignment=TA_LEFT
)
metodologia_exemplo_conta = ParagraphStyle(
    "metodologia_exemplo_conta", fontName=FONTE_PADRAO, fontSize=12, leading=15,
    alignment=TA_CENTER, textColor=colors.HexColor("#555555"),
)
metodologia_exemplo_valor = ParagraphStyle(
    "metodologia_exemplo_valor", fontName=FONTE_PADRAO_NEGRITO, fontSize=12, leading=15,
    alignment=TA_CENTER
)

# As duas tabelas de apoio (composição e escala de Status) usam um corpo
# menor que a tabela principal do exemplo: são detalhamento de uma linha
# dela, e em 12 pt as duas juntas não caberiam na página.
metodologia_apoio_desc = ParagraphStyle(
    "metodologia_apoio_desc", parent=metodologia_exemplo_desc, fontSize=10.5, leading=13
)
metodologia_apoio_conta = ParagraphStyle(
    "metodologia_apoio_conta", parent=metodologia_exemplo_conta, fontSize=10.5, leading=13
)
metodologia_apoio_valor = ParagraphStyle(
    "metodologia_apoio_valor", parent=metodologia_exemplo_valor, fontSize=10.5, leading=13
)
metodologia_apoio_header = ParagraphStyle(
    "metodologia_apoio_header", parent=metodologia_tabela_header, fontSize=11.5, leading=14
)
metodologia_apoio_header_centro = ParagraphStyle(
    "metodologia_apoio_header_centro", parent=metodologia_apoio_header, alignment=TA_CENTER
)

def _num_metodologia(valor):
    # Número com uma casa decimal e vírgula, como o resto do relatório.
    return f"{valor:.1f}".replace(".", ",")


capa_projeto = ParagraphStyle(
    "capa_projeto",
    fontName=FONTE_PADRAO_NEGRITO,
    fontSize=110,
    leading=110,
    alignment=TA_LEFT,
    textColor=COR_MARCA_TEAL_ESCURO,
)
capa_relatorio = ParagraphStyle(
    "capa_relatorio",
    fontName=FONTE_PADRAO,
    fontSize=55,
    leading=60,
    alignment=TA_LEFT,
    textColor=colors.HexColor("#444444"),
)
capa_data = ParagraphStyle(
    "capa_data",
    fontName=FONTE_PADRAO_NEGRITO,
    fontSize=32,
    leading=25,
    alignment=TA_RIGHT,
    textColor=COR_MARCA_GOLD_ESCURO,
)
capa_cgape = ParagraphStyle(
    "capa_cgape",
    fontName=FONTE_PADRAO_NEGRITO,
    fontSize=25,
    leading=29,
    alignment=TA_RIGHT,
    textColor=COR_MARCA_TEAL_ESCURO,
)
# Versão alinhada à esquerda do texto do CGAPE — usada só na capa, desenhada
# direto no canvas (não como flowable) pra poder posicionar na mesma altura
# do emblema do estado, que também é desenhado por coordenada fixa ali.
capa_cgape_esquerda = ParagraphStyle(
    "capa_cgape_esquerda", parent=capa_cgape, alignment=TA_LEFT,
)
# Linha "SECRETARIA | EXECUTOR | FASE | STATUS" que aparece na capa só
# quando o relatório está filtrado a um único valor de cada categoria —
# 4pt maior que a linha do CGAPE (25+4=29).
capa_filtro_unico = ParagraphStyle(
    "capa_filtro_unico",
    fontName=FONTE_PADRAO_NEGRITO,
    fontSize=29,
    leading=33,
    alignment=TA_RIGHT,
    textColor=COR_MARCA_TEAL_ESCURO,
)

# =====================================================
# 2B. CONFIGURAÇÃO DAS TABELAS DE RESUMO (FONTE / COLUNAS / LINHAS)
# =====================================================
# Concentra os parâmetros visuais das tabelas de resumo — EIXO/STATUS/FASE
# por QTD e INVESTIMENTO, a tabela por Secretaria/Fase e os cards de QTD
# TOTAL/INVESTIMENTO — para ajustar fonte, largura de coluna e altura de
# linha num único lugar, sem precisar procurar em cada função.

# --- Fonte ---
RESUMO_FONTE_NORMAL = FONTE_PADRAO
RESUMO_FONTE_NEGRITO = FONTE_PADRAO_NEGRITO
RESUMO_FONTE_TAM_CABECALHO = 14  # cabeçalhos das colunas e linha TOTAL
RESUMO_FONTE_TAM_CORPO = 14      # demais células

# --- Larguras de coluna: tabela por Secretaria/Fase ---
# Colunas fixas, sobre a largura útil total da página (LARGURA_UTIL). O
# espaço restante (1 - soma abaixo) é dividido igualmente entre as colunas
# de fase ativas (CAPTAÇÃO DE RECURSO, LICITAÇÃO, EXECUÇÃO DO OBJETO,
# CONCLUÍDA), para a tabela sempre ocupar 100% da largura da página.
SEC_FASE_LARGURA_SECRETARIA = 0.14
SEC_FASE_LARGURA_EXECUTOR = 0.14
SEC_FASE_LARGURA_OBJETO = 0.32

# Fonte e altura de linha exclusivas da tabela por Secretaria/Fase (maiores
# que as demais tabelas de resumo, que continuam usando RESUMO_FONTE_TAM_*).
# Este é o tamanho PREFERIDO, não o final: quando a base tem muitas
# secretarias, o autofit abaixo encolhe o que for preciso para a tabela
# caber em uma página. Subir aqui só melhora as bases pequenas e médias —
# nas grandes o autofit continua mandando.
SEC_FASE_FONTE_TAM = 16
SEC_FASE_PADDING_VERTICAL = 8
# Pisos mínimos usados pelo autofit: se a tabela não couber em uma única
# página com o tamanho preferido acima, ela reduz primeiro o preenchimento
# vertical e, se ainda não bastar, a fonte — nunca abaixo destes valores.
SEC_FASE_FONTE_MINIMA = 9
SEC_FASE_PADDING_MINIMA = 2

# --- Cards QTD TOTAL / INVESTIMENTO ---
CARDS_FATOR_LARGURA = 1 / 7  # antes era 1/5 — cards mais estreitos
CARDS_ALTURA_CABECALHO = 26
CARDS_ALTURA_VALOR = 42
CARDS_FONTE_TAM_CABECALHO = 11
CARDS_FONTE_TAM_VALOR = 26

# --- Estilos derivados, usados pelas tabelas de resumo acima ---
resumo_header_style = ParagraphStyle(
    "resumo_header",
    fontName=RESUMO_FONTE_NEGRITO,
    fontSize=RESUMO_FONTE_TAM_CABECALHO,
    leading=RESUMO_FONTE_TAM_CABECALHO + 2,
    alignment=TA_CENTER,
    textColor=COR_MARCA_TEAL_ESCURO,
)
resumo_cell_style = ParagraphStyle(
    "resumo_cell",
    fontName=RESUMO_FONTE_NORMAL,
    fontSize=RESUMO_FONTE_TAM_CORPO,
    leading=RESUMO_FONTE_TAM_CORPO + 2,
    alignment=TA_LEFT,
)
resumo_cell_center_style = ParagraphStyle(
    "resumo_cell_center", parent=resumo_cell_style, alignment=TA_CENTER
)
resumo_valor_style = ParagraphStyle(
    "resumo_valor", parent=resumo_cell_style, alignment=TA_RIGHT
)

# Resumos por Eixo e por Status: corpo maior que o padrão. Com os valores
# abreviados ("R$ 5,70 Bi") a linha ficou curta, e essas duas tabelas são a
# leitura principal da página — sobrava espaço para crescer.
RESUMO_BARRAS_FONTE_TAM = 17
resumo_barras_header_style = ParagraphStyle(
    "resumo_barras_header", parent=resumo_header_style,
    fontSize=RESUMO_BARRAS_FONTE_TAM, leading=RESUMO_BARRAS_FONTE_TAM + 3,
)
resumo_barras_cell_style = ParagraphStyle(
    "resumo_barras_cell", parent=resumo_cell_style,
    fontSize=RESUMO_BARRAS_FONTE_TAM, leading=RESUMO_BARRAS_FONTE_TAM + 3,
)
resumo_barras_center_style = ParagraphStyle(
    "resumo_barras_center", parent=resumo_barras_cell_style, alignment=TA_CENTER
)
resumo_barras_valor_style = ParagraphStyle(
    "resumo_barras_valor", parent=resumo_barras_cell_style, alignment=TA_RIGHT
)

# Rótulo (secretaria + qtd + investimento) do gráfico de barras empilhadas
# de Detalhamento Financeiro por Secretaria — fonte maior que o padrão das
# demais tabelas de resumo, só pra essa seção.
detalhamento_secretaria_rotulo_style = ParagraphStyle(
    "detalhamento_secretaria_rotulo",
    fontName=FONTE_PADRAO,
    fontSize=15,
    leading=18,
    alignment=TA_LEFT,
)

# Estilos exclusivos da tabela por Secretaria/Fase (fonte maior que as
# demais tabelas de resumo, que continuam com resumo_header_style/
# resumo_cell_style/resumo_cell_center_style sem alteração).
sec_fase_header_style = ParagraphStyle(
    "sec_fase_header",
    fontName=RESUMO_FONTE_NEGRITO,
    fontSize=SEC_FASE_FONTE_TAM,
    leading=SEC_FASE_FONTE_TAM + 3,
    alignment=TA_CENTER,
)
sec_fase_cell_style = ParagraphStyle(
    "sec_fase_cell",
    fontName=RESUMO_FONTE_NORMAL,
    fontSize=SEC_FASE_FONTE_TAM,
    leading=SEC_FASE_FONTE_TAM + 3,
    alignment=TA_LEFT,
)
sec_fase_cell_center_style = ParagraphStyle(
    "sec_fase_cell_center", parent=sec_fase_cell_style, alignment=TA_CENTER
)

# =====================================================
# 3. MÉTODOS DE TRATAMENTO E FORMATAÇÃO DE DADOS
# =====================================================

def limpar_executor(txt):
    if pd.isna(txt):
        return ""
    texto = str(txt)
    texto = re.sub(r"^\s*\d+\s*-\s*GOV.*?-\s*", "", texto, flags=re.IGNORECASE)
    return texto.strip()

def remover_acentos(txt):
    if pd.isna(txt):
        return ""
    return "".join(
        c
        for c in unicodedata.normalize("NFD", str(txt))
        if unicodedata.category(c) != "Mn"
    )

def normalizar_item(item_bruto):
    # Mesma normalização já usada no controle de qualidade: o ITEM pode vir
    # como número (int/float do Excel) ou texto — sempre devolve texto
    # comparável, sem ".0" sobrando em números inteiros.
    if isinstance(item_bruto, (int, float, np.integer, np.floating)) and not pd.isna(item_bruto):
        return str(int(item_bruto)) if float(item_bruto).is_integer() else str(item_bruto)
    return str(item_bruto).strip() if item_bruto is not None else ""

def converter_valor(valor):
    if pd.isna(valor):
        return 0.0
    if isinstance(valor, (int, float)):
        return float(valor)
    valor = str(valor).replace("R$", "").replace(" ", "")
    if "," in valor:
        valor = valor.replace(".", "").replace(",", ".")
    try:
        return float(valor)
    except:
        return 0.0

def moeda_sem_quebra(valor):
    txt = f"R$ {valor:,.2f}"
    return (
        txt.replace(",", "X")
        .replace(".", ",")
        .replace("X", ".")
        .replace(" ", "&nbsp;")
    )

def moeda_texto_puro(valor):
    # Mesma formatação de moeda_sem_quebra, mas com espaço normal em vez
    # de "&nbsp;" — usada em contextos que NÃO interpretam entidades HTML
    # (como a Ficha Cadastral, montada em JS puro), onde o "&nbsp;" apareceria
    # como texto literal em vez de virar espaço.
    txt = f"R$ {valor:,.2f}"
    return txt.replace(",", "X").replace(".", ",").replace("X", ".")

def formatar_mi_bi(valor):
    if valor >= 1_000_000_000:
        res = f"R$&nbsp;{valor/1_000_000_000:,.1f}&nbsp;Bi"
    elif valor >= 1_000_000:
        res = f"R$&nbsp;{valor/1_000_000:,.1f}&nbsp;Mi"
    else:
        # moeda_sem_quebra já devolve o valor no formato brasileiro (ponto
        # milhar, vírgula decimal) — não pode passar pela troca de
        # separador de novo abaixo, senão volta pro formato americano
        # (era exatamente isso que estava acontecendo com valores abaixo
        # de R$ 1 milhão, como o menor valor de uma faixa do mapa).
        return moeda_sem_quebra(valor)
    return res.replace(",", "X").replace(".", ",").replace("X", ".")

def formatar_contagem_opcional(contagem):
    return f" ({contagem})" if contagem >= 2 else ""

def formatar_municipios_limpo(x):
    contagem = x.dropna().astype(str).value_counts()
    itens_ordenados = sorted(contagem.items(), key=lambda item: remover_acentos(item[0]))
    items = []
    for municipio, quantidade in itens_ordenados:
        sufixo_qtd = formatar_contagem_opcional(quantidade)
        items.append(f"{municipio}{sufixo_qtd}")
    return ", ".join(items)

def classificar_termo_compromisso(valor):
    if pd.isna(valor):
        return "NÃO"
    txt = str(valor).strip()
    if txt == "" or txt.upper() == "NAN" or txt == "-" or txt == "0":
        return "NÃO"
    return "SIM"

def extrair_ordem_clausula_suspensiva(valor):
    m = re.match(r"^\s*(\d+)", str(valor))
    return int(m.group(1)) if m else 999

def tratar_clausula_suspensiva(valor):
    # Usa o texto que já vem preenchido na própria coluna CLÁUSULA SUSPENSIVA
    # da planilha (ex: "01 - Retirada Total"), removendo o número do início
    # — igual ao tratamento já usado em STATUS e FASE — pois esse número
    # serve só para ordenação, não para exibição no filtro/relatório.
    if pd.isna(valor):
        return "NÃO DEFINIDA"
    txt = str(valor).strip()
    if txt == "" or txt.upper() == "NAN":
        return "NÃO DEFINIDA"
    texto_limpo = re.sub(r"^\s*\d+\s*-\s*", "", txt).strip()
    return texto_limpo if texto_limpo else txt

# =====================================================
# ÍNDICE DE DESEMPENHO POR SECRETARIA (gráfico de medidor)
# =====================================================
# Metodologia (definida em conversa com o CGAPE):
# - Cada AÇÃO recebe uma nota de 0 a 100 combinando 4 critérios, com pesos
#   diferentes: STATUS e FASE pesam mais (30% cada) por serem o retrato
#   mais direto do andamento; CLÁUSULA SUSPENSIVA e o TEMPO do ciclo
#   aviso-de-licitação → O.S. → previsão de conclusão pesam menos (10%
#   cada); a QUANTIDADE de ações da secretaria (10%) — quanto mais ações
#   pra administrar, maior a dificuldade de atingir as metas, então mais
#   ações rende uma nota maior nesse critério (é um "grau de dificuldade",
#   não um "quanto pior mais ações, pior a nota"); e a PROPORÇÃO DE VALOR
#   CONTRATADO (10%) — o quanto do VALOR CONTRATADO da secretaria/executor
#   representa do investimento total DA PRÓPRIA GESTÃO (Estadual ou
#   Federal, nunca as duas somadas) — quem lida com valores menores dentro
#   da sua gestão tem menos peso nesse critério; quanto maior essa
#   proporção, melhor a nota.
# - O critério FINANCIAMENTO/OGU (proporção entre recurso reembolsável e
#   não reembolsável) foi REMOVIDO do cálculo a pedido do CGAPE. Os 10%
#   que ele ocupava foram redistribuídos entre STATUS e FASE (5% pra
#   cada), que passaram de 25% para 30% — assim a soma dos pesos continua
#   valendo exatamente 100 e a escala do medidor (0 a 100) não muda.
# - O índice de cada SECRETARIA é a média das notas das suas ações nos
#   critérios por ação, PONDERADA pelo investimento de cada uma (ações
#   maiores pesam mais), somada às notas de quantidade e de proporção de
#   valor contratado (essas duas são calculadas uma vez por secretaria,
#   não por ação).
# - A unidade avaliada MUDA conforme a gestão (ver
#   CHAVES_AGRUPAMENTO_INDICE_POR_GESTAO): na GESTÃO ESTADUAL a avaliação
#   é só por SECRETARIA (o executor não separa os medidores); na GESTÃO
#   FEDERAL continua por SECRETARIA | EXECUTOR.
# - O número final (0-100) é convertido numa categoria pra exibição no
#   medidor: Insatisfatório / Regular / Bom / Ótimo.
PESOS_INDICE_DESEMPENHO = {
    "status": 0.30,
    "fase": 0.30,
    "clausula": 0.10,
    "tempo": 0.10,
    "quantidade": 0.10,
    "proporcao_contratado": 0.10,
}

# Como cada gestão agrupa as ações pra virar um medidor: a Estadual só
# por secretaria, a Federal pela combinação secretaria + executor. Se
# alguma gestão não estiver listada aqui, cai no comportamento antigo
# (secretaria + executor).
CHAVES_AGRUPAMENTO_INDICE_POR_GESTAO = {
    "GESTÃO ESTADUAL": ["SECRETARIA_LIMPA"],
    "GESTÃO FEDERAL": ["SECRETARIA_LIMPA", "EXECUTOR"],
}

# Do pior (PARALISADA) pro melhor (INAUGURADA) — ordem informada pelo
# CGAPE. Status fora dessa lista recebem nota neutra (50).
ORDEM_STATUS_DESEMPENHO = [
    "PARALISADA",
    "NÃO HABILITADA",
    "ESTUDO",
    "CADASTRADA",
    "ELABORAÇÃO DE PROJETO",
    "ENVIADA PARA ANÁLISE",
    "EM COMPLEMENTAÇÃO",
    "COMPLEMENTADA E ENVIADA",
    "AGUARDANDO PUBLICAÇÃO",
    "HABILITADA",
    "AGUARDANDO AUTORIZO",
    "SELECIONADA",
    "À LICITAR",
    "LICITAÇÃO /CONTRATAÇÃO",
    "AGUARDANDO ORDEM DE SERVIÇO",
    "ANDAMENTO",
    "CONCLUÍDA",
    "INAUGURADA",
]
# Os 16 status "em andamento" (tudo antes de Concluída) ficam numa faixa
# mais apertada (0 a 65), enquanto CONCLUÍDA e INAUGURADA dão um salto bem
# maior (85 e 100) — pra pesar mais do que só "mais um passo à frente",
# recompensando de verdade a entrega efetiva da ação.
_STATUS_EM_ANDAMENTO_DESEMPENHO = ORDEM_STATUS_DESEMPENHO[:-2]
_MAPA_SCORE_STATUS_DESEMPENHO = {
    nome: (indice / (len(_STATUS_EM_ANDAMENTO_DESEMPENHO) - 1)) * 65
    for indice, nome in enumerate(_STATUS_EM_ANDAMENTO_DESEMPENHO)
}
_MAPA_SCORE_STATUS_DESEMPENHO["CONCLUÍDA"] = 85.0
_MAPA_SCORE_STATUS_DESEMPENHO["INAUGURADA"] = 100.0

def _score_status_desempenho(status_texto):
    return _MAPA_SCORE_STATUS_DESEMPENHO.get(str(status_texto).strip().upper(), 50.0)

# Além de já pontuarem mais alto (ver _MAPA_SCORE_STATUS_DESEMPENHO acima),
# o INVESTIMENTO de ações Concluídas e Inauguradas pesa mais na média
# ponderada da secretaria/executor — Inaugurada pesa ainda mais que
# Concluída, já que é o estágio final da entrega.
_MULTIPLICADOR_PESO_INVESTIMENTO_STATUS = {
    "CONCLUÍDA": 1.5,
    "INAUGURADA": 3.0,
}

def _multiplicador_peso_investimento_status(status_texto):
    return _MULTIPLICADOR_PESO_INVESTIMENTO_STATUS.get(str(status_texto).strip().upper(), 1.0)

def _score_fase_desempenho(fase_texto):
    # Usa o texto já limpo da fase (não o número extraído do texto bruto,
    # que usa 999 como valor "não reconhecido" e explodiria o cálculo) —
    # ORDEM_FASES vai do pior (Captação de Recurso) pro melhor (Concluída).
    try:
        posicao = ORDEM_FASES.index(str(fase_texto).strip().upper())
        return (posicao / (len(ORDEM_FASES) - 1)) * 100
    except (ValueError, AttributeError):
        return 50.0

def _score_clausula_desempenho(situacao_clausula_texto):
    # Usa o texto já tratado (SITUACAO_CLAUSULA_SUSPENSIVA), que já
    # distingue "NÃO DEFINIDA" (sem cláusula nenhuma) de uma cláusula
    # ainda vigente — o número bruto (CLAUSULA_SUSPENSIVA_ORDEM) usa o
    # mesmo valor sentinela (999) pros dois casos, o que juntaria errado
    # "sem cláusula" (deveria ser ótimo) com "cláusula vigente" (deveria
    # ser péssimo).
    texto = str(situacao_clausula_texto).strip().upper()
    if texto in ("NÃO DEFINIDA", "", "NAN"):
        return 100.0
    if "RETIRADA TOTAL" in texto:
        return 100.0
    if "RETIRADA PARCIAL" in texto:
        return 50.0
    return 0.0

def _dias_ciclo_desempenho(aviso, emissao_os, prazo_atual):
    # Soma os dois trechos do ciclo (aviso→O.S. e O.S.→previsão de
    # conclusão atual) quando as datas envolvidas existem — usado depois
    # como base pra um ranking relativo entre as ações (não como valor
    # absoluto), já que nem toda ação tem as 3 datas preenchidas.
    dias = 0
    tem_dado = False
    try:
        if pd.notna(aviso) and pd.notna(emissao_os):
            d = (pd.to_datetime(emissao_os) - pd.to_datetime(aviso)).days
            if d >= 0:
                dias += d
                tem_dado = True
    except Exception:
        pass
    try:
        if pd.notna(emissao_os) and pd.notna(prazo_atual):
            d = (pd.to_datetime(prazo_atual) - pd.to_datetime(emissao_os)).days
            if d >= 0:
                dias += d
                tem_dado = True
    except Exception:
        pass
    return dias if tem_dado else None

def _categoria_indice_desempenho(indice):
    if indice < 25:
        return "Insatisfatório"
    if indice < 50:
        return "Regular"
    if indice < 75:
        return "Bom"
    return "Ótimo"

# Cor por categoria — mesmo espírito das 4 faixas do mapa coroplético
# (vermelho/laranja/verde/azul), mas aqui numa progressão só de
# ruim→bom, do vermelho ao verde escuro.
CORES_CATEGORIA_DESEMPENHO = {
    "Insatisfatório": cor_grafico("#BB6060"),
    "Regular": cor_grafico("#D9A441"),
    "Bom": cor_grafico("#9FCE9B"),
    "Ótimo": cor_grafico("#3F8F52"),
}

def _score_quantidade_desempenho(qtd_secretaria, qtd_minima, qtd_maxima):
    # Quanto mais ações a secretaria tem pra administrar, maior a
    # dificuldade de atingir as metas — por isso mais ações rende nota
    # MAIOR nesse critério (não é "quantidade ruim", é "grau de
    # dificuldade"). Escala relativa: a secretaria com menos ações no
    # recorte atual fica em 0, a com mais ações fica em 100.
    if qtd_maxima <= qtd_minima:
        return 50.0
    return ((qtd_secretaria - qtd_minima) / (qtd_maxima - qtd_minima)) * 100.0

def _score_proporcao_contratado_desempenho(proporcao_secretaria, proporcao_minima, proporcao_maxima):
    # Proporção = VALOR CONTRATADO da secretaria/executor dividido pelo
    # investimento total da GESTÃO (Estadual ou Federal, nunca as duas
    # juntas). Quanto maior essa proporção, melhor a nota — escala
    # relativa dentro da própria gestão: quem tem a menor proporção fica
    # em 0, quem tem a maior fica em 100.
    if proporcao_maxima <= proporcao_minima:
        return 50.0
    return ((proporcao_secretaria - proporcao_minima) / (proporcao_maxima - proporcao_minima)) * 100.0

def calcular_indice_desempenho_secretarias(df_base, combos_ativos=None):
    # Recebe o DataFrame já filtrado (precisa ter FASE_TEXTO, STATUS_TEXTO,
    # SITUACAO_CLAUSULA_SUSPENSIVA, SECRETARIA_LIMPA, EXECUTOR, GESTAO e a
    # coluna _INVESTIMENTO_AJUSTADO já calculadas) e devolve uma lista de
    # grupos — um por GESTÃO (Estadual, depois Federal) — cada um com os
    # itens já ORDENADOS da melhor pra pior. O que é um "item" depende da
    # gestão: na ESTADUAL é uma SECRETARIA; na FEDERAL, uma combinação
    # SECRETARIA | EXECUTOR. A avaliação nunca mistura as duas gestões: tanto o
    # ranking quanto o "grau de dificuldade" da quantidade de ações (ver
    # _score_quantidade_desempenho) são calculados SEPARADAMENTE dentro de
    # cada gestão — comparar um ministério federal com uma secretaria
    # estadual lado a lado não faria sentido.
    #
    # combos_ativos: opcional — um set de tuplas (secretaria, executor)
    # que aparecem no recorte ATUALMENTE FILTRADO do relatório (diferente
    # de df_base, que deve ser sempre a base COMPLETA, sem filtro). Serve
    # só pra marcar visualmente ("esmaecido") os itens que os filtros
    # aplicados nessa geração específica deixariam de fora — o valor do
    # índice em si nunca muda por causa disso, só o sinalizador.
    if df_base.empty:
        return []

    df_calc = df_base.copy()
    df_calc["_SCORE_STATUS"] = df_calc["STATUS_TEXTO"].apply(_score_status_desempenho)
    df_calc["_SCORE_FASE"] = df_calc["FASE_TEXTO"].apply(_score_fase_desempenho)
    df_calc["_SCORE_CLAUSULA"] = df_calc["SITUACAO_CLAUSULA_SUSPENSIVA"].apply(_score_clausula_desempenho)

    df_calc["_DIAS_CICLO"] = df_calc.apply(
        lambda r: _dias_ciclo_desempenho(
            r.get(col_aviso_licitacao), r.get(col_emissao_os), r.get(col_prazo_atual)
        ),
        axis=1,
    )
    df_calc["_SCORE_TEMPO"] = 50.0
    mascara_com_dias = df_calc["_DIAS_CICLO"].notna()
    if mascara_com_dias.sum() > 0:
        percentil_ascendente = df_calc.loc[mascara_com_dias, "_DIAS_CICLO"].rank(pct=True, ascending=True)
        df_calc.loc[mascara_com_dias, "_SCORE_TEMPO"] = 100.0 - (percentil_ascendente * 100.0)

    pesos = PESOS_INDICE_DESEMPENHO
    df_calc["_INDICE_ACAO"] = (
        df_calc["_SCORE_STATUS"] * pesos["status"]
        + df_calc["_SCORE_FASE"] * pesos["fase"]
        + df_calc["_SCORE_CLAUSULA"] * pesos["clausula"]
        + df_calc["_SCORE_TEMPO"] * pesos["tempo"]
    )
    # Peso de investimento usado na média ponderada — o investimento de
    # ações Concluídas/Inauguradas conta mais (ver
    # _multiplicador_peso_investimento_status), então esse peso EFETIVO é
    # diferente do investimento REAL (que continua sendo o valor mostrado
    # nos cards de resumo, sem multiplicador nenhum).
    df_calc["_PESO_INVESTIMENTO_EFETIVO"] = df_calc["_INVESTIMENTO_AJUSTADO"] * df_calc[
        "STATUS_TEXTO"
    ].apply(_multiplicador_peso_investimento_status)

    # Secretarias presentes no recorte filtrado — usado pelo "esmaecido"
    # da GESTÃO ESTADUAL, onde a unidade avaliada é só a secretaria (não a
    # combinação com o executor).
    secretarias_ativas = (
        {str(sec) for sec, _ in combos_ativos} if combos_ativos is not None else None
    )

    grupos_resultado = []
    for gestao in ["GESTÃO ESTADUAL", "GESTÃO FEDERAL"]:
        df_gestao = df_calc[df_calc["GESTAO"] == gestao]
        if df_gestao.empty:
            continue

        # Estadual agrupa só por secretaria; Federal por secretaria +
        # executor (ver CHAVES_AGRUPAMENTO_INDICE_POR_GESTAO). Todas as
        # estatísticas relativas da gestão (mínimo/máximo de quantidade e
        # de proporção de valor contratado) usam a MESMA chave, pra que a
        # comparação seja sempre entre unidades do mesmo tipo.
        chaves_agrupamento = CHAVES_AGRUPAMENTO_INDICE_POR_GESTAO.get(
            gestao, ["SECRETARIA_LIMPA", "EXECUTOR"]
        )

        contagens = df_gestao.groupby(chaves_agrupamento).size()
        qtd_minima = float(contagens.min()) if len(contagens) else 0
        qtd_maxima = float(contagens.max()) if len(contagens) else 0

        # Proporção VALOR CONTRATADO / investimento total DA PRÓPRIA GESTÃO.
        # Antes o denominador era o PAC inteiro, as duas gestões somadas, o
        # que fazia a gestão menor ter proporções sempre pequenas: todas as
        # unidades dela ficavam comprimidas na base da escala por causa do
        # tamanho da outra gestão, e não do que elas de fato contrataram.
        # Com o total da própria gestão como denominador, a proporção passa
        # a medir o peso da unidade dentro do universo em que ela realmente
        # é comparada — o mesmo universo já usado na quantidade de ações e
        # na normalização min/max deste critério.
        investimento_total_gestao = float(df_gestao["_INVESTIMENTO_AJUSTADO"].sum())
        soma_contratado_por_grupo = df_gestao.groupby(chaves_agrupamento)[col_valor_contratado].sum()
        if investimento_total_gestao > 0:
            proporcoes_contratado = soma_contratado_por_grupo / investimento_total_gestao
        else:
            proporcoes_contratado = soma_contratado_por_grupo * 0.0
        proporcao_minima = float(proporcoes_contratado.min()) if len(proporcoes_contratado) else 0.0
        proporcao_maxima = float(proporcoes_contratado.max()) if len(proporcoes_contratado) else 0.0

        itens = []
        for chave_grupo, grupo in df_gestao.groupby(chaves_agrupamento):
            # Com uma única chave de agrupamento o pandas pode devolver o
            # valor "cru" em vez de uma tupla — normaliza os dois casos.
            if not isinstance(chave_grupo, tuple):
                chave_grupo = (chave_grupo,)
            secretaria = chave_grupo[0]
            executor = chave_grupo[1] if len(chave_grupo) > 1 else ""
            if not str(secretaria).strip():
                continue
            investimento_real = float(grupo["_INVESTIMENTO_AJUSTADO"].sum())
            peso_efetivo_total = float(grupo["_PESO_INVESTIMENTO_EFETIVO"].sum())
            if peso_efetivo_total > 0:
                indice_por_acao = float(
                    (grupo["_INDICE_ACAO"] * grupo["_PESO_INVESTIMENTO_EFETIVO"]).sum() / peso_efetivo_total
                )
            else:
                indice_por_acao = float(grupo["_INDICE_ACAO"].mean())
            score_quantidade = _score_quantidade_desempenho(len(grupo), qtd_minima, qtd_maxima)
            valor_contratado_grupo = float(grupo[col_valor_contratado].sum())
            proporcao_grupo = (
                valor_contratado_grupo / investimento_total_gestao if investimento_total_gestao > 0 else 0.0
            )
            score_proporcao_contratado = _score_proporcao_contratado_desempenho(
                proporcao_grupo, proporcao_minima, proporcao_maxima
            )
            indice = (
                indice_por_acao
                + (score_quantidade * pesos["quantidade"])
                + (score_proporcao_contratado * pesos["proporcao_contratado"])
            )
            indice = max(0.0, min(100.0, indice))
            # O rótulo (e a checagem do "esmaecido") acompanha o nível de
            # agrupamento da gestão: só a secretaria na Estadual, a
            # combinação com o executor na Federal.
            if str(executor).strip():
                rotulo = f"{secretaria} | {executor}"
                esmaecido = (
                    combos_ativos is not None
                    and (str(secretaria), str(executor)) not in combos_ativos
                )
            else:
                rotulo = str(secretaria)
                esmaecido = (
                    secretarias_ativas is not None and str(secretaria) not in secretarias_ativas
                )
            itens.append(
                {
                    "secretaria": str(secretaria),
                    "executor": str(executor),
                    "rotulo": rotulo,
                    "indice": round(indice, 1),
                    "categoria": _categoria_indice_desempenho(indice),
                    "qtd": int(len(grupo)),
                    "investimento": investimento_real,
                    "esmaecido": esmaecido,
                    # Peças intermediárias do cálculo, guardadas para o
                    # exemplo passo a passo da página de Metodologia. Não
                    # são usadas em nenhum outro lugar do relatório: o
                    # medidor mostra só o "indice" final.
                    "detalhe_calculo": {
                        "media_status": float((grupo["_SCORE_STATUS"] * grupo["_PESO_INVESTIMENTO_EFETIVO"]).sum() / peso_efetivo_total) if peso_efetivo_total > 0 else float(grupo["_SCORE_STATUS"].mean()),
                        # Abertura da média de Status por status: quantas
                        # ações, que nota cada status vale, e quanto de peso
                        # (investimento efetivo) elas carregam. É com isso
                        # que a página de Metodologia mostra de onde saiu a
                        # média, em vez de apresentá-la como um número dado.
                        "composicao_status": [
                            {
                                "status": str(status_nome),
                                "nota": float(sub["_SCORE_STATUS"].iloc[0]),
                                "qtd": int(len(sub)),
                                "peso": float(sub["_PESO_INVESTIMENTO_EFETIVO"].sum()),
                            }
                            for status_nome, sub in grupo.groupby("STATUS_TEXTO")
                            if len(sub) > 0
                        ],
                        "peso_efetivo_total": peso_efetivo_total,
                        "media_fase": float((grupo["_SCORE_FASE"] * grupo["_PESO_INVESTIMENTO_EFETIVO"]).sum() / peso_efetivo_total) if peso_efetivo_total > 0 else float(grupo["_SCORE_FASE"].mean()),
                        "media_clausula": float((grupo["_SCORE_CLAUSULA"] * grupo["_PESO_INVESTIMENTO_EFETIVO"]).sum() / peso_efetivo_total) if peso_efetivo_total > 0 else float(grupo["_SCORE_CLAUSULA"].mean()),
                        "media_tempo": float((grupo["_SCORE_TEMPO"] * grupo["_PESO_INVESTIMENTO_EFETIVO"]).sum() / peso_efetivo_total) if peso_efetivo_total > 0 else float(grupo["_SCORE_TEMPO"].mean()),
                        "indice_por_acao": indice_por_acao,
                        "score_quantidade": score_quantidade,
                        "score_proporcao": score_proporcao_contratado,
                        "qtd_minima": qtd_minima,
                        "qtd_maxima": qtd_maxima,
                        "valor_contratado": valor_contratado_grupo,
                        "investimento_gestao": investimento_total_gestao,
                        "proporcao": proporcao_grupo,
                        "proporcao_minima": proporcao_minima,
                        "proporcao_maxima": proporcao_maxima,
                    },
                }
            )
        itens.sort(key=lambda x: x["indice"], reverse=True)
        grupos_resultado.append({"gestao": gestao, "itens": itens})

    return grupos_resultado

# Dimensões do medidor compacto que acompanha o nome da secretaria no
# cabeçalho do Detalhamento Analítico. A proporção segue a dos cards da
# página do índice (altura ≈ 0,85 × largura), que é o que mantém o
# semicírculo inteiro dentro do desenho.
LARGURA_MEDIDOR_CABECALHO_SECRETARIA = 58
ALTURA_MEDIDOR_CABECALHO_SECRETARIA = 50

# A categoria ("Regular") ao lado do medidor no cabeçalho da secretaria — a
# nota fica dentro do anel, então aqui sobra só a palavra, sem parênteses.
# Menor que o nome da secretaria de propósito: é informação de apoio, não
# deve competir com o título da seção. A cor é definida na hora, pela
# categoria.
nota_indice_cabecalho_style = ParagraphStyle(
    "nota_indice_cabecalho",
    fontName=FONTE_PADRAO_NEGRITO,
    fontSize=13,
    leading=15,
    textColor=colors.HexColor("#2B2B2B"),
)

def _indice_desempenho_por_secretaria(grupos):
    # Achata os grupos do Índice de Desempenho num mapa
    # (gestão, secretaria) -> (índice, categoria), para o cabeçalho de cada
    # secretaria no Detalhamento Analítico.
    #
    # Na Estadual a unidade avaliada já é a própria secretaria, então a
    # correspondência é direta. Na Federal a unidade é SECRETARIA |
    # EXECUTOR: a mesma secretaria tem um índice por executor, e o
    # cabeçalho do detalhamento não separa por executor. Nesse caso o
    # medidor mostra a MÉDIA dos índices dos executores, ponderada pelo
    # investimento de cada um — é um resumo, não um índice recalculado para
    # a secretaria inteira. Quem precisa ver executor por executor continua
    # tendo a página do Índice de Desempenho.
    mapa = {}
    for grupo in grupos:
        por_secretaria = {}
        for item in grupo["itens"]:
            por_secretaria.setdefault(str(item["secretaria"]).strip(), []).append(item)
        for secretaria, itens in por_secretaria.items():
            pesos = [max(0.0, float(item.get("investimento") or 0.0)) for item in itens]
            total_peso = sum(pesos)
            if total_peso > 0:
                indice = sum(i["indice"] * p for i, p in zip(itens, pesos)) / total_peso
            else:
                # Todas as unidades com investimento zerado: média simples,
                # já que a ponderação não teria como distinguir uma da outra.
                indice = sum(i["indice"] for i in itens) / len(itens)
            mapa[(str(grupo["gestao"]).strip(), secretaria)] = (
                round(indice, 1),
                _categoria_indice_desempenho(indice),
            )
    return mapa

def extrair_ordem_status(status):
    m = re.match(r"^\s*(\d+)", str(status))
    return int(m.group(1)) if m else 999

def limpar_texto_status(status):
    return re.sub(r"^\s*\d+\s*-\s*", "", str(status)).strip()

def extrair_ordem_fase(fase):
    m = re.match(r"F\s*(\d+)", str(fase), re.IGNORECASE)
    return int(m.group(1)) if m else 999

def limpar_texto_fase(fase):
    return re.sub(r"F\s*\d+\s*-\s*", "", str(fase), flags=re.IGNORECASE).strip()

def formatar_prazo(valor):
    # Datas do corpo do relatório (prazos, emissão de O.S.) em dd/mm/aa —
    # ano com dois dígitos, para ocupar menos largura nas colunas. A data da
    # CAPA e o carimbo de "base atualizada em" continuam com o ano cheio, de
    # propósito: ali o ano completo é informação de registro do documento.
    if pd.isna(valor):
        return ""
    if isinstance(valor, (pd.Timestamp, datetime)):
        return valor.strftime("%d/%m/%y")
    texto = str(valor).strip()
    if texto == "" or texto.upper() in ("NAN", "NAT"):
        return ""
    return texto

def formatar_percentual(valor):
    if pd.isna(valor):
        return ""
    if isinstance(valor, (int, float, np.integer, np.floating)):
        # Só multiplica por 100 quando o valor é uma fração propriamente dita
        # (ex: 0,58 vindo de célula formatada como porcentagem no Excel).
        # O valor exato 1 (sem casas decimais, sem símbolo de %) NÃO é
        # tratado como fração — é o próprio número do percentual (1%), e não
        # 100%, evitando a ambiguidade que gerava "100,0%" incorretamente.
        pct = valor * 100 if 0 < valor < 1 else valor
        return f"{pct:.2f}".replace(".", ",") + "%"
    texto = str(valor).strip()
    if texto == "" or texto.upper() == "NAN":
        return ""
    return texto if texto.endswith("%") else f"{texto}%"

def _texto_campo_ficha(row, coluna):
    # Extrai um campo de texto de forma defensiva pra Ficha Cadastral: se a
    # coluna não existir na planilha (nome ainda não 100% confirmado) ou
    # estiver vazia, devolve string vazia — nunca quebra o programa.
    if coluna not in row.index:
        return ""
    valor = row[coluna]
    if valor is None or pd.isna(valor):
        # pd.isna cobre None, NaN e também pd.NaT (data vazia) — importante
        # checar ANTES do isinstance de datetime logo abaixo, porque NaT
        # passa no isinstance(..., datetime) mas não tem .strftime().
        return ""
    if isinstance(valor, (pd.Timestamp, datetime)):
        # Mesmo formato do corpo do relatório: dd/mm/aa.
        return valor.strftime("%d/%m/%y")
    texto = str(valor).strip()
    if texto.upper() in ("NAN", "NAT", "NONE"):
        return ""
    return texto

def _normalizar_link_ficha(texto):
    # Deixa o valor pronto para virar um link clicável. Muita gente cola o
    # endereço sem o "https://" na frente ("www.google.com/maps/...",
    # "monitora.ba.gov.br/..."), e nesse formato ele não é reconhecido como
    # link nem na tela nem no PDF — vira texto morto. Aqui o esquema é
    # acrescentado quando o valor claramente já é um endereço. O que não
    # parecer endereço volta intacto, para continuar aparecendo como texto.
    texto = str(texto or "").strip()
    if not texto:
        return ""
    if re.match(r"^https?://", texto, re.IGNORECASE):
        return texto
    if re.match(r"^www\.", texto, re.IGNORECASE):
        return "https://" + texto
    # domínio.tld seguido de barra ou fim (ex: monitora.ba.gov.br/acao/123)
    if re.match(r"^[\w-]+(\.[\w-]+){1,}(/|$)", texto) and " " not in texto:
        return "https://" + texto
    return texto

# --- Controle de Qualidade do campo LINK LOCALIZAÇÃO ---
# O campo existe para guardar o link do Google Maps do ponto da obra, e é
# comum vir preenchido com o endereço escrito ("Rua Dr. Rodrigues Lima,
# S/N, Centro - 44420000"). Nesse formato ele não abre mapa nenhum: na
# Ficha Cadastral o campo vira texto morto, e quem consulta fica sem saber
# onde a obra está. Os domínios abaixo cobrem as formas em que o Maps
# distribui um link hoje: o endereço completo do site, o encurtado do
# botão "Compartilhar" e o encurtado antigo.
DOMINIOS_LINK_MAPS = (
    "google.com/maps",
    "google.com.br/maps",
    "maps.google.com",
    "maps.google.com.br",
    "maps.app.goo.gl",
    "goo.gl/maps",
)

def _texto_vazio(valor):
    # Verdadeiro quando o campo está em branco: None/NaN, string vazia, ou os
    # textos que o pandas usa pra representar "nada" ("NAN"/"NAT"/"NONE").
    if valor is None or (isinstance(valor, float) and pd.isna(valor)):
        return True
    texto = str(valor).strip()
    return not texto or texto.upper() in ("NAN", "NAT", "NONE")

def _motivo_link_localizacao(valor):
    # Devolve o motivo do alerta, ou None quando o campo está vazio (campo
    # em branco não é erro de preenchimento aqui — quem cobra o preenchimento
    # em obras ANDAMENTO é _campos_alerta_qualidade) ou já traz um link do
    # Maps.
    if _texto_vazio(valor):
        return None
    texto = str(valor).strip()
    # Passa pela mesma normalização da ficha antes de julgar: um valor
    # colado sem o "https://" ("maps.app.goo.gl/abc") é um link válido, e
    # seria injusto acusá-lo só por causa do esquema ausente.
    endereco = _normalizar_link_ficha(texto).lower()
    if not endereco.startswith(("http://", "https://")):
        return "Link Localização preenchido com endereço em vez do link do Google Maps"
    if not any(dominio in endereco for dominio in DOMINIOS_LINK_MAPS):
        return "Link Localização com endereço que não é do Google Maps"
    return None

def _montar_dados_ficha_acao(row):
    def campo(coluna):
        return _texto_campo_ficha(row, coluna)

    def valor_moeda(coluna):
        if coluna not in row.index:
            return ""
        v = row[coluna]
        if v is None or (isinstance(v, float) and pd.isna(v)):
            return ""
        try:
            return moeda_texto_puro(float(v))
        except (TypeError, ValueError):
            return ""

    return {
        "item": normalizar_item(row.get(col_item)),
        "objeto": campo(col_objeto),
        "descricao": campo(col_descricao),
        "secretaria": campo(col_orgao),
        "executor": limpar_executor(campo(col_executor)),
        "gestao": campo("GESTAO"),
        "eixo": campo(col_eixo),
        "municipio": campo(col_municipio),
        "fase": campo("FASE_TEXTO"),
        "status": campo("STATUS_TEXTO"),
        "fonte": campo(col_fonte),
        "clausula_suspensiva": limpar_texto_status(campo(col_clausula_suspensiva)),
        "motivo_clausula_suspensiva": campo(col_motivo_clausula_suspensiva),
        "vigencia": campo(col_vigencia),
        "prazo_fase": campo(col_prazo),
        "prazo_atual": campo(col_prazo_atual),
        "avanco": formatar_percentual(row.get(col_avanco)) if col_avanco in row.index else "",
        "valor_contratado": valor_moeda(col_valor_contratado),
        "financiamento": valor_moeda(col_financiamento),
        "apoiado": valor_moeda(col_apoiado),
        "contrapartida": valor_moeda(col_contrapartida),
        "complementar": valor_moeda(col_complementar),
        "investimento_total": valor_moeda(col_invest),
        "pendencia": campo(col_pendencia),
        "providencias": campo(col_providencias),
        "prazo_pendencia": campo(col_prazo_pendencia),
        "proximos_passos": campo(col_proximos_passos),
        "link_monitora": _normalizar_link_ficha(campo(col_link_monitora)),
        "link_localizacao": _normalizar_link_ficha(campo(col_link_localizacao)),
        "aviso_licitacao": campo(col_aviso_licitacao),
        "abertura_licitacao": campo(col_abertura_licitacao),
        "emissao_os": campo(col_emissao_os),
        # Campos que o Controle de Qualidade sinalizou pra essa ação —
        # mesma lógica/fonte de verdade, usada pra destacar (contorno
        # vermelho) o campo na Ficha Cadastral, tanto na tela quanto na
        # versão impressa.
        "_campos_alerta": list(_campos_alerta_qualidade(row).keys()),
    }

def _definicao_secoes_ficha():
    # Mesma estrutura (seções/campos) usada pra montar a ficha no JS —
    # replicada aqui pra gerar o PDF de UMA ação com o Save em PDF.
    return [
        # O número do ITEM não entra como campo da seção: ele já aparece no
        # cabeçalho da ficha, ao lado do objeto ("Drenagem Urbana — Item 49"),
        # e repeti-lo aqui só gastava uma coluna larga com um número curto.
        {"titulo": "IDENTIFICAÇÃO", "largo": True, "campos": [
            ("objeto", "Objeto"), ("descricao", "Descrição"),
        ]},
        {"titulo": "EXECUÇÃO", "campos": [
            ("secretaria", "Secretaria/Órgão"), ("executor", "Órgão Executor"), ("gestao", "Gestão"),
            ("eixo", "Eixo"), ("municipio", "Município"), ("fonte", "Fonte de Recurso"),
        ]},
        {"titulo": "SITUAÇÃO", "campos": [
            ("fase", "Fase"), ("status", "Status"),
            ("clausula_suspensiva", "Cláusula Suspensiva"), ("motivo_clausula_suspensiva", "Motivo da Cláusula Suspensiva"),
        ]},
        {"titulo": "PRAZOS", "campos": [
            ("vigencia", "Vigência"), ("prazo_atual", "Previsão de Conclusão Atual"),
            ("prazo_fase", "Prazo de Conclusão da Fase"), ("avanco", "Avanço da Obra"),
        ]},
        {"titulo": "LICITAÇÃO", "campos": [
            ("aviso_licitacao", "Aviso de Licitação"), ("abertura_licitacao", "Abertura de Licitação"), ("emissao_os", "Emissão de O.S."),
        ]},
        {"titulo": "FINANCEIRO", "campos": [
            ("valor_contratado", "Valor Contratado"), ("financiamento", "Financiamento"), ("apoiado", "Apoiado (OGU)"),
            ("contrapartida", "Contrapartida"), ("complementar", "Complementar"), ("investimento_total", "Investimento Total"),
        ]},
        {"titulo": "ACOMPANHAMENTO", "largo": True, "campos": [
            ("pendencia", "Pendências / Tarefa"), ("providencias", "Providências (Datas)"),
            ("prazo_pendencia", "Prazo da Pendência / Tarefa"), ("proximos_passos", "Próximos Passos"),
        ]},
        # "colunas": 2 só entra aqui (não no JS) porque essa seção, na
        # versão PDF, é a única com QR Code — precisa de mais largura por
        # campo do que as 3 colunas padrão dariam, mas cabe lado a lado, ao
        # contrário das demais seções "largo" (texto corrido). Ordem
        # esquerda→direita: Localização primeiro, Monitora depois.
        {"titulo": "LINKS", "largo": True, "colunas": 2, "campos": [
            ("link_localizacao", "Link Localização"), ("link_monitora", "Link Monitora"),
        ]},
    ]

ficha_pdf_titulo_style = ParagraphStyle("ficha_pdf_titulo", fontName=FONTE_PADRAO_NEGRITO, fontSize=18, leading=21)
ficha_pdf_subtitulo_style = ParagraphStyle("ficha_pdf_subtitulo", fontName=FONTE_PADRAO, fontSize=9, textColor=colors.HexColor("#666666"))
ficha_pdf_secao_titulo_style = ParagraphStyle(
    "ficha_pdf_secao_titulo", fontName=FONTE_PADRAO_NEGRITO, fontSize=10, textColor=colors.HexColor("#1565A3")
)
ficha_pdf_rotulo_style = ParagraphStyle("ficha_pdf_rotulo", fontName=FONTE_PADRAO, fontSize=7.5, textColor=colors.HexColor("#777777"))
ficha_pdf_valor_style = ParagraphStyle("ficha_pdf_valor", fontName=FONTE_PADRAO, fontSize=10, leading=13, textColor=colors.HexColor("#1A1A1A"))
ficha_pdf_valor_vazio_style = ParagraphStyle("ficha_pdf_valor_vazio", fontName=FONTE_PADRAO, fontSize=10, textColor=colors.HexColor("#999999"))
ficha_pdf_qr_legenda_style = ParagraphStyle(
    "ficha_pdf_qr_legenda", fontName=FONTE_PADRAO, fontSize=7, leading=9, textColor=colors.HexColor("#777777")
)

# Lado do QR Code impresso na Ficha Cadastral, em pontos (1 mm = 2,8346 pt).
# 24 mm dá margem suficiente pra maioria dos leitores de câmera de celular
# ler mesmo numa folha A4 impressa em preto e branco.
TAMANHO_QR_FICHA = 24 * mm

class QRCodeLink(Flowable):
    # QR Code vetorial e clicável: no PDF digital funciona como um link
    # (clique abre o endereço, igual ao texto ao lado); na versão impressa
    # a mesma imagem é o que dá acesso ao link, apontando a câmera do
    # celular — por isso o link_monitora e o link_localizacao da Ficha
    # Cadastral ganham um QR Code em vez de só o texto.
    def __init__(self, url, tamanho=TAMANHO_QR_FICHA):
        Flowable.__init__(self)
        self.url = url
        self.tamanho = tamanho
        self.width = tamanho
        self.height = tamanho

    def draw(self):
        widget = QrCodeWidget(self.url)
        x0, y0, x1, y1 = widget.getBounds()
        largura_nativa = x1 - x0
        altura_nativa = y1 - y0
        desenho = Drawing(
            self.tamanho, self.tamanho,
            transform=[self.tamanho / largura_nativa, 0, 0, self.tamanho / altura_nativa, 0, 0],
        )
        desenho.add(widget)
        renderPDF.draw(desenho, self.canv, 0, 0)
        # "relative=1": a área clicável usa o mesmo sistema de coordenadas
        # do draw() (origem no canto do próprio Flowable), não a página
        # inteira — senão o link ficaria deslocado da imagem do QR.
        self.canv.linkURL(self.url, (0, 0, self.tamanho, self.tamanho), relative=1, thickness=0)

def gerar_pdf_ficha_acao(dados, caminho):
    # Gera um PDF de UMA página A4 com os dados completos de uma ação —
    # usado pelo botão "Salvar em PDF" da Ficha Cadastral. Campos
    # sinalizados pelo Controle de Qualidade (dados["_campos_alerta"])
    # recebem a mesma borda vermelha de destaque usada na tela.
    largura_pagina, altura_pagina = pagesizes.A4
    margem = 16 * mm
    largura_util = largura_pagina - 2 * margem
    campos_alerta = set(dados.get("_campos_alerta", []))

    # Campos da ficha que são endereços — no PDF eles viram link clicável,
    # em azul e sublinhado, em vez de um texto longo qualquer.
    CHAVES_LINK_FICHA = {"link_monitora", "link_localizacao"}

    def celula_campo(chave, rotulo, largura):
        valor = str(dados.get(chave, "") or "").strip()
        estilo_valor = ficha_pdf_valor_style if valor else ficha_pdf_valor_vazio_style
        texto_valor = valor.replace("\n", "<br/>") if valor else "—"
        eh_link = bool(valor) and chave in CHAVES_LINK_FICHA and re.match(r"^https?://", valor, re.IGNORECASE)

        bloco_rotulo_valor = [[Paragraph(rotulo.upper(), ficha_pdf_rotulo_style)]]
        if not eh_link:
            bloco_rotulo_valor.append([Paragraph(texto_valor, estilo_valor)])
        estilo_cel = [
            ("LEFTPADDING", (0, 0), (-1, -1), 6),
            ("RIGHTPADDING", (0, 0), (-1, -1), 6),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]
        if chave in campos_alerta:
            estilo_cel.append(("BOX", (0, 0), (-1, -1), 1, colors.HexColor("#D64545")))
            estilo_cel.append(("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#FCEDED")))

        if not eh_link:
            conteudo = Table(bloco_rotulo_valor, colWidths=[largura - 12])
            conteudo.setStyle(TableStyle(estilo_cel))
            return conteudo

        # Campo de link (Monitora / Localização): em vez do endereço como
        # texto, só o rótulo e o QR Code — clicável no PDF digital, e o que
        # dá acesso ao link na versão impressa, bastando apontar a câmera
        # do celular.
        bloco_rotulo_valor.append([Spacer(1, 3)])
        bloco_rotulo_valor.append(
            [Paragraph("Aponte a câmera do celular para acessar", ficha_pdf_qr_legenda_style)]
        )
        largura_texto = largura - TAMANHO_QR_FICHA - 24
        texto_tbl = Table(bloco_rotulo_valor, colWidths=[largura_texto])
        texto_tbl.setStyle(
            TableStyle(
                [
                    ("LEFTPADDING", (0, 0), (-1, -1), 0),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 0),
                    ("TOPPADDING", (0, 0), (-1, -1), 0),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
                ]
            )
        )
        conteudo = Table(
            [[QRCodeLink(valor), texto_tbl]],
            colWidths=[TAMANHO_QR_FICHA + 12, largura_texto],
        )
        estilo_cel.append(("VALIGN", (0, 0), (-1, -1), "MIDDLE"))
        conteudo.setStyle(TableStyle(estilo_cel))
        return conteudo

    elementos = [
        Paragraph(str(dados.get("objeto") or "Ficha Cadastral"), ficha_pdf_titulo_style),
        Paragraph(f"PAC - Ficha Cadastral da Ação — Item {dados.get('item', '')}", ficha_pdf_subtitulo_style),
        Spacer(1, 10),
    ]
    for secao in _definicao_secoes_ficha():
        # Cada seção (título + todos os campos dela) fica agrupada num
        # KeepTogether — se não couber no resto da página atual, a seção
        # INTEIRA passa pra próxima, em vez de quebrar no meio dela.
        flu_secao = [Paragraph(secao["titulo"], ficha_pdf_secao_titulo_style), Spacer(1, 3)]
        num_colunas = secao.get("colunas") or (1 if secao.get("largo") else 3)
        campos = secao["campos"]
        largura_coluna = largura_util / num_colunas
        for i in range(0, len(campos), num_colunas):
            linha_campos = campos[i:i + num_colunas]
            linha = [celula_campo(chave, rotulo, largura_coluna) for chave, rotulo in linha_campos]
            while len(linha) < num_colunas:
                linha.append("")
            tabela_linha = Table([linha], colWidths=[largura_coluna] * num_colunas)
            tabela_linha.setStyle(
                TableStyle(
                    [
                        ("VALIGN", (0, 0), (-1, -1), "TOP"),
                        ("LEFTPADDING", (0, 0), (-1, -1), 0),
                        ("RIGHTPADDING", (0, 0), (-1, -1), 0),
                        ("TOPPADDING", (0, 0), (-1, -1), 0),
                        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                    ]
                )
            )
            flu_secao.append(tabela_linha)
        flu_secao.append(Spacer(1, 8))
        elementos.append(KeepTogether(flu_secao))

    def desenhar_marca_ficha(canvas, doc_ficha):
        # Mesma marca do relatório gerencial (ver desenhar_cabecalho): os
        # dois logos no topo, a linha de 2 pt dividida ao meio entre teal e
        # gold separando o cabeçalho do corpo, e a data de atualização
        # centralizada no rodapé. Assim a ficha e o relatório se reconhecem
        # como peças do mesmo documento.
        #
        # As medidas são as do relatório; só as margens mudam, porque a
        # ficha é A4 retrato e o relatório é A3 paisagem.
        canvas.saveState()
        largura_pagina_marca, altura_pagina_marca = pagesizes.A4

        largura_pac, altura_pac = 40, 25
        largura_gov, altura_gov = 25, 35
        pos_y_logos = altura_pagina_marca - 37

        try:
            canvas.drawImage(
                caminho_recurso("GOVERNO.PNG"), margem, pos_y_logos,
                width=largura_gov, height=altura_gov, mask="auto",
            )
            canvas.drawImage(
                caminho_recurso("PAC.png"),
                largura_pagina_marca - margem - largura_pac, pos_y_logos,
                width=largura_pac, height=altura_pac, mask="auto",
            )
        except Exception:
            pass

        pos_y_linha = altura_pagina_marca - 46
        meia_largura = largura_pagina_marca / 2
        canvas.setLineWidth(2)
        canvas.setStrokeColor(COR_MARCA_TEAL)
        canvas.line(margem, pos_y_linha, meia_largura, pos_y_linha)
        canvas.setStrokeColor(COR_MARCA_GOLD)
        canvas.line(meia_largura, pos_y_linha, largura_pagina_marca - margem, pos_y_linha)

        # Rodapé: mesma frase de atualização do relatório gerencial, em
        # corpo menor porque a A4 é bem mais estreita que a A3 paisagem.
        canvas.setFont(FONTE_PADRAO, 8)
        canvas.setFillColor(colors.black)
        canvas.drawCentredString(meia_largura, 25, ultima_atualizacao_txt)
        canvas.restoreState()

    doc = SimpleDocTemplate(
        caminho,
        pagesize=pagesizes.A4,
        leftMargin=margem,
        rightMargin=margem,
        # Topo maior que a margem lateral para o conteúdo começar abaixo da
        # linha teal/gold; base o suficiente para não encostar na data.
        topMargin=58,
        bottomMargin=40,
    )
    doc.build(elementos, onFirstPage=desenhar_marca_ficha, onLaterPages=desenhar_marca_ficha)

def calcular_trimestre(data, hoje=None):
    if pd.isna(data) or not isinstance(data, (pd.Timestamp, datetime)):
        return "A definir"
    if hoje is None:
        hoje = datetime.now()
    # Trimestre a trimestre só no ano vigente e no seguinte — são os que
    # importam pra fiscalização de perto. Datas mais distantes (passadas ou
    # futuras) ficam agrupadas por ano inteiro, senão o eixo do gráfico
    # acumula dezenas de barras minúsculas sem informação útil.
    if data.year in (hoje.year, hoje.year + 1):
        trimestre = (data.month - 1) // 3 + 1
        ano_curto = data.strftime("%y")
        return f"{trimestre}ºTri/{ano_curto}"
    return str(data.year)

def ordenar_trimestre(rotulo):
    if rotulo == "A definir":
        return (9999, 9)
    texto = str(rotulo)
    m = re.match(r"(\d)ºTri/(\d+)", texto)
    if m:
        # +2000 no ano curto do trimestre pra ordenar corretamente junto dos
        # rótulos de ano cheio (4 dígitos) dos anos agrupados.
        return (2000 + int(m.group(2)), int(m.group(1)))
    if texto.isdigit():
        return (int(texto), 0)
    return (9999, 9)

# =====================================================
# 3.1. CATÁLOGO DE COLUNAS DO DETALHAMENTO ANALÍTICO
# =====================================================
# Quais colunas a tabela de DETALHAMENTO pode mostrar, em que ordem, com
# que largura relativa e — o ponto delicado — como cada uma participa das
# MESCLAS VERTICAIS.
#
# Existem dois tipos de mescla na tabela, e a diferença importa:
#
#   1) HIERARQUIA (OBJETO > FASE > STATUS): as linhas são ordenadas nessa
#      ordem, e cada nível só pode ser mesclado porque o nível de cima já
#      garante que as linhas iguais estão contíguas. Por isso essas três
#      colunas têm DEPENDÊNCIA declarada em "requer": FASE só pode ser
#      exibida com OBJETO, STATUS só com OBJETO e FASE. Sem isso, uma
#      seleção como "STATUS sem FASE" produziria blocos mesclados que
#      parecem quebrados (o mesmo status aparecendo várias vezes seguidas,
#      sem nada visível explicando a divisão).
#
#   2) REPETIÇÃO ("mescla_repetido"): FONTE, TERMO, PRAZO e as três
#      colunas de acompanhamento só juntam células vizinhas de valor igual
#      DENTRO do bloco hierárquico mais interno que estiver visível. Elas
#      nunca criam blocos próprios, então entram e saem da seleção sem
#      afetar as demais mesclas.
#
# "peso" é a largura relativa da coluna (a soma dos pesos das colunas
# escolhidas é normalizada para a largura útil da página, então as colunas
# se reacomodam sozinhas quando a seleção muda).
# "min_pt" é a largura mínima da coluna, em pontos: o suficiente para o
# cabeçalho e o conteúdo típico dela caberem sem virar uma coluna espremida.
# "peso" é a fatia que a coluna recebe do espaço que SOBRA depois dos
# mínimos — colunas de conteúdo curto e previsível (datas, percentual,
# SIM/NÃO) têm peso 0 e ficam no mínimo, devolvendo todo o excedente para
# as colunas de texto, que é onde a leitura realmente ganha.
CATALOGO_COLUNAS_DETALHAMENTO = [
    {"chave": "OBJETO", "titulo": "OBJETO", "min_pt": 95, "peso": 1.4,
     "obrigatoria": True, "requer": [], "padrao": True, "nivel_hierarquia": 1},
    {"chave": "FASE", "titulo": "FASE", "min_pt": 85, "peso": 0.9,
     "requer": ["OBJETO"], "padrao": True, "nivel_hierarquia": 2},
    {"chave": "STATUS", "titulo": "STATUS", "min_pt": 100, "peso": 1.0,
     "requer": ["OBJETO", "FASE"], "padrao": True, "nivel_hierarquia": 3},
    {"chave": "EMISSAO_OS", "titulo": "EMISSÃO DE O.S.", "min_pt": 80, "peso": 0.0,
     "padrao": True, "mescla_repetido": True},
    {"chave": "AVANCO", "titulo": "AVANÇO FÍSICO (%)", "min_pt": 70, "peso": 0.0, "padrao": True},
    {"chave": "MUNICIPIOS", "titulo": "MUNICÍPIOS", "min_pt": 120, "peso": 2.0, "padrao": True},
    # O mínimo aqui foi MEDIDO, não estimado: "R$ 495.000.000,00" ocupa
    # 102 pt no corpo da tabela e a célula ainda gasta 12 pt de padding, ou
    # seja, os 100 pt anteriores quebravam o valor em duas linhas. Com 136
    # cabe em uma linha só até a casa dos bilhões ("R$ 1.234.567.890,00"),
    # com folga para a diferença de métrica entre a fonte DIN do relatório
    # e a de referência. Peso 0: o conteúdo tem largura previsível, então a
    # coluna fica no tamanho de que precisa e devolve o excedente às
    # colunas de texto.
    {"chave": "INVESTIMENTO", "titulo": "INVESTIMENTO", "min_pt": 136, "peso": 0.0, "padrao": True},
    {"chave": "FONTE", "titulo": "FONTE DE RECURSO", "min_pt": 110, "peso": 1.0,
     "padrao": True, "mescla_repetido": True},
    {"chave": "TERMO", "titulo": "TERMO DE COMPROMISSO/ FINANCIAMENTO", "min_pt": 85, "peso": 0.2,
     "padrao": True, "mescla_repetido": True},
    {"chave": "PRAZO_FASE", "titulo": "PRAZO DE CONCLUSÃO DA FASE", "min_pt": 85, "peso": 0.2,
     "padrao": True, "mescla_repetido": True},
    {"chave": "PENDENCIA", "titulo": "PENDÊNCIAS / TAREFA", "min_pt": 110, "peso": 1.7,
     "padrao": False, "mescla_repetido": True},
    # PROVIDÊNCIAS (DATAS) é, na prática, a mais extensa das três colunas de
    # acompanhamento: costuma trazer o histórico corrido com datas e números
    # de processo ("Publicado aviso em 03/04, abertura em 22/04, análise
    # finalizada..."), enquanto PENDÊNCIAS traz uma frase curta ("Elaboração
    # TR e Edital") e PRÓXIMOS PASSOS uma ou duas linhas. Por isso ela tem o
    # maior mínimo E o maior peso das três — na prática sai com cerca de
    # 1,8× a largura de PENDÊNCIAS.
    {"chave": "PROVIDENCIAS", "titulo": "PROVIDÊNCIAS (DATAS)", "min_pt": 165, "peso": 4.4,
     "padrao": False, "mescla_repetido": True},
    {"chave": "PROXIMOS_PASSOS", "titulo": "PRÓXIMOS PASSOS", "min_pt": 120, "peso": 2.1,
     "padrao": False, "mescla_repetido": True},
]

COLUNAS_DETALHAMENTO_POR_CHAVE = {c["chave"]: c for c in CATALOGO_COLUNAS_DETALHAMENTO}
ORDEM_COLUNAS_DETALHAMENTO = [c["chave"] for c in CATALOGO_COLUNAS_DETALHAMENTO]
COLUNAS_DETALHAMENTO_PADRAO = [c["chave"] for c in CATALOGO_COLUNAS_DETALHAMENTO if c.get("padrao")]

# Teto de colunas na tabela. Com onze colunas a página A3 paisagem ainda
# comporta larguras mínimas razoáveis para todas: a soma dos mínimos das
# onze colunas mais largas do catálogo dá 1211 pt contra os 1395 pt de
# largura útil, então nenhuma seleção de onze colunas fica espremida.
# Acima disso as colunas de texto começam a quebrar em tantas linhas que
# as células mescladas estouram a altura da página. É um limite de
# segurança, não só estético — se um dia subir daqui, revisar os "min_pt"
# do catálogo junto.
#
# Foi de 10 para 11 quando EMISSÃO DE O.S. entrou na seleção padrão: com o
# padrão ocupando 10 vagas, o teto anterior deixava as colunas de
# acompanhamento (Pendências, Providências, Próximos Passos) bloqueadas
# até desmarcar alguma outra.
LIMITE_COLUNAS_DETALHAMENTO = 11

# A soma dos pesos das 9 colunas originais dava 0.971 da largura útil —
# mantido para que a seleção padrão saia exatamente com as larguras de
# sempre, e para deixar uma folga em relação ao padding interno do Frame.
FRACAO_LARGURA_DETALHAMENTO = 0.971

# Colunas de texto livre que a customização pode acrescentar à tabela de
# detalhamento. Ficam separadas porque precisam de tratamento próprio: não
# existem obrigatoriamente na planilha e atravessam o agrupamento do
# relatório por junção de textos, não por soma nem como chave.
COLUNAS_ACOMPANHAMENTO_DETALHAMENTO = [col_pendencia, col_providencias, col_proximos_passos]

# Nome da coluna DERIVADA (criada pelo próprio relatório, não vem da
# planilha) com a data de emissão da O.S. já formatada em dd/mm/aa.
COLUNA_EMISSAO_OS_TEXTO = "EMISSAO_OS_TEXTO"

# Ligação entre a chave da coluna no catálogo do detalhamento e o nome real
# da coluna no dataframe. Usada para decidir, na hora do agrupamento, quais
# dessas colunas entram como CHAVE (uma linha por valor distinto) em vez de
# agregação por junção. EMISSÃO DE O.S. entra aqui pelo mesmo motivo das
# três de acompanhamento: duas datas de O.S. empilhadas na mesma célula não
# dizem qual pertence a qual ação.
CHAVE_COLUNA_TRAVESSIA = {
    "PENDENCIA": col_pendencia,
    "PROVIDENCIAS": col_providencias,
    "PROXIMOS_PASSOS": col_proximos_passos,
    "EMISSAO_OS": COLUNA_EMISSAO_OS_TEXTO,
}

# Todas as colunas que atravessam o agrupamento — as que não viraram chave
# continuam sendo agregadas por junção dos textos distintos.
COLUNAS_TRAVESSIA_DETALHAMENTO = COLUNAS_ACOMPANHAMENTO_DETALHAMENTO + [COLUNA_EMISSAO_OS_TEXTO]

def texto_acompanhamento_chave(valor):
    # Normalização mínima para essas colunas quando elas viram chave de
    # agrupamento: tira espaços das pontas e trata NaN / "nan" como vazio.
    # Sem isso, o mesmo texto com um espaço a mais no fim viraria uma linha
    # separada na tabela, e células vazias apareceriam como "nan".
    if valor is None or (isinstance(valor, float) and pd.isna(valor)):
        return ""
    texto = str(valor).strip()
    return "" if texto.upper() == "NAN" else texto


def juntar_textos_distintos(serie, separador=" | "):
    # Junta os textos DIFERENTES de um grupo, na ordem em que aparecem,
    # ignorando vazios, NaN e repetições. Usada na agregação das colunas de
    # acompanhamento: como várias linhas da planilha viram uma linha só na
    # tabela, a célula precisa mostrar tudo que foi registrado naquele
    # grupo, sem repetir o mesmo texto várias vezes.
    vistos, partes = set(), []
    for valor in serie:
        if valor is None or (isinstance(valor, float) and pd.isna(valor)):
            continue
        texto = str(valor).strip()
        if not texto or texto.upper() == "NAN" or texto in vistos:
            continue
        vistos.add(texto)
        partes.append(texto)
    return separador.join(partes)

# Prazo de conclusão da fase VENCIDO — destacado em vermelho na tabela de
# detalhamento.
COR_PRAZO_VENCIDO = "#C00000"

# Status em que a fase já foi entregue: um prazo com data passada nesses
# casos não é atraso, é simplesmente uma etapa concluída, e não deve sair
# em vermelho.
STATUS_FASE_ENTREGUE = {"CONCLUÍDA", "CONCLUIDA", "INAUGURADA"}

def prazo_fase_vencido(texto_prazo, status_texto, hoje):
    # Lê a data já formatada da própria célula, em vez de buscar a data
    # crua: a coluna PRAZO DE CONCLUSÃO DA FASE chega na tabela como texto,
    # e a planilha às vezes traz ali coisas que não são data ("A DEFINIR",
    # "-"). Qualquer valor que não seja uma data válida simplesmente não é
    # considerado vencido.
    #
    # Aceita dd/mm/aa (o formato atual) e dd/mm/aaaa — assim a regra
    # continua valendo se o formato do relatório mudar de novo, ou se a
    # própria planilha trouxer o ano cheio nessa coluna.
    texto = str(texto_prazo or "").strip()
    correspondencia = re.match(r"^(\d{2})/(\d{2})/(\d{2}|\d{4})$", texto)
    if not correspondencia:
        return False
    if str(status_texto or "").strip().upper() in STATUS_FASE_ENTREGUE:
        return False
    try:
        dia, mes, ano = (int(parte) for parte in correspondencia.groups())
        if len(correspondencia.group(3)) == 2:
            ano += 2000
        return datetime(ano, mes, dia).date() < hoje
    except ValueError:
        return False  # data impossível na planilha (ex: 31/02)

def normalizar_colunas_detalhamento(chaves_selecionadas):
    # Recebe o que veio do painel (ou None) e devolve uma lista de colunas
    # SEMPRE válida. Toda a segurança da customização mora aqui, e não no
    # JavaScript: mesmo que o painel mande lixo, chaves desconhecidas, uma
    # lista vazia, colunas fora de ordem ou uma seleção que quebraria as
    # mesclas, o que sai daqui é uma tabela que se sustenta.
    #
    # As regras, na ordem em que são aplicadas:
    #   1. chaves desconhecidas são descartadas;
    #   2. lista vazia vira a seleção padrão (as 9 colunas de sempre);
    #   3. colunas obrigatórias entram sempre;
    #   4. dependências de hierarquia são satisfeitas automaticamente
    #      (pediu STATUS, ganha OBJETO e FASE junto);
    #   5. a ordem final é SEMPRE a do catálogo — o painel escolhe quais
    #      colunas, nunca a ordem, o que mantém a hierarquia de mesclas
    #      (OBJETO > FASE > STATUS) da esquerda pra direita;
    #   6. o limite de colunas é aplicado por último, cortando da direita
    #      pra esquerda e nunca deixando cair uma coluna da qual outra
    #      selecionada dependa.
    if not chaves_selecionadas:
        return list(COLUNAS_DETALHAMENTO_PADRAO)

    escolhidas = {str(c).strip().upper() for c in chaves_selecionadas}
    escolhidas = {c for c in escolhidas if c in COLUNAS_DETALHAMENTO_POR_CHAVE}
    if not escolhidas:
        return list(COLUNAS_DETALHAMENTO_PADRAO)

    for chave, config in COLUNAS_DETALHAMENTO_POR_CHAVE.items():
        if config.get("obrigatoria"):
            escolhidas.add(chave)

    # Fecho das dependências: repete até estabilizar, porque uma
    # dependência pode arrastar outra (STATUS puxa FASE, que puxa OBJETO).
    while True:
        faltando = set()
        for chave in escolhidas:
            for requerida in COLUNAS_DETALHAMENTO_POR_CHAVE[chave].get("requer", []):
                if requerida not in escolhidas:
                    faltando.add(requerida)
        if not faltando:
            break
        escolhidas |= faltando

    ordenadas = [c for c in ORDEM_COLUNAS_DETALHAMENTO if c in escolhidas]

    # Corte pelo limite: tira sempre a coluna mais à direita que ninguém
    # depende e que não seja obrigatória. Como a hierarquia fica à
    # esquerda, na prática caem primeiro as colunas de acompanhamento.
    while len(ordenadas) > LIMITE_COLUNAS_DETALHAMENTO:
        removivel = None
        for chave in reversed(ordenadas):
            config = COLUNAS_DETALHAMENTO_POR_CHAVE[chave]
            if config.get("obrigatoria"):
                continue
            dependentes = [
                outra for outra in ordenadas
                if chave in COLUNAS_DETALHAMENTO_POR_CHAVE[outra].get("requer", [])
            ]
            if not dependentes:
                removivel = chave
                break
        if removivel is None:
            break
        ordenadas.remove(removivel)

    return ordenadas

def calcular_larguras(largura_total, colunas=None):
    # Larguras absolutas das colunas da tabela de detalhamento, em duas
    # camadas: primeiro cada coluna recebe a sua largura MÍNIMA (o que ela
    # precisa para o cabeçalho e o conteúdo típico não ficarem espremidos);
    # o espaço que sobra é então distribuído pelos pesos, que privilegiam as
    # colunas de texto corrido.
    #
    # Isso substitui a divisão puramente proporcional que havia antes, em
    # que uma coluna de percentual ou de data crescia junto com as demais
    # quando havia poucas colunas selecionadas — desperdiçando largura que
    # fazia falta em MUNICÍPIOS ou em PENDÊNCIAS.
    colunas = colunas or COLUNAS_DETALHAMENTO_PADRAO
    largura_disponivel = largura_total * FRACAO_LARGURA_DETALHAMENTO

    minimos = [COLUNAS_DETALHAMENTO_POR_CHAVE[c]["min_pt"] for c in colunas]
    pesos = [COLUNAS_DETALHAMENTO_POR_CHAVE[c]["peso"] for c in colunas]
    soma_minimos = sum(minimos)

    if soma_minimos >= largura_disponivel:
        # Não caberia nem no mínimo (só aconteceria se o limite de colunas
        # subisse muito): reparte proporcionalmente aos mínimos, para pelo
        # menos não estourar a largura da página.
        fator = largura_disponivel / soma_minimos
        return [minimo * fator for minimo in minimos]

    sobra = largura_disponivel - soma_minimos
    soma_pesos = sum(pesos)
    if soma_pesos <= 0:
        # Seleção só de colunas de conteúdo curto: distribui o excedente por
        # igual, em vez de deixar uma faixa vazia à direita da tabela.
        return [minimo + sobra / len(colunas) for minimo in minimos]

    return [minimo + sobra * peso / soma_pesos for minimo, peso in zip(minimos, pesos)]

def estimar_altura_linha(celulas):
    # Estima a altura real da linha (pior caso, com o texto completo em cada
    # célula) usando a própria engine de quebra de texto do ReportLab, para
    # decidir com precisão onde uma célula mesclada pode ser cortada com
    # segurança sem estourar a altura da página. Recebe a lista de células
    # (texto, estilo, largura) que a linha realmente tem — como as colunas
    # agora são configuráveis, medir só um conjunto fixo delas deixaria a
    # conta errada justamente quando entra uma coluna de texto longo.
    maior_altura = 0
    for texto, estilo, largura in celulas:
        p = Paragraph(texto if str(texto).strip() else "&nbsp;", estilo)
        _, altura = p.wrap(largura, 3000)
        maior_altura = max(maior_altura, altura)
    return maior_altura + 16  # margem de segurança do padding vertical da célula (cobre variações de métrica entre fontes)

# Altura máxima de UMA linha da tabela. Uma linha não pode ser dividida
# entre páginas pelo ReportLab: se uma única linha ficar mais alta que a
# página, o build inteiro morre com LayoutError. As colunas de texto livre
# (Pendências, Providências, Próximos Passos) podem trazer parágrafos
# enormes da planilha, então o texto é aparado até caber nesse teto.
ALTURA_MAX_LINHA_DETALHAMENTO = 330

def limitar_texto_para_altura(texto, estilo, largura, altura_max=ALTURA_MAX_LINHA_DETALHAMENTO):
    # Encurta o texto (com reticências) até a célula caber na altura máxima.
    # Busca binária no número de caracteres em vez de cortar num tamanho
    # fixo: um texto de 400 caracteres pode caber numa coluna larga e
    # estourar numa estreita, e a seleção de colunas muda essa largura.
    texto = str(texto or "")
    if not texto.strip():
        return texto
    if Paragraph(texto, estilo).wrap(largura, 5000)[1] <= altura_max:
        return texto

    baixo, alto, melhor = 0, len(texto), ""
    while baixo <= alto:
        meio = (baixo + alto) // 2
        tentativa = texto[:meio].rstrip() + "…"
        if Paragraph(tentativa, estilo).wrap(largura, 5000)[1] <= altura_max:
            melhor = tentativa
            baixo = meio + 1
        else:
            alto = meio - 1
    return melhor or "…"

# =====================================================
# 4. ELEMENTOS VISUAIS E CALLBACKS DE PÁGINA
# =====================================================

def desenhar_background_capa(canvas, doc):
    canvas.saveState()
    largura, altura = PAGINA
    largura_img = largura * 1
    altura_img = altura * 1
    pos_x = -largura * 0.10
    pos_y = -altura * 0.10
    
    try:
        canvas.setFillAlpha(0.15)
        canvas.setStrokeAlpha(0.15)
        canvas.drawImage(
            caminho_recurso("PAC.png"),
            pos_x,
            pos_y,
            width=largura_img,
            height=altura_img,
            mask="auto",
        )
    except Exception:
        pass
        
    canvas.setFillAlpha(1.0)
    canvas.setStrokeAlpha(1.0)
    
    largura_gov, altura_gov = 100, 110
    pos_x_gov = largura - MARGEM_DIR - largura_gov
    pos_y_gov = MARGEM_ESQ
    
    try:
        canvas.drawImage(
            caminho_recurso("GOVERNO.PNG"),
            pos_x_gov,
            pos_y_gov,
            width=largura_gov,
            height=altura_gov,
            mask="auto",
        )
    except Exception:
        pass

    # Texto do CGAPE do lado ESQUERDO da folha, centralizado verticalmente
    # na mesma faixa de altura do emblema do estado (GOVERNO.PNG, desenhado
    # acima) — por isso é desenhado aqui, com coordenada própria, em vez de
    # entrar no fluxo normal de parágrafos da capa.
    paragrafo_cgape_esq = Paragraph(
        "Fonte: CASA CIVIL/CGAPE - Coordenação Geral de Acompanhamento de Políticas Estratégicas - Planilha Panorama",
        capa_cgape_esquerda,
    )
    largura_disponivel_cgape = largura * 0.55
    _, altura_paragrafo_cgape = paragrafo_cgape_esq.wrap(largura_disponivel_cgape, altura)
    pos_y_cgape = pos_y_gov + (altura_gov - altura_paragrafo_cgape) / 2
    paragrafo_cgape_esq.drawOn(canvas, MARGEM_ESQ, pos_y_cgape)

    canvas.restoreState()

def desenhar_cabecalho(canvas, doc):
    canvas.saveState()
    largura, altura = PAGINA

    largura_pac, altura_pac = 40, 25
    largura_gov, altura_gov = 25, 35
    pos_y_logos = altura - 37

    pos_x_gov = MARGEM_ESQ
    pos_x_pac = largura - MARGEM_DIR - largura_pac

    try:
        canvas.drawImage(
            caminho_recurso("GOVERNO.PNG"),
            pos_x_gov,
            pos_y_logos,
            width=largura_gov,
            height=altura_gov,
            mask="auto",
        )
        canvas.drawImage(
            caminho_recurso("PAC.png"),
            pos_x_pac,
            pos_y_logos,
            width=largura_pac,
            height=altura_pac,
            mask="auto",
        )
    except Exception:
        pass

    pos_y_linha = altura - 46
    meia_largura = largura / 2

    canvas.setLineWidth(2)
    canvas.setStrokeColor(COR_MARCA_TEAL)
    canvas.line(MARGEM_ESQ, pos_y_linha, meia_largura, pos_y_linha)

    canvas.setStrokeColor(COR_MARCA_GOLD)
    canvas.line(meia_largura, pos_y_linha, largura - MARGEM_DIR, pos_y_linha)

    canvas.setFont(FONTE_PADRAO, 13)
    canvas.setFillColor(colors.black)
    canvas.drawCentredString(largura / 2, 25, ultima_atualizacao_txt)
    canvas.restoreState()

# Linha de filtros únicos ("SECRETARIA | EXECUTOR | FASE | STATUS") que o
# rodapé das páginas de gestão desenha à esquerda, na mesma altura e com a
# mesma formatação do "GESTÃO ESTADUAL - Página 01 de 15" da direita. Fica
# em variável de módulo porque o canvas é instanciado pelo ReportLab lá
# dentro do doc.build() e não recebe o dataframe — quem gera o PDF preenche
# esta variável antes de montar o documento. Vazia = nada é desenhado.
TEXTO_FILTROS_RODAPE = ""

# --- Numeração de páginas por gestão ("Gestão Estadual - Página 01 de 15"),
# reiniciando a contagem em cada gestão, mesmo quando as páginas de cada
# gestão não ficam fisicamente seguidas no PDF (ex: se no futuro alguma
# seção passar a intercalar gestões). Funciona em duas passadas: 1) cada
# página, ao ser fechada, guarda uma cópia do estado do canvas junto com
# a gestão marcada nela (via _MarcadorGestao); 2) só depois de processar
# TODAS as páginas — quando já se sabe quantas cada gestão tem — é que o
# texto "Página X de Y" é de fato desenhado, uma página de cada vez.
class NumeradorPaginasGestaoCanvas(reportlab_canvas.Canvas):
    def __init__(self, *args, **kwargs):
        reportlab_canvas.Canvas.__init__(self, *args, **kwargs)
        self._paginas_salvas = []
        self._gestao_pagina_atual = None

    def showPage(self):
        self._paginas_salvas.append(dict(self.__dict__))
        self._startPage()

    def save(self):
        # Conta quantas páginas cada gestão tem, na ordem em que aparecem.
        contagem_por_gestao = {}
        for estado in self._paginas_salvas:
            g = estado.get("_gestao_pagina_atual")
            contagem_por_gestao[g] = contagem_por_gestao.get(g, 0) + 1

        indice_por_gestao = {}
        for estado in self._paginas_salvas:
            self.__dict__.update(estado)
            g = self._gestao_pagina_atual
            if g:
                indice_por_gestao[g] = indice_por_gestao.get(g, 0) + 1
                total_da_gestao = contagem_por_gestao[g]
                texto_pagina = f"{g} - Página {indice_por_gestao[g]:02d} de {total_da_gestao:02d}"
                largura, _altura = PAGINA
                self.saveState()
                self.setFont(FONTE_PADRAO_NEGRITO, 15)
                self.setFillColor(COR_MARCA_TEAL_ESCURO)
                self.drawRightString(largura - MARGEM_DIR, 25, texto_pagina)
                # Filtros únicos do recorte, à esquerda, na mesma linha de
                # base e com a mesma fonte/cor. O espaço disponível é o que
                # sobra até onde começa o texto da direita, com um vão de
                # 30 pt entre os dois; se o nome da secretaria ou do
                # executor for longo demais, a fonte encolhe até 9 pt e, no
                # limite, o texto é cortado com reticências, em vez de
                # encavalar na numeração.
                if TEXTO_FILTROS_RODAPE:
                    largura_pagina_txt = self.stringWidth(texto_pagina, FONTE_PADRAO_NEGRITO, 15)
                    espaco = largura - MARGEM_ESQ - MARGEM_DIR - largura_pagina_txt - 30
                    texto_filtros = TEXTO_FILTROS_RODAPE
                    corpo = 15
                    while corpo > 9 and self.stringWidth(texto_filtros, FONTE_PADRAO_NEGRITO, corpo) > espaco:
                        corpo -= 1
                    if self.stringWidth(texto_filtros, FONTE_PADRAO_NEGRITO, corpo) > espaco:
                        while texto_filtros and self.stringWidth(
                            texto_filtros + "...", FONTE_PADRAO_NEGRITO, corpo
                        ) > espaco:
                            texto_filtros = texto_filtros[:-1]
                        texto_filtros = texto_filtros.rstrip(" |") + "..."
                    if espaco > 0:
                        self.setFont(FONTE_PADRAO_NEGRITO, corpo)
                        self.drawString(MARGEM_ESQ, 25, texto_filtros)
                self.restoreState()
            reportlab_canvas.Canvas.showPage(self)
        reportlab_canvas.Canvas.save(self)

class _MarcadorGestao(Flowable):
    # Flowable "invisível" (não ocupa espaço nem desenha nada visível) —
    # serve só pra marcar, no momento em que é desenhado, qual gestão a
    # página atual representa. Usado pelo NumeradorPaginasGestaoCanvas
    # acima pra saber a que gestão cada página pertence.
    def __init__(self, gestao_label):
        Flowable.__init__(self)
        self.gestao_label = gestao_label
        self.width = 0
        self.height = 0

    def wrap(self, *args):
        return (0, 0)

    def draw(self):
        self.canv._gestao_pagina_atual = self.gestao_label

# =====================================================
# 5. FÁBRICA DE TABELAS E CARDS (FABRICATOR)
# =====================================================

def gerar_cards_resumo(qtd, investimento):
    data = [
        [
            Paragraph(
                "QTD TOTAL",
                ParagraphStyle(
                    "card_header_qtd",
                    fontName=RESUMO_FONTE_NEGRITO,
                    fontSize=CARDS_FONTE_TAM_CABECALHO,
                    alignment=TA_CENTER,
                    textColor=colors.white,
                ),
            ),
            Paragraph(
                "INVESTIMENTO",
                ParagraphStyle(
                    "card_header_inv",
                    fontName=RESUMO_FONTE_NEGRITO,
                    fontSize=CARDS_FONTE_TAM_CABECALHO,
                    alignment=TA_CENTER,
                    textColor=colors.white,
                ),
            ),
        ],
        [
            Paragraph(
                str(qtd),
                ParagraphStyle(
                    "card_value_qtd",
                    fontName=FONTE_BAHNSCHRIFT,
                    fontSize=CARDS_FONTE_TAM_VALOR,
                    leading=CARDS_FONTE_TAM_VALOR,
                    alignment=TA_CENTER,
                    textColor=colors.black,
                ),
            ),
            Paragraph(
                formatar_mi_bi(investimento),
                ParagraphStyle(
                    "card_value_inv",
                    fontName=FONTE_BAHNSCHRIFT,
                    fontSize=CARDS_FONTE_TAM_VALOR,
                    leading=CARDS_FONTE_TAM_VALOR,
                    alignment=TA_CENTER,
                    textColor=colors.black,
                ),
            ),
        ],
    ]
    tabela = Table(
        data,
        colWidths=[LARGURA_UTIL * CARDS_FATOR_LARGURA, LARGURA_UTIL * CARDS_FATOR_LARGURA],
        rowHeights=[CARDS_ALTURA_CABECALHO, CARDS_ALTURA_VALOR],
    )
    tabela.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), COR_MARCA_TEAL_ESCURO),
                ("BACKGROUND", (0, 1), (-1, -1), colors.whitesmoke),
                ("BOX", (0, 0), (-1, -1), 1.0, colors.grey),
                ("INNERGRID", (0, 0), (-1, -1), 0.5, colors.lightgrey),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ]
        )
    )
    return tabela

CORES_DETALHAMENTO_SECRETARIA = {
    "Valor Contratado": cor_grafico("#49925C"),
    "Valor Apoiado OGU": cor_grafico("#4E92BA"),
    "Recurso Estadual": cor_grafico("#BC9E2C"),
    "Financiamento": cor_grafico("#BB6060"),
}

def desenhar_legenda_detalhamento_secretaria(largura=LARGURA_UTIL):
    # Uma linha só, com um quadradinho colorido + nome por componente —
    # mesma legenda usada uma única vez no topo da seção equivalente do
    # dashboard (não repete em cada barra).
    #
    # Valor Apoiado OGU, Recurso Estadual e Financiamento levam "*" — são
    # valores de PREVISÃO ORÇAMENTÁRIA (ainda não contratados), mesmo
    # critério da legenda do RESUMO FINANCEIRO do Painel Geral (ver
    # NOMES_PREVISAO_ORCAMENTARIA em gerar_grafico_financeiro), mas com um
    # único "*" — nesta página específica não há outra marcação em disputa
    # pelo símbolo. Valor Contratado é firme e por isso fica sem marcação.
    NOMES_PREVISAO_ORCAMENTARIA = {"Valor Apoiado OGU", "Recurso Estadual", "Financiamento"}
    altura = 26
    d = Drawing(largura, altura)
    x = 0
    for nome, cor in CORES_DETALHAMENTO_SECRETARIA.items():
        rotulo = f"{nome}*" if nome in NOMES_PREVISAO_ORCAMENTARIA else nome
        d.add(Rect(x, 6, 14, 14, fillColor=cor, strokeColor=None))
        d.add(String(x + 19, 6.5, rotulo, fontName=FONTE_PADRAO, fontSize=16, fillColor=colors.HexColor("#555555")))
        x += 19 + len(rotulo) * 9.0 + 30
    return d

def _formatar_mi_bi_texto_puro(valor):
    # Versão em TEXTO PURO de formatar_mi_bi — usada dentro de String (do
    # reportlab.graphics.shapes), que não interpreta "&nbsp;" como o
    # Paragraph faz; usar formatar_mi_bi aqui mostraria o "&nbsp;" literal.
    if valor >= 1_000_000_000:
        return f"R$ {valor/1_000_000_000:,.1f}".replace(",", "X").replace(".", ",").replace("X", ".") + " Bi"
    if valor >= 1_000_000:
        return f"R$ {valor/1_000_000:,.1f}".replace(",", "X").replace(".", ",").replace("X", ".") + " Mi"
    return f"R$ {valor:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")

def formas_barra_arredondada(x, y, largura, altura, cor, ponta_inicio=True, ponta_fim=True,
                             horizontal=True, raio=None):
    # Retângulo de barra com as pontas EXTERNAS arredondadas, na mesma
    # terminação do arco do medidor de desempenho: raio igual à metade da
    # espessura, o que faz a ponta virar um semicírculo perfeito.
    #
    # O ReportLab só arredonda os quatro cantos de uma vez (rx/ry do Rect),
    # então as pontas que precisam continuar retas — a base de uma coluna,
    # ou o lado de um segmento que encosta no segmento vizinho — são
    # quadradas de volta por um retângulo comum, da mesma cor, cobrindo a
    # faixa do raio. Como a cobertura usa a cor da própria barra, funciona
    # sobre qualquer fundo.
    #
    # horizontal=True: ponta_inicio é a esquerda e ponta_fim é a direita.
    # horizontal=False: ponta_inicio é a base e ponta_fim é o topo.
    espessura = altura if horizontal else largura
    comprimento = largura if horizontal else altura
    raio = espessura / 2 if raio is None else raio
    raio = max(0.0, min(raio, espessura / 2, comprimento / 2))
    if raio <= 0 or (not ponta_inicio and not ponta_fim):
        # Nenhuma ponta externa: é um segmento do meio de uma barra
        # empilhada, encostado nos vizinhos dos dois lados. Sai como
        # retângulo comum, sem passar pelo arredondar-e-quadrar-de-volta.
        return [Rect(x, y, largura, altura, fillColor=cor, strokeColor=None)]

    formas = [Rect(x, y, largura, altura, fillColor=cor, strokeColor=None, rx=raio, ry=raio)]
    if not ponta_inicio:
        formas.append(
            Rect(x, y, raio, altura, fillColor=cor, strokeColor=None) if horizontal
            else Rect(x, y, largura, raio, fillColor=cor, strokeColor=None)
        )
    if not ponta_fim:
        formas.append(
            Rect(x + largura - raio, y, raio, altura, fillColor=cor, strokeColor=None) if horizontal
            else Rect(x, y + altura - raio, largura, raio, fillColor=cor, strokeColor=None)
        )
    return formas

def desenhar_barra_detalhamento_secretaria(item, largura, altura_barra=13, mostrar_detalhe=True,
                                          texto_acima=None, fonte_acima=12):
    # Barra empilhada PERCENTUAL (sempre 100% de largura, como no dashboard)
    # — cada segmento representa a fatia de Valor Contratado/OGU/Recurso
    # Estadual/Financiamento dentro do investimento daquela secretaria, sem
    # comparar o comprimento total entre secretarias diferentes.
    base = item["investimento"] if item["investimento"] > 0 else 1
    segmentos = [
        ("Valor Contratado", item["valorContratado"]),
        ("Valor Apoiado OGU", item["valorApoiadoOgu"]),
        ("Recurso Estadual", item["recursoEstadual"]),
        ("Financiamento", item["financiamento"]),
    ]
    siglas = {"Valor Contratado": "VC", "Valor Apoiado OGU": "OGU", "Recurso Estadual": "RE", "Financiamento": "FIN"}
    partes_texto = []
    for nome, valor in segmentos:
        if valor <= 0:
            continue
        texto_segmento = f"{siglas[nome]}: {_formatar_mi_bi_texto_puro(valor)}"
        if nome == "Financiamento" and item.get("fontesFinanciamento"):
            texto_segmento += " (" + ", ".join(item["fontesFinanciamento"]) + ")"
        partes_texto.append(texto_segmento)
    texto_detalhe = "     ".join(partes_texto)

    # Texto do detalhamento numa linha só, na mesma fonte/tamanho da
    # legenda (13pt) — altura do desenho fixa, já que não varia mais
    # conforme a quantidade de categorias presentes.
    #
    # mostrar_detalhe=False: miniatura, usada no cabeçalho de secretaria do
    # Detalhamento Analítico. Só a barra colorida, sem a linha "VC: ... OGU:
    # ..." embaixo — naquele tamanho o texto não caberia, e a legenda da
    # página de Detalhamento Financeiro já explica as cores.
    altura_texto_detalhe = 22 if mostrar_detalhe else 0
    espaco_acima_detalhe = 4 if mostrar_detalhe else 0
    # texto_acima: linha curta centralizada em cima da barra (o percentual
    # já contratado, no uso do cabeçalho de secretaria). Fica ACIMA e não
    # dentro porque a barra tem 11 pt de altura na miniatura — não há como
    # escrever ali dentro sem cobrir os segmentos.
    altura_texto_acima = (fonte_acima + 3) if texto_acima else 0
    altura_total = altura_barra + espaco_acima_detalhe + altura_texto_detalhe + altura_texto_acima
    topo_barra = altura_total - altura_texto_acima
    d = Drawing(largura, altura_total)
    if texto_acima:
        d.add(
            String(
                largura / 2, topo_barra + 3, texto_acima,
                fontName=FONTE_PADRAO_NEGRITO, fontSize=fonte_acima, textAnchor="middle",
                fillColor=colors.HexColor("#555555"),
            )
        )
    for forma in formas_barra_arredondada(
        0, topo_barra - altura_barra, largura, altura_barra, colors.HexColor("#E4E4E4")
    ):
        d.add(forma)
    # Só o PRIMEIRO segmento arredonda à esquerda e só o ÚLTIMO à direita —
    # as junções internas continuam retas, senão apareceria uma falha em
    # forma de lente entre um segmento e o vizinho.
    segmentos_visiveis = [(nome, valor) for nome, valor in segmentos if valor > 0]
    x = 0
    limites_internos = []
    for posicao, (nome, valor) in enumerate(segmentos_visiveis):
        largura_segmento = (valor / base) * largura
        for forma in formas_barra_arredondada(
            x, topo_barra - altura_barra, largura_segmento, altura_barra,
            CORES_DETALHAMENTO_SECRETARIA[nome],
            ponta_inicio=(posicao == 0),
            ponta_fim=(posicao == len(segmentos_visiveis) - 1),
            raio=altura_barra / 2,
        ):
            d.add(forma)
        x += largura_segmento
        limites_internos.append(x)

    # Linha VERTICAL branca separando cada segmento colorido do seguinte —
    # design da barra em si (não confundir com as linhas entre categorias
    # de texto abaixo, nem com a linha entre secretarias na tabela).
    for x_divisao in limites_internos[:-1]:
        d.add(
            Line(
                x_divisao, topo_barra - altura_barra,
                x_divisao, topo_barra,
                strokeColor=colors.white, strokeWidth=1.6,
            )
        )

    if mostrar_detalhe:
        d.add(String(0, 5, texto_detalhe, fontName=FONTE_PADRAO, fontSize=16, fillColor=colors.HexColor("#666666")))
    return d

def gerar_grafico_detalhamento_secretaria(df_base):
    # Mesma regra de valores ajustados usada no dashboard: o INVESTIMENTO de
    # cada secretaria é a soma de Valor Contratado + OGU + Recurso Estadual
    # (Contrapartida + Complementar) + Financiamento, todos já ajustados
    # pra não contar duas vezes — garante que a barra empilhada sempre fecha
    # em 100% exatos.
    df_resumo = (
        df_base.groupby(["GESTAO", "SECRETARIA_LIMPA"])
        .agg(
            QTD=(col_objeto, "count"),
            VALOR_CONTRATADO=(col_valor_contratado, "sum"),
            APOIADO_AJ=(col_apoiado_ajustado, "sum"),
            CONTRAPARTIDA_AJ=(col_contrapartida_ajustado, "sum"),
            COMPLEMENTAR_AJ=(col_complementar_ajustado, "sum"),
            FINANCIAMENTO_AJ=(col_financiamento_ajustado, "sum"),
        )
        .reset_index()
    )
    df_resumo["RECURSO_ESTADUAL"] = df_resumo["CONTRAPARTIDA_AJ"] + df_resumo["COMPLEMENTAR_AJ"]
    df_resumo["INVESTIMENTO"] = (
        df_resumo["VALOR_CONTRATADO"]
        + df_resumo["APOIADO_AJ"]
        + df_resumo["RECURSO_ESTADUAL"]
        + df_resumo["FINANCIAMENTO_AJ"]
    )
    df_resumo = df_resumo[df_resumo["QTD"] > 0]
    if df_resumo.empty:
        return None

    # Fontes de Financiamento distintas (ex: FGTS, CAIXA) usadas em cada
    # secretaria, só entre as ações que de fato têm Financiamento > 0 —
    # acesso defensivo: se a coluna não existir na planilha, simplesmente
    # não aparece nada extra no detalhe da barra.
    if col_fonte_financiamento in df_base.columns:
        fontes_fin_por_grupo = (
            df_base[df_base[col_financiamento_ajustado] > 0]
            .groupby(["GESTAO", "SECRETARIA_LIMPA"])[col_fonte_financiamento]
            .apply(lambda s: sorted(set(str(v).strip() for v in s.dropna() if str(v).strip())))
            .to_dict()
        )
    else:
        fontes_fin_por_grupo = {}

    largura_rotulo = LARGURA_UTIL * 0.30
    largura_barra = LARGURA_UTIL * 0.68

    # Gestão Estadual sempre primeiro, Federal logo abaixo — cada uma numa
    # seção separada por um cabeçalho, em vez de secretarias das duas
    # gestões misturadas numa lista só ordenada por investimento.
    ORDEM_GESTAO = ["GESTÃO ESTADUAL", "GESTÃO FEDERAL"]
    gestoes_presentes = sorted(
        df_resumo["GESTAO"].unique(),
        key=lambda g: ORDEM_GESTAO.index(g) if g in ORDEM_GESTAO else 99,
    )

    data = []
    estilos_extra = []
    linhas_cabecalho_gestao = []
    linha = 0
    for gestao in gestoes_presentes:
        df_gestao = df_resumo[df_resumo["GESTAO"] == gestao].sort_values(
            "INVESTIMENTO", ascending=False
        )
        if df_gestao.empty:
            continue

        cabecalho_gestao = Table(
            [[Paragraph(f"<b>{gestao}</b>", detalhamento_gestao_header_style)]],
            colWidths=[largura_rotulo + largura_barra],
        )
        cabecalho_gestao.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, -1), COR_MARCA_TEAL_CLARA),
                    ("LEFTPADDING", (0, 0), (-1, -1), 10),
                    ("TOPPADDING", (0, 0), (-1, -1), 10),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 10),
                ]
            )
        )
        data.append([cabecalho_gestao, ""])
        estilos_extra.append(("SPAN", (0, linha), (1, linha)))
        linhas_cabecalho_gestao.append(linha)
        linha += 1

        for _, row in df_gestao.iterrows():
            item = {
                "secretaria": str(row["SECRETARIA_LIMPA"]),
                "qtd": int(row["QTD"]),
                "investimento": float(row["INVESTIMENTO"]),
                "valorContratado": float(row["VALOR_CONTRATADO"]),
                "valorApoiadoOgu": float(row["APOIADO_AJ"]),
                "recursoEstadual": float(row["RECURSO_ESTADUAL"]),
                "financiamento": float(row["FINANCIAMENTO_AJ"]),
                "fontesFinanciamento": fontes_fin_por_grupo.get((gestao, str(row["SECRETARIA_LIMPA"])), []),
            }
            rotulo = Paragraph(
                f'<font name="{FONTE_PADRAO_NEGRITO}">{item["secretaria"]}</font> '
                f"({item['qtd']}) — {formatar_mi_bi(item['investimento'])}",
                detalhamento_secretaria_rotulo_style,
            )
            barra = desenhar_barra_detalhamento_secretaria(item, largura_barra)
            data.append([rotulo, barra])
            linha += 1

    tabela = Table(data, colWidths=[largura_rotulo, largura_barra], repeatRows=0)
    estilo_base = [
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ("LEFTPADDING", (0, 0), (0, -1), 0),
        ("RIGHTPADDING", (0, 0), (0, -1), 8),
        ("LEFTPADDING", (1, 0), (1, -1), 0),
        ("LINEBELOW", (0, 0), (-1, -2), 0.4, colors.HexColor("#DDDDDD")),
    ]
    # Mais respiro acima do cabeçalho de cada gestão seguinte (a partir da
    # segunda), separando visualmente uma gestão da outra — a primeira não
    # precisa, já que vem logo depois da legenda.
    for idx_linha_cabecalho in linhas_cabecalho_gestao[1:]:
        estilo_base.append(("TOPPADDING", (0, idx_linha_cabecalho), (-1, idx_linha_cabecalho), 14))
    estilo_base.extend(estilos_extra)
    tabela.setStyle(TableStyle(estilo_base))
    tabela.hAlign = "CENTER"
    return tabela

# Resumo por Eixo / por Status: cada linha traz o número e, ao lado, uma
# barra do tamanho proporcional ao maior valor da coluna. A barra não
# acrescenta informação nova — ela torna a comparação imediata, que numa
# lista de dezesseis linhas de números alinhados não acontece: dá para ver
# de relance que Prevenção a Desastres tem quase quatro vezes mais ações
# que Cidades Sustentáveis e, ao mesmo tempo, menos de um quarto do
# investimento dela.
#
# Uma cor só para todas as barras, de propósito: cor por status sugeriria
# que uns são melhores que outros, e este é um resumo de composição, não de
# desempenho.
COR_BARRA_RESUMO = colors.HexColor("#5E9C96")
ALTURA_BARRA_RESUMO = 9.0

def _barra_resumo(valor, maximo, largura):
    # Barra proporcional ao maior valor da coluna. Devolve None quando não
    # há o que comparar (coluna toda zerada) — melhor a célula vazia do que
    # uma barra que não significa nada.
    if maximo <= 0 or valor <= 0 or largura <= 0:
        return None
    largura_barra = max(1.5, (valor / maximo) * largura)
    d = Drawing(largura, ALTURA_BARRA_RESUMO)
    for forma in formas_barra_arredondada(
        0, 0, largura_barra, ALTURA_BARRA_RESUMO, COR_BARRA_RESUMO
    ):
        d.add(forma)
    return d

def _moeda_abreviada_resumo(valor):
    # "R$ 5,70 Bi" / "R$ 439,16 Mi" — duas casas decimais, ao contrário de
    # formatar_mi_bi (que usa uma), porque aqui a coluna é a informação
    # principal da linha e uma casa perderia diferença entre valores
    # próximos. Abaixo de um milhão, mostra o valor cheio.
    if valor >= 1_000_000_000:
        texto = f"R$&nbsp;{valor/1_000_000_000:,.2f}&nbsp;Bi"
    elif valor >= 1_000_000:
        texto = f"R$&nbsp;{valor/1_000_000:,.2f}&nbsp;Mi"
    else:
        return moeda_sem_quebra(valor)
    return texto.replace(",", "X").replace(".", ",").replace("X", ".")

def _tabela_resumo_com_barras(df_base, gestao, coluna_nome, coluna_ordem, titulo, largura):
    # Uma função para os dois resumos (Eixo e Status): a estrutura é a
    # mesma, muda só a coluna de agrupamento e a de ordenação.
    target_df = df_base[df_base["GESTAO"] == gestao] if gestao != "GERAL" else df_base
    target_df = target_df.assign(_INV_AJ=_serie_investimento_ajustado(target_df))
    df_resumo = (
        target_df.groupby([coluna_ordem, coluna_nome])
        .agg({col_objeto: "count", "_INV_AJ": "sum"})
        .reset_index()
        .sort_values(coluna_ordem)
    )
    if df_resumo.empty:
        return None

    max_inv = float(df_resumo["_INV_AJ"].max())

    # Sem barra de quantidade: os números de QTD são curtos e já se
    # comparam bem lidos lado a lado numa coluna estreita. A barra fica só
    # no investimento, onde a leitura dos valores é mais difícil.
    largura_nome = largura * 0.42
    largura_qtd = largura * 0.10
    largura_inv = largura * 0.22
    largura_barra_inv = largura * 0.26

    # O cabeçalho de INVESTIMENTO abrange também a coluna da barra ao lado
    # (SPAN, mais abaixo): a barra é a mesma informação, não uma coluna
    # nova a ser rotulada.
    data = [
        [
            Paragraph(f"<b>{titulo}</b>", resumo_barras_header_style),
            Paragraph("<b>QTD</b>", resumo_barras_header_style),
            Paragraph("<b>INVESTIMENTO</b>", resumo_barras_header_style),
            "",
        ]
    ]
    for _, row in df_resumo.iterrows():
        data.append(
            [
                Paragraph(str(row[coluna_nome]), resumo_barras_cell_style),
                Paragraph(str(int(row[col_objeto])), resumo_barras_center_style),
                Paragraph(_moeda_abreviada_resumo(row["_INV_AJ"]), resumo_barras_valor_style),
                _barra_resumo(float(row["_INV_AJ"]), max_inv, largura_barra_inv - 16) or "",
            ]
        )

    linha_total = len(data)
    data.append(
        [
            Paragraph("<b>TOTAL</b>", resumo_barras_cell_style),
            Paragraph(f"<b>{int(df_resumo[col_objeto].sum())}</b>", resumo_barras_center_style),
            Paragraph(
                f"<b>{_moeda_abreviada_resumo(df_resumo['_INV_AJ'].sum())}</b>",
                resumo_barras_valor_style,
            ),
            "",
        ]
    )

    tabela = Table(
        data,
        colWidths=[largura_nome, largura_qtd, largura_inv, largura_barra_inv],
    )
    tabela.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), COR_MARCA_TEAL_CLARA),
                ("SPAN", (2, 0), (3, 0)),
                ("BACKGROUND", (0, linha_total), (-1, linha_total), colors.HexColor("#F2F2F2")),
                # Sem grade completa: só uma linha fina abaixo de cada item.
                # Com barras na mesma célula, a grade fechada vira ruído e
                # briga com o desenho.
                ("LINEBELOW", (0, 0), (-1, -1), 0.35, colors.HexColor("#D8D8D8")),
                ("LINEABOVE", (0, linha_total), (-1, linha_total), 0.8, colors.grey),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("ALIGN", (3, 1), (3, -1), "LEFT"),
                ("LEFTPADDING", (0, 0), (-1, -1), 6),
                ("RIGHTPADDING", (0, 0), (-1, -1), 6),
                # A barra encosta um pouco mais no número que ela ilustra.
                ("LEFTPADDING", (3, 1), (3, -1), 10),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ]
        )
    )
    tabela.hAlign = "LEFT"
    return tabela

def gerar_faixa_resumo_eixo_status(df_base, gestao, largura_total=None):
    # Eixo à esquerda, Status à direita, ocupando a largura da página. Antes
    # as duas eram empilhadas numa coluna de 60% da largura útil, o que
    # deixava as laterais vazias na A3 paisagem e empurrava a de Status para
    # o pé da folha.
    largura_total = LARGURA_UTIL if largura_total is None else largura_total
    largura_coluna = (largura_total - 24) / 2

    tabela_eixo = _tabela_resumo_com_barras(
        df_base, gestao, col_eixo, "EIXO_SORT", "EIXO", largura_coluna
    )
    tabela_status = _tabela_resumo_com_barras(
        df_base, gestao, "STATUS_TEXTO", "STATUS_ORDEM", "STATUS", largura_coluna
    )
    if tabela_eixo is None and tabela_status is None:
        return Spacer(1, 0)

    faixa = Table(
        [[tabela_eixo or "", tabela_status or ""]],
        colWidths=[largura_coluna, largura_coluna],
        hAlign="CENTER",
    )
    faixa.setStyle(
        TableStyle(
            [
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING", (0, 0), (0, 0), 0),
                ("LEFTPADDING", (1, 0), (1, 0), 24),
                ("RIGHTPADDING", (0, 0), (-1, -1), 0),
                ("TOPPADDING", (0, 0), (-1, -1), 0),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
            ]
        )
    )
    return faixa

def gerar_tabela_secretaria_fase(df_base, gestao, altura_maxima=None):
    df_filtrado = df_base[df_base["GESTAO"] == gestao]
    df_filtrado = df_filtrado.assign(_INV_AJ=_serie_investimento_ajustado(df_filtrado))
    fases = ["CAPTAÇÃO DE RECURSO", "LICITAÇÃO", "EXECUÇÃO DO OBJETO", "CONCLUÍDA"]
    pares = df_filtrado[["SECRETARIA_LIMPA", "EXECUTOR"]].drop_duplicates().copy()
    pares["_SEC_SORT"] = pares["SECRETARIA_LIMPA"].apply(remover_acentos)
    pares["_EXE_SORT"] = pares["EXECUTOR"].apply(remover_acentos)
    pares = pares.sort_values(["_SEC_SORT", "_EXE_SORT"])

    totais_qtd, totais_invest = {f: 0 for f in fases}, {f: 0.0 for f in fases}
    linhas_base = []

    for _, par in pares.iterrows():
        sec, exe = str(par["SECRETARIA_LIMPA"]), str(par["EXECUTOR"])
        df_sec = df_filtrado[df_filtrado["SECRETARIA_LIMPA"] == sec]
        df_exec = df_sec[df_sec["EXECUTOR"] == exe]
        contagem_objetos = df_exec[col_objeto].value_counts()

        lista_formatada = [
            f"{str(obj)}{formatar_contagem_opcional(cont)}"
            for obj, cont in contagem_objetos.items()
        ]
        texto_obj = ", ".join(lista_formatada)

        sufixo_sec = formatar_contagem_opcional(len(df_sec))
        sufixo_exe = formatar_contagem_opcional(len(df_exec))

        valores_fase = {}
        for f in fases:
            df_temp = df_exec[
                df_exec["FASE_TEXTO"].str.contains(f, case=False, na=False)
            ]
            q, inv = len(df_temp), df_temp["_INV_AJ"].sum()
            totais_qtd[f] += q
            totais_invest[f] += inv
            valores_fase[f] = (q, inv)

        linhas_base.append(
            {
                "sec": sec,
                "exe": exe,
                "sec_label": f"{sec}{sufixo_sec} ({formatar_mi_bi(df_sec['_INV_AJ'].sum())})",
                "exe_label": f"{exe}{sufixo_exe} ({formatar_mi_bi(df_exec['_INV_AJ'].sum())})",
                "texto_obj": texto_obj,
                "valores_fase": valores_fase,
            }
        )

    # Remove automaticamente qualquer fase sem nenhuma ação (0 itens), seja
    # porque não existe na base, seja porque os filtros da interface
    # eliminaram todos os registros daquela fase.
    fases_ativas = [f for f in fases if totais_qtd[f] > 0]

    largura_fixa = (
        SEC_FASE_LARGURA_SECRETARIA + SEC_FASE_LARGURA_EXECUTOR + SEC_FASE_LARGURA_OBJETO
    )
    largura_fases_total = 1.0 - largura_fixa
    largura_por_fase = largura_fases_total / max(len(fases_ativas), 1)
    col_widths = [
        LARGURA_UTIL * SEC_FASE_LARGURA_SECRETARIA,
        LARGURA_UTIL * SEC_FASE_LARGURA_EXECUTOR,
        LARGURA_UTIL * SEC_FASE_LARGURA_OBJETO,
    ] + [LARGURA_UTIL * largura_por_fase for _ in fases_ativas]

    def _montar_tabela(fonte_tam, padding_vertical):
        # Estilos remontados a cada tentativa de autofit, já que fonte e
        # leading mudam conforme o tamanho testado.
        estilo_header = ParagraphStyle(
            "sec_fase_header_dyn",
            fontName=RESUMO_FONTE_NEGRITO,
            fontSize=fonte_tam,
            leading=fonte_tam + 3,
            alignment=TA_CENTER,
        )
        estilo_cell = ParagraphStyle(
            "sec_fase_cell_dyn",
            fontName=RESUMO_FONTE_NORMAL,
            fontSize=fonte_tam,
            leading=fonte_tam + 3,
            alignment=TA_LEFT,
        )
        estilo_cell_center = ParagraphStyle(
            "sec_fase_cell_center_dyn", parent=estilo_cell, alignment=TA_CENTER
        )
        # Mesma cor exata do cabeçalho (fundo teal claro + texto teal
        # escuro) na linha TOTAL, pra ficar idêntica — sem usar um tom de
        # verde separado.
        estilo_total = ParagraphStyle(
            "sec_fase_total_dyn", parent=estilo_header, textColor=COR_MARCA_TEAL_ESCURO
        )

        data = [
            [
                Paragraph("<b>SECRETARIA/ ÓRGÃO</b>", estilo_header),
                Paragraph("<b>EXECUTOR</b>", estilo_header),
                Paragraph("<b>OBJETO</b>", estilo_header),
            ]
            + [Paragraph(f"<b>{f}</b>", estilo_header) for f in fases_ativas]
        ]

        spans, controle_sec, controle_exe = [], {}, {}
        linha_idx = 1

        for item in linhas_base:
            linha = [
                Paragraph(item["sec_label"], estilo_cell),
                Paragraph(item["exe_label"], estilo_cell),
                Paragraph(item["texto_obj"], estilo_cell),
            ]

            for f in fases_ativas:
                q, inv = item["valores_fase"][f]
                sufixo_fase = formatar_contagem_opcional(q) if q > 0 else ""
                linha.append(
                    Paragraph(
                        "-"
                        if q == 0
                        else f"{sufixo_fase.strip()} {formatar_mi_bi(inv)}".strip(),
                        estilo_cell_center,
                    )
                )

            data.append(linha)
            controle_sec.setdefault(item["sec"], []).append(linha_idx)
            controle_exe.setdefault((item["sec"], item["exe"]), []).append(linha_idx)
            linha_idx += 1

        total_l = [
            Paragraph("<b>TOTAL</b>", estilo_total),
            Paragraph("", estilo_cell),
            Paragraph("", estilo_cell),
        ]
        for f in fases_ativas:
            total_l.append(
                Paragraph(
                    f"<b>({totais_qtd[f]}) {formatar_mi_bi(totais_invest[f])}</b>",
                    estilo_total,
                )
            )
        data.append(total_l)

        for l in controle_sec.values():
            if len(l) > 1:
                spans.append(("SPAN", (0, l[0]), (0, l[-1])))
        for l in controle_exe.values():
            if len(l) > 1:
                spans.append(("SPAN", (1, l[0]), (1, l[-1])))

        tabela_montada = Table(data, colWidths=col_widths, repeatRows=1)
        tabela_montada.setStyle(
            TableStyle(
                [
                    ("GRID", (0, 0), (-1, -1), 0.4, colors.grey),
                    ("BACKGROUND", (0, 0), (-1, 0), COR_MARCA_TEAL_CLARA),
                    ("BACKGROUND", (0, -1), (-1, -1), COR_MARCA_TEAL_CLARA),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ("TOPPADDING", (0, 0), (-1, -1), padding_vertical),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), padding_vertical),
                ]
                + spans
            )
        )
        tabela_montada.hAlign = "CENTER"
        return tabela_montada

    fonte_tam = SEC_FASE_FONTE_TAM
    padding_vertical = SEC_FASE_PADDING_VERTICAL
    tabela = _montar_tabela(fonte_tam, padding_vertical)

    # Autofit: garante que TODO o conteúdo (cabeçalho + linhas + TOTAL) caiba
    # em uma única página, mesmo com bases maiores. Primeiro reduz o
    # preenchimento vertical das células; se ainda não for suficiente,
    # reduz também a fonte — sempre até os pisos mínimos configurados acima.
    #
    # Otimização: em vez de testar padding de 1 em 1 (cada teste reconstrói
    # a tabela inteira, caro), calcula direto quanto precisa reduzir — a
    # altura muda de forma exatamente linear com o padding (2pt por unidade,
    # por linha), então uma única conta substitui vários reconstruções. Já a
    # fonte muda a quebra de texto de forma não-linear (não dá pra calcular
    # direto), mas ainda dá pra usar busca binária em vez de ir de 1 em 1.
    if altura_maxima is not None:
        _, altura_tabela = tabela.wrap(LARGURA_UTIL, 100000)
        if altura_tabela > altura_maxima and padding_vertical > SEC_FASE_PADDING_MINIMA:
            num_linhas_tabela = len(linhas_base) + 2  # cabeçalho + TOTAL
            excesso = altura_tabela - altura_maxima
            reducao = -(-excesso // (2 * num_linhas_tabela))  # ceil
            padding_vertical = max(SEC_FASE_PADDING_MINIMA, padding_vertical - reducao)
            tabela = _montar_tabela(fonte_tam, padding_vertical)
            _, altura_tabela = tabela.wrap(LARGURA_UTIL, 100000)
            # Margem de segurança: se a conta linear não bastou (raro, só
            # quando o padding já bateu no piso), completa com um ajuste
            # fino de 1 em 1 até o piso.
            while altura_tabela > altura_maxima and padding_vertical > SEC_FASE_PADDING_MINIMA:
                padding_vertical -= 1
                tabela = _montar_tabela(fonte_tam, padding_vertical)
                _, altura_tabela = tabela.wrap(LARGURA_UTIL, 100000)

        if altura_tabela > altura_maxima and fonte_tam > SEC_FASE_FONTE_MINIMA:
            fonte_min_testada = SEC_FASE_FONTE_MINIMA
            fonte_max_testada = fonte_tam
            tabela_min = _montar_tabela(fonte_min_testada, padding_vertical)
            _, altura_no_piso = tabela_min.wrap(LARGURA_UTIL, 100000)
            if altura_no_piso > altura_maxima:
                # Nem no piso de fonte cabe — fica no menor tamanho possível.
                fonte_tam, tabela, altura_tabela = fonte_min_testada, tabela_min, altura_no_piso
            else:
                # Busca binária: a altura diminui (ou mantém) conforme a
                # fonte diminui, então dá pra descartar metade do intervalo
                # a cada tentativa em vez de testar fonte por fonte.
                melhor_fonte, melhor_tabela, melhor_altura = (
                    fonte_min_testada, tabela_min, altura_no_piso
                )
                baixo, alto = fonte_min_testada, fonte_max_testada
                while baixo < alto:
                    meio = (baixo + alto + 1) // 2
                    tabela_meio = _montar_tabela(meio, padding_vertical)
                    _, altura_meio = tabela_meio.wrap(LARGURA_UTIL, 100000)
                    if altura_meio <= altura_maxima:
                        melhor_fonte, melhor_tabela, melhor_altura = meio, tabela_meio, altura_meio
                        baixo = meio
                    else:
                        alto = meio - 1
                fonte_tam, tabela, altura_tabela = melhor_fonte, melhor_tabela, melhor_altura

    return tabela

# =====================================================
# 5B. GRÁFICOS DO PAINEL GERAL
# =====================================================

COR_AZUL = cor_grafico("#4E92BA")
COR_VERMELHO = cor_grafico("#BB6060")
COR_VERDE = cor_grafico("#49925C")
COR_AMARELO = cor_grafico("#BC9E2C")

# Ordem e cores padrão das FASES, usadas tanto no gráfico PANORAMA GERAL DAS
# FASES quanto na nova página PANORAMA POR SECRETARIA | EXECUTOR — mantém as
# duas visualizações com a mesma paleta.
ORDEM_FASES = ["CAPTAÇÃO DE RECURSO", "LICITAÇÃO", "EXECUÇÃO DO OBJETO", "CONCLUÍDA"]
CORES_FASE = {
    "CAPTAÇÃO DE RECURSO": COR_VERMELHO,
    "LICITAÇÃO": COR_AMARELO,
    "EXECUÇÃO DO OBJETO": COR_VERDE,
    "CONCLUÍDA": COR_AZUL,
}

GRUPOS_STATUS_POR_FASE = {
    # "Concluída" (azul) — mesma cor da fatia CONCLUÍDA no gráfico de FASE
    "CONCLUÍDA": (3, COR_AZUL),
    "INAUGURADA": (3, COR_AZUL),
    # "Execução do Objeto" (verde)
    "AGUARDANDO ORDEM DE SERVIÇO": (2, COR_VERDE),
    "ANDAMENTO": (2, COR_VERDE),
    # "Licitação" (amarelo)
    "LICITAÇÃO /CONTRATAÇÃO": (1, COR_AMARELO),
    # Todo o resto — incluindo "À LICITAR" — cai em "Captação de Recurso"
    # (vermelho), que é o valor padrão devolvido pelo .get() abaixo, não
    # precisa listar aqui.
}

def _cor_status_por_grupo_fase(status_texto):
    # Esquema de cores ORGANIZADO por qual FASE aquele STATUS representa na
    # prática — não é mais um gradiente numérico solto, é a MESMA cor da
    # fase correspondente, pra ler os dois gráficos (FASE e STATUS) com a
    # mesma linguagem visual: um status vermelho no gráfico da direita
    # sempre corresponde a uma ação que ainda está na fase de Captação de
    # Recurso no gráfico da esquerda, e assim por diante.
    #   Concluída (azul): Concluída, Inaugurada
    #   Execução do Objeto (verde): Aguardando Ordem de Serviço, Andamento
    #   Licitação (amarelo): Licitação/Contratação
    #   Captação de Recurso (vermelho): À Licitar e todos os demais status
    return GRUPOS_STATUS_POR_FASE.get(str(status_texto).strip().upper(), (0, COR_VERMELHO))

def _grafico_pizza_igual_com_legenda(titulo, categorias_cores, diametro=130, max_linhas_por_coluna=8, largura_coluna_legenda=175):
    # Gráfico de pizza DIDÁTICO: todas as fatias do MESMO tamanho (não
    # representam proporção de dados — só ilustram "essas são as
    # categorias e essa é a cor de cada uma"), com legenda ao lado. Usado
    # na página de Observação pra explicar visualmente FASE e STATUS.
    #
    # A legenda quebra em MAIS DE UMA COLUNA quando passa de
    # "max_linhas_por_coluna" itens — o STATUS pode ter bem mais valores
    # distintos que a FASE (fixa em até 4), e sem isso a legenda cresceria
    # sem limite conforme o filtro trouxesse mais categorias, estourando a
    # página. Com colunas, a altura fica sempre travada no mesmo teto,
    # não importa se são 5 ou 15 status diferentes.
    n = len(categorias_cores)
    altura_titulo = 20
    fonte_legenda = 12
    altura_linha_legenda = 17

    if n == 0:
        d = Drawing(diametro, altura_titulo + 20)
        d.add(String(0, altura_titulo + 2, titulo, fontName=FONTE_PADRAO_NEGRITO, fontSize=13, fillColor=colors.HexColor("#333333")))
        d.add(String(0, 2, "Sem dados no filtro atual", fontName=FONTE_PADRAO, fontSize=10, fillColor=colors.grey))
        return d

    n_colunas = -(-n // max_linhas_por_coluna)  # arredonda pra cima
    linhas_por_coluna = -(-n // n_colunas)
    altura_legenda = max(linhas_por_coluna * altura_linha_legenda, diametro)
    altura_total = altura_titulo + altura_legenda
    largura_total = diametro + 16 + n_colunas * largura_coluna_legenda

    d = Drawing(largura_total, altura_total)
    d.add(String(0, altura_total - 16, titulo, fontName=FONTE_PADRAO_NEGRITO, fontSize=14, fillColor=colors.HexColor("#333333")))

    y_topo_area = altura_total - altura_titulo
    grafico = Pie()
    grafico.x = 0
    grafico.y = y_topo_area - altura_legenda + (altura_legenda - diametro) / 2
    grafico.width = diametro
    grafico.height = diametro
    grafico.data = [1] * n
    grafico.labels = [""] * n
    grafico.simpleLabels = 0
    grafico.slices.strokeWidth = 1.4
    grafico.slices.strokeColor = colors.white
    for idx, (_, cor) in enumerate(categorias_cores):
        grafico.slices[idx].fillColor = cor
    d.add(grafico)

    x_base_legenda = diametro + 16
    for idx, (rotulo, cor) in enumerate(categorias_cores):
        coluna = idx // linhas_por_coluna
        linha = idx % linhas_por_coluna
        x_legenda = x_base_legenda + coluna * largura_coluna_legenda
        y = y_topo_area - 12 - linha * altura_linha_legenda
        d.add(Rect(x_legenda, y - 8, 11, 11, fillColor=cor, strokeColor=None))
        d.add(String(x_legenda + 16, y - 7, rotulo, fontName=FONTE_PADRAO, fontSize=fonte_legenda, fillColor=colors.HexColor("#333333")))
    return d

fase_status_header_style = ParagraphStyle(
    "fase_status_header", fontName=FONTE_PADRAO_NEGRITO, fontSize=15, leading=18, alignment=TA_LEFT,
)
fase_status_cell_style = ParagraphStyle(
    "fase_status_cell", fontName=FONTE_PADRAO, fontSize=15, leading=18, alignment=TA_LEFT,
)

def _tabela_fase_status(df_base, max_linhas_por_bloco=8):
    # Tabela de referência (SEM nenhum elemento gráfico) mostrando o
    # agrupamento de cada STATUS em relação à FASE correspondente — mesma
    # regra de GRUPOS_STATUS_POR_FASE usada em outras partes do relatório,
    # só que aqui apresentada como texto/tabela em vez de cor. Dinâmica:
    # só lista os status que aparecem de fato no filtro atual.
    NOME_GRUPO_FASE = {
        0: "CAPTAÇÃO DE RECURSO",
        1: "LICITAÇÃO",
        2: "EXECUÇÃO DO OBJETO",
        3: "CONCLUÍDA",
    }
    status_presentes = (
        df_base[["STATUS_ORDEM", "STATUS_TEXTO"]]
        .drop_duplicates()
        .sort_values("STATUS_ORDEM")["STATUS_TEXTO"]
        .tolist()
    )
    if not status_presentes:
        return None

    status_com_grupo = sorted(
        ((_cor_status_por_grupo_fase(s)[0], s) for s in status_presentes),
        key=lambda item: item[0],
    )
    grupos = [
        (NOME_GRUPO_FASE.get(grupo, ""), [nome for _, nome in itens_grupo])
        for grupo, itens_grupo in itertools.groupby(status_com_grupo, key=lambda item: item[0])
    ]
    total_linhas = sum(len(nomes) for _, nomes in grupos)

    # Com poucos status (linha só cabe tudo), NÃO divide em blocos — só
    # divide em dois blocos lado a lado quando passar do limite, pra
    # altura da tabela não crescer sem controle conforme mais status
    # aparecem no filtro. Nunca quebra um grupo de fase no meio: cada
    # grupo inteiro vai pra um bloco só.
    if total_linhas <= max_linhas_por_bloco:
        blocos = [grupos]
    else:
        metade = total_linhas / 2
        bloco_esq, bloco_dir, acumulado = [], [], 0
        for fase_nome, nomes in grupos:
            (bloco_esq if acumulado < metade else bloco_dir).append((fase_nome, nomes))
            acumulado += len(nomes)
        blocos = [b for b in (bloco_esq, bloco_dir) if b]

    def _montar_bloco(grupos_bloco):
        data = [
            [
                Paragraph("<b>FASE</b>", fase_status_header_style),
                Paragraph("<b>STATUS</b>", fase_status_header_style),
            ]
        ]
        spans = []
        linha = 1
        for fase_nome, nomes_do_grupo in grupos_bloco:
            inicio_linha = linha
            for i, nome in enumerate(nomes_do_grupo):
                data.append(
                    [
                        Paragraph(fase_nome if i == 0 else "", fase_status_cell_style),
                        Paragraph(nome, fase_status_cell_style),
                    ]
                )
                linha += 1
            # Uma FASE com mais de um STATUS agrupado embaixo dela mescla a
            # célula da coluna FASE verticalmente, deixando claro
            # visualmente que aqueles status pertencem todos à mesma fase.
            if len(nomes_do_grupo) > 1:
                spans.append(("SPAN", (0, inicio_linha), (0, inicio_linha + len(nomes_do_grupo) - 1)))

        tabela_bloco = Table(data, colWidths=[160, 215])
        estilo_bloco = [
            ("GRID", (0, 0), (-1, -1), 0.6, colors.HexColor("#CCCCCC")),
            ("BACKGROUND", (0, 0), (-1, 0), COR_MARCA_TEAL_CLARA),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ("LEFTPADDING", (0, 0), (-1, -1), 10),
        ]
        estilo_bloco.extend(spans)
        tabela_bloco.setStyle(TableStyle(estilo_bloco))
        return tabela_bloco

    tabelas_blocos = [_montar_bloco(g) for g in blocos]
    if len(tabelas_blocos) == 1:
        tabela_final = tabelas_blocos[0]
    else:
        espaco_entre_blocos = 30
        larguras_blocos = [t.wrap(600, 2000)[0] for t in tabelas_blocos]
        tabela_final = Table(
            [tabelas_blocos],
            colWidths=[larguras_blocos[0] + espaco_entre_blocos] + larguras_blocos[1:],
        )
        tabela_final.setStyle(
            TableStyle(
                [
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("LEFTPADDING", (0, 0), (0, -1), 0),
                    ("RIGHTPADDING", (0, 0), (0, -1), espaco_entre_blocos),
                    ("LEFTPADDING", (1, 0), (1, -1), 0),
                    ("RIGHTPADDING", (1, 0), (1, -1), 0),
                    ("TOPPADDING", (0, 0), (-1, -1), 0),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
                ]
            )
        )
    tabela_final.hAlign = "CENTER"
    return tabela_final

def gerar_graficos_observacao(df_base):
    # Monta o gráfico de pizza da FASE (mantido) e, ao lado, a tabela de
    # referência FASE/STATUS (substitui o gráfico de pizza do STATUS —
    # decisão de não usar nenhum elemento gráfico pra representar status).
    fases_presentes = (
        df_base[["FASE_ORDEM", "FASE_TEXTO"]]
        .drop_duplicates()
        .sort_values("FASE_ORDEM")["FASE_TEXTO"]
        .tolist()
    )
    fase_cores = [(f, CORES_FASE.get(f, colors.grey)) for f in fases_presentes]
    grafico_fase = _grafico_pizza_igual_com_legenda("FASE", fase_cores, diametro=175, max_linhas_por_coluna=8, largura_coluna_legenda=180)

    tabela_status = _tabela_fase_status(df_base)
    if tabela_status is None:
        return grafico_fase

    # O par (gráfico de FASE + tabela de STATUS) fica CENTRALIZADO na
    # página: colunas do tamanho real de cada elemento (não esticadas pra
    # ocupar a largura toda), com um espaço fixo entre os dois, e o
    # conjunto inteiro centralizado horizontalmente — mantendo os dois
    # lado a lado.
    largura_tabela_status, _ = tabela_status.wrap(600, 2000)
    espaco_entre_elementos = 40
    tabela = Table(
        [[grafico_fase, tabela_status]],
        colWidths=[grafico_fase.width, largura_tabela_status],
    )
    tabela.setStyle(
        TableStyle(
            [
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                ("LEFTPADDING", (0, 0), (0, -1), 0),
                ("RIGHTPADDING", (0, 0), (0, -1), espaco_entre_elementos),
                ("LEFTPADDING", (1, 0), (1, -1), 0),
                ("RIGHTPADDING", (1, 0), (1, -1), 0),
                ("TOPPADDING", (0, 0), (-1, -1), 0),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
            ]
        )
    )
    tabela.hAlign = "CENTER"
    # tabela.hAlign = "CENTER" sozinho NÃO funciona de forma confiável
    # aqui: essa tabela acaba ficando aninhada dentro da célula de outra
    # tabela (a "tabela_observacoes" que envolve toda a página) — e hAlign
    # de um flowable só é respeitado quando ele está direto na história
    # principal do documento, não quando está dentro da célula de uma
    # tabela. Por isso envolvemos num wrapper com ALIGN CENTER definido
    # via TableStyle (esse sim funciona em qualquer nível de aninhamento)
    # ocupando a largura toda disponível pro conteúdo da página.
    wrapper_centralizado = Table([[tabela]], colWidths=[LARGURA_UTIL * 0.98])
    wrapper_centralizado.setStyle(
        TableStyle(
            [
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("LEFTPADDING", (0, 0), (-1, -1), 0),
                ("RIGHTPADDING", (0, 0), (-1, -1), 0),
                ("TOPPADDING", (0, 0), (-1, -1), 0),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
            ]
        )
    )
    return wrapper_centralizado

def desenho_sem_dados(largura, altura, mensagem="Sem investimento no filtro atual"):
    # Placeholder exibido no lugar de um gráfico de pizza quando, depois de
    # aplicado o filtro, sobra quantidade de ações mas o investimento total
    # soma zero (célula em branco/zero na planilha) — sem isso, o ReportLab
    # tenta dividir pelo total (zero) pra calcular o tamanho de cada fatia e
    # trava com "float division by zero".
    desenho = Drawing(largura, altura)
    desenho.add(
        String(
            largura / 2, altura / 2, mensagem,
            fontName=FONTE_PADRAO, fontSize=12, fillColor=colors.grey,
            textAnchor="middle",
        )
    )
    return desenho

def desenhar_legenda_topo(desenho, itens, y, centro_x, espacamento=18, fonte_tam=8, raio_bolinha=3.5):
    # Posiciona a legenda em uma única linha horizontal, centralizada em
    # centro_x, na ordem em que os itens forem passados (bolinha + nome,
    # sem valor).
    larguras = [
        raio_bolinha * 2 + 4 + pdfmetrics.stringWidth(nome, FONTE_PADRAO_NEGRITO, fonte_tam)
        for nome, _ in itens
    ]
    largura_total = sum(larguras) + espacamento * (len(itens) - 1)
    x = centro_x - largura_total / 2
    for (nome, cor), largura_item in zip(itens, larguras):
        centro_y = y + fonte_tam * 0.32
        desenho.add(Circle(x + raio_bolinha, centro_y, raio_bolinha, fillColor=cor, strokeColor=None))
        desenho.add(
            String(
                x + raio_bolinha * 2 + 4, y, nome,
                fontName=FONTE_PADRAO_NEGRITO, fontSize=fonte_tam, fillColor=colors.black,
            )
        )
        x += largura_item + espacamento

# Raio do canto no topo das colunas verticais. Vale para o PDF e tem o
# equivalente no CSS do dashboard (.grafico-barra).
RAIO_CANTO_COLUNA = 4.0

def arredondar_topo_barras(no, cores_barras, y_base):
    # Percorre a árvore de formas que o VerticalBarChart produz e troca cada
    # retângulo de barra por uma versão de topo arredondado. O ReportLab não
    # tem opção de canto arredondado nas barras do gráfico, então o caminho
    # é desenhar o gráfico (grafico.draw() devolve um Group de formas) e
    # editar as formas antes de colocá-las no Drawing.
    #
    # Uma barra é reconhecida por dois sinais juntos: a cor de preenchimento
    # é uma das que foram atribuídas às barras E a base está na linha do
    # eixo. Só a cor não bastaria — a legenda e o fundo podem repetir o
    # mesmo tom.
    conteudo = getattr(no, "contents", None)
    if not conteudo:
        return
    novo_conteudo = []
    trocou = False
    for filho in conteudo:
        eh_barra = (
            isinstance(filho, Rect)
            and any(filho.fillColor == cor for cor in cores_barras)
            and abs(filho.y - y_base) < 0.5
        )
        if eh_barra:
            novo_conteudo.extend(
                formas_barra_arredondada(
                    filho.x, filho.y, filho.width, filho.height, filho.fillColor,
                    ponta_inicio=False, ponta_fim=True, horizontal=False,
                    # Canto arredondado, não semicírculo: com o raio padrão
                    # (metade da espessura) o topo da coluna virava uma
                    # cúpula que alongava visualmente a barra e atrapalhava
                    # a comparação de alturas entre trimestres. O raio fixo
                    # suaviza o canto sem mexer na leitura.
                    raio=RAIO_CANTO_COLUNA,
                )
            )
            trocou = True
        else:
            arredondar_topo_barras(filho, cores_barras, y_base)
            novo_conteudo.append(filho)
    # Substitui NA PRÓPRIA LISTA (fatia) em vez de reatribuir o atributo
    # "contents": o Group valida o que é atribuído nele e recusa os widgets
    # de rótulo que o gráfico mantém ali dentro, que precisam continuar
    # onde estão para saberem a que gráfico pertencem.
    if trocou:
        conteudo[:] = novo_conteudo

class VerticalBarChartArredondado(VerticalBarChart):
    # VerticalBarChart com o topo das barras arredondado. A rodada de
    # arredondamento acontece dentro de makeBars, e não sobre o gráfico já
    # desenhado, porque os rótulos de valor em cima das barras são widgets
    # que precisam continuar pendurados no gráfico para conseguirem se
    # desenhar — destacá-los quebra a renderização.
    def makeBars(self):
        grupo = VerticalBarChart.makeBars(self)
        arredondar_topo_barras(grupo, self._cores_barras_arredondadas, self.y)
        return grupo

def gerar_grafico_prazo(df_base):
    contagem = df_base["PRAZO_TRIMESTRE"].value_counts().reset_index()
    contagem.columns = ["TRIMESTRE", "QTD"]
    contagem["ORDEM"] = contagem["TRIMESTRE"].apply(ordenar_trimestre)
    contagem = contagem.sort_values("ORDEM")

    categorias = contagem["TRIMESTRE"].tolist()
    valores = contagem["QTD"].tolist()

    desenho = Drawing(520, 240)
    grafico = VerticalBarChartArredondado()
    grafico.x = 30
    grafico.y = 55
    grafico.height = 145
    grafico.width = 475
    grafico.data = [valores]
    grafico.categoryAxis.categoryNames = categorias
    grafico.categoryAxis.labels.fontName = FONTE_PADRAO_NEGRITO
    grafico.categoryAxis.labels.fontSize = 13
    grafico.categoryAxis.labels.angle = 35
    grafico.categoryAxis.labels.boxAnchor = "ne"
    grafico.categoryAxis.labels.dx = 2
    grafico.categoryAxis.labels.dy = -2
    grafico.valueAxis.valueMin = 0
    grafico.valueAxis.visible = False
    grafico.bars.strokeColor = None
    # Trava a LARGURA da barra em um valor fixo (não deixa esticar quando há
    # poucas categorias, como no caso de só existir "A definir"). O
    # ESPAÇAMENTO entre barras (groupSpacing), porém, é calculado aqui em
    # função da largura do gráfico e da quantidade de categorias — o
    # ReportLab, no modo "bg" (largura e espaçamento fixos), usa metade do
    # groupSpacing como margem à esquerda de cada barra dentro da fatia da
    # categoria; se o groupSpacing for um valor fixo pequeno (ex: 12) e a
    # fatia da categoria for bem mais larga que isso (poucas categorias),
    # a barra fica encostada à esquerda da fatia em vez de centralizada.
    # Calculando o groupSpacing como "largura da fatia − largura da barra",
    # a margem (metade do groupSpacing) fica igual dos dois lados da barra,
    # centralizando-a corretamente, seja com 1 ou com várias categorias.
    # barWidth reduzido (de 22 para 14) para abrir mais espaço entre as
    # barras — o groupSpacing cresce automaticamente já que é calculado a
    # partir da diferença entre a largura da fatia e a largura da barra.
    grafico.useAbsolute = "bg"
    grafico.barWidth = 14
    largura_categoria = grafico.width / max(len(categorias), 1)
    grafico.groupSpacing = max(largura_categoria - grafico.barWidth, 0)
    grafico.barLabelFormat = "%d"
    grafico.barLabels.fontName = FONTE_BAHNSCHRIFT
    grafico.barLabels.fontSize = 12
    grafico.barLabels.dy = 10

    for idx, categoria in enumerate(categorias):
        cor = COR_VERMELHO if categoria == "A definir" else COR_AZUL
        grafico.bars[(0, idx)].fillColor = cor

    grafico._cores_barras_arredondadas = [COR_AZUL, COR_VERMELHO]
    desenho.add(grafico)
    return desenho

def gerar_grafico_financeiro(df_base):
    categorias_valores = [
        ("VALOR CONTRATADO", df_base[col_valor_contratado].sum(), COR_VERDE),
        ("OGU", df_base[col_apoiado_ajustado].sum(), COR_AZUL),
        ("FINANCIAMENTO", df_base[col_financiamento_ajustado].sum(), COR_VERMELHO),
        (
            "RECURSO ESTADUAL",
            df_base[col_contrapartida_ajustado].sum()
            + df_base[col_complementar_ajustado].sum(),
            COR_AMARELO,
        ),
    ]
    categorias_valores = [
        (nome, valor, cor) for nome, valor, cor in categorias_valores if valor > 0
    ]
    if not categorias_valores:
        return desenho_sem_dados(480, 260)
    total = sum(valor for _, valor, _ in categorias_valores)

    desenho = Drawing(480, 260)
    grafico = Pie()
    grafico.x, grafico.y = 165, 35
    grafico.width, grafico.height = 150, 150
    grafico.data = [valor for _, valor, _ in categorias_valores]
    grafico.labels = [
        f"{formatar_mi_bi(valor).replace('&nbsp;', ' ')} ({valor / total * 100:.1f}%)"
        for _, valor, _ in categorias_valores
    ]
    grafico.slices.strokeWidth = 1
    grafico.slices.strokeColor = colors.white
    # Rótulos externos com linha de chamada (evita sobreposição automaticamente).
    grafico.simpleLabels = 0
    grafico.sideLabels = 1
    grafico.sideLabelsOffset = 0.15
    # Ativa o ajuste automático de sobreposição do ReportLab: quando dois
    # rótulos laterais ficam próximos demais (fatias pequenas, como no caso
    # do amarelo/verde), ele afasta um do outro verticalmente.
    grafico.checkLabelOverlap = 1

    for idx, (_, _, cor) in enumerate(categorias_valores):
        grafico.slices[idx].fillColor = cor
        grafico.slices[idx].fontName = FONTE_BAHNSCHRIFT
        grafico.slices[idx].fontSize = 16
        # Linha de chamada (pointer) do rótulo: mais grossa e na cor da fatia
        # (o valor de cor aqui não é respeitado pelo ReportLab no modo
        # sideLabels — corrigido manualmente logo abaixo).
        grafico.slices[idx].label_pointer_strokeColor = cor
        grafico.slices[idx].label_pointer_strokeWidth = 1.5

    # O ReportLab tem uma falha no modo sideLabels: todas as linhas de
    # chamada acabam usando a cor/espessura da ÚLTIMA fatia processada (uma
    # variável interna reaproveitada por engano), em vez da cor de cada
    # fatia correspondente. Por isso desenhamos o gráfico manualmente aqui
    # (grafico.draw()) e corrigimos a cor/espessura de cada linha, na ordem
    # em que elas são geradas internamente (duas linhas por fatia, na mesma
    # ordem de categorias_valores).
    elemento_grafico = grafico.draw()
    linhas_chamada = [e for e in elemento_grafico.contents if isinstance(e, Line)]
    for idx_linha, linha in enumerate(linhas_chamada):
        cor_fatia = categorias_valores[idx_linha // 2][2]
        linha.strokeColor = cor_fatia
        linha.strokeWidth = 1.5

    desenho.add(elemento_grafico)

    # Legenda horizontal centralizada no topo do gráfico, na mesma ordem das
    # categorias (CONTRATADO, OGU, FINANCIAMENTO, RECURSO ESTADUAL). OGU,
    # FINANCIAMENTO e RECURSO ESTADUAL levam "**" — são valores de PREVISÃO
    # ORÇAMENTÁRIA (ainda não contratados), diferente de VALOR CONTRATADO,
    # que é firme. A explicação por extenso da marcação ("** Previsão
    # Orçamentária") fica no título do gráfico, não aqui no desenho.
    NOMES_PREVISAO_ORCAMENTARIA = {"OGU", "FINANCIAMENTO", "RECURSO ESTADUAL"}
    itens_legenda = [
        (f"{nome}**" if nome in NOMES_PREVISAO_ORCAMENTARIA else nome, cor)
        for nome, _, cor in categorias_valores
    ]
    desenhar_legenda_topo(desenho, itens_legenda, y=245, centro_x=240, espacamento=20, fonte_tam=12)

    return desenho

def gerar_grafico_panorama(df_base):
    df_base = df_base.assign(_INV_AJ=_serie_investimento_ajustado(df_base))
    resumo = (
        df_base.groupby(["FASE_ORDEM", "FASE_TEXTO"])
        .agg(QTD=(col_objeto, "count"), VALOR=("_INV_AJ", "sum"))
        .reset_index()
        .sort_values("FASE_ORDEM")
    )
    # Filtra por QTD>0 (fase precisa ter ao menos uma ação) E por VALOR>0
    # (o tamanho da fatia é o investimento — uma fase com investimento zero
    # deixaria o ReportLab dividir pelo total zero e travar com
    # "float division by zero").
    resumo = resumo[(resumo["QTD"] > 0) & (resumo["VALOR"] > 0)]
    if resumo.empty:
        return desenho_sem_dados(480, 235)

    desenho = Drawing(480, 235)
    grafico = Pie()
    grafico.x, grafico.y = 165, 30
    grafico.width, grafico.height = 150, 150
    # O tamanho de cada fatia representa o INVESTIMENTO (a grandeza
    # principal do gráfico); a quantidade de ações só acompanha o
    # investimento no rótulo de dados — não influencia o tamanho da fatia.
    grafico.data = resumo["VALOR"].tolist()
    grafico.labels = [
        f"({int(qtd)}) {formatar_mi_bi(valor).replace('&nbsp;', ' ')}"
        for qtd, valor in zip(resumo["QTD"], resumo["VALOR"])
    ]
    grafico.slices.strokeWidth = 1
    grafico.slices.strokeColor = colors.white
    # Rótulos externos com linha de chamada e ajuste automático de
    # sobreposição — mesma configuração usada no gráfico Resumo Financeiro.
    grafico.simpleLabels = 0
    grafico.sideLabels = 1
    grafico.sideLabelsOffset = 0.15
    grafico.checkLabelOverlap = 1

    cores_slices = [CORES_FASE.get(fase, colors.grey) for fase in resumo["FASE_TEXTO"]]
    for idx, cor in enumerate(cores_slices):
        grafico.slices[idx].fillColor = cor
        grafico.slices[idx].fontName = FONTE_BAHNSCHRIFT
        grafico.slices[idx].fontSize = 14
        # Linha de chamada do rótulo de dados — mesma configuração usada no
        # gráfico Resumo Financeiro (mais grossa, na cor da fatia).
        grafico.slices[idx].label_pointer_strokeColor = cor
        grafico.slices[idx].label_pointer_strokeWidth = 1.5

    # Mesma correção da falha do ReportLab usada no gráfico Resumo
    # Financeiro (todas as linhas saem com a cor da ÚLTIMA fatia processada
    # por padrão): desenha o gráfico e recolore cada linha na ordem em que
    # é gerada internamente. No modo sideLabels o ReportLab desenha DUAS
    # linhas por fatia (um "cotovelo" de dois segmentos), por isso o índice
    # é dividido por 2 — na mesma ordem de "cores_slices".
    elemento_grafico = grafico.draw()
    linhas_chamada = [e for e in elemento_grafico.contents if isinstance(e, Line)]
    for idx_linha, linha in enumerate(linhas_chamada):
        cor_fatia = cores_slices[idx_linha // 2]
        linha.strokeColor = cor_fatia
        linha.strokeWidth = 1.5

    desenho.add(elemento_grafico)

    # Legenda horizontal no topo (mesmo padrão do gráfico RESUMO FINANCEIRO):
    # só o nome da fase, sem o investimento — o investimento já aparece no
    # rótulo de dados de cada fatia, junto da quantidade.
    itens_legenda = [
        (linha["FASE_TEXTO"], CORES_FASE.get(linha["FASE_TEXTO"], colors.grey))
        for _, linha in resumo.iterrows()
    ]
    desenhar_legenda_topo(desenho, itens_legenda, y=220, centro_x=240, espacamento=20, fonte_tam=12)

    return desenho

def gerar_grafico_termo_compromisso(df_base):
    cores_situacao = {"SIM": COR_VERDE, "NÃO": COR_VERMELHO}
    df_base = df_base.assign(_INV_AJ=_serie_investimento_ajustado(df_base))
    resumo = (
        df_base.groupby("SINALIZACAO_TC")
        .agg(QTD=(col_objeto, "count"), VALOR=("_INV_AJ", "sum"))
        .reindex(["SIM", "NÃO"])
        .fillna(0)
        .reset_index()
    )
    resumo.columns = ["SITUACAO", "QTD", "VALOR"]
    # Filtra por QTD>0 E por VALOR>0 (o tamanho da fatia é o investimento —
    # uma situação com investimento zero deixaria o ReportLab dividir pelo
    # total zero e travar com "float division by zero").
    resumo = resumo[(resumo["QTD"] > 0) & (resumo["VALOR"] > 0)]
    if resumo.empty:
        return desenho_sem_dados(480, 235)

    desenho = Drawing(480, 235)
    grafico = Pie()
    grafico.x, grafico.y = 165, 30
    grafico.width, grafico.height = 150, 150
    # O tamanho de cada fatia representa o INVESTIMENTO (a grandeza
    # principal do gráfico); a quantidade só acompanha o investimento no
    # rótulo de dados — não influencia o tamanho da fatia.
    grafico.data = resumo["VALOR"].tolist()
    grafico.labels = [
        f"({int(qtd)}) {formatar_mi_bi(valor).replace('&nbsp;', ' ')}"
        for qtd, valor in zip(resumo["QTD"], resumo["VALOR"])
    ]
    grafico.slices.strokeWidth = 1
    grafico.slices.strokeColor = colors.white
    # Rótulos externos com linha de chamada e ajuste automático de
    # sobreposição — mesma configuração usada no gráfico Resumo Financeiro.
    grafico.simpleLabels = 0
    grafico.sideLabels = 1
    grafico.sideLabelsOffset = 0.15
    grafico.checkLabelOverlap = 1

    cores_slices = [cores_situacao.get(sit, colors.grey) for sit in resumo["SITUACAO"]]
    for idx, cor in enumerate(cores_slices):
        grafico.slices[idx].fillColor = cor
        grafico.slices[idx].fontName = FONTE_BAHNSCHRIFT
        grafico.slices[idx].fontSize = 14
        # Linha de chamada do rótulo de dados — mesma configuração usada no
        # gráfico Resumo Financeiro (mais grossa, na cor da fatia).
        grafico.slices[idx].label_pointer_strokeColor = cor
        grafico.slices[idx].label_pointer_strokeWidth = 1.5

    # Mesma correção da falha do ReportLab usada no gráfico Resumo
    # Financeiro: recolore cada linha de chamada na ordem em que é gerada
    # internamente. No modo sideLabels o ReportLab desenha DUAS linhas por
    # fatia (um "cotovelo" de dois segmentos), por isso o índice é dividido
    # por 2 — na mesma ordem de "cores_slices".
    elemento_grafico = grafico.draw()
    linhas_chamada = [e for e in elemento_grafico.contents if isinstance(e, Line)]
    for idx_linha, linha in enumerate(linhas_chamada):
        cor_fatia = cores_slices[idx_linha // 2]
        linha.strokeColor = cor_fatia
        linha.strokeWidth = 1.5

    desenho.add(elemento_grafico)

    # Legenda horizontal no topo (mesmo padrão do gráfico RESUMO FINANCEIRO):
    # só o rótulo da situação, sem investimento — que já aparece no rótulo
    # de dados de cada fatia, junto da quantidade. Mantém a ordem NÃO/SIM
    # usada antes.
    itens_legenda = [
        (
            "Termo Não Assinado (NÃO)" if linha["SITUACAO"] == "NÃO" else "Termo Assinado (SIM)",
            cores_situacao.get(linha["SITUACAO"], colors.grey),
        )
        for _, linha in resumo.iloc[::-1].iterrows()
    ]
    desenhar_legenda_topo(desenho, itens_legenda, y=220, centro_x=240, espacamento=20, fonte_tam=12)

    return desenho

# =====================================================
# 5C. PANORAMA POR SECRETARIA | EXECUTOR (GRADE DE MINI-GRÁFICOS)
# =====================================================

def gerar_mini_grafico_fases(resumo_fase, largura=150, altura=132):
    # Mesma configuração de fatias/rótulos/linhas de chamada do gráfico
    # PANORAMA GERAL DAS FASES (fatia proporcional ao INVESTIMENTO, rótulo
    # "(qtd) investimento abreviado", rótulos externos com ajuste automático
    # de sobreposição, mesmas cores por fase) — só em tamanho reduzido e sem
    # legenda própria, já que a legenda da página é única (ver
    # gerar_legenda_fases_unica).
    raio = min(largura, altura) * 0.5
    desenho = Drawing(largura, altura)
    grafico = Pie()
    grafico.x = (largura - raio) / 2
    grafico.y = (altura - raio) / 2 - 6
    grafico.width = grafico.height = raio
    grafico.data = resumo_fase["VALOR"].tolist()
    grafico.labels = [
        f"({int(qtd)}) {formatar_mi_bi(valor).replace('&nbsp;', ' ')}"
        for qtd, valor in zip(resumo_fase["QTD"], resumo_fase["VALOR"])
    ]
    grafico.slices.strokeWidth = 1
    grafico.slices.strokeColor = colors.white
    grafico.simpleLabels = 0
    grafico.sideLabels = 1
    grafico.sideLabelsOffset = 0.15
    grafico.checkLabelOverlap = 1

    cores_slices = [CORES_FASE.get(fase, colors.grey) for fase in resumo_fase["FASE_TEXTO"]]
    for idx, cor in enumerate(cores_slices):
        grafico.slices[idx].fillColor = cor
        grafico.slices[idx].fontName = FONTE_BAHNSCHRIFT
        grafico.slices[idx].fontSize = 10
        grafico.slices[idx].label_pointer_strokeColor = cor
        grafico.slices[idx].label_pointer_strokeWidth = 1.2

    # Mesma correção de cor das linhas de chamada usada no gráfico Panorama
    # Geral (2 linhas por fatia no modo sideLabels).
    elemento_grafico = grafico.draw()
    linhas_chamada = [e for e in elemento_grafico.contents if isinstance(e, Line)]
    for idx_linha, linha in enumerate(linhas_chamada):
        cor_fatia = cores_slices[idx_linha // 2]
        linha.strokeColor = cor_fatia
        linha.strokeWidth = 1.2

    desenho.add(elemento_grafico)
    return desenho

def gerar_legenda_fases_unica(fases_presentes, largura=380, altura=22, fonte_tam=11):
    # Legenda única da página (uma só, não uma por gráfico) — mesmo
    # comportamento dinâmico da legenda do gráfico PANORAMA GERAL DAS FASES:
    # só mostra as fases que realmente aparecem nos dados já filtrados desta
    # página (ex: se o filtro aplicado não deixa nenhuma ação CONCLUÍDA, essa
    # fase não entra na legenda).
    desenho = Drawing(largura, altura)
    itens_legenda = [(fase, CORES_FASE[fase]) for fase in fases_presentes]
    desenhar_legenda_topo(desenho, itens_legenda, y=6, centro_x=largura / 2, espacamento=22, fonte_tam=fonte_tam)
    return desenho

cabecalho_sec_exec_style = ParagraphStyle(
    "cabecalho_sec_exec",
    parent=cell_nowrap,
    fontName=FONTE_PADRAO_NEGRITO,
    fontSize=11,
    leading=13,
    textColor=COR_MARCA_TEAL_ESCURO,
    alignment=TA_LEFT,
)
cabecalho_sec_exec_esmaecido_style = ParagraphStyle(
    "cabecalho_sec_exec_esmaecido", parent=cabecalho_sec_exec_style, textColor=colors.HexColor("#8C8C8C"),
)
detalhamento_gestao_header_style = ParagraphStyle(
    "detalhamento_gestao_header",
    fontName=FONTE_PADRAO_NEGRITO,
    fontSize=17,
    leading=20,
    textColor=COR_MARCA_TEAL_ESCURO,
    alignment=TA_LEFT,
)
cabecalho_grafico_painel_style = ParagraphStyle(
    "cabecalho_grafico_painel",
    parent=cell_nowrap,
    fontName=FONTE_PADRAO_NEGRITO,
    fontSize=14,
    leading=17,
    textColor=COR_MARCA_TEAL_ESCURO,
    alignment=TA_LEFT,
)
cabecalho_grafico_painel_style_direita = ParagraphStyle(
    "cabecalho_grafico_painel_direita",
    parent=cell_nowrap,
    fontName=FONTE_PADRAO,
    fontSize=11,
    leading=17,
    textColor=colors.HexColor("#666666"),
    alignment=TA_RIGHT,
)

def montar_pagina_panorama_secretaria(df_pagina, titulo_pagina, num_colunas=4):
    # Monta os elementos de UMA página do Panorama por Secretaria/Executor:
    # título, cards de resumo, legenda única e uma grade de mini-gráficos —
    # um por combinação SECRETARIA/EXECUTOR presente em df_pagina. Retorna
    # None se não houver nenhum par (ex: gestão vazia após split).
    df_pagina = df_pagina.assign(_INV_AJ=_serie_investimento_ajustado(df_pagina))
    pares = df_pagina[["SECRETARIA_LIMPA", "EXECUTOR"]].drop_duplicates().copy()
    pares["_SEC_SORT"] = pares["SECRETARIA_LIMPA"].apply(remover_acentos)
    pares["_EXE_SORT"] = pares["EXECUTOR"].apply(remover_acentos)
    pares = pares.sort_values(["_SEC_SORT", "_EXE_SORT"])

    largura_coluna = LARGURA_UTIL / num_colunas
    celulas = []
    for _, par in pares.iterrows():
        sec, exe = str(par["SECRETARIA_LIMPA"]), str(par["EXECUTOR"])
        df_par = df_pagina[
            (df_pagina["SECRETARIA_LIMPA"] == sec) & (df_pagina["EXECUTOR"] == exe)
        ]
        resumo_fase = (
            df_par.groupby("FASE_TEXTO")
            .agg(QTD=(col_objeto, "count"), VALOR=("_INV_AJ", "sum"))
            .reindex(ORDEM_FASES)
            .fillna(0)
            .reset_index()
        )
        # Filtra por QTD>0 E por VALOR>0 (o tamanho da fatia é o
        # investimento — uma fase com investimento zero deixaria o
        # ReportLab dividir pelo total zero e travar com
        # "float division by zero").
        resumo_fase = resumo_fase[(resumo_fase["QTD"] > 0) & (resumo_fase["VALOR"] > 0)]
        if resumo_fase.empty:
            continue

        cabecalho_tab = Table(
            [[Paragraph(f"{sec} | {exe}", cabecalho_sec_exec_style)]],
            colWidths=[largura_coluna - 8],
        )
        cabecalho_tab.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, -1), COR_MARCA_TEAL_CLARA),
                    ("LEFTPADDING", (0, 0), (-1, -1), 6),
                    ("TOPPADDING", (0, 0), (-1, -1), 4),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                ]
            )
        )
        grafico_mini = gerar_mini_grafico_fases(resumo_fase, largura=largura_coluna - 8)
        celulas.append([cabecalho_tab, Spacer(1, 4), grafico_mini])

    if not celulas:
        return None

    linhas_grid, linha_atual = [], []
    for celula in celulas:
        linha_atual.append(celula)
        if len(linha_atual) == num_colunas:
            linhas_grid.append(linha_atual)
            linha_atual = []
    if linha_atual:
        while len(linha_atual) < num_colunas:
            linha_atual.append("")
        linhas_grid.append(linha_atual)

    grid = Table(linhas_grid, colWidths=[largura_coluna] * num_colunas)
    grid.setStyle(
        TableStyle(
            [
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("BOX", (0, 0), (-1, -1), 0.5, colors.grey),
                ("INNERGRID", (0, 0), (-1, -1), 0.5, colors.lightgrey),
                ("TOPPADDING", (0, 0), (-1, -1), 8),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
                ("LEFTPADDING", (0, 0), (-1, -1), 4),
                ("RIGHTPADDING", (0, 0), (-1, -1), 4),
            ]
        )
    )

    fases_presentes_pagina = [
        fase for fase in ORDEM_FASES if fase in df_pagina["FASE_TEXTO"].unique()
    ]

    return [
        Paragraph(titulo_pagina, titulo_style),
        Spacer(1, 13),
        gerar_cards_resumo(len(df_pagina), _investimento_ajustado(df_pagina)),
        Spacer(1, 6),
        Paragraph(
            "* O tamanho de cada fatia do gráfico corresponde ao valor financeiro "
            "(R$), e os valores entre parênteses ( ) indicam a quantidade de ações.",
            orientacao_leitura_style,
        ),
        Spacer(1, 6),
        gerar_legenda_fases_unica(fases_presentes_pagina),
        Spacer(1, 12),
        grid,
    ]

def estimar_altura_pagina_panorama_secretaria(df_pagina, num_colunas=4):
    # Estima a altura da página SEM desenhar nenhum mini-gráfico — só conta
    # quantos pares Secretaria/Executor existem e calcula a altura esperada
    # da grade matematicamente. Desenhar cada mini-gráfico (rótulos com
    # ajuste automático de sobreposição) não é barato, e antes essa decisão
    # de "cabe numa página ou divide por gestão" desenhava a página inteira
    # só para medir e, no caso comum de não caber, jogava tudo fora e
    # desenhava de novo — dobrando o tempo de geração à toa.
    n_pares = df_pagina[["SECRETARIA_LIMPA", "EXECUTOR"]].drop_duplicates().shape[0]
    if n_pares == 0:
        return 0, 0
    n_linhas = -(-n_pares // num_colunas)  # ceil(n_pares / num_colunas)
    ALTURA_LINHA_GRID = 173  # cabeçalho + spacer + gráfico (132) + padding da célula
    altura_grid = n_linhas * ALTURA_LINHA_GRID
    # título + cards + nota de orientação + legenda + grid
    altura_pagina = 25 + 13 + 68 + 6 + 15 + 6 + 22 + 12 + altura_grid
    return n_pares, altura_pagina

# =====================================================
# 5D. MAPA COROPLÉTICO DOS MUNICÍPIOS (BAHIA)
# =====================================================
# Contorno dos 417 municípios baixado do IBGE via
# github.com/tbrugz/geodata-br (Creative Commons CC0 — domínio público).
# Precisa estar salvo como "municipios_bahia.geojson" na mesma pasta do
# script/.exe (igual PAC.png, GOVERNO.PNG e as fontes). Se o arquivo não
# for encontrado, a página do mapa é omitida do relatório sem travar a
# geração — mesmo padrão de tolerância a falta de recurso usado nas fontes.
ARQUIVO_GEOJSON_MUNICIPIOS = "municipios_bahia.geojson"

# Classes discretas (não degradê contínuo) — cada faixa tem uma cor bem
# diferente da vizinha, o que facilita muito mais a diferenciação visual
# entre municípios do que um gradiente de uma cor só. As bordas de cada
# classe são FIXAS (ver BORDAS_FIXAS_MAPA / _calcular_bordas_classes_mapa),
# definidas manualmente em vez de calculadas por quantil — permite
# comparar relatórios diferentes (filtros/datas diferentes) na mesma
# escala de valores.
NUM_CLASSES_MAPA = 4
CORES_CLASSES_MAPA = [
    cor_grafico("#4184AE"),
    cor_grafico("#9FCE9B"),
    cor_grafico("#EA936B"),
    cor_grafico("#C44F5C"),
]
COR_MAPA_SEM_DADO = colors.HexColor("#E0E0E0")

def carregar_geojson_municipios():
    try:
        with open(caminho_recurso(ARQUIVO_GEOJSON_MUNICIPIOS), encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return None

def _bbox_projecao_municipios(features):
    # Projeção equirretangular simples, com correção de achatamento
    # leste-oeste pela latitude média — suficiente para um mapa estadual
    # como o da Bahia, sem depender de bibliotecas de projeção cartográfica.
    lons = [lon for feat in features for anel in feat["geometry"]["coordinates"] for lon, _ in anel]
    lats = [lat for feat in features for anel in feat["geometry"]["coordinates"] for _, lat in anel]
    minlon, maxlon = min(lons), max(lons)
    minlat, maxlat = min(lats), max(lats)
    correcao_x = math.cos(math.radians((minlat + maxlat) / 2))
    return minlon, minlat, maxlon, maxlat, correcao_x

# Faixas de valor FIXAS (não mais por quantil): azul até R$ 50 Mi, verde de
# R$ 50 Mi a R$ 500 Mi, laranja de R$ 500 Mi a R$ 2 Bi, vermelho a partir
# de R$ 2 Bi.
BORDAS_FIXAS_MAPA = [50_000_000.0, 500_000_000.0, 2_000_000_000.0]

def _calcular_bordas_classes_mapa(valores):
    # Começa no MENOR valor real do filtro atual (não mais fixo em R$ 0) e
    # termina no maior — os limites fixos intermediários só entram se
    # ficarem dentro desse intervalo (evita uma faixa "vazia" no topo
    # quando o filtro deixa só municípios com investimento pequeno).
    if not valores:
        return [0.0, 0.0]
    valor_min = min(valores)
    valor_max = max(valores)
    bordas = [valor_min]
    for limite in BORDAS_FIXAS_MAPA:
        if valor_min < limite < valor_max:
            bordas.append(limite)
    bordas.append(valor_max)
    return bordas

def _indice_classe_mapa(valor, bordas):
    # Encontra em qual classe o valor cai (última borda é inclusiva).
    for i in range(len(bordas) - 1):
        if valor <= bordas[i + 1] or i == len(bordas) - 2:
            return i
    return len(bordas) - 2


# Geometria do medidor: arco aberto de 250°, começando às 215° (embaixo à
# esquerda) e terminando às -35° (embaixo à direita), com o índice 50
# caindo exatamente no topo. O vão embaixo é o que abre espaço para a
# categoria e dá ao desenho a silhueta de anel, em vez do semicírculo de
# velocímetro que havia antes.
ANGULO_INICIAL_MEDIDOR = 215.0
VARREDURA_MEDIDOR = 250.0
COR_TRILHO_MEDIDOR = colors.HexColor("#E8E8E8")
COR_TRILHO_MEDIDOR_ESMAECIDO = colors.HexColor("#EDEDED")

def _cor_texto_categoria(cor):
    # As cores das categorias foram escolhidas para preencher área (arco,
    # barra, mapa), e as claras — "Bom" é um verde bem pálido — não têm
    # contraste para virar texto sobre fundo branco. Escurece 30% só para o
    # rótulo, mantendo o tom reconhecível ao lado do arco.
    return colors.Color(cor.red * 0.7, cor.green * 0.7, cor.blue * 0.7)

def gerar_medidor_desempenho(secretaria, indice, categoria, largura=200, altura=165, esmaecido=False,
                             mostrar_rotulos=True, mostrar_categoria=True):
    # Medidor do Índice de Desempenho: anel de 250° com o trecho percorrido
    # pintado na cor da categoria, um marcador redondo no ponto exato do
    # índice, a nota grande no centro e a categoria embaixo, na mesma cor.
    #
    # O desenho anterior era um velocímetro: semicírculo dividido nas quatro
    # faixas de cor, com ponteiro e pino central. As faixas mostravam a
    # escala inteira o tempo todo, o que competia com a informação que
    # importa (onde ESTA secretaria caiu). Aqui a cor é uma só, a da
    # categoria em que o índice caiu, e a leitura de "quanto" fica por conta
    # do quanto do anel está preenchido. Os limites das faixas continuam
    # explicados na página de Metodologia.
    #
    # Não desenha o rótulo "SECRETARIA | EXECUTOR" aqui dentro — quem chama
    # essa função já mostra isso no cabeçalho teal logo acima do medidor
    # (ver montar_pagina_indice_desempenho).
    #
    # esmaecido=True: usado quando essa combinação SECRETARIA | EXECUTOR não
    # aparece no recorte filtrado da geração atual do relatório — o índice em
    # si nunca muda (sempre calculado com a base completa), só o desenho fica
    # em tons de cinza pra indicar visualmente que os filtros aplicados
    # deixariam essa combinação de fora do resto do relatório.
    #
    # mostrar_rotulos=False: sai só o anel com o marcador, sem texto nenhum.
    #
    # mostrar_categoria=False: é o caso do cabeçalho de secretaria no
    # Detalhamento Analítico — a nota fica no centro do anel, igual à página
    # do Índice de Desempenho, e a categoria é escrita ao lado do desenho
    # por quem chama, porque ali embaixo do anel não há altura para ela.
    desenho = Drawing(largura, altura)
    cor_categoria = (
        colors.HexColor("#B0B0B0") if esmaecido
        else CORES_CATEGORIA_DESEMPENHO.get(categoria, colors.HexColor("#8C8C8C"))
    )
    cor_trilho = COR_TRILHO_MEDIDOR_ESMAECIDO if esmaecido else COR_TRILHO_MEDIDOR

    # Espaço reservado embaixo para a categoria (só quando ela é desenhada).
    # A altura do arco sai do que sobra: 1,574 é a altura de um arco de 250°
    # em função do raio (1 acima do centro, 0,574 abaixo, que é o seno de
    # 215°), então o raio é o maior que cabe na largura E na altura.
    espaco_categoria = altura * 0.22 if (mostrar_rotulos and mostrar_categoria) else 0.0
    # Espessura do anel: proporcional, com piso — no medidor de 46 pt do
    # cabeçalho do detalhamento, a proporção pura deixava um fio fino demais
    # para se ler impresso.
    espessura = max(4.0, min(largura, altura) * 0.12)
    raio = min((largura - espessura) / 2, (altura - espaco_categoria - espessura) / 1.574)
    cx = largura / 2
    cy = espaco_categoria + espessura / 2 + 0.574 * raio

    indice_limitado = max(0.0, min(100.0, float(indice)))
    angulo_valor = ANGULO_INICIAL_MEDIDOR - (indice_limitado / 100.0) * VARREDURA_MEDIDOR

    def _ponto(angulo_graus, raio_ponto=None):
        rad = math.radians(angulo_graus)
        r = raio if raio_ponto is None else raio_ponto
        return cx + r * math.cos(rad), cy + r * math.sin(rad)

    # Trilho completo, com as pontas arredondadas por dois círculos — o Wedge
    # do ReportLab só produz cortes retos.
    desenho.add(
        Wedge(
            cx, cy, raio + espessura / 2,
            ANGULO_INICIAL_MEDIDOR - VARREDURA_MEDIDOR, ANGULO_INICIAL_MEDIDOR,
            radius1=raio - espessura / 2, fillColor=cor_trilho, strokeColor=None,
        )
    )
    for angulo_ponta in (ANGULO_INICIAL_MEDIDOR, ANGULO_INICIAL_MEDIDOR - VARREDURA_MEDIDOR):
        x_ponta, y_ponta = _ponto(angulo_ponta)
        desenho.add(Circle(x_ponta, y_ponta, espessura / 2, fillColor=cor_trilho, strokeColor=None))

    # Trecho percorrido. A ponta arredondada do início só entra quando o
    # trecho é longo o bastante para ela não se fundir com o marcador e
    # virar um borrão — em índices muito baixos o marcador sozinho já diz
    # tudo.
    if raio * math.radians(ANGULO_INICIAL_MEDIDOR - angulo_valor) > espessura * 1.5:
        desenho.add(
            Wedge(
                cx, cy, raio + espessura / 2, angulo_valor, ANGULO_INICIAL_MEDIDOR,
                radius1=raio - espessura / 2, fillColor=cor_categoria, strokeColor=None,
            )
        )
        x_ini, y_ini = _ponto(ANGULO_INICIAL_MEDIDOR)
        desenho.add(Circle(x_ini, y_ini, espessura / 2, fillColor=cor_categoria, strokeColor=None))

    # Marcador: círculo colorido sobre um anel branco que o destaca do
    # trilho, como no desenho de referência.
    x_marcador, y_marcador = _ponto(angulo_valor)
    desenho.add(Circle(x_marcador, y_marcador, espessura * 0.86, fillColor=colors.white, strokeColor=None))
    desenho.add(Circle(x_marcador, y_marcador, espessura * 0.58, fillColor=cor_categoria, strokeColor=None))

    if not mostrar_rotulos:
        return desenho

    # Nota no centro do anel, o maior elemento do desenho. O limite pela
    # largura evita que três dígitos ("100") encostem no anel.
    fonte_numero = min(raio * 0.85, largura * 0.34)
    desenho.add(
        String(
            cx, cy - fonte_numero * 0.34, f"{indice:.0f}",
            fontName=FONTE_PADRAO_NEGRITO, fontSize=fonte_numero, textAnchor="middle",
            fillColor=colors.HexColor("#9A9A9A") if esmaecido else colors.HexColor("#2B2B2B"),
        )
    )
    if not mostrar_categoria:
        return desenho

    # Categoria proporcional à nota, e não a um valor fixo: com a nota em
    # torno de 55 pt e a categoria travada em 15, a diferença entre as duas
    # era de quase quatro vezes, e o rótulo sumia debaixo do número. O teto
    # pela largura garante que "Insatisfatório", a palavra mais longa,
    # continue cabendo dentro do desenho.
    fonte_categoria = max(7, fonte_numero * 0.38)
    largura_categoria = pdfmetrics.stringWidth(categoria, FONTE_PADRAO_NEGRITO, fonte_categoria)
    if largura_categoria > largura * 0.94:
        fonte_categoria = fonte_categoria * (largura * 0.94) / largura_categoria
    desenho.add(
        String(
            cx, espaco_categoria * 0.30, categoria,
            fontName=FONTE_PADRAO_NEGRITO, fontSize=fonte_categoria, textAnchor="middle",
            fillColor=colors.HexColor("#9A9A9A") if esmaecido else _cor_texto_categoria(cor_categoria),
        )
    )
    return desenho

ESPACO_TITULO_MEDIDOR = 9

# Título de cada card do Índice de Desempenho (barra teal com "1º —
# SECRETARIA"). Fonte maior que a dos mini-gráficos do Panorama por
# Secretaria — aqui cada card é bem maior, e o rótulo estava ficando
# pequeno demais em relação ao medidor.
cabecalho_card_indice_style = ParagraphStyle(
    "cabecalho_card_indice",
    parent=cabecalho_sec_exec_style,
    fontSize=14,
    leading=17,
)
cabecalho_card_indice_esmaecido_style = ParagraphStyle(
    "cabecalho_card_indice_esmaecido",
    parent=cabecalho_card_indice_style,
    textColor=colors.HexColor("#8C8C8C"),
)

# Paddings verticais aplicados pelo TableStyle da grade (TOP + BOTTOM) —
# entram na conta da altura de cada célula.
PADDING_VERTICAL_CELULA_INDICE = 16
# Folga de segurança pra arredondamento: sem ela, um erro de fração de
# ponto na medição pode empurrar a grade pra uma segunda página.
FOLGA_SEGURANCA_INDICE = 4
# Limites do medidor: o mínimo evita um desenho ilegível quando há muitas
# secretarias; o máximo evita um medidor gigantesco e desproporcional
# quando há pouquíssimas (com 2 ou 3 itens, uma coluna ocuparia meia
# página de largura).
LARGURA_MIN_MEDIDOR_INDICE = 70
LARGURA_MAX_MEDIDOR_INDICE = 330
ALTURA_MIN_MEDIDOR_INDICE = 52
# Piso absoluto: abaixo disso o desenho não é mais um medidor, é um borrão.
# Só entra em cena numa gestão com uma quantidade de itens fora do normal,
# e ainda assim é preferível a quebrar a gestão em duas páginas.
LARGURA_ABSOLUTA_MIN_MEDIDOR_INDICE = 26

def _texto_nota_indice(texto_agrupamento):
    return (
        "* O índice combina Status, Fase, Cláusula Suspensiva, o tempo "
        "entre aviso de licitação, O.S. e conclusão prevista, a "
        "quantidade de ações administradas e a proporção de Valor "
        "Contratado sobre o investimento total da própria gestão — ponderado "
        f"pelo investimento de cada ação. Avaliação {texto_agrupamento}, "
        "separada por gestão, ordenada da melhor pra pior."
    )

def _altura_cabecalho_card_indice(itens, largura_cabecalho):
    # Mede a altura REAL do maior cabeçalho (barra teal com "1º — SECRETARIA
    # | EXECUTOR") na largura de coluna considerada — rótulos longos quebram
    # em 2 ou 3 linhas nas grades mais estreitas, e chutar um valor fixo
    # aqui era justamente o que sobrava/faltava espaço no fim da página.
    # +8 = TOPPADDING (4) + BOTTOMPADDING (4) da tabela do cabeçalho.
    largura_texto = max(10, largura_cabecalho - 12)
    altura = 0
    for item in itens:
        paragrafo = Paragraph(
            f"{item.get('posicao', 1)}º — {item['rotulo']}", cabecalho_card_indice_style
        )
        altura = max(altura, paragrafo.wrap(largura_texto, ALTURA_UTIL)[1])
    return altura + 8

def dimensionar_grade_indice_desempenho(itens, altura_disponivel, largura_disponivel):
    # Escolhe o nº de colunas E a altura do medidor que melhor APROVEITAM o
    # espaço recebido: percorre todas as grades possíveis (de 1 coluna até
    # uma coluna por item) e fica com a que produz o maior medidor visível,
    # respeitando ao mesmo tempo a largura da coluna e a altura que sobra
    # pra cada linha.
    #
    # REGRA INEGOCIÁVEL: uma gestão SEMPRE cabe numa página só. Todo
    # candidato considerado aqui já nasce cabendo (a altura do medidor é
    # sempre o que sobra na linha, nunca um valor fixo), e quando nem o
    # layout mais apertado atende aos tamanhos mínimos confortáveis, o
    # código encolhe o medidor em vez de partir a grade em duas páginas.
    #
    # O "tamanho visível" é o raio externo do arco, que o desenho calcula
    # como min(largura * 0.46, altura * 0.68) — usar o mesmo critério aqui
    # garante que a escolha reflita o que a pessoa realmente enxerga, e não
    # só a largura da célula.
    n_itens = len(itens)
    candidatos = []
    for num_colunas in range(1, n_itens + 1):
        n_linhas = math.ceil(n_itens / num_colunas)
        largura_coluna = largura_disponivel / num_colunas
        largura_medidor = min(largura_coluna - 8, LARGURA_MAX_MEDIDOR_INDICE)
        if largura_medidor < LARGURA_ABSOLUTA_MIN_MEDIDOR_INDICE:
            # Daqui pra frente as colunas só ficam mais estreitas — nem o
            # cabeçalho caberia.
            break
        altura_cabecalho = _altura_cabecalho_card_indice(itens, largura_coluna - 8)
        altura_linha = altura_disponivel / n_linhas
        altura_medidor = (
            altura_linha - altura_cabecalho - ESPACO_TITULO_MEDIDOR - PADDING_VERTICAL_CELULA_INDICE
        )
        if altura_medidor <= 0:
            # Nessa grade nem sobra espaço pro desenho depois do cabeçalho.
            continue
        # Acima da proporção natural (85% da largura) o arco pararia de
        # crescer — o raio fica limitado pela largura e o excedente viraria
        # só espaço vazio dentro do desenho. Melhor devolver essa sobra pra
        # grade, que distribui entre as linhas.
        altura_medidor = min(altura_medidor, largura_medidor * 0.85)
        candidatos.append(
            {
                "num_colunas": num_colunas,
                "n_linhas": n_linhas,
                "largura_medidor": largura_medidor,
                "altura_medidor": altura_medidor,
                "altura_cabecalho": altura_cabecalho,
                "raio": min(largura_medidor * 0.46, altura_medidor * 0.68),
            }
        )

    # Preferência: entre os layouts que respeitam os tamanhos mínimos
    # confortáveis, o de maior medidor. Se NENHUM respeitar (gestão com
    # itens demais pra página), cai pro melhor entre todos os que cabem —
    # medidor pequeno, mas tudo numa página só, que é o que importa.
    confortaveis = [
        c
        for c in candidatos
        if c["largura_medidor"] >= LARGURA_MIN_MEDIDOR_INDICE
        and c["altura_medidor"] >= ALTURA_MIN_MEDIDOR_INDICE
    ]
    elegiveis = confortaveis or candidatos
    if elegiveis:
        melhor = max(elegiveis, key=lambda c: c["raio"])
    else:
        # Só chegaria aqui com uma quantidade absurda de itens; ainda assim,
        # uma página só: a grade mais larga possível com o menor desenho
        # aceitável.
        num_colunas = max(1, int(largura_disponivel // (LARGURA_ABSOLUTA_MIN_MEDIDOR_INDICE + 8)))
        n_linhas = math.ceil(n_itens / num_colunas)
        largura_coluna = largura_disponivel / num_colunas
        altura_cabecalho = _altura_cabecalho_card_indice(itens, largura_coluna - 8)
        melhor = {
            "num_colunas": num_colunas,
            "n_linhas": n_linhas,
            "largura_medidor": largura_coluna - 8,
            "altura_medidor": max(
                10.0,
                altura_disponivel / n_linhas
                - altura_cabecalho
                - ESPACO_TITULO_MEDIDOR
                - PADDING_VERTICAL_CELULA_INDICE,
            ),
            "altura_cabecalho": altura_cabecalho,
            "raio": 0,
        }

    # Altura de cada linha: SEMPRE a altura disponível dividida pelo número
    # de linhas. Como a tabela recebe essas alturas explicitamente, o total
    # da grade é, por construção, exatamente o espaço que sobrou na página —
    # é isso que garante ao mesmo tempo o preenchimento até o pé da página e
    # a regra de uma gestão por página. (A sobra, quando o medidor bate no
    # teto da proporção natural, vira respiro distribuído entre as linhas,
    # com os cards centralizados.)
    melhor["altura_linha"] = altura_disponivel / melhor["n_linhas"]
    return melhor

def montar_tabela_grade_indice_desempenho(itens, layout, largura_disponivel):
    # Monta a tabela de medidores em si, com as dimensões já decididas por
    # dimensionar_grade_indice_desempenho.
    num_colunas = layout["num_colunas"]
    largura_coluna = largura_disponivel / num_colunas
    celulas = []
    for item in itens:
        esmaecido = item.get("esmaecido", False)
        estilo_cabecalho = (
            cabecalho_card_indice_esmaecido_style if esmaecido else cabecalho_card_indice_style
        )
        cabecalho_tab = Table(
            [[Paragraph(f"{item['posicao']}º — {item['rotulo']}", estilo_cabecalho)]],
            colWidths=[largura_coluna - 8],
            # Todos os cabeçalhos com a MESMA altura (a do maior rótulo, já
            # medida no dimensionamento) — sem isso, um rótulo que quebra em
            # duas linhas empurra o medidor daquele card pra baixo e ele sai
            # desalinhado dos vizinhos da mesma fileira.
            rowHeights=[layout["altura_cabecalho"]],
        )
        cabecalho_tab.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#EDEDED") if esmaecido else COR_MARCA_TEAL_CLARA),
                    ("LEFTPADDING", (0, 0), (-1, -1), 6),
                    ("TOPPADDING", (0, 0), (-1, -1), 4),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                ]
            )
        )
        medidor = gerar_medidor_desempenho(
            item["rotulo"], item["indice"], item["categoria"],
            largura=layout["largura_medidor"], altura=layout["altura_medidor"],
            esmaecido=esmaecido,
        )
        # Respiro entre o cabeçalho (barra teal com o rótulo) e o medidor —
        # se mudar aqui, mudar também a conta de altura da célula em
        # dimensionar_grade_indice_desempenho.
        celulas.append([cabecalho_tab, Spacer(1, ESPACO_TITULO_MEDIDOR), medidor])

    linhas_grid, linha_atual = [], []
    for celula in celulas:
        linha_atual.append(celula)
        if len(linha_atual) == num_colunas:
            linhas_grid.append(linha_atual)
            linha_atual = []
    if linha_atual:
        while len(linha_atual) < num_colunas:
            linha_atual.append("")
        linhas_grid.append(linha_atual)

    grid = Table(
        linhas_grid,
        colWidths=[largura_coluna] * num_colunas,
        rowHeights=[layout["altura_linha"]] * len(linhas_grid),
    )
    grid.setStyle(
        TableStyle(
            [
                # MIDDLE (e não TOP): com as linhas esticadas pra preencher
                # a página, os cards ficam distribuídos verticalmente em vez
                # de grudados no topo de cada linha.
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("BOX", (0, 0), (-1, -1), 0.5, colors.grey),
                ("INNERGRID", (0, 0), (-1, -1), 0.5, colors.lightgrey),
                ("TOPPADDING", (0, 0), (-1, -1), 8),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
                ("LEFTPADDING", (0, 0), (-1, -1), 4),
                ("RIGHTPADDING", (0, 0), (-1, -1), 4),
            ]
        )
    )
    return grid

class GradeIndiceDesempenho(Flowable):
    # Grade de medidores que só decide o próprio tamanho na hora em que o
    # ReportLab a posiciona na página.
    #
    # Por que assim: o wrap() recebe do Frame a largura e a altura REAIS que
    # sobraram depois do título e da nota de leitura já colocados. Calcular
    # isso por fora (altura da página menos margens, menos altura estimada
    # do título...) obriga a adivinhar coisas que só o ReportLab sabe — o
    # padding interno do Frame, a altura exata do texto com a fonte DIN, o
    # espaço tomado por outros flowables. Qualquer erro nessas contas
    # empurrava a última fileira de medidores pra uma segunda página, que é
    # exatamente o que não pode acontecer. Perguntando ao Frame, não sobra
    # estimativa nenhuma pra errar.
    def __init__(self, itens):
        Flowable.__init__(self)
        self.itens = itens
        self._tabela = None
        self.width = 0
        self.height = 0

    def wrap(self, largura_disponivel, altura_disponivel):
        # Se o que sobrou na página for pouco (a grade caiu numa página já
        # ocupada por outro conteúdo), não encolhe os medidores pra caber na
        # sobra: devolve uma altura maior que a disponível, o que faz o
        # ReportLab jogar a grade inteira pra próxima página, onde ela terá a
        # folha toda.
        if altura_disponivel < ALTURA_UTIL * 0.5:
            self._tabela = None
            self.width, self.height = largura_disponivel, altura_disponivel + 1
            return self.width, self.height

        altura_util = max(1.0, altura_disponivel - FOLGA_SEGURANCA_INDICE)
        layout = dimensionar_grade_indice_desempenho(self.itens, altura_util, largura_disponivel)
        self._tabela = montar_tabela_grade_indice_desempenho(
            self.itens, layout, largura_disponivel
        )
        self.width, self.height = self._tabela.wrap(largura_disponivel, altura_util)
        return self.width, self.height

    def split(self, largura_disponivel, altura_disponivel):
        # Nunca se divide: uma gestão inteira numa página só.
        return []

    def draw(self):
        if self._tabela is not None:
            self._tabela.drawOn(self.canv, 0, 0)

def montar_pagina_indice_desempenho(itens_pagina, titulo_pagina, texto_agrupamento="por SECRETARIA | EXECUTOR"):
    # Monta os elementos de UMA página do Índice de Desempenho: título, nota
    # de leitura e a grade de medidores — um por unidade avaliada (secretaria
    # na Estadual, secretaria + executor na Federal), já vindo ORDENADA da
    # melhor pra pior (a ordenação e a separação por gestão já aconteceram
    # antes, em calcular_indice_desempenho_secretarias — essa função só
    # desenha). Sem os cards de QTD Total/Investimento, de propósito — dá
    # mais espaço vertical pra grade de medidores em si.
    if not itens_pagina:
        return None

    return [
        Paragraph(titulo_pagina, titulo_style),
        Spacer(1, 13),
        Paragraph(_texto_nota_indice(texto_agrupamento), orientacao_leitura_style),
        Spacer(1, 12),
        GradeIndiceDesempenho(itens_pagina),
    ]

def montar_paginas_indice_desempenho_gestao(itens_gestao, titulo_base, texto_agrupamento="por SECRETARIA | EXECUTOR"):
    # Recebe os itens JÁ CALCULADOS de uma única gestão (ordenados da melhor
    # pra pior) e devolve os elementos de UMA ÚNICA página — nunca divide em
    # mais de uma. O nº de colunas e o tamanho dos medidores são decididos
    # pela própria grade, no espaço que realmente sobrar na página.
    if not itens_gestao:
        return []

    itens_numerados = [dict(item, posicao=i + 1) for i, item in enumerate(itens_gestao)]
    elementos = montar_pagina_indice_desempenho(itens_numerados, titulo_base, texto_agrupamento)
    return [elementos] if elementos is not None else []

def gerar_mapa_coropletico_municipios(df_base, geojson_data, largura_mapa=500):
    # Soma o investimento e a quantidade de ações por município (nome
    # normalizado — sem acento, maiúsculo, igual ao resto do relatório) e
    # pinta cada município do mapa numa escala de azul proporcional ao
    # investimento; cinza-claro para municípios sem nenhuma ação no filtro
    # atual. Desenhado como vetor (Path) — continua editável ao abrir o PDF
    # num editor gráfico (Illustrator, Inkscape, Corel etc.), não é imagem.
    resumo = (
        df_base.assign(
            _MUN_NORM=df_base[col_municipio].apply(lambda x: remover_acentos(x).strip().upper()),
            _INV_AJ=_serie_investimento_ajustado(df_base),
        )
        .groupby("_MUN_NORM")
        .agg(QTD=(col_objeto, "count"), VALOR=("_INV_AJ", "sum"))
    )

    minlon, minlat, maxlon, maxlat, correcao_x = _bbox_projecao_municipios(geojson_data["features"])
    largura_bruta = (maxlon - minlon) * correcao_x
    altura_bruta = maxlat - minlat
    escala = largura_mapa / largura_bruta
    altura_mapa = altura_bruta * escala

    def projetar(lon, lat):
        return (lon - minlon) * correcao_x * escala, (lat - minlat) * escala

    dados_por_municipio = {}
    for feat in geojson_data["features"]:
        nome_norm = remover_acentos(feat["properties"]["name"]).strip().upper()
        if nome_norm in resumo.index:
            dados_por_municipio[feat["properties"]["name"]] = (
                int(resumo.loc[nome_norm, "QTD"]),
                float(resumo.loc[nome_norm, "VALOR"]),
            )

    valores_presentes = [valor for _, valor in dados_por_municipio.values()]
    bordas_classes = _calcular_bordas_classes_mapa(valores_presentes) if valores_presentes else [0.0, 0.0]

    desenho = Drawing(largura_mapa, altura_mapa)
    for feat in geojson_data["features"]:
        dado = dados_por_municipio.get(feat["properties"]["name"])
        if dado:
            cor = CORES_CLASSES_MAPA[_indice_classe_mapa(dado[1], bordas_classes) % len(CORES_CLASSES_MAPA)]
        else:
            cor = COR_MAPA_SEM_DADO
        for anel in feat["geometry"]["coordinates"]:
            caminho = Path(fillColor=cor, strokeColor=colors.white, strokeWidth=0.4)
            x0, y0 = projetar(*anel[0])
            caminho.moveTo(x0, y0)
            for lon, lat in anel[1:]:
                x, y = projetar(lon, lat)
                caminho.lineTo(x, y)
            caminho.closePath()
            desenho.add(caminho)

    qtd_municipios_com_dado = len(dados_por_municipio)
    qtd_municipios_total = len(geojson_data["features"])
    return desenho, bordas_classes, qtd_municipios_com_dado, qtd_municipios_total

def gerar_legenda_mapa_municipios(bordas_classes, largura=340, altura=290):
    # Uma amostra de cor sólida por classe (não degradê) — cada faixa mostra
    # o intervalo de investimento que ela representa, da maior (topo) pra
    # menor (embaixo), mais o indicador cinza de "sem ação no filtro atual".
    desenho = Drawing(largura, altura)
    num_classes = len(bordas_classes) - 1
    tam_swatch = 26
    espaco_swatch = 34
    x_swatch = 10
    y_topo = altura - 24
    fonte_tam = 13

    for i in reversed(range(num_classes)):
        y = y_topo - (num_classes - 1 - i) * espaco_swatch
        cor = CORES_CLASSES_MAPA[i % len(CORES_CLASSES_MAPA)]
        desenho.add(
            Rect(
                x_swatch, y, tam_swatch, tam_swatch,
                fillColor=cor, strokeColor=colors.grey, strokeWidth=0.5,
            )
        )
        rotulo = (
            f"{formatar_mi_bi(bordas_classes[i]).replace('&nbsp;', ' ')} a "
            f"{formatar_mi_bi(bordas_classes[i + 1]).replace('&nbsp;', ' ')}"
        )
        desenho.add(
            String(
                x_swatch + tam_swatch + 8, y + tam_swatch / 2 - 3, rotulo,
                fontName=FONTE_PADRAO, fontSize=fonte_tam, fillColor=colors.black,
            )
        )

    y_cinza = y_topo - num_classes * espaco_swatch - 10
    desenho.add(
        Rect(
            x_swatch, y_cinza, tam_swatch, tam_swatch,
            fillColor=COR_MAPA_SEM_DADO, strokeColor=colors.grey, strokeWidth=0.5,
        )
    )
    desenho.add(
        String(
            x_swatch + tam_swatch + 8, y_cinza + tam_swatch / 2 - 3,
            "Sem ação no filtro atual",
            fontName=FONTE_PADRAO, fontSize=fonte_tam, fillColor=colors.black,
        )
    )

    return desenho

# =====================================================
# 6. ENGENHARIA DE DADOS (ETL PANDAS)
# =====================================================

try:
    # Antes lia a planilha INTEIRA (header=None) só pra descobrir em qual
    # linha fica o cabeçalho (procurando "EIXO"), e depois lia de novo,
    # inteira, pra carregar os dados de verdade — ou seja, processava todo
    # o arquivo duas vezes. Aqui a busca do cabeçalho olha só as primeiras
    # linhas (onde ele quase sempre está); se não achar nesse intervalo,
    # cai de volta para a varredura completa, como antes, garantindo que
    # planilhas fora do padrão continuem funcionando.
    #
    # A busca usa VÁRIAS colunas-âncora (não só "EIXO") — se um dia uma
    # coluna for excluída ou renomeada na planilha, a detecção do
    # cabeçalho continua funcionando enquanto pelo menos uma das âncoras
    # (normalmente bem estáveis, dificilmente excluídas) ainda existir.
    ANCORAS_CABECALHO = ["EIXO", "OBJETO", "ITEM", "SECRETARIA"]

    def _linha_do_cabecalho(df_bruto):
        for i in range(len(df_bruto)):
            texto_linha = df_bruto.iloc[i].astype(str).str.upper()
            if any(texto_linha.str.contains(ancora, case=False).any() for ancora in ANCORAS_CABECALHO):
                return i
        return None

    LINHAS_BUSCA_CABECALHO = 20
    df_temp = pd.read_excel(arquivo_excel, header=None, nrows=LINHAS_BUSCA_CABECALHO)
    linha_header = _linha_do_cabecalho(df_temp)
    if linha_header is None:
        df_temp = pd.read_excel(arquivo_excel, header=None)
        linha_header = _linha_do_cabecalho(df_temp)
    if linha_header is None:
        _erro_fatal_inicializacao(
            "Não foi possível identificar a linha do cabeçalho na planilha — "
            "nenhuma das colunas de referência (EIXO, OBJETO, ITEM, SECRETARIA) "
            "foi encontrada.\n\n"
            "Verifique se a planilha não teve os cabeçalhos removidos ou "
            "renomeados por engano.\n\n"
            f"Arquivo:\n{arquivo_excel}"
        )
    df = pd.read_excel(arquivo_excel, header=linha_header)
except PermissionError:
    _erro_fatal_inicializacao(
        "Não foi possível abrir a planilha porque ela está sendo usada por outro "
        "programa (provavelmente aberta no Excel, ou ainda sincronizando no OneDrive).\n\n"
        "Feche a planilha e tente novamente.\n\n"
        f"Arquivo:\n{arquivo_excel}"
    )
except FileNotFoundError:
    _erro_fatal_inicializacao(
        f"Planilha não encontrada:\n{arquivo_excel}\n\n"
        "Verifique se o arquivo está na mesma pasta do programa (.exe) e se o "
        "nome está exatamente igual."
    )

df.columns = [
    re.sub(r"\s+", " ", remover_acentos(c).strip().upper())
    for c in df.columns
]

# --- Colunas de LINK da Ficha Cadastral (Monitora Bahia / localização) ---
# São as duas colunas que mais mudam de nome de uma versão da planilha
# para outra ("LINK MONITORA", "LINK DO MONITORA BAHIA", "LOCALIZAÇÃO
# (GOOGLE MAPS)"...). O casamento era por nome EXATO, então um cabeçalho
# diferente fazia o campo aparecer vazio na ficha, em silêncio — nem erro,
# nem aviso. Agora o nome é resolvido aqui na carga: primeiro pelos nomes
# conhecidos, depois por palavra-chave no cabeçalho.
def _resolver_coluna_por_apelido(colunas, candidatos, grupos_palavras):
    normalizadas = {
        coluna: re.sub(r"\s+", " ", remover_acentos(str(coluna)).strip().upper())
        for coluna in colunas
    }
    for candidato in candidatos:
        alvo = re.sub(r"\s+", " ", remover_acentos(str(candidato)).strip().upper())
        for coluna, normalizada in normalizadas.items():
            if normalizada == alvo:
                return coluna
    # Nenhum nome conhecido bateu: aceita qualquer cabeçalho que contenha
    # todas as palavras de algum dos grupos (ex: "LINK DA LOCALIZACAO DA
    # OBRA" casa pelo grupo ["LOCALIZ"]).
    for palavras in grupos_palavras:
        for coluna, normalizada in normalizadas.items():
            if all(palavra in normalizada for palavra in palavras):
                return coluna
    return None

_col_monitora_encontrada = _resolver_coluna_por_apelido(
    df.columns,
    [col_link_monitora, "LINK DO MONITORA", "LINK MONITORA BAHIA",
     "LINK DO MONITORA BAHIA", "MONITORA BAHIA", "MONITORA"],
    [["MONITORA"]],
)
if _col_monitora_encontrada:
    col_link_monitora = _col_monitora_encontrada

_col_localizacao_encontrada = _resolver_coluna_por_apelido(
    df.columns,
    [col_link_localizacao, "LINK DE LOCALIZACAO", "LINK DA LOCALIZACAO",
     "LINK GOOGLE MAPS", "LINK DO GOOGLE MAPS", "GOOGLE MAPS",
     "LOCALIZACAO", "LINK DO MAPA", "LINK MAPA"],
    [["MAPS"], ["LOCALIZ"], ["MAPA"]],
)
if _col_localizacao_encontrada:
    col_link_localizacao = _col_localizacao_encontrada

def _preencher_hiperlinks_embutidos(arquivo, linha_do_cabecalho, coluna_alvo):
    # Recupera os endereços dos hiperlinks EMBUTIDOS nas células.
    #
    # Quando alguém cola um link no Excel escrevendo um texto por cima
    # ("Ver no mapa", "Monitora", o nome da obra), a célula guarda duas
    # coisas: o texto exibido e o endereço de destino. O pandas lê apenas o
    # TEXTO — o endereço fica invisível para ele, e a ficha acaba mostrando
    # uma palavra que não leva a lugar nenhum.
    #
    # Só roda quando a coluna existe mas nenhuma das suas células tem um
    # endereço visível, porque abrir a planilha uma segunda vez (agora pelo
    # openpyxl, que enxerga os hiperlinks) custa tempo. Falha em silêncio:
    # se não der para ler, a ficha continua mostrando o texto de sempre.
    if coluna_alvo not in df.columns:
        return
    valores = df[coluna_alvo].astype(str)
    if valores.str.contains(r"https?://", case=False, na=False).any():
        return  # os endereços já estão visíveis no texto; nada a fazer

    try:
        from openpyxl import load_workbook

        planilha = load_workbook(arquivo, data_only=True)
        aba = planilha.worksheets[0]
        linha_cabecalho_excel = linha_do_cabecalho + 1

        indice_coluna_excel = None
        for celula in aba[linha_cabecalho_excel]:
            if celula.value is None:
                continue
            nome = re.sub(r"\s+", " ", remover_acentos(str(celula.value)).strip().upper())
            if nome == coluna_alvo:
                indice_coluna_excel = celula.column
                break
        if indice_coluna_excel is None:
            planilha.close()
            return

        enderecos = {}
        for posicao, linha in enumerate(
            aba.iter_rows(
                min_row=linha_cabecalho_excel + 1,
                min_col=indice_coluna_excel,
                max_col=indice_coluna_excel,
            )
        ):
            celula = linha[0]
            if celula.hyperlink is not None and celula.hyperlink.target:
                enderecos[posicao] = celula.hyperlink.target
        planilha.close()

        if not enderecos:
            return
        # As posições do openpyxl e os índices do dataframe só correspondem
        # porque nada foi filtrado entre a leitura e este ponto.
        for posicao, endereco in enderecos.items():
            if posicao < len(df):
                df.iat[posicao, df.columns.get_loc(coluna_alvo)] = endereco
    except Exception:
        return

for _coluna_link in (col_link_monitora, col_link_localizacao):
    _preencher_hiperlinks_embutidos(arquivo_excel, linha_header, _coluna_link)

# Diagnóstico rápido no console: se um dos campos de link aparecer vazio na
# ficha, essas linhas dizem se o problema é o cabeçalho da planilha (coluna
# não localizada) ou o conteúdo das células (coluna achada, mas sem
# endereço nenhum).
for _rotulo_link, _coluna_link, _achou in (
    ("Monitora", col_link_monitora, bool(_col_monitora_encontrada)),
    ("Localização", col_link_localizacao, bool(_col_localizacao_encontrada)),
):
    if not _achou:
        print(f"[Ficha Cadastral] Coluna de link '{_rotulo_link}' NÃO localizada na planilha "
              f"(procurei por '{_coluna_link}' e variações). O campo ficará vazio.")
    else:
        _com_endereco = int(
            df[_coluna_link].astype(str).str.contains(r"https?://", case=False, na=False).sum()
        )
        print(f"[Ficha Cadastral] Link '{_rotulo_link}' → coluna '{_coluna_link}': "
              f"{_com_endereco} de {len(df)} linhas com endereço.")

# Confere logo aqui, ANTES de qualquer processamento, se as colunas que o
# programa realmente não consegue funcionar sem estão todas presentes —
# se uma planilha futura excluir ou renomear (por engano) alguma dessas
# colunas, o aviso já mostra exatamente qual é o nome faltando, em vez de
# travar mais adiante com um erro genérico difícil de entender. As colunas
# "opcionais" (usadas só na Ficha Cadastral, como DESCRIÇÃO ou os campos de
# acompanhamento) NÃO entram aqui de propósito — a ausência delas já é
# tratada de forma silenciosa (campo em branco) em cada função que as usa.
COLUNAS_ESSENCIAIS = {
    col_objeto: "OBJETO",
    col_orgao: "SECRETARIA/ ORGAO",
    col_executor: "ORGAO EXECUTOR",
    col_eixo: "EIXO",
    col_municipio: "MUNICIPIO",
    col_fase: "FASE",
    col_status: "STATUS",
    col_fonte: "FONTE DE RECURSO",
    col_tc: "Nº DO TERMO DE COMP. (TC) / CT FINANCIAMENTO",
    col_prazo: "PRAZO DE CONCLUSAO DA FASE",
    col_avanco: "AVANCO DA OBRA (%)",
    col_vigencia: "VIGENCIA",
    col_prazo_atual: "PREVISAO DE CONCLUSAO ATUAL",
    col_clausula_suspensiva: "CLAUSULA SUSPENSIVA",
    col_item: "ITEM",
    col_financiamento: "FINANCIAMENTO",
    col_apoiado: "APOIADO",
    col_contrapartida: "CONTRAPARTIDA",
    col_complementar: "COMPLEMENTAR",
    col_valor_contratado: "VALOR CONTRATADO",
    col_invest: "INVESTIMENTO TOTAL",
}
_colunas_faltando = [nome for coluna, nome in COLUNAS_ESSENCIAIS.items() if coluna not in df.columns]
if _colunas_faltando:
    _plural = len(_colunas_faltando) > 1
    _erro_fatal_inicializacao(
        f"A planilha está sem {'as colunas' if _plural else 'a coluna'} abaixo, "
        f"que {'são' if _plural else 'é'} essenci{'ais' if _plural else 'al'} "
        "pro programa funcionar:\n\n"
        + "\n".join(f"• {nome}" for nome in _colunas_faltando)
        + "\n\nVerifique se o(s) cabeçalho(s) não foi(ram) renomeado(s) ou "
        "excluído(s) por engano na planilha.\n\n"
        f"Arquivo:\n{arquivo_excel}"
    )

df = df[df[col_objeto].notna() & (df[col_objeto].astype(str).str.strip() != "")]

df["GESTAO"] = df[col_orgao].apply(
    lambda x: "GESTÃO FEDERAL" if "FEDERAL" in str(x).upper() else "GESTÃO ESTADUAL"
)
df["SECRETARIA_LIMPA"] = df[col_orgao].fillna("").astype(str).str.strip()
df["EXECUTOR"] = df[col_executor].apply(limpar_executor)
df[col_invest] = df[col_invest].apply(converter_valor)
df[col_financiamento] = df[col_financiamento].apply(converter_valor)
df[col_apoiado] = df[col_apoiado].apply(converter_valor)
df[col_contrapartida] = df[col_contrapartida].apply(converter_valor)
df[col_complementar] = df[col_complementar].apply(converter_valor)
df[col_valor_contratado] = df[col_valor_contratado].apply(converter_valor)

# Colunas auxiliares condicionais: se VALOR CONTRATADO > 0, o valor ajustado é zero;
# caso contrário, mantém o valor original da coluna.
df[col_financiamento_ajustado] = np.where(
    df[col_valor_contratado] > 0, 0, df[col_financiamento]
)
df[col_apoiado_ajustado] = np.where(
    df[col_valor_contratado] > 0, 0, df[col_apoiado]
)
df[col_contrapartida_ajustado] = np.where(
    df[col_valor_contratado] > 0, 0, df[col_contrapartida]
)
df[col_complementar_ajustado] = np.where(
    df[col_valor_contratado] > 0, 0, df[col_complementar]
)

df["STATUS_ORDEM"] = df[col_status].apply(extrair_ordem_status)
df["STATUS_TEXTO"] = df[col_status].apply(limpar_texto_status).astype(str).str.strip()
df["FASE_ORDEM"] = df[col_fase].apply(extrair_ordem_fase)
df["FASE_TEXTO"] = df[col_fase].apply(limpar_texto_fase).astype(str).str.strip()
df["SINALIZACAO_TC"] = df[col_tc].apply(classificar_termo_compromisso)
df["EIXO_SORT"] = df[col_eixo].apply(remover_acentos)
df["PRAZO_FASE_TEXTO"] = df[col_prazo].apply(formatar_prazo)
df["PRAZO_TRIMESTRE"] = df[col_prazo].apply(calcular_trimestre)
df[col_vigencia] = pd.to_datetime(df[col_vigencia], errors="coerce")
df["CLAUSULA_SUSPENSIVA_ORDEM"] = df[col_clausula_suspensiva].apply(extrair_ordem_clausula_suspensiva)
df["SITUACAO_CLAUSULA_SUSPENSIVA"] = df[col_clausula_suspensiva].apply(tratar_clausula_suspensiva)
df["AVANCO_OBRA_TEXTO"] = df[col_avanco].apply(formatar_percentual)

df[col_objeto] = df[col_objeto].astype(str).str.strip()
df[col_eixo] = df[col_eixo].astype(str).str.strip()
df[col_fonte] = (
    df[col_fonte].astype(str).str.strip().str.replace(r"\s+", " ", regex=True)
)

df = df[~df["STATUS_TEXTO"].isin(status_excluir)]

df_original = df.copy()  # cópia intacta pós-ETL, usada para reaplicar os filtros a cada geração

# =====================================================
# 6B. INTERFACE GRÁFICA DE FILTROS (TKINTER)
# =====================================================

# Status pré-selecionados por padrão ao abrir o painel de filtros (os demais
# vêm desmarcados; o usuário pode ajustar livremente com "Marcar tudo"/"Limpar").
STATUS_PADRAO_SELECIONADOS = {
    "AGUARDANDO ORDEM DE SERVIÇO",
    "ANDAMENTO",
    "ELABORAÇÃO DE PROJETO",
    "ESTUDO",
    "HABILITADA",
    "LICITAÇÃO /CONTRATAÇÃO",
    "SELECIONADA",
    "À LICITAR",
}

def _serie_investimento_ajustado(df):
    # Investimento AJUSTADO por linha — mesma regra usada em todo o
    # relatório: Valor Contratado substitui Financiamento/Apoiado/
    # Contrapartida/Complementar quando preenchido, evitando contar duas
    # vezes. Devolve uma Series alinhada ao índice de df, pra usar em
    # qualquer groupby/agg no lugar da coluna bruta "INVESTIMENTO TOTAL"
    # (que pode não estar sincronizada linha a linha com esses
    # componentes).
    return (
        df[col_valor_contratado]
        + df[col_apoiado_ajustado]
        + df[col_financiamento_ajustado]
        + df[col_contrapartida_ajustado]
        + df[col_complementar_ajustado]
    )

def _investimento_ajustado(df):
    # Mesma regra de valores ajustados já usada no dashboard: soma os
    # componentes (Valor Contratado + OGU + Financiamento + Contrapartida +
    # Complementar, já ajustados pra não contar duas vezes quando há Valor
    # Contratado preenchido) em vez da coluna bruta "INVESTIMENTO TOTAL",
    # que pode não estar sincronizada linha a linha com essas colunas —
    # garante que o card "INVESTIMENTO" do relatório impresso sempre bate
    # com o do dashboard, para o mesmo filtro.
    return float(_serie_investimento_ajustado(df).sum())

# =====================================================
# ÍNDICE DE DESEMPENHO POR SECRETARIA (gráfico de medidor)
# =====================================================
# Metodologia: cada AÇÃO recebe uma nota de 0 a 1 em cada um dos 5
# critérios abaixo; essas 5 notas são combinadas numa nota única da ação
# (média ponderada, com STATUS e FASE pesando mais); depois, a nota de
# CADA SECRETARIA é a média das notas de suas ações, ponderada pelo
# INVESTIMENTO AJUSTADO de cada uma (ações maiores pesam mais) — por fim,
# essa nota de 0 a 1 vira uma categoria (Insatisfatório/Regular/Bom/Ótimo).
#
# Pesos dos 5 critérios na nota da ação — Status e Fase valem mais, por
# serem o retrato mais direto do andamento da ação:
PESO_INDICE_STATUS = 0.30
PESO_INDICE_FASE = 0.30
PESO_INDICE_CLAUSULA = 0.15
PESO_INDICE_FINANCIAMENTO = 0.15
PESO_INDICE_TEMPORAL = 0.10

# Ordem de STATUS do melhor pro pior (índice 0 = melhor) — usada só pra
# calcular a nota, não altera a ordem de exibição do status em nenhum
# outro lugar do relatório.
ORDEM_STATUS_INDICE = [
    "INAUGURADA",
    "CONCLUÍDA",
    "ANDAMENTO",
    "AGUARDANDO ORDEM DE SERVIÇO",
    "LICITAÇÃO /CONTRATAÇÃO",
    "À LICITAR",
    "SELECIONADA",
    "AGUARDANDO AUTORIZO",
    "HABILITADA",
    "AGUARDANDO PUBLICAÇÃO",
    "COMPLEMENTADA E ENVIADA",
    "EM COMPLEMENTAÇÃO",
    "ENVIADA PARA ANÁLISE",
    "ELABORAÇÃO DE PROJETO",
    "CADASTRADA",
    "ESTUDO",
    "NÃO HABILITADA",
    "PARALISADA",
]
_MAPA_ORDEM_STATUS_INDICE = {s: i for i, s in enumerate(ORDEM_STATUS_INDICE)}

def _dias_entre(data_inicio, data_fim):
    if not isinstance(data_inicio, (pd.Timestamp, datetime)) or pd.isna(data_inicio):
        return None
    if not isinstance(data_fim, (pd.Timestamp, datetime)) or pd.isna(data_fim):
        return None
    dias = (data_fim - data_inicio).days
    return dias if dias >= 0 else None

# Faixas da nota (0 a 1) pra cada categoria — ajustável se as faixas
# precisarem ficar mais/menos rígidas no futuro.
FAIXAS_CATEGORIA_INDICE = [
    (0.80, "ÓTIMO"),
    (0.60, "BOM"),
    (0.35, "REGULAR"),
    (0.0, "INSATISFATÓRIO"),
]
def _categoria_indice(nota):
    for limite, nome in FAIXAS_CATEGORIA_INDICE:
        if nota >= limite:
            return nome
    return FAIXAS_CATEGORIA_INDICE[-1][1]

def _filtrar_dataframe(filtros_selecionados):
    # Aplica os filtros selecionados sempre a partir de uma cópia limpa dos
    # dados originais (df_original) — permite chamar isso várias vezes
    # seguidas, com filtros diferentes, sem reabrir a planilha. Pode
    # devolver um DataFrame vazio; quem chama decide como avisar o usuário.
    df = df_original.copy()

    if filtros_selecionados["GESTAO"]:
        df = df[df["GESTAO"].isin(filtros_selecionados["GESTAO"])]
    if filtros_selecionados["EIXO"]:
        df = df[df[col_eixo].isin(filtros_selecionados["EIXO"])]
    if filtros_selecionados["STATUS"]:
        df = df[df["STATUS_TEXTO"].isin(filtros_selecionados["STATUS"])]
    if filtros_selecionados["FASE"]:
        df = df[df["FASE_TEXTO"].isin(filtros_selecionados["FASE"])]
    if filtros_selecionados["ORGAO"]:
        df = df[df["SECRETARIA_LIMPA"].isin(filtros_selecionados["ORGAO"])]
    if filtros_selecionados["EXECUTOR"]:
        df = df[df["EXECUTOR"].isin(filtros_selecionados["EXECUTOR"])]
    if filtros_selecionados["OBJETO"]:
        df = df[df[col_objeto].isin(filtros_selecionados["OBJETO"])]
    if filtros_selecionados["MUNICIPIO"]:
        df = df[df[col_municipio].astype(str).isin(filtros_selecionados["MUNICIPIO"])]
    if filtros_selecionados["FONTE"]:
        df = df[df[col_fonte].astype(str).isin(filtros_selecionados["FONTE"])]
    if filtros_selecionados["CLAUSULA_SUSPENSIVA"]:
        df = df[df["SITUACAO_CLAUSULA_SUSPENSIVA"].isin(filtros_selecionados["CLAUSULA_SUSPENSIVA"])]
    if filtros_selecionados["TERMO_COMPROMISSO"]:
        df = df[df["SINALIZACAO_TC"].isin(filtros_selecionados["TERMO_COMPROMISSO"])]

    df = _aplicar_filtro_arvore_data(df, col_prazo, filtros_selecionados.get("DATAS_CONCLUSAO_FASE"))
    df = _aplicar_filtro_arvore_data(df, col_vigencia, filtros_selecionados.get("DATAS_VIGENCIA"))
    df = _aplicar_filtro_arvore_data(df, col_prazo_atual, filtros_selecionados.get("DATAS_CONCLUSAO_ATUAL"))

    return df


def _aplicar_filtro_arvore_data(df, coluna, datas_selecionadas):
    # Filtra pelas datas exatas marcadas na árvore Ano > Mês > Dia (formato
    # "AAAA-MM-DD"). Sem seleção (lista vazia/None) = sem filtro, mesmo
    # padrão usado nos demais blocos. Defensivo: se a coluna não existir na
    # planilha (nome ainda não confirmado) ou não houver seleção, devolve o
    # df sem alterar — não trava o relatório.
    if not datas_selecionadas or coluna not in df.columns:
        return df
    datas_str = pd.to_datetime(df[coluna], errors="coerce").dt.strftime("%Y-%m-%d")
    return df[datas_str.isin(datas_selecionadas)]


def _construir_arvore_datas(df, coluna):
    # Monta a árvore Ano > "MM-nome do mês" > [dias] usada pelo filtro de
    # datas em formato de calendário (igual ao filtro de datas do Excel).
    # O mês fica com o número na frente ("09-setembro") só para o
    # JavaScript conseguir montar a data exata sem precisar traduzir nome
    # de mês — o texto exibido no painel usa só a parte depois do "-".
    # Defensivo: se a coluna não existir na planilha, devolve árvore vazia
    # em vez de travar (protege especialmente col_prazo_atual, cujo nome
    # exato ainda precisa ser confirmado).
    if coluna not in df.columns:
        return {}
    serie = pd.to_datetime(df[coluna], errors="coerce").dropna()
    arvore = {}
    for valor in sorted(serie.unique()):
        ts = pd.Timestamp(valor)
        ano = str(ts.year)
        mes_chave = f"{ts.month:02d}-{MESES_PT[ts.month]}"
        dia = int(ts.day)
        dias_do_mes = arvore.setdefault(ano, {}).setdefault(mes_chave, [])
        if dia not in dias_do_mes:
            dias_do_mes.append(dia)
    for ano in arvore:
        for mes in arvore[ano]:
            arvore[ano][mes].sort()
    return arvore


# Quantas ações a lista da lupa mostra de uma vez. Com os campos de
# busca vazios o recorte pode ter centenas de ações, e jogar todas
# numa lista só deixaria a janela lenta e impossível de ler — a tela
# avisa quantas ficaram de fora.
LIMITE_LISTA_BUSCA_FICHA = 50

CAMPOS_ALERTA_STATUS_IGNORAR_VENCIMENTO = {"CONCLUÍDA", "INAUGURADA"}
# Fase em que a ação também já está entregue — usada junto com o conjunto
# de status acima para não acusar prazo vencido em obra que acabou.
CAMPOS_ALERTA_FASE_IGNORAR_VENCIMENTO = {"CONCLUÍDA"}
CAMPOS_ALERTA_FASE_SEM_PRAZO = "EXECUÇÃO DO OBJETO"
CAMPOS_ALERTA_STATUS_SEM_PRAZO = {"ANDAMENTO", "CONCLUÍDA", "INAUGURADA"}

def _extrair_data_alerta(valor):
    if isinstance(valor, (pd.Timestamp, datetime)) and not pd.isna(valor):
        return valor.date()
    return None

def _campos_alerta_qualidade(row, hoje=None):
    # Mesma lógica de detecção usada no painel de Controle de Qualidade
    # (_montar_aviso_qualidade), só que aplicada a UMA linha só — devolve
    # um dict {chave_do_campo_na_ficha: motivo} com os campos que
    # precisam de atenção, usado tanto pra montar o aviso do painel quanto
    # pra destacar o campo (contorno vermelho) na Ficha Cadastral. Única
    # fonte de verdade: os dois lugares nunca ficam dessincronizados.
    if hoje is None:
        hoje = datetime.now().date()
    status_atual = str(row.get("STATUS_TEXTO", "")).strip().upper()
    fase_atual = str(row.get("FASE_TEXTO", "")).strip().upper()
    data_fase = _extrair_data_alerta(row.get(col_prazo)) if col_prazo in row.index else None
    alertas = {}

    if status_atual not in CAMPOS_ALERTA_STATUS_IGNORAR_VENCIMENTO:
        data_vigencia = _extrair_data_alerta(row.get(col_vigencia)) if col_vigencia in row.index else None
        if data_fase is not None and data_fase < hoje:
            alertas["prazo_fase"] = "Prazo de Conclusão da Fase vencido"
        if data_vigencia is not None and data_vigencia < hoje:
            alertas["vigencia"] = "Prazo da Cláusula Suspensiva (Vigência) vencido"

    if (
        fase_atual == CAMPOS_ALERTA_FASE_SEM_PRAZO
        and status_atual in CAMPOS_ALERTA_STATUS_SEM_PRAZO
        and data_fase is None
    ):
        alertas["prazo_fase"] = "Execução do Objeto sem Prazo de Conclusão da Fase definido"

    # Fase LICITAÇÃO sem Aviso de Licitação publicado — mas o campo culpado
    # depende do que mais está preenchido, e são dois problemas diferentes:
    #
    #   1. Existe ABERTURA DE LICITAÇÃO: se a abertura já tem data, o aviso
    #      necessariamente foi publicado (não se abre licitação sem avisar).
    #      A fase está CERTA; o que falta é lançar a data do aviso na
    #      planilha. O destaque vai no campo do AVISO.
    #
    #   2. Não existe abertura nenhuma: aí não há sinal de que a licitação
    #      tenha começado, e a leitura mais provável é que a ação ainda
    #      esteja em Captação de Recurso. O destaque vai na FASE.
    #
    # Antes os dois casos caíam no segundo, e a ficha acusava fase errada em
    # ações cuja fase estava correta — o problema era só o campo em branco.
    if col_aviso_licitacao in row.index:
        data_aviso_licitacao = _extrair_data_alerta(row.get(col_aviso_licitacao))
        data_abertura_licitacao = (
            _extrair_data_alerta(row.get(col_abertura_licitacao))
            if col_abertura_licitacao in row.index
            else None
        )
        if fase_atual == "LICITAÇÃO" and data_aviso_licitacao is None:
            if data_abertura_licitacao is not None:
                alertas["aviso_licitacao"] = (
                    "Abertura de Licitação preenchida sem a data do Aviso de "
                    "Licitação — a fase está correta, falta lançar o aviso"
                )
            else:
                alertas["fase"] = (
                    "Fase Licitação sem Aviso de Licitação publicado — deveria "
                    "permanecer em Captação de Recurso"
                )

    # Prazo da Pendência / Tarefa vencido — uma tarefa/pendência em aberto
    # com prazo já ultrapassado precisa de atenção.
    #
    # Só que ela deixa de estar "em aberto" quando a ação já foi entregue:
    # com status CONCLUÍDA/INAUGURADA, ou com a FASE já CONCLUÍDA, o prazo
    # da pendência no passado é o registro de uma tarefa que se resolveu
    # junto com a obra — não um atraso. Antes a regra valia para todas as
    # linhas, e obras concluídas apareciam no Controle de Qualidade e com o
    # campo contornado de vermelho na Ficha Cadastral sem ter problema
    # nenhum.
    acao_ja_entregue = (
        status_atual in CAMPOS_ALERTA_STATUS_IGNORAR_VENCIMENTO
        or fase_atual in CAMPOS_ALERTA_FASE_IGNORAR_VENCIMENTO
    )
    if col_prazo_pendencia in row.index and not acao_ja_entregue:
        data_prazo_pendencia = _extrair_data_alerta(row.get(col_prazo_pendencia))
        if data_prazo_pendencia is not None and data_prazo_pendencia < hoje:
            alertas["prazo_pendencia"] = "Prazo da Pendência / Tarefa vencido"

    # Link Localização preenchido com endereço em vez do link do Google
    # Maps. Vale para qualquer fase/status, inclusive obra concluída: o
    # link é o que permite localizar a obra depois, então um endereço
    # digitado ali continua sendo um defeito de cadastro mesmo com a ação
    # entregue. Campo em branco não entra aqui — o alerta é sobre o que está
    # preenchido errado, não sobre o que falta preencher.
    if col_link_localizacao in row.index:
        valor_link_localizacao = row.get(col_link_localizacao)
        motivo_link_localizacao = _motivo_link_localizacao(valor_link_localizacao)
        # A cobrança do campo em branco em ANDAMENTO só faz sentido pra OBRA:
        # é o único TIPO que tem execução em campo pra localizar. EQUIPAMENTOS
        # e PROJETO ficam de fora da checagem — cobrar o link deles aqui geraria
        # alerta sobre algo que não se aplica ao tipo de ação.
        tipo_atual = str(row.get(col_tipo, "")).strip().upper() if col_tipo in row.index else ""
        if motivo_link_localizacao:
            alertas["link_localizacao"] = motivo_link_localizacao
        elif (
            status_atual == "ANDAMENTO"
            and tipo_atual not in {"EQUIPAMENTOS", "PROJETO"}
            and _texto_vazio(valor_link_localizacao)
        ):
            # Obra em ANDAMENTO é o único caso em que o campo em branco É o
            # problema: enquanto a obra está sendo executada, o link é o que
            # permite localizá-la em campo — sem ele, não dá pra fiscalizar.
            alertas["link_localizacao"] = "Obra em Andamento sem Link Localização preenchido"

    return alertas

def _montar_aviso_qualidade(df):
    # Retorna o texto do aviso de Controle de Qualidade (prazos vencidos ou
    # ações sem prazo definido), ou None se não houver nenhum problema. Não
    # mostra mais nenhuma janela aqui — quem chama decide como exibir (o
    # painel web mostra isso num modal em HTML).
    # --- CONTROLE DE QUALIDADE: prazos vencidos e inconsistências na base ---
    # Duas verificações independentes, cada ação (linha) pode acumular os
    # motivos que se aplicarem a ela:
    # 1) PRAZO DE CONCLUSÃO DA FASE e/ou PRAZO DA CLÁUSULA SUSPENSIVA
    #    (VIGÊNCIA) já vencidos em relação à data de hoje — não se aplica a
    #    ações já CONCLUÍDA/INAUGURADA.
    # 2) Ação na FASE "Execução do Objeto" com STATUS ANDAMENTO, CONCLUÍDA ou
    #    INAUGURADA, mas SEM o PRAZO DE CONCLUSÃO DA FASE preenchido — essa
    #    verificação se aplica mesmo às ações CONCLUÍDA/INAUGURADA.
    # É apenas um aviso — não impede a geração do PDF, salvo se o usuário
    # optar por Cancelar no painel exibido.
    hoje = datetime.now().date()

    # TODAS as ações da base passam pelo Controle de Qualidade. Não existe
    # mais lista de itens dispensados: a checagem é uniforme, e um item com
    # prazo vencido aparece no aviso independentemente do objeto ou do órgão
    # responsável.
    problemas_por_objeto = {}
    for _, row in df.iterrows():
        gestao_atual = str(row.get("GESTAO", "")).strip()
        obj = str(row[col_objeto])
        item = normalizar_item(row.get(col_item))

        motivos = set(_campos_alerta_qualidade(row, hoje).values())

        if motivos:
            chave = (gestao_atual, obj)
            registro = problemas_por_objeto.setdefault(
                chave, {"itens": set(), "motivos": set(), "itens_por_motivo": {}}
            )
            registro["motivos"].update(motivos)
            item_valido = item if (item and item.upper() != "NAN") else None
            if item_valido:
                registro["itens"].add(item_valido)
            # Cada motivo guarda os SEUS itens. Antes o objeto acumulava
            # todos os motivos numa lista só e todos os itens noutra, sem
            # ligação entre as duas — quem lia via "itens 47, 49, 51..." e
            # quatro tipos de pendência, mas não sabia qual item tinha qual.
            for motivo in motivos:
                registro["itens_por_motivo"].setdefault(motivo, set())
                if item_valido:
                    registro["itens_por_motivo"][motivo].add(item_valido)

    if problemas_por_objeto:
        # Agrupa os objetos com problema por GESTÃO (Estadual/Federal), na
        # mesma ordem usada no resto do relatório, e dentro de cada gestão
        # ordena os objetos alfabeticamente.
        ORDEM_GESTAO_QC = ["GESTÃO ESTADUAL", "GESTÃO FEDERAL"]
        objetos_por_gestao = {}
        for gestao_chave, obj in problemas_por_objeto:
            objetos_por_gestao.setdefault(gestao_chave, []).append(obj)
        gestoes_com_problema = sorted(
            objetos_por_gestao.keys(),
            key=lambda g: ORDEM_GESTAO_QC.index(g) if g in ORDEM_GESTAO_QC else 99,
        )

        LIMITE_LINHAS_AVISO = 30

        def _chave_ordenacao_item(valor):
            try:
                return (0, float(valor))
            except ValueError:
                return (1, valor)

        # Dados ESTRUTURADOS (não mais um texto único) — o JS monta o HTML
        # do modal a partir disso, transformando cada número de item num
        # link clicável que abre a Ficha Cadastral direto, sem precisar
        # digitar o item de novo no campo de busca.
        grupos = []
        for gestao_chave in gestoes_com_problema:
            objs_ordenados = sorted(objetos_por_gestao[gestao_chave], key=remover_acentos)
            linhas_grupo = []
            for obj in objs_ordenados[:LIMITE_LINHAS_AVISO]:
                info = problemas_por_objeto[(gestao_chave, obj)]
                itens_ordenados = sorted(info["itens"], key=_chave_ordenacao_item) if info["itens"] else []
                # Hierarquia OBJETO > ALERTA > ITENS: cada objeto vira um
                # tópico e, dentro dele, um subtópico por tipo de pendência,
                # com os itens que têm exatamente aquela pendência.
                alertas_do_objeto = [
                    {
                        "motivo": motivo,
                        "itens": sorted(itens_do_motivo, key=_chave_ordenacao_item),
                    }
                    for motivo, itens_do_motivo in sorted(info["itens_por_motivo"].items())
                ]
                linhas_grupo.append(
                    {
                        "objeto": obj,
                        "itens": itens_ordenados,
                        "alertas": alertas_do_objeto,
                        # Mantido para quem só precisa da lista corrida de
                        # motivos do objeto (compatibilidade).
                        "motivos": ", ".join(sorted(info["motivos"])),
                    }
                )
            qtd_restante = len(objs_ordenados) - LIMITE_LINHAS_AVISO
            grupos.append(
                {
                    "gestao": gestao_chave,
                    "linhas": linhas_grupo,
                    "qtd_restante": qtd_restante if qtd_restante > 0 else 0,
                }
            )

        qtd_total_itens = sum(
            len(info["itens"]) if info["itens"] else 1
            for info in problemas_por_objeto.values()
        )
        texto_item_total = "item" if qtd_total_itens == 1 else "itens"
        texto_verbo_total = "Foi encontrado" if qtd_total_itens == 1 else "Foram encontrados"
        return {
            "cabecalho": (
                f"Hoje: {hoje.strftime('%d/%m/%Y')}\n"
                f"Base de dados atualizada em: {ultima_atualizacao.strftime('%d/%m/%Y às %Hh%Mmin')}\n\n"
                f"{texto_verbo_total} {qtd_total_itens} {texto_item_total} com pendência(s) de qualidade na base:"
            ),
            "grupos": grupos,
            # Número de ações (itens) com alguma pendência de qualidade no
            # recorte — o painel web usa isso no badge do botão de Controle
            # de Qualidade ao lado da lupa.
            "total_itens": qtd_total_itens,
        }

    return None


def _dados_pre_visualizacao(df):
    # Prepara os MESMOS agrupamentos usados nos gráficos do Painel Geral do
    # PDF (gerar_grafico_prazo/financeiro/panorama/termo_compromisso), só
    # que devolvendo números simples (não desenhos do ReportLab) — o painel
    # web desenha os gráficos em CSS/JS a partir desses números.

    # Investimento AJUSTADO por linha — mesma regra usada em todo o resto
    # do relatório e do dashboard (Valor Contratado + OGU + Financiamento +
    # Contrapartida + Complementar, todos já ajustados pra não contar duas
    # vezes) — calculado uma vez aqui e reaproveitado em TODOS os
    # agrupamentos abaixo (panorama por fase, termo de compromisso, total
    # geral), em vez da coluna bruta "INVESTIMENTO TOTAL" da planilha, que
    # pode não estar sincronizada linha a linha com essas colunas.
    df = df.assign(
        _INVESTIMENTO_AJUSTADO=(
            df[col_valor_contratado]
            + df[col_apoiado_ajustado]
            + df[col_financiamento_ajustado]
            + df[col_contrapartida_ajustado]
            + df[col_complementar_ajustado]
        )
    )

    # --- Previsão de conclusão por trimestre ---
    contagem_prazo = df["PRAZO_TRIMESTRE"].value_counts().reset_index()
    contagem_prazo.columns = ["TRIMESTRE", "QTD"]
    contagem_prazo["ORDEM"] = contagem_prazo["TRIMESTRE"].apply(ordenar_trimestre)
    contagem_prazo = contagem_prazo.sort_values("ORDEM")
    prazo = []
    for _, r in contagem_prazo.iterrows():
        trimestre_nome = str(r["TRIMESTRE"])
        # Mesmo detalhamento OBJETO/INVESTIMENTO já usado no tooltip dos
        # mini-cards de Secretaria/Executor — aqui, um por trimestre em vez
        # de por fase.
        resumo_objetos_prazo = (
            df[df["PRAZO_TRIMESTRE"] == trimestre_nome]
            .groupby(col_objeto)
            .agg(QTD=(col_objeto, "count"), VALOR=("_INVESTIMENTO_AJUSTADO", "sum"))
            .reset_index()
            .sort_values("VALOR", ascending=False)
        )
        objetos_prazo = [
            {"objeto": str(o[col_objeto]), "qtd": int(o["QTD"]), "valor": float(o["VALOR"])}
            for _, o in resumo_objetos_prazo.iterrows()
        ]
        prazo.append({"rotulo": trimestre_nome, "qtd": int(r["QTD"]), "objetos": objetos_prazo})

    # --- Resumo financeiro ---
    financeiro_bruto = [
        ("VALOR CONTRATADO", float(df[col_valor_contratado].sum())),
        ("OGU", float(df[col_apoiado_ajustado].sum())),
        ("FINANCIAMENTO", float(df[col_financiamento_ajustado].sum())),
        (
            "RECURSO ESTADUAL",
            float(df[col_contrapartida_ajustado].sum() + df[col_complementar_ajustado].sum()),
        ),
    ]
    financeiro = [{"rotulo": nome, "valor": valor} for nome, valor in financeiro_bruto if valor > 0]

    # --- Panorama geral das fases ---
    resumo_fase = (
        df.groupby(["FASE_ORDEM", "FASE_TEXTO"])
        .agg(QTD=(col_objeto, "count"), VALOR=("_INVESTIMENTO_AJUSTADO", "sum"))
        .reset_index()
        .sort_values("FASE_ORDEM")
    )
    resumo_fase = resumo_fase[(resumo_fase["QTD"] > 0) & (resumo_fase["VALOR"] > 0)]
    panorama = []
    for _, r in resumo_fase.iterrows():
        fase_nome = str(r["FASE_TEXTO"])
        # Mesmo detalhamento OBJETO/INVESTIMENTO usado no tooltip do gráfico
        # de prazo — aqui, um por fase em vez de por trimestre.
        resumo_objetos_fase = (
            df[df["FASE_TEXTO"] == fase_nome]
            .groupby(col_objeto)
            .agg(QTD=(col_objeto, "count"), VALOR=("_INVESTIMENTO_AJUSTADO", "sum"))
            .reset_index()
            .sort_values("VALOR", ascending=False)
        )
        objetos_fase = [
            {"objeto": str(o[col_objeto]), "qtd": int(o["QTD"]), "valor": float(o["VALOR"])}
            for _, o in resumo_objetos_fase.iterrows()
        ]
        panorama.append(
            {"rotulo": fase_nome, "qtd": int(r["QTD"]), "valor": float(r["VALOR"]), "objetos": objetos_fase}
        )

    # --- Situação do termo de compromisso ---
    resumo_tc = (
        df.groupby("SINALIZACAO_TC")
        .agg(QTD=(col_objeto, "count"), VALOR=("_INVESTIMENTO_AJUSTADO", "sum"))
        .reindex(["SIM", "NÃO"])
        .fillna(0)
        .reset_index()
    )
    resumo_tc.columns = ["SITUACAO", "QTD", "VALOR"]
    resumo_tc = resumo_tc[(resumo_tc["QTD"] > 0) & (resumo_tc["VALOR"] > 0)]
    termo = []
    for _, r in resumo_tc.iterrows():
        situacao = str(r["SITUACAO"])
        # Mesmo detalhamento OBJETO/INVESTIMENTO usado no tooltip do gráfico
        # de prazo — aqui, um por situação do termo (SIM/NÃO).
        resumo_objetos_tc = (
            df[df["SINALIZACAO_TC"] == situacao]
            .groupby(col_objeto)
            .agg(QTD=(col_objeto, "count"), VALOR=("_INVESTIMENTO_AJUSTADO", "sum"))
            .reset_index()
            .sort_values("VALOR", ascending=False)
        )
        objetos_tc = [
            {"objeto": str(o[col_objeto]), "qtd": int(o["QTD"]), "valor": float(o["VALOR"])}
            for _, o in resumo_objetos_tc.iterrows()
        ]
        termo.append(
            {
                "rotulo": "Termo Assinado" if situacao == "SIM" else "Termo Não Assinado",
                "situacao": situacao,
                "qtd": int(r["QTD"]),
                "valor": float(r["VALOR"]),
                "objetos": objetos_tc,
            }
        )

    # --- Secretaria/Executor único no filtro atual ---
    # Quando o filtro deixa só uma combinação de SECRETARIA/EXECUTOR, esse
    # dado é destacado no topo da pré-visualização (fica fácil esquecer
    # qual secretaria está sendo vista quando o filtro já restringiu tanto).
    pares_unicos = df[["SECRETARIA_LIMPA", "EXECUTOR"]].drop_duplicates()
    secretaria_unica = None
    if len(pares_unicos) == 1:
        linha_unica = pares_unicos.iloc[0]
        secretaria_unica = {
            "secretaria": str(linha_unica["SECRETARIA_LIMPA"]),
            "executor": str(linha_unica["EXECUTOR"]),
        }

    # O total do card "INVESTIMENTO" usa a MESMA regra de valores ajustados
    # do Resumo Financeiro logo acima (VALOR CONTRATADO substitui
    # Financiamento/Apoiado/Contrapartida/Complementar quando preenchido,
    # evitando contar duas vezes) — mesma coluna _INVESTIMENTO_AJUSTADO já
    # calculada no início da função, reaproveitada aqui.
    investimento_total = float(df["_INVESTIMENTO_AJUSTADO"].sum())

    # --- Detalhamento financeiro por Secretaria (matriz: secretaria x
    # investimento/valor contratado/OGU/financiamento/recurso estadual) ---
    # O "investimento" aqui é CALCULADO como a soma dos componentes (não a
    # coluna bruta "INVESTIMENTO TOTAL" da planilha), pra garantir que o
    # total da barra bate exatamente com a soma dos segmentos empilhados —
    # mesmo princípio já usado no card "INVESTIMENTO" do topo do dashboard.
    df_sec_grp = (
        df.groupby("SECRETARIA_LIMPA")
        .agg(
            QTD=(col_objeto, "count"),
            VALOR_CONTRATADO=(col_valor_contratado, "sum"),
            APOIADO_AJ=(col_apoiado_ajustado, "sum"),
            CONTRAPARTIDA_AJ=(col_contrapartida_ajustado, "sum"),
            COMPLEMENTAR_AJ=(col_complementar_ajustado, "sum"),
            FINANCIAMENTO_AJ=(col_financiamento_ajustado, "sum"),
        )
        .reset_index()
    )
    df_sec_grp["RECURSO_ESTADUAL"] = df_sec_grp["CONTRAPARTIDA_AJ"] + df_sec_grp["COMPLEMENTAR_AJ"]
    df_sec_grp["INVESTIMENTO_CALC"] = (
        df_sec_grp["VALOR_CONTRATADO"]
        + df_sec_grp["APOIADO_AJ"]
        + df_sec_grp["RECURSO_ESTADUAL"]
        + df_sec_grp["FINANCIAMENTO_AJ"]
    )
    df_sec_grp = df_sec_grp[df_sec_grp["QTD"] > 0].sort_values("INVESTIMENTO_CALC", ascending=False)
    detalhamento_secretaria = [
        {
            "secretaria": str(r["SECRETARIA_LIMPA"]),
            "qtd": int(r["QTD"]),
            "investimento": float(r["INVESTIMENTO_CALC"]),
            "valorContratado": float(r["VALOR_CONTRATADO"]),
            "valorApoiadoOgu": float(r["APOIADO_AJ"]),
            "recursoEstadual": float(r["RECURSO_ESTADUAL"]),
            "financiamento": float(r["FINANCIAMENTO_AJ"]),
        }
        for _, r in df_sec_grp.iterrows()
    ]

    return {
        "qtd": int(len(df)),
        "investimento": investimento_total,
        "prazo": prazo,
        "financeiro": financeiro,
        "panorama": panorama,
        "termo": termo,
        "secretaria": _dados_panorama_secretaria(df),
        "detalhamentoSecretaria": detalhamento_secretaria,
        "desempenho": calcular_indice_desempenho_secretarias(
            df_original.assign(_INVESTIMENTO_AJUSTADO=_serie_investimento_ajustado(df_original)),
            combos_ativos=set(
                df[["SECRETARIA_LIMPA", "EXECUTOR"]]
                .dropna()
                .apply(lambda r: (str(r["SECRETARIA_LIMPA"]), str(r["EXECUTOR"])), axis=1)
            ),
        ),
        "mapa": _dados_mapa_municipios(df),
        "secretaria_unica": secretaria_unica,
    }


def _dados_panorama_secretaria(df):
    # Mesmo agrupamento usado em montar_pagina_panorama_secretaria (um
    # mini-gráfico de fases por combinação SECRETARIA/EXECUTOR), só que
    # devolvendo números simples em vez de montar o Drawing do ReportLab —
    # o painel web desenha cada mini-gráfico em CSS a partir desses números.
    # Separado por GESTÃO (Estadual primeiro, depois Federal), mesma ordem
    # usada no restante do relatório.
    ORDEM_GESTAO_PREVIEW = ["GESTÃO ESTADUAL", "GESTÃO FEDERAL"]

    # Investimento AJUSTADO por linha — mesma regra usada em todo o resto
    # do relatório e do dashboard, em vez da coluna bruta "INVESTIMENTO
    # TOTAL" (que pode não estar sincronizada linha a linha com os
    # componentes). Calculado aqui dentro pra função ficar auto-contida,
    # sem depender de quem a chama já ter feito isso antes.
    if "_INVESTIMENTO_AJUSTADO" not in df.columns:
        df = df.assign(
            _INVESTIMENTO_AJUSTADO=(
                df[col_valor_contratado]
                + df[col_apoiado_ajustado]
                + df[col_financiamento_ajustado]
                + df[col_contrapartida_ajustado]
                + df[col_complementar_ajustado]
            )
        )

    grupos = []
    for gestao in ORDEM_GESTAO_PREVIEW:
        df_gestao = df[df["GESTAO"] == gestao]
        if df_gestao.empty:
            continue

        pares = df_gestao[["SECRETARIA_LIMPA", "EXECUTOR"]].drop_duplicates().copy()
        pares["_SEC_SORT"] = pares["SECRETARIA_LIMPA"].apply(remover_acentos)
        pares["_EXE_SORT"] = pares["EXECUTOR"].apply(remover_acentos)
        pares = pares.sort_values(["_SEC_SORT", "_EXE_SORT"])

        cartoes = []
        for _, par in pares.iterrows():
            sec, exe = str(par["SECRETARIA_LIMPA"]), str(par["EXECUTOR"])
            df_par = df_gestao[(df_gestao["SECRETARIA_LIMPA"] == sec) & (df_gestao["EXECUTOR"] == exe)]
            resumo_fase = (
                df_par.groupby("FASE_TEXTO")
                .agg(QTD=(col_objeto, "count"), VALOR=("_INVESTIMENTO_AJUSTADO", "sum"))
                .reindex(ORDEM_FASES)
                .fillna(0)
                .reset_index()
            )
            resumo_fase = resumo_fase[(resumo_fase["QTD"] > 0) & (resumo_fase["VALOR"] > 0)]
            if resumo_fase.empty:
                continue

            fases_lista = []
            for _, r in resumo_fase.iterrows():
                fase_nome = str(r["FASE_TEXTO"])
                df_fase = df_par[df_par["FASE_TEXTO"] == fase_nome]
                resumo_objetos = (
                    df_fase.groupby(col_objeto)
                    .agg(QTD=(col_objeto, "count"), VALOR=("_INVESTIMENTO_AJUSTADO", "sum"))
                    .reset_index()
                    .sort_values("VALOR", ascending=False)
                )
                objetos_lista = [
                    {"objeto": str(o[col_objeto]), "qtd": int(o["QTD"]), "valor": float(o["VALOR"])}
                    for _, o in resumo_objetos.iterrows()
                ]
                fases_lista.append(
                    {
                        "rotulo": fase_nome,
                        "qtd": int(r["QTD"]),
                        "valor": float(r["VALOR"]),
                        "objetos": objetos_lista,
                    }
                )

            cartoes.append(
                {
                    "secretaria": sec,
                    "executor": exe,
                    "fases": fases_lista,
                }
            )
        if cartoes:
            grupos.append({"gestao": gestao, "itens": cartoes})

    return grupos


def _cor_reportlab_para_hex(cor):
    return "#%02X%02X%02X" % (round(cor.red * 255), round(cor.green * 255), round(cor.blue * 255))


def _dados_mapa_municipios(df, largura_mapa=760):
    # Mesma lógica de gerar_mapa_coropletico_municipios (projeção
    # equirretangular + classes fixas de cor por faixa de investimento), só
    # que devolvendo os polígonos já projetados (coordenadas x/y prontas)
    # em vez de desenhar num Drawing do ReportLab — o painel web desenha
    # cada polígono como um <polygon> de SVG a partir desses pontos. Se o
    # arquivo do mapa não for encontrado, devolve None (mesma tolerância a
    # falta de recurso usada no restante do relatório).
    geojson_data = carregar_geojson_municipios()
    if geojson_data is None:
        return None

    df_com_norm = df.assign(
        _MUN_NORM=df[col_municipio].apply(lambda x: remover_acentos(x).strip().upper()),
        _INV_AJ=_serie_investimento_ajustado(df),
    )
    resumo = df_com_norm.groupby("_MUN_NORM").agg(QTD=(col_objeto, "count"), VALOR=("_INV_AJ", "sum"))

    # Detalhamento por OBJETO dentro de cada município — usado no tooltip
    # (tabela OBJETO | INVESTIMENTO TOTAL ao passar o mouse), mesmo padrão
    # já usado no tooltip dos mini-cards de Secretaria/Executor.
    resumo_objetos = (
        df_com_norm.groupby(["_MUN_NORM", col_objeto])
        .agg(VALOR=("_INV_AJ", "sum"))
        .reset_index()
        .sort_values("VALOR", ascending=False)
    )
    objetos_por_municipio_norm = {}
    for _, linha in resumo_objetos.iterrows():
        objetos_por_municipio_norm.setdefault(linha["_MUN_NORM"], []).append(
            {"objeto": str(linha[col_objeto]), "valor": float(linha["VALOR"])}
        )

    minlon, minlat, maxlon, maxlat, correcao_x = _bbox_projecao_municipios(geojson_data["features"])
    largura_bruta = (maxlon - minlon) * correcao_x
    altura_bruta = maxlat - minlat
    escala = largura_mapa / largura_bruta
    altura_mapa = altura_bruta * escala

    def projetar_svg(lon, lat):
        x = (lon - minlon) * correcao_x * escala
        # SVG cresce para baixo (eixo Y invertido em relação ao PDF) — por
        # isso o "altura_mapa - y" abaixo, para o norte ficar em cima.
        y = altura_mapa - (lat - minlat) * escala
        return x, y

    dados_por_municipio = {}
    for feat in geojson_data["features"]:
        nome_norm = remover_acentos(feat["properties"]["name"]).strip().upper()
        if nome_norm in resumo.index:
            dados_por_municipio[feat["properties"]["name"]] = (
                int(resumo.loc[nome_norm, "QTD"]),
                float(resumo.loc[nome_norm, "VALOR"]),
                objetos_por_municipio_norm.get(nome_norm, []),
            )

    valores_presentes = [valor for _, valor, _ in dados_por_municipio.values()]
    bordas_classes = _calcular_bordas_classes_mapa(valores_presentes) if valores_presentes else [0.0, 0.0]
    cor_sem_dado_hex = _cor_reportlab_para_hex(COR_MAPA_SEM_DADO)
    cores_classes_hex = [_cor_reportlab_para_hex(c) for c in CORES_CLASSES_MAPA]

    poligonos = []
    for feat in geojson_data["features"]:
        nome_municipio = feat["properties"]["name"]
        dado = dados_por_municipio.get(nome_municipio)
        if dado:
            cor = cores_classes_hex[_indice_classe_mapa(dado[1], bordas_classes) % len(cores_classes_hex)]
        else:
            cor = cor_sem_dado_hex
        for anel in feat["geometry"]["coordinates"]:
            pontos = " ".join(
                f"{x:.2f},{y:.2f}" for x, y in (projetar_svg(lon, lat) for lon, lat in anel)
            )
            poligono = {"pontos": pontos, "cor": cor, "municipio": nome_municipio}
            if dado:
                poligono["qtd"] = dado[0]
                poligono["valor"] = dado[1]
                poligono["objetos"] = dado[2]
            poligonos.append(poligono)

    legenda = []
    for i in range(len(bordas_classes) - 1):
        legenda.append(
            {
                "cor": cores_classes_hex[i % len(cores_classes_hex)],
                "de": formatar_mi_bi(bordas_classes[i]).replace("&nbsp;", " "),
                "ate": formatar_mi_bi(bordas_classes[i + 1]).replace("&nbsp;", " "),
            }
        )

    return {
        "largura": largura_mapa,
        "altura": altura_mapa,
        "poligonos": poligonos,
        "legenda": legenda,
        "cor_sem_dado": cor_sem_dado_hex,
        "qtd_municipios_com_dado": len(dados_por_municipio),
        "qtd_municipios_total": len(geojson_data["features"]),
    }


def _linha_filtro_unico(df, colunas):
    # Junta, por " | ", o valor das categorias que têm um ÚNICO valor no
    # recorte atual. Categoria com mais de um valor fica de fora, já que não
    # haveria um nome só pra mostrar.
    #
    # Valor repetido em sequência entra uma vez só: é comum a secretaria ser
    # o próprio executor (SEINFRA executando obra da SEINFRA), e escrever
    # "SEINFRA | SEINFRA" não diz nada além do que "SEINFRA" já diz. A
    # comparação ignora caixa e acento, porque as duas colunas vêm de campos
    # diferentes da planilha e nem sempre são digitadas igual.
    partes = []
    for coluna in colunas:
        if coluna not in df.columns:
            continue
        valores = df[coluna].dropna().astype(str).str.strip()
        valores = valores[valores != ""].unique()
        if len(valores) != 1:
            continue
        valor = valores[0]
        if partes and remover_acentos(valor).upper() == remover_acentos(partes[-1]).upper():
            continue
        partes.append(valor)
    return " | ".join(partes)

# Ordem fixa das categorias na linha de filtros. A capa mostra a lista
# inteira; o rodapé das páginas de gestão repete a mesma linha SEM a
# gestão, que já aparece ali do lado direito no "GESTÃO ESTADUAL - Página
# 01 de 15" e ficaria escrita duas vezes na mesma faixa.
COLUNAS_LINHA_FILTRO_UNICO = ["GESTAO", "SECRETARIA_LIMPA", "EXECUTOR", "FASE_TEXTO", "STATUS_TEXTO"]
COLUNAS_LINHA_FILTRO_RODAPE = [c for c in COLUNAS_LINHA_FILTRO_UNICO if c != "GESTAO"]

def _linha_capa_filtro_unico(df):
    # Monta "GESTÃO | SECRETARIA | EXECUTOR | FASE | STATUS" pra capa.
    return _linha_filtro_unico(df, COLUNAS_LINHA_FILTRO_UNICO)


def _miniatura_detalhamento_financeiro(df_secretaria, largura, altura_barra=11):
    # Miniatura da barra empilhada do Detalhamento Financeiro para uma
    # única secretaria — mesmos componentes e a mesma regra de valores
    # ajustados da página cheia, só que sem texto nenhum. Devolve None
    # quando não há investimento a decompor: uma barra inteira cinza não
    # informaria nada.
    valor_contratado = float(df_secretaria[col_valor_contratado].sum())
    apoiado_ogu = float(df_secretaria[col_apoiado_ajustado].sum())
    recurso_estadual = float(
        df_secretaria[col_contrapartida_ajustado].sum()
        + df_secretaria[col_complementar_ajustado].sum()
    )
    financiamento = float(df_secretaria[col_financiamento_ajustado].sum())
    investimento = valor_contratado + apoiado_ogu + recurso_estadual + financiamento
    if investimento <= 0 or largura <= 0:
        return None
    # Aqui a barra não decompõe por fonte de recurso — só verde (Valor
    # Contratado) sobre o fundo cinza claro, mostrando a fatia já
    # contratada do investimento total da secretaria. Por isso OGU/Recurso
    # Estadual/Financiamento entram no denominador (investimento) mas não
    # como segmentos coloridos: o fundo cinza que a própria função já
    # desenha cobre o restante sozinho.
    item = {
        "investimento": investimento,
        "valorContratado": valor_contratado,
        "valorApoiadoOgu": 0.0,
        "recursoEstadual": 0.0,
        "financiamento": 0.0,
        "fontesFinanciamento": [],
    }
    percentual_contratado = valor_contratado / investimento * 100
    return desenhar_barra_detalhamento_secretaria(
        item, largura, altura_barra=altura_barra, mostrar_detalhe=False,
        # Percentual do investimento da secretaria que já virou contrato —
        # é a leitura que a barra permite de relance, e escrevê-la evita
        # que quem olha tenha que estimar no olho o tamanho do primeiro
        # segmento.
        texto_acima=f"{percentual_contratado:.0f}% contratado",
    )

def _cabecalho_secretaria_detalhamento(texto, indice_categoria, miniatura=None, recuo_miniatura=0):
    # Cabeçalho de cada secretaria no Detalhamento Analítico: o nome com a
    # contagem e o investimento, o medidor do Índice de Desempenho, a nota
    # em texto ("44 (Regular)") e, mais à direita, a miniatura da barra do
    # Detalhamento Financeiro alinhada à coluna INVESTIMENTO da tabela que
    # vem logo abaixo.
    #
    # Sem índice e sem miniatura, o cabeçalho volta a ser só o parágrafo de
    # antes — nada de espaço vazio reservado para um gráfico que não existe.
    if not indice_categoria and miniatura is None:
        return Paragraph(texto, eixo_style)

    celulas = [Paragraph(texto, eixo_style)]
    larguras = [0]  # a largura da primeira coluna é definida no fim
    estilo = [
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 0),
        ("RIGHTPADDING", (0, 0), (-1, -1), 0),
        ("TOPPADDING", (0, 0), (-1, -1), 0),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]

    if indice_categoria:
        indice, categoria = indice_categoria
        celulas.append(
            gerar_medidor_desempenho(
                texto, indice, categoria,
                largura=LARGURA_MEDIDOR_CABECALHO_SECRETARIA,
                altura=ALTURA_MEDIDOR_CABECALHO_SECRETARIA,
                mostrar_categoria=False,
            )
        )
        larguras.append(LARGURA_MEDIDOR_CABECALHO_SECRETARIA)
        cor_texto = colors.HexColor("#9A9A9A")
        if categoria in CORES_CATEGORIA_DESEMPENHO:
            cor_texto = _cor_texto_categoria(CORES_CATEGORIA_DESEMPENHO[categoria])
        # Estilo clonado só para trocar a cor — mais seguro que embutir uma
        # tag <font> no texto, que exigiria escapar o valor hexadecimal.
        celulas.append(
            Paragraph(
                categoria,
                ParagraphStyle(
                    f"nota_indice_{categoria}",
                    parent=nota_indice_cabecalho_style,
                    textColor=cor_texto,
                ),
            )
        )
        larguras.append(
            pdfmetrics.stringWidth(
                categoria, FONTE_PADRAO_NEGRITO, nota_indice_cabecalho_style.fontSize
            ) + 8
        )
        estilo.append(("LEFTPADDING", (2, 0), (2, 0), 6))

    # A coluna do texto é medida pelo próprio texto, e não pela largura da
    # página: assim o medidor encosta no fim do nome da secretaria em vez de
    # ficar isolado na margem direita. O "&nbsp;" que moeda_sem_quebra usa
    # vira espaço só para a medição — no Paragraph ele continua sendo a
    # entidade, que é o que impede a quebra do valor.
    largura_texto = pdfmetrics.stringWidth(
        texto.replace("&nbsp;", " "), FONTE_PADRAO_NEGRITO, eixo_style.fontSize
    ) + 6
    largura_ocupada = sum(larguras)

    if miniatura is not None:
        # recuo_miniatura é a distância da margem esquerda até onde começa a
        # coluna INVESTIMENTO da tabela de baixo. A miniatura precisa cair
        # exatamente ali, senão o alinhamento que motiva ela deixa de
        # existir — então, se o nome da secretaria mais o medidor passarem
        # desse ponto, é a coluna do texto que cede (o Paragraph quebra em
        # duas linhas). Só quando nem assim sobra espaço razoável é que a
        # miniatura vem logo depois, sem alinhamento.
        folga = recuo_miniatura - largura_ocupada - largura_texto
        if folga >= 0:
            celulas.append("")
            larguras.append(folga)
        elif recuo_miniatura - largura_ocupada > 120:
            largura_texto = recuo_miniatura - largura_ocupada
        else:
            celulas.append("")
            larguras.append(8)
        celulas.append(miniatura)
        larguras.append(miniatura.width)

    larguras[0] = largura_texto
    cabecalho = Table([celulas], colWidths=larguras, hAlign="LEFT")
    cabecalho.setStyle(TableStyle(estilo))
    # keepWithNext do eixo_style não atravessa a tabela — sem isso, o nome
    # da secretaria poderia ficar sozinho no pé de uma página, com a tabela
    # dela começando só na seguinte.
    cabecalho.keepWithNext = True
    return cabecalho

# --- Seções do relatório, na ordem em que aparecem no PDF ---
# "chave" é o que trafega entre o painel e a geração; "titulo" é o que a
# janela de seleção mostra; "layout" escolhe o desenho da miniatura.
#
# Nem toda seção existe em toda geração: a de Mapa depende do geojson, a de
# Resumo Geral só sai com as duas gestões no recorte, e assim por diante.
# Quem decide isso é _secoes_disponiveis_relatorio, logo abaixo, para a
# janela não oferecer página que não vai existir.
SECOES_RELATORIO = [
    {"chave": "CAPA", "titulo": "Capa", "layout": "capa"},
    {"chave": "OBSERVACAO", "titulo": "Observações", "layout": "texto"},
    {"chave": "PAINEL_GERAL", "titulo": "Painel Geral", "layout": "graficos"},
    {"chave": "PANORAMA", "titulo": "Panorama por Secretaria", "layout": "tabela"},
    {"chave": "METODOLOGIA", "titulo": "Metodologia do Índice", "layout": "texto_tabelas"},
    {"chave": "INDICE_DESEMPENHO", "titulo": "Índice de Desempenho", "layout": "medidores"},
    {"chave": "DETALHAMENTO_FINANCEIRO", "titulo": "Detalhamento Financeiro", "layout": "barras"},
    {"chave": "MAPA", "titulo": "Mapa dos Municípios", "layout": "mapa"},
    {"chave": "RESUMO_GERAL", "titulo": "Resumo Executivo Geral", "layout": "cards_tabelas"},
    {"chave": "POR_STATUS", "titulo": "Visão por Status", "layout": "cards_tabelas"},
    {"chave": "POR_FASE", "titulo": "Visão por Fase", "layout": "cards_tabela"},
    {"chave": "DETALHAMENTO", "titulo": "Detalhamento Analítico", "layout": "detalhamento"},
]
SECOES_RELATORIO_POR_CHAVE = {sec["chave"]: sec for sec in SECOES_RELATORIO}

def normalizar_secoes_relatorio(secoes):
    # None ou vazio = relatório completo. Ignora chaves desconhecidas, para
    # uma versão futura do painel não quebrar a geração.
    if not secoes:
        return set(SECOES_RELATORIO_POR_CHAVE)
    validas = {str(c) for c in secoes if str(c) in SECOES_RELATORIO_POR_CHAVE}
    return validas or set(SECOES_RELATORIO_POR_CHAVE)

def _secoes_disponiveis_relatorio(df):
    # Quais seções o recorte ATUAL realmente produziria, com a contagem de
    # páginas de cada uma. É o que a janela de seleção lista.
    #
    # Cada regra aqui espelha uma condição que existe dentro de _gerar_pdf.
    # Elas precisam andar juntas: se a geração passar a suprimir mais uma
    # página em algum caso, a regra correspondente entra aqui também —
    # senão a janela oferece uma página que o PDF não vai conter, e quem
    # marcou fica sem entender por que ela não saiu.
    if df is None or df.empty:
        return []

    gestoes = sorted({str(g).strip() for g in df["GESTAO"].dropna().unique() if str(g).strip()})
    n_gestoes = max(1, len(gestoes))

    # DETALHAMENTO FINANCEIRO: num recorte só de ANDAMENTO a página não é
    # mais suprimida — ela apenas chega desmarcada na janela de seleção.
    status_no_recorte = {
        str(st).strip().upper()
        for st in df["STATUS_TEXTO"].dropna().unique()
        if str(st).strip()
    }
    so_andamento = status_no_recorte == {"ANDAMENTO"}

    # PANORAMA: existe só se houver par SECRETARIA | EXECUTOR no recorte, e
    # pode virar uma página por gestão quando a grade não cabe numa só.
    try:
        n_pares_panorama, altura_panorama = estimar_altura_pagina_panorama_secretaria(df)
    except Exception:
        n_pares_panorama, altura_panorama = 1, 0

    # MAPA: depende do geojson e de haver município reconhecido no recorte.
    tem_mapa = carregar_geojson_municipios() is not None

    # ÍNDICE DE DESEMPENHO: com uma gestão só no recorte, a página da outra
    # é suprimida (mesma regra aplicada na geração).
    try:
        grupos_indice = calcular_indice_desempenho_secretarias(
            df_original.assign(_INVESTIMENTO_AJUSTADO=_serie_investimento_ajustado(df_original))
        )
    except Exception:
        grupos_indice = []
    if len(gestoes) == 1:
        grupos_indice = [g for g in grupos_indice if str(g["gestao"]).strip() in set(gestoes)]
    paginas_indice = len([g for g in grupos_indice if g.get("itens")])

    disponiveis = []
    for secao in SECOES_RELATORIO:
        chave = secao["chave"]
        paginas = 1

        if chave == "MAPA" and not tem_mapa:
            continue
        if chave == "PANORAMA":
            if n_pares_panorama <= 0:
                continue
            paginas = 1 if altura_panorama <= ALTURA_UTIL else n_gestoes
        if chave == "RESUMO_GERAL" and len(gestoes) <= 1:
            # Com uma gestão só, esta página fica idêntica à primeira da
            # Visão por Status — o próprio relatório já a omite.
            continue
        if chave == "INDICE_DESEMPENHO":
            if paginas_indice <= 0:
                continue
            paginas = paginas_indice
        if chave in ("POR_STATUS", "POR_FASE"):
            paginas = n_gestoes
        if chave == "DETALHAMENTO":
            # Muitas páginas, quantidade imprevisível: a janela mostra uma
            # miniatura só, como representação da seção inteira.
            paginas = None

        # Detalhamento Financeiro num recorte só de ANDAMENTO: a página
        # continua disponível, mas chega desmarcada, porque a decomposição
        # do investimento acrescenta pouco nesse corte. Antes ela era
        # suprimida à força e não havia como pedi-la.
        marcada = not (chave == "DETALHAMENTO_FINANCEIRO" and so_andamento)

        disponiveis.append(
            {
                "chave": chave,
                "titulo": secao["titulo"],
                "layout": secao["layout"],
                "paginas": paginas,
                "marcada": marcada,
            }
        )
    return disponiveis

def _gerar_pdf(df, arquivo_pdf, colunas_detalhamento=None, secoes=None):
    # Gera o PDF em si (agrupamento + montagem via ReportLab) para o
    # dataframe já filtrado, salvando no caminho informado. Levanta exceção
    # em caso de erro — quem chama decide como reportar ao usuário (o painel
    # web mostra isso como um alerta); não usa mais messagebox aqui.
    #
    # colunas_detalhamento: seleção de colunas da tabela de DETALHAMENTO
    # vinda do painel. Passa por normalizar_colunas_detalhamento aqui, na
    # entrada, e não lá na interface: qualquer caminho que chame esta função
    # (painel, teste, chamada direta) recebe o mesmo tratamento de
    # dependências, ordem e limite.
    colunas_detalhamento = normalizar_colunas_detalhamento(colunas_detalhamento)
    # secoes: chaves de SECOES_RELATORIO que devem entrar no PDF. None = tudo.
    secoes_ativas = normalizar_secoes_relatorio(secoes)

    def secao_ativa(chave):
        return chave in secoes_ativas

    # Trabalha numa cópia: mais abaixo esta função CRIA colunas derivadas
    # (EMISSAO_OS_TEXTO) e preenche vazios com fillna direto no dataframe.
    # Hoje quem chama sempre passa um recorte recém-filtrado, então ninguém
    # se machuca; mas o dia em que alguém passar o df_original, ele voltaria
    # alterado dessa chamada — e o efeito só apareceria no relatório
    # seguinte, que é o tipo de bug que custa caro pra encontrar.
    df = df.copy()
    if True:

        colunas_preenchimento = [
            "GESTAO",
            "FASE_TEXTO",
            col_eixo,
            "EIXO_SORT",
            col_objeto,
            "STATUS_TEXTO",
            col_fonte,
            "SINALIZACAO_TC",
            "PRAZO_FASE_TEXTO",
            "AVANCO_OBRA_TEXTO",
        ]
        for c in colunas_preenchimento:
            df[c] = df[c].fillna("").astype(str)

        # Chaves FIXAS: identificam a seção do relatório (gestão, secretaria,
        # eixo) e a hierarquia da tabela (objeto > fase > status), que também
        # comanda a ordenação e as mesclas. Ficam sempre, mesmo que FASE ou
        # STATUS não estejam entre as colunas escolhidas.
        _colunas_agrupamento_analitico = [
            "GESTAO",
            "SECRETARIA_LIMPA",
            "EIXO_SORT",
            col_eixo,
            col_objeto,
            "FASE_ORDEM",
            "FASE_TEXTO",
            "STATUS_ORDEM",
            "STATUS_TEXTO",
        ]
        # Chaves de ATRIBUTO: entram só quando a coluna correspondente está
        # sendo exibida. A regra passa a ser simples de enunciar — duas ações
        # ocupam a mesma linha quando tudo o que a tabela MOSTRA sobre elas é
        # igual. Antes, fonte, termo, prazo e avanço eram chave sempre, então
        # ações separadas por uma diferença invisível (uma coluna fora da
        # seleção) ficavam em linhas distintas sem que nada na página
        # explicasse o motivo.
        #
        # MUNICÍPIOS e INVESTIMENTO não entram aqui de propósito: são as duas
        # colunas que sabem se juntar sozinhas, uma pela contagem
        # ("Salvador (3)") e a outra pela soma.
        for _chave_atributo, _colunas_atributo in (
            ("FONTE", [col_fonte, col_fonte_financiamento]),
            ("TERMO", ["SINALIZACAO_TC"]),
            ("PRAZO_FASE", ["PRAZO_FASE_TEXTO"]),
            ("AVANCO", ["AVANCO_OBRA_TEXTO"]),
        ):
            if _chave_atributo not in colunas_detalhamento:
                continue
            for _coluna_atributo in _colunas_atributo:
                # FONTE DE FINANCIAMENTO é impressa junto de FONTE DE
                # RECURSO na mesma célula, por isso acompanha essa chave —
                # e, como toda coluna opcional, só entra se existir mesmo
                # na planilha.
                if _coluna_atributo in df.columns:
                    _colunas_agrupamento_analitico.append(_coluna_atributo)
        # EMISSÃO DE O.S. entra como TEXTO já formatado (dd/mm/aa) antes do
        # agrupamento, e não como data crua: assim ela atravessa o groupby
        # do mesmo jeito que as colunas de acompanhamento.
        df[COLUNA_EMISSAO_OS_TEXTO] = (
            df[col_emissao_os].apply(formatar_prazo)
            if col_emissao_os in df.columns
            else ""
        )

        # As colunas de acompanhamento (Pendências, Providências, Próximos
        # Passos) e a EMISSÃO DE O.S. precisam ATRAVESSAR o agrupamento
        # acima, senão chegam vazias na tabela.
        #
        # Quando a coluna ESTÁ na seleção do painel, ela entra como CHAVE de
        # agrupamento: duas ações do mesmo objeto/fase/status com pendências
        # (ou datas de O.S.) diferentes viram duas linhas, cada uma com a
        # sua providência ao lado. Antes elas viravam uma linha só, com os
        # textos distintos unidos por " | " na mesma célula, o que
        # embaralhava qual providência pertencia a qual pendência e
        # empilhava duas datas de O.S. dentro do mesmo quadrinho.
        #
        # Quando a coluna NÃO está na seleção, ela continua sendo agregada
        # por junção dos textos distintos: dividir linhas por um conteúdo
        # que não aparece na tabela produziria linhas visualmente idênticas
        # e repetidas, com o investimento fatiado entre elas sem explicação
        # visível.
        _colunas_travessia_chave = []
        for _chave_travessia, _coluna_travessia in CHAVE_COLUNA_TRAVESSIA.items():
            if _coluna_travessia not in df.columns:
                continue
            if _chave_travessia in colunas_detalhamento:
                df[_coluna_travessia] = df[_coluna_travessia].apply(
                    texto_acompanhamento_chave
                )
                _colunas_agrupamento_analitico.append(_coluna_travessia)
                _colunas_travessia_chave.append(_coluna_travessia)

        agregacoes_analiticas = {
            col_municipio: formatar_municipios_limpo,
            col_valor_contratado: "sum",
            col_apoiado_ajustado: "sum",
            col_financiamento_ajustado: "sum",
            col_contrapartida_ajustado: "sum",
            col_complementar_ajustado: "sum",
        }
        for _coluna_travessia in COLUNAS_TRAVESSIA_DETALHAMENTO:
            if (
                _coluna_travessia in df.columns
                and _coluna_travessia not in _colunas_travessia_chave
            ):
                agregacoes_analiticas[_coluna_travessia] = juntar_textos_distintos

        df_agrupado = (
            df.groupby(_colunas_agrupamento_analitico, dropna=False)
            .agg(agregacoes_analiticas)
            .reset_index()
        )

        df_counts_objeto = (
            df.groupby(["GESTAO", col_eixo, col_objeto])[col_objeto].count().to_dict()
        )
        df_counts_objeto = {(str(k[0]), str(k[1]), str(k[2])): v for k, v in df_counts_objeto.items()}

        df_counts_fase = (
            df.groupby(["GESTAO", col_eixo, col_objeto, "FASE_TEXTO"])["FASE_TEXTO"]
            .count()
            .to_dict()
        )
        df_counts_fase = {(str(k[0]), str(k[1]), str(k[2]), str(k[3])): v for k, v in df_counts_fase.items()}

        df_counts_status = (
            df.groupby(["GESTAO", col_eixo, col_objeto, "FASE_TEXTO", "STATUS_TEXTO"])["STATUS_TEXTO"]
            .count()
            .to_dict()
        )
        df_counts_status = {(str(k[0]), str(k[1]), str(k[2]), str(k[3]), str(k[4])): v for k, v in df_counts_status.items()}
        doc = SimpleDocTemplate(
            arquivo_pdf,
            pagesize=PAGINA,
            topMargin=MARGEM_SUP,
            bottomMargin=MARGEM_INF,
            leftMargin=MARGEM_ESQ,
            rightMargin=MARGEM_DIR,
        )
        elements = []
        # Estado compartilhado entre seções, inicializado ANTES de qualquer
        # bloco condicional. Com a seleção de páginas, uma seção pode ser
        # desmarcada e outra que depende dela continuar ligada — e uma
        # variável que só nascesse dentro do bloco desmarcado derrubaria a
        # geração com UnboundLocalError. Os valores de verdade são
        # calculados logo abaixo, fora dos "ifs"; estes aqui só garantem que
        # o nome exista em qualquer caminho.
        grupos_indice_desempenho = []
        indices_por_secretaria = {}

        # Precisa vir cedo (antes até da capa) porque também é usado como
        # rótulo padrão de gestão na numeração de páginas — todas as
        # páginas "combinadas" (que não pertencem a uma gestão específica:
        # capa, observação, painel geral etc.) usam esse rótulo pra
        # título — mas não pra numeração de página, já que a capa e as
        # páginas de gráfico não devem ser numeradas.
        gestoes_presentes_titulo = sorted(df["GESTAO"].dropna().unique().tolist())
        if len(gestoes_presentes_titulo) == 1:
            sufixo_titulo_painel = gestoes_presentes_titulo[0]
        else:
            sufixo_titulo_painel = "GESTÃO FEDERAL E ESTADUAL"
        elements.append(_MarcadorGestao(None))

        if secao_ativa("CAPA"):
            # --- ETAPA A: CAPA DO RELATÓRIO ---
            # O texto do CGAPE não entra aqui — é desenhado à parte, do lado
            # esquerdo da folha, em desenhar_background_capa (ver mais acima),
            # pra ficar alinhado com a altura do emblema do estado.
            elements.append(Spacer(1, ALTURA_UTIL * 0.18))
            elements.append(Paragraph("<b>PAC</b>", capa_projeto))
            elements.append(Spacer(1, 20))
            elements.append(Paragraph("RELATÓRIO GERENCIAL - BAHIA", capa_relatorio))
            elements.append(Spacer(1, 142))
            linha_filtro_unico_capa = _linha_capa_filtro_unico(df)
            # Mesma linha da capa, sem a gestão, para o rodapé de todas as
            # páginas de gestão. Definida aqui e não dentro do canvas porque é
            # o dataframe já filtrado deste PDF que decide o conteúdo.
            global TEXTO_FILTROS_RODAPE
            TEXTO_FILTROS_RODAPE = _linha_filtro_unico(df, COLUNAS_LINHA_FILTRO_RODAPE)
            if linha_filtro_unico_capa:
                elements.append(Spacer(1, 6))
                elements.append(Paragraph(linha_filtro_unico_capa, capa_filtro_unico))
            elements.append(Spacer(1, 8))
            elements.append(Paragraph(data_capa_txt, capa_data))
            elements.append(PageBreak())

        if secao_ativa("OBSERVACAO"):
            # --- ETAPA B: PÁGINA DE OBSERVAÇÕES ---
            # Coluna mais estreita que a largura total (88%) — dá um respiro
            # visível nas laterais e evita parágrafos esticados de ponta a
            # ponta numa página bem larga (16:9), além de sobrar altura pra
            # espaçar mais os elementos verticalmente.
            LARGURA_OBSERVACOES = LARGURA_UTIL * 0.88

            conteudo_observacoes = [
                Paragraph(
                    "O relatório mostra o monitoramento das entregas NOVO PAC em dois âmbitos: por <b>FASES</b> e por <b>STATUS</b>.",
                    obs_texto,
                ),
                Spacer(1, 15),
                Paragraph(
                    "<b>FASE:</b> Representa a situação atual da formalização da proposta de cada obra/projeto junto ao Ministério, Mandatária ou Ente Financiador.",
                    obs_texto,
                ),
                Spacer(1, 15),
                Paragraph(
                    "<b>STATUS:</b> Representa a situação da obra/projeto junto ao órgão executor ou da seleção da proposta junto ao Governo Federal.",
                    obs_texto,
                ),
                Spacer(1, 15),
                Paragraph(
                    "Portanto, algumas ações podem estar com status mais avançados, no entanto, em relação às fases algumas etapas ainda serão cumpridas.",
                    obs_texto,
                ),
                Spacer(1, 15),
                Paragraph(
                    "Exemplo: Temos obras que já estão com STATUS de licitação ou em andamento que ainda não venceram a FASE de captação de recurso.",
                    obs_texto,
                ),
                Spacer(1, 26),
                # Gráfico de FASE (mantido) + tabela de referência FASE/STATUS
                # (substitui o gráfico de pizza do STATUS) — dinâmico: mostra
                # só as fases/status que aparecem no filtro atual do relatório,
                # centralizados na página, lado a lado: FASE à esquerda, tabela
                # de STATUS à direita.
                gerar_graficos_observacao(df),
                Spacer(1, 26),
                Paragraph("<b>OUTRAS OBSERVAÇÕES:</b>", obs_texto),
                Spacer(1, 15),
                Paragraph(
                    " - Para o equipamento Aeroporto de Barreiras estão sendo contabilizadas 4 ações referentes às licitações que compõem a execução.",
                    obs_texto,
                ),
                Spacer(1, 15),
                Paragraph(
                    " - Para alguns lotes da FIOL ainda não obtivemos informação de expectativa de valores à licitar.",
                    obs_texto,
                ),
                Spacer(1, 15),
                Paragraph(
                    " - Para as ações sob a Gestão Federal não obtivemos informação quanto a formalização de contrato para captação do recurso federal.",
                    obs_texto,
                ),
            ]

            tabela_observacoes = Table(
                [[conteudo_observacoes]], colWidths=[LARGURA_OBSERVACOES]
            )
            tabela_observacoes.setStyle(
                TableStyle(
                    [
                        ("VALIGN", (0, 0), (-1, -1), "TOP"),
                        ("LEFTPADDING", (0, 0), (-1, -1), 0),
                        ("RIGHTPADDING", (0, 0), (-1, -1), 0),
                        ("TOPPADDING", (0, 0), (-1, -1), 0),
                        ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
                    ]
                )
            )
            # Título + conteúdo inteiro centralizados como um bloco único na
            # página (não cada linha de texto individualmente) — mesma técnica
            # do gráfico de FASE/tabela de STATUS logo abaixo: um wrapper com
            # ALIGN CENTER via TableStyle, já que hAlign de um flowable sozinho
            # não é respeitado quando aninhado dentro de outra tabela.
            bloco_observacoes = Table(
                [[Paragraph("OBSERVAÇÃO", obs_titulo)], [Spacer(1, 20)], [tabela_observacoes]],
                colWidths=[LARGURA_OBSERVACOES],
            )
            bloco_observacoes.setStyle(
                TableStyle(
                    [
                        ("LEFTPADDING", (0, 0), (-1, -1), 0),
                        ("RIGHTPADDING", (0, 0), (-1, -1), 0),
                        ("TOPPADDING", (0, 0), (-1, -1), 0),
                        ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
                    ]
                )
            )
            wrapper_observacoes = Table([[bloco_observacoes]], colWidths=[LARGURA_UTIL])
            wrapper_observacoes.setStyle(
                TableStyle(
                    [
                        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                        ("LEFTPADDING", (0, 0), (-1, -1), 0),
                        ("RIGHTPADDING", (0, 0), (-1, -1), 0),
                        ("TOPPADDING", (0, 0), (-1, -1), 0),
                        ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
                    ]
                )
            )

            elements.append(wrapper_observacoes)
            elements.append(PageBreak())

        if secao_ativa("PAINEL_GERAL"):
            # --- ETAPA B2: PAINEL GERAL (GRÁFICOS) ---
            # sufixo_titulo_painel e gestoes_presentes_titulo já foram
            # calculados mais acima (logo no início da função) — usa o mesmo
            # valor aqui, sem recalcular.

            def _titulo_grafico_painel(texto, marcador_direita=None):
                # Mesma formatação (barra azul clara, texto à esquerda) usada nos
                # títulos "SECRETARIA | EXECUTOR" do Panorama por Secretaria —
                # também mais baixa que o título centralizado usado antes aqui.
                # Quando há um marcador (ex: "**"), ele fica numa coluna própria
                # à direita da mesma barra, sem disputar alinhamento com o texto
                # principal do título.
                if marcador_direita:
                    linha = [
                        Paragraph(texto, cabecalho_grafico_painel_style),
                        Paragraph(marcador_direita, cabecalho_grafico_painel_style_direita),
                    ]
                    larguras = [LARGURA_UTIL / 2 - 8 - 155, 155]
                else:
                    linha = [Paragraph(texto, cabecalho_grafico_painel_style)]
                    larguras = [LARGURA_UTIL / 2 - 8]
                tabela_titulo = Table([linha], colWidths=larguras)
                tabela_titulo.setStyle(
                    TableStyle(
                        [
                            ("BACKGROUND", (0, 0), (-1, -1), COR_MARCA_TEAL_CLARA),
                            ("LEFTPADDING", (0, 0), (0, -1), 6),
                            ("RIGHTPADDING", (-1, 0), (-1, -1), 8),
                            ("TOPPADDING", (0, 0), (-1, -1), 3),
                            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
                        ]
                    )
                )
                return tabela_titulo

            tabela_painel_geral = Table(
                [
                    [
                        _titulo_grafico_painel("PREVISÃO DE CONCLUSÃO DA FASE"),
                        _titulo_grafico_painel("RESUMO FINANCEIRO", marcador_direita="** Previsão Orçamentária"),
                    ],
                    [gerar_grafico_prazo(df), gerar_grafico_financeiro(df)],
                    [
                        _titulo_grafico_painel("PANORAMA GERAL DAS FASES*"),
                        _titulo_grafico_painel("SITUAÇÃO DO TERMO DE COMPROMISSO*"),
                    ],
                    [gerar_grafico_panorama(df), gerar_grafico_termo_compromisso(df)],
                ],
                colWidths=[LARGURA_UTIL / 2, LARGURA_UTIL / 2],
            )
            tabela_painel_geral.setStyle(
                TableStyle(
                    [
                        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                        ("BOX", (0, 0), (-1, -1), 0.5, colors.grey),
                        ("INNERGRID", (0, 0), (-1, -1), 0.5, colors.lightgrey),
                        ("TOPPADDING", (0, 0), (-1, -1), 4),
                        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                    ]
                )
            )
            elements.append(
                KeepTogether(
                    [
                        Paragraph(f"PAC - PAINEL GERAL - {sufixo_titulo_painel}", titulo_style),
                        Spacer(1, 10),
                        gerar_cards_resumo(len(df), _investimento_ajustado(df)),
                        Spacer(1, 4),
                        Paragraph(
                            "* O tamanho de cada fatia do gráfico corresponde ao valor financeiro "
                            "(R$), e os valores entre parênteses ( ) indicam a quantidade de ações.",
                            orientacao_leitura_style,
                        ),
                        Spacer(1, 4),
                        tabela_painel_geral,
                    ]
                )
            )
            elements.append(PageBreak())

        if secao_ativa("PANORAMA"):
            # --- ETAPA B3: PANORAMA POR SECRETARIA | EXECUTOR ---
            # Tenta primeiro uma única página combinando as duas gestões (mesmo
            # texto dinâmico de título usado no Painel Geral: "GESTÃO FEDERAL E
            # ESTADUAL" quando há as duas, ou só o nome da gestão quando há
            # apenas uma). Se a grade de mini-gráficos não couber inteira numa
            # página, divide em uma página por gestão (ESTADUAL, depois FEDERAL).
            # A decisão usa uma ESTIMATIVA de altura (sem desenhar nenhum
            # gráfico) — só depois de decidido é que a página escolhida é
            # realmente montada e desenhada, uma única vez.
            n_pares_total, altura_estimada_combinada = estimar_altura_pagina_panorama_secretaria(df)
            if n_pares_total > 0:
                titulo_secretaria_exec = f"PAC - PANORAMA POR SECRETARIA | EXECUTOR* - {sufixo_titulo_painel}"
                if altura_estimada_combinada <= ALTURA_UTIL:
                    elementos_panorama_secretaria = montar_pagina_panorama_secretaria(
                        df, titulo_secretaria_exec
                    )
                    if elementos_panorama_secretaria is not None:
                        elements.extend(elementos_panorama_secretaria)
                        elements.append(PageBreak())
                else:
                    for gestao_pagina in ["GESTÃO ESTADUAL", "GESTÃO FEDERAL"]:
                        df_gestao_pagina = df[df["GESTAO"] == gestao_pagina]
                        if df_gestao_pagina.empty:
                            continue
                        titulo_pagina_gestao = (
                            f"PAC - PANORAMA POR SECRETARIA | EXECUTOR* - {gestao_pagina}"
                        )
                        elementos_pagina_gestao = montar_pagina_panorama_secretaria(
                            df_gestao_pagina, titulo_pagina_gestao
                        )
                        if elementos_pagina_gestao is not None:
                            elements.extend(elementos_pagina_gestao)
                            elements.append(PageBreak())

        # Índice calculado AQUI, e não mais só na etapa das páginas de
        # medidores logo adiante: a página de Metodologia precisa dos
        # números de uma secretaria de verdade para o exemplo passo a
        # passo, e ela vem antes. O resultado é reaproveitado nas duas
        # etapas — a função não é chamada duas vezes.
        #
        # FORA do "if" de qualquer seção: Metodologia, Índice de Desempenho
        # e o medidor do cabeçalho do Detalhamento Analítico dependem dele,
        # e qualquer uma delas pode ser gerada sem o Panorama.
        combos_ativos_indice = set(
            df[["SECRETARIA_LIMPA", "EXECUTOR"]]
            .dropna()
            .apply(lambda r: (str(r["SECRETARIA_LIMPA"]), str(r["EXECUTOR"])), axis=1)
        )
        df_completo_com_investimento_indice = df_original.assign(
            _INVESTIMENTO_AJUSTADO=_serie_investimento_ajustado(df_original)
        )
        grupos_indice_desempenho = calcular_indice_desempenho_secretarias(
            df_completo_com_investimento_indice, combos_ativos=combos_ativos_indice
        )
        # Mapa (gestão, secretaria) -> índice, usado no cabeçalho de cada
        # secretaria do Detalhamento Analítico. Também fica fora dos "ifs":
        # o Detalhamento pode ser gerado sem a página do Índice.
        indices_por_secretaria = _indice_desempenho_por_secretaria(grupos_indice_desempenho)

        if secao_ativa("METODOLOGIA"):
            # --- ETAPA B3.55: METODOLOGIA DO ÍNDICE DE DESEMPENHO ---
            # Mesmo padrão visual da página de OBSERVAÇÃO — título + coluna de
            # texto centralizada (88% da largura) — explicando os critérios,
            # pesos e fórmula usados no cálculo do Índice de Desempenho, antes
            # das páginas que de fato mostram os medidores.
            LARGURA_METODOLOGIA = LARGURA_UTIL * 0.88

            # A faixa abaixo do texto tem duas colunas, alinhadas pelo topo: o
            # exemplo do cálculo com números reais à esquerda, os quatro
            # medidores de categoria à direita.
            #
            # A tabela "Critério × Peso" e o quadro da fórmula abstrata saíram:
            # o exemplo mostra os mesmos pesos aplicados a valores de verdade,
            # e ler "58,3 × 0,30 = 17,5" ensina mais do que ler "Status ... 30%"
            # em uma tabela e "0,30 × Status" em outra. Assim toda a explicação
            # do índice cabe numa página só.
            LARGURA_EXEMPLO_CALCULO_METODOLOGIA = LARGURA_METODOLOGIA * 0.54
            LARGURA_EXEMPLOS_METODOLOGIA = LARGURA_METODOLOGIA - LARGURA_EXEMPLO_CALCULO_METODOLOGIA

            # Quatro medidores de exemplo, um por categoria, com a nota escolhida
            # no meio de cada faixa. Servem de legenda visual: quem vê o medidor
            # de uma secretaria nas páginas seguintes já sabe reconhecer a que
            # categoria ele corresponde sem voltar ao texto.
            #
            # São menores que os das páginas de resultado (até 135 pt contra 196):
            # o tamanho sai do espaço livre à direita da tabela e da fórmula,
            # limitado por um teto para o bloco inteiro continuar cabendo numa
            # página só junto do texto.
            exemplos_categorias_metodologia = [
                ("Insatisfatório", 12),
                ("Regular", 37),
                ("Bom", 62),
                ("Ótimo", 88),
            ]
            # O teto de 120 pt não é estético: a coluna da direita empilha os
            # medidores, o texto de apoio e a tabela de composição, e o bloco
            # inteiro tem cerca de 8 pt de folga na página. Medidor maior que
            # isso empurra a coluna para além da folha.
            LARGURA_EXEMPLO_MEDIDOR = min(120.0, LARGURA_EXEMPLOS_METODOLOGIA / 4 - 6)
            celulas_exemplos = [
                gerar_medidor_desempenho(
                    categoria_exemplo, nota_exemplo, categoria_exemplo,
                    largura=LARGURA_EXEMPLO_MEDIDOR,
                    altura=LARGURA_EXEMPLO_MEDIDOR * 0.82,
                )
                for categoria_exemplo, nota_exemplo in exemplos_categorias_metodologia
            ]
            faixas_exemplos = [
                Paragraph(f"<b>{rotulo}</b>", metodologia_exemplo_faixa)
                for rotulo in ("0 a 24", "25 a 49", "50 a 74", "75 a 100")
            ]
            tabela_exemplos_metodologia = Table(
                [
                    [Paragraph("<b>4 · O índice vira uma categoria</b>", metodologia_tabela_header)] + [""] * 3,
                    celulas_exemplos,
                    faixas_exemplos,
                ],
                colWidths=[LARGURA_EXEMPLOS_METODOLOGIA / 4] * 4,
            )
            tabela_exemplos_metodologia.setStyle(
                TableStyle(
                    [
                        ("SPAN", (0, 0), (-1, 0)),
                        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                        ("VALIGN", (0, 1), (-1, 1), "MIDDLE"),
                        ("LEFTPADDING", (0, 0), (-1, -1), 0),
                        ("RIGHTPADDING", (0, 0), (-1, -1), 0),
                        ("TOPPADDING", (0, 0), (-1, 0), 0),
                        ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
                        ("TOPPADDING", (0, 1), (-1, -1), 0),
                        ("BOTTOMPADDING", (0, 1), (-1, -1), 0),
                    ]
                )
            )

            # --- Exemplo passo a passo com uma unidade REAL da base ---
            # Escolhe uma unidade de índice mediano dentro do maior grupo: uma
            # nota do meio da tabela ilustra melhor que a primeira ou a última,
            # onde os scores relativos ficam travados em 100 ou em 0 e o leitor
            # não vê a conta funcionando de verdade. Os números vêm do mesmo
            # cálculo que alimenta os medidores, então batem com o que aparece
            # nas páginas seguintes.
            item_exemplo_metodologia = None
            grupo_exemplo_metodologia = max(
                (g for g in grupos_indice_desempenho if g["itens"]),
                key=lambda g: len(g["itens"]),
                default=None,
            )
            if grupo_exemplo_metodologia:
                itens_com_detalhe = [
                    i for i in grupo_exemplo_metodologia["itens"] if i.get("detalhe_calculo")
                ]
                if itens_com_detalhe:
                    item_exemplo_metodologia = itens_com_detalhe[len(itens_com_detalhe) // 2]

            if item_exemplo_metodologia:
                det = item_exemplo_metodologia["detalhe_calculo"]
                pesos_txt = PESOS_INDICE_DESEMPENHO

                def _linha_exemplo(descricao, conta, resultado):
                    return [
                        Paragraph(descricao, metodologia_exemplo_desc),
                        Paragraph(conta, metodologia_exemplo_conta),
                        Paragraph(f"<b>{resultado}</b>", metodologia_exemplo_valor),
                    ]

                parcela_status = det["media_status"] * pesos_txt["status"]
                parcela_fase = det["media_fase"] * pesos_txt["fase"]
                parcela_clausula = det["media_clausula"] * pesos_txt["clausula"]
                parcela_tempo = det["media_tempo"] * pesos_txt["tempo"]
                parcela_quantidade = det["score_quantidade"] * pesos_txt["quantidade"]
                parcela_proporcao = det["score_proporcao"] * pesos_txt["proporcao_contratado"]

                linhas_exemplo = [
                    [
                        Paragraph("<b>3 · Os seis critérios viram o índice</b>", metodologia_tabela_header),
                        Paragraph("<b>Conta</b>", metodologia_tabela_header_centro),
                        Paragraph("<b>Resultado</b>", metodologia_tabela_header_centro),
                    ],
                    _linha_exemplo(
                        "Status (30%) — média ponderada das ações",
                        f"{_num_metodologia(det['media_status'])} × 0,30",
                        _num_metodologia(parcela_status),
                    ),
                    _linha_exemplo(
                        "Fase (30%) — mesma média ponderada",
                        f"{_num_metodologia(det['media_fase'])} × 0,30",
                        _num_metodologia(parcela_fase),
                    ),
                    _linha_exemplo(
                        "Cláusula Suspensiva (10%)",
                        f"{_num_metodologia(det['media_clausula'])} × 0,10",
                        _num_metodologia(parcela_clausula),
                    ),
                    _linha_exemplo(
                        "Tempo do ciclo (10%) — aviso → O.S. → conclusão",
                        f"{_num_metodologia(det['media_tempo'])} × 0,10",
                        _num_metodologia(parcela_tempo),
                    ),
                    _linha_exemplo(
                        f"Quantidade de ações (10%) — {item_exemplo_metodologia['qtd']} ações, de "
                        f"{int(det['qtd_minima'])} a {int(det['qtd_maxima'])} na gestão",
                        f"{_num_metodologia(det['score_quantidade'])} × 0,10",
                        _num_metodologia(parcela_quantidade),
                    ),
                    _linha_exemplo(
                        # Valores abreviados (Mi/Bi): por extenso a linha
                        # quebraria em duas na coluna mais estreita.
                        f"Proporção de Valor Contratado (10%) — {_formatar_mi_bi_texto_puro(det['valor_contratado'])} "
                        f"de {_formatar_mi_bi_texto_puro(det['investimento_gestao'])} = "
                        f"{_num_metodologia(det['proporcao'] * 100)}%",
                        f"{_num_metodologia(det['score_proporcao'])} × 0,10",
                        _num_metodologia(parcela_proporcao),
                    ),
                    [
                        Paragraph("<b>Índice final</b>", metodologia_exemplo_desc),
                        Paragraph("soma das parcelas", metodologia_exemplo_conta),
                        Paragraph(
                            f"<b>{_num_metodologia(item_exemplo_metodologia['indice'])} "
                            f"({item_exemplo_metodologia['categoria']})</b>",
                            metodologia_exemplo_valor,
                        ),
                    ],
                ]
                tabela_exemplo_calculo = Table(
                    linhas_exemplo,
                    colWidths=[
                        LARGURA_EXEMPLO_CALCULO_METODOLOGIA * 0.55,
                        LARGURA_EXEMPLO_CALCULO_METODOLOGIA * 0.24,
                        LARGURA_EXEMPLO_CALCULO_METODOLOGIA * 0.21,
                    ],
                )
                tabela_exemplo_calculo.setStyle(
                    TableStyle(
                        [
                            ("BACKGROUND", (0, 0), (-1, 0), COR_MARCA_TEAL_CLARA),
                            ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#F2F2F2")),
                            ("BOX", (0, 0), (-1, -1), 0.5, colors.grey),
                            ("INNERGRID", (0, 0), (-1, -1), 0.5, colors.lightgrey),
                            ("LINEABOVE", (0, -1), (-1, -1), 0.9, colors.grey),
                            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                            ("LEFTPADDING", (0, 0), (-1, -1), 8),
                            ("RIGHTPADDING", (0, 0), (-1, -1), 8),
                            ("TOPPADDING", (0, 0), (-1, -1), 3),
                            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
                        ]
                    )
                )
                # Abertura da primeira linha do exemplo: de onde saiu a média de
                # Status. É a etapa que mais gera dúvida, porque o número não
                # está em lugar nenhum da planilha — ele nasce de uma nota fixa
                # por status, ponderada pelo investimento das ações. As outras
                # três notas de ação (Fase, Cláusula, Tempo) seguem exatamente a
                # mesma mecânica, então basta destrinchar uma.
                composicao = det.get("composicao_status") or []
                peso_total_exemplo = det.get("peso_efetivo_total") or 0.0
                linhas_composicao = [
                    [
                        Paragraph("<b>2 · As notas viram a nota da unidade</b>", metodologia_apoio_header),
                        Paragraph("<b>Nota</b>", metodologia_apoio_header_centro),
                        Paragraph("<b>Ações</b>", metodologia_apoio_header_centro),
                        Paragraph("<b>Peso</b>", metodologia_apoio_header_centro),
                        Paragraph("<b>Contribui</b>", metodologia_apoio_header_centro),
                    ]
                ]
                for parcela in sorted(composicao, key=lambda c: c["peso"], reverse=True):
                    fatia_peso = parcela["peso"] / peso_total_exemplo if peso_total_exemplo > 0 else 0.0
                    linhas_composicao.append(
                        [
                            Paragraph(parcela["status"].title(), metodologia_apoio_desc),
                            Paragraph(_num_metodologia(parcela["nota"]), metodologia_apoio_conta),
                            Paragraph(str(parcela["qtd"]), metodologia_apoio_conta),
                            Paragraph(f"{_num_metodologia(fatia_peso * 100)}%", metodologia_apoio_conta),
                            Paragraph(
                                f"<b>{_num_metodologia(parcela['nota'] * fatia_peso)}</b>",
                                metodologia_apoio_valor,
                            ),
                        ]
                    )
                linhas_composicao.append(
                    [
                        Paragraph("<b>Nota de Status da unidade</b>", metodologia_apoio_desc),
                        Paragraph("", metodologia_apoio_conta),
                        Paragraph("", metodologia_apoio_conta),
                        Paragraph("<b>100%</b>", metodologia_apoio_conta),
                        Paragraph(
                            f"<b>{_num_metodologia(det['media_status'])}</b>", metodologia_apoio_valor
                        ),
                    ]
                )
                tabela_composicao_status = Table(
                    linhas_composicao,
                    colWidths=[
                        LARGURA_EXEMPLOS_METODOLOGIA * 0.34,
                        LARGURA_EXEMPLOS_METODOLOGIA * 0.15,
                        LARGURA_EXEMPLOS_METODOLOGIA * 0.14,
                        LARGURA_EXEMPLOS_METODOLOGIA * 0.16,
                        LARGURA_EXEMPLOS_METODOLOGIA * 0.21,
                    ],
                )
                tabela_composicao_status.setStyle(
                    TableStyle(
                        [
                            ("BACKGROUND", (0, 0), (-1, 0), COR_MARCA_TEAL_CLARA),
                            ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#F2F2F2")),
                            ("BOX", (0, 0), (-1, -1), 0.5, colors.grey),
                            ("INNERGRID", (0, 0), (-1, -1), 0.5, colors.lightgrey),
                            ("LINEABOVE", (0, -1), (-1, -1), 0.9, colors.grey),
                            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                            ("LEFTPADDING", (0, 0), (-1, -1), 6),
                            ("RIGHTPADDING", (0, 0), (-1, -1), 6),
                            ("TOPPADDING", (0, 0), (-1, -1), 3),
                            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
                        ]
                    )
                )
                # Escala de nota por Status. Explica a origem dos números da
                # coluna "Nota" da tabela ao lado — por que ANDAMENTO vale 65 e
                # não 100, que é a dúvida mais frequente: os status em andamento
                # se distribuem proporcionalmente de 0 a 65 conforme a posição na
                # ordem, e o teto de 65 fica reservado para que Concluída (85) e
                # Inaugurada (100) tenham um salto claro acima de tudo que ainda
                # não foi entregue.
                #
                # Mostra pontos da escala, não os 18 status: a tabela inteira não
                # caberia e a regra fica igualmente clara com o primeiro, alguns
                # do meio e os três últimos.
                total_em_andamento = len(_STATUS_EM_ANDAMENTO_DESEMPENHO)
                posicoes_escala = [0, total_em_andamento // 2, total_em_andamento - 1]
                linhas_escala = [
                    [
                        Paragraph("<b>1 · Cada status vale uma nota</b>", metodologia_apoio_header),
                        Paragraph("<b>Posição</b>", metodologia_apoio_header_centro),
                        Paragraph("<b>Nota</b>", metodologia_apoio_header_centro),
                    ]
                ]
                for posicao in sorted(set(posicoes_escala)):
                    nome_status = _STATUS_EM_ANDAMENTO_DESEMPENHO[posicao]
                    linhas_escala.append(
                        [
                            Paragraph(str(nome_status).title(), metodologia_apoio_desc),
                            Paragraph(f"{posicao + 1}ª de {total_em_andamento}", metodologia_apoio_conta),
                            Paragraph(
                                _num_metodologia(_MAPA_SCORE_STATUS_DESEMPENHO[nome_status]),
                                metodologia_apoio_valor,
                            ),
                        ]
                    )
                for nome_status in ("CONCLUÍDA", "INAUGURADA"):
                    linhas_escala.append(
                        [
                            Paragraph(f"<b>{nome_status.title()}</b>", metodologia_apoio_desc),
                            Paragraph("entregue", metodologia_apoio_conta),
                            Paragraph(
                                f"<b>{_num_metodologia(_MAPA_SCORE_STATUS_DESEMPENHO[nome_status])}</b>",
                                metodologia_apoio_valor,
                            ),
                        ]
                    )
                tabela_escala_status = Table(
                    linhas_escala,
                    colWidths=[
                        LARGURA_EXEMPLO_CALCULO_METODOLOGIA * 0.55,
                        LARGURA_EXEMPLO_CALCULO_METODOLOGIA * 0.24,
                        LARGURA_EXEMPLO_CALCULO_METODOLOGIA * 0.21,
                    ],
                )
                tabela_escala_status.setStyle(
                    TableStyle(
                        [
                            ("BACKGROUND", (0, 0), (-1, 0), COR_MARCA_TEAL_CLARA),
                            ("BACKGROUND", (0, -2), (-1, -1), colors.HexColor("#F2F2F2")),
                            ("BOX", (0, 0), (-1, -1), 0.5, colors.grey),
                            ("INNERGRID", (0, 0), (-1, -1), 0.5, colors.lightgrey),
                            ("LINEABOVE", (0, -2), (-1, -2), 0.9, colors.grey),
                            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                            ("LEFTPADDING", (0, 0), (-1, -1), 6),
                            ("RIGHTPADDING", (0, 0), (-1, -1), 6),
                            ("TOPPADDING", (0, 0), (-1, -1), 3),
                            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
                        ]
                    )
                )

                # As quatro tabelas em ORDEM DE RACIOCÍNIO, numeradas de 1 a 4 no
                # próprio cabeçalho — a nota nasce do status de cada ação (1),
                # vira a nota da unidade naquele critério (2), entra na conta dos
                # seis critérios (3) e por fim vira uma categoria (4).
                #
                # A disposição anterior colocava o resultado antes da origem, o
                # que obrigava a ler a página de trás para frente para entender
                # de onde vinha cada número.
                faixa_pesos_e_exemplos_metodologia = Table(
                    [
                        [tabela_escala_status, tabela_composicao_status],
                        [Spacer(1, 8), Spacer(1, 8)],
                        [tabela_exemplo_calculo, tabela_exemplos_metodologia],
                    ],
                    colWidths=[LARGURA_EXEMPLO_CALCULO_METODOLOGIA, LARGURA_EXEMPLOS_METODOLOGIA],
                )
                faixa_pesos_e_exemplos_metodologia.setStyle(
                    TableStyle(
                        [
                            ("VALIGN", (0, 0), (-1, -1), "TOP"),
                            ("LEFTPADDING", (0, 0), (0, -1), 0),
                            ("LEFTPADDING", (1, 0), (1, -1), 14),
                            ("RIGHTPADDING", (0, 0), (-1, -1), 0),
                            ("TOPPADDING", (0, 0), (-1, -1), 0),
                            ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
                        ]
                    )
                )
            else:
                # Base sem nenhuma unidade com detalhe de cálculo: a faixa fica
                # só com os medidores de categoria.
                faixa_pesos_e_exemplos_metodologia = tabela_exemplos_metodologia

            conteudo_metodologia = [
                Paragraph(
                    "O Índice de Desempenho resume, numa nota de 0 a 100, a qualidade do andamento — sempre "
                    "calculado com a base completa de dados (nunca influenciado pelos filtros aplicados na "
                    "geração deste relatório específico) e sempre separado por Gestão (Estadual e Federal nunca "
                    "são comparadas entre si). Os itens que os filtros atuais deixariam de fora do restante do "
                    "relatório aparecem esmaecidos nas páginas seguintes, mas o valor do índice não muda por "
                    "causa disso.",
                    metodologia_texto,
                ),
                Spacer(1, 15),
                Paragraph(
                    "<b>Unidade avaliada:</b> na <b>Gestão Estadual</b>, a avaliação é feita por SECRETARIA — "
                    "todas as ações da secretaria entram no mesmo medidor, independentemente do executor. Na "
                    "<b>Gestão Federal</b>, a avaliação continua sendo feita por combinação SECRETARIA | "
                    "EXECUTOR. As comparações relativas (quantidade de ações e proporção de Valor Contratado) "
                    "são sempre feitas entre unidades do mesmo tipo, dentro da mesma gestão.",
                    metodologia_texto,
                ),
                Spacer(1, 15),
                Paragraph(
                    "<b>Como o índice é montado:</b> as quatro tabelas abaixo estão na ordem do cálculo"
                    + (
                        f", com os números de <b>{item_exemplo_metodologia['rotulo']}</b> "
                        f"({grupo_exemplo_metodologia['gestao']}) — uma unidade real desta base"
                        if item_exemplo_metodologia else ""
                    )
                    + ". Cada AÇÃO recebe uma nota por critério <b>(1)</b>; a nota da unidade naquele "
                    "critério é a média dessas notas, PONDERADA pelo investimento de cada ação <b>(2)</b>, e "
                    "o investimento de Concluídas pesa 1,5× mais, Inauguradas 3,0× mais. As notas dos seis "
                    "critérios, cada uma com seu peso, somam o índice <b>(3)</b>, que enfim vira uma categoria "
                    "<b>(4)</b>. Os dois últimos critérios são calculados uma vez por unidade, não por ação.",
                    metodologia_texto,
                ),
                Spacer(1, 12),
                faixa_pesos_e_exemplos_metodologia,
                Spacer(1, 12),
                Paragraph(
                    "<b>Status (30%)</b> e <b>Fase (30%):</b> os status que ainda não representam entrega se "
                    "distribuem de 0 a <b>65</b> conforme a posição na ordem — por isso Andamento, a última "
                    "delas, vale 65 e não 100: o que sobra acima disso fica reservado para a entrega de fato "
                    "(Concluída 85, Inaugurada 100). Fase segue a mesma ideia, de Captação de Recurso até "
                    "Concluída.",
                    metodologia_texto,
                ),
                Spacer(1, 12),
                Paragraph(
                    "<b>Cláusula Suspensiva (10%)</b> e <b>Tempo do ciclo (10%):</b> na cláusula, Retirada "
                    "Total ou nenhuma cláusula definida = melhor nota, Retirada Parcial = nota intermediária, "
                    "cláusula ainda vigente = pior nota. O tempo do ciclo é a soma dos dias entre Aviso de "
                    "Licitação → Emissão de O.S. → Previsão de Conclusão Atual, comparado com as demais ações "
                    "do recorte — quanto menos tempo, melhor.",
                    metodologia_texto,
                ),
                Spacer(1, 12),
                Paragraph(
                    "<b>Quantidade de ações (10%)</b> e <b>Proporção de Valor Contratado (10%):</b> são notas "
                    "RELATIVAS — vêm da posição da unidade entre as demais da mesma gestão, não do valor em si. "
                    "Quem administra menos ações fica com 0 e quem administra mais fica com 100 (mais ações = "
                    "mais dificuldade); o mesmo vale para a proporção de Valor Contratado.",
                    metodologia_texto,
                ),
            ]

            tabela_conteudo_metodologia = Table([[conteudo_metodologia]], colWidths=[LARGURA_METODOLOGIA])
            tabela_conteudo_metodologia.setStyle(
                TableStyle(
                    [
                        ("VALIGN", (0, 0), (-1, -1), "TOP"),
                        ("LEFTPADDING", (0, 0), (-1, -1), 0),
                        ("RIGHTPADDING", (0, 0), (-1, -1), 0),
                        ("TOPPADDING", (0, 0), (-1, -1), 0),
                        ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
                    ]
                )
            )
            bloco_metodologia = Table(
                [[Paragraph("METODOLOGIA — ÍNDICE DE DESEMPENHO", metodologia_titulo)], [Spacer(1, 20)], [tabela_conteudo_metodologia]],
                colWidths=[LARGURA_METODOLOGIA],
            )
            bloco_metodologia.setStyle(
                TableStyle(
                    [
                        ("LEFTPADDING", (0, 0), (-1, -1), 0),
                        ("RIGHTPADDING", (0, 0), (-1, -1), 0),
                        ("TOPPADDING", (0, 0), (-1, -1), 0),
                        ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
                    ]
                )
            )
            wrapper_metodologia = Table([[bloco_metodologia]], colWidths=[LARGURA_UTIL])
            wrapper_metodologia.setStyle(
                TableStyle(
                    [
                        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                        ("LEFTPADDING", (0, 0), (-1, -1), 0),
                        ("RIGHTPADDING", (0, 0), (-1, -1), 0),
                        ("TOPPADDING", (0, 0), (-1, -1), 0),
                        ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
                    ]
                )
            )
            elements.append(wrapper_metodologia)
            elements.append(PageBreak())

        if secao_ativa("INDICE_DESEMPENHO"):
            # --- ETAPA B3.6: ÍNDICE DE DESEMPENHO POR SECRETARIA | EXECUTOR ---
            # SEMPRE separado por gestão (Estadual, depois Federal) — nunca
            # combinado numa única página/ranking, já que comparar uma
            # secretaria estadual com um ministério federal lado a lado não
            # faz sentido. calcular_indice_desempenho_secretarias já devolve
            # os dados agrupados e ordenados dessa forma; aqui só percorremos
            # os grupos e desenhamos — SEMPRE numa página só (a grade se
            # adapta, ver montar_paginas_indice_desempenho_gestao).
            #
            # IMPORTANTE: usa df_original (a base completa, sem os filtros
            # aplicados na geração deste relatório específico) — o índice é
            # uma medida de qualidade da base como um todo, e não pode mudar
            # dependendo de quais filtros a pessoa marcou pra gerar ESSE
            # relatório. Isso é intencional e não deve ser alterado mesmo que
            # o resto da página use `df` (filtrado) normalmente.
            #
            # A ÚNICA influência que os filtros têm aqui: combinações
            # SECRETARIA | EXECUTOR que os filtros aplicados deixariam de fora
            # do restante do relatório aparecem "esmaecidas" (em tons de
            # cinza) — não somem, não recalculam, só ficam visualmente
            # diferenciadas.
            # (o cálculo e o mapa por secretaria já foram feitos mais acima,
            # fora dos "ifs" de seção)

            # Quando os filtros deixam uma única gestão no recorte, a página da
            # OUTRA gestão sai do relatório. O índice continua sendo calculado
            # sobre a base inteira (ver acima), então os números da gestão que
            # ficou não mudam nada com isso — some só a folha que não tem
            # nenhuma ação do recorte para explicar. Com as duas gestões
            # presentes, as duas páginas saem como antes.
            gestoes_no_recorte = {
                str(g).strip() for g in df["GESTAO"].dropna().unique() if str(g).strip()
            }
            if len(gestoes_no_recorte) == 1:
                grupos_indice_desempenho = [
                    grupo for grupo in grupos_indice_desempenho
                    if str(grupo["gestao"]).strip() in gestoes_no_recorte
                ]
            for grupo_gestao in grupos_indice_desempenho:
                # O título e a nota de leitura acompanham o nível de
                # agrupamento da gestão: só SECRETARIA na Estadual,
                # SECRETARIA | EXECUTOR na Federal.
                agrupa_por_executor = "EXECUTOR" in CHAVES_AGRUPAMENTO_INDICE_POR_GESTAO.get(
                    grupo_gestao["gestao"], ["SECRETARIA_LIMPA", "EXECUTOR"]
                )
                unidade_indice = "SECRETARIA | EXECUTOR" if agrupa_por_executor else "SECRETARIA"
                titulo_base_indice = (
                    f"PAC - ÍNDICE DE DESEMPENHO POR {unidade_indice} - {grupo_gestao['gestao']}"
                )
                paginas_indice = montar_paginas_indice_desempenho_gestao(
                    grupo_gestao["itens"], titulo_base_indice, f"por {unidade_indice}"
                )
                for pagina_elementos in paginas_indice:
                    elements.extend(pagina_elementos)
                    elements.append(PageBreak())

        if secao_ativa("DETALHAMENTO_FINANCEIRO"):
            # --- ETAPA B3.5: DETALHAMENTO FINANCEIRO POR SECRETARIA ---
            # Uma barra empilhada percentual por secretaria, com o investimento
            # decomposto em Valor Contratado / OGU / Recurso Estadual /
            # Financiamento (mesma regra de valores ajustados usada em todo o
            # resto do relatório e no dashboard, e o mesmo gráfico do painel
            # interativo — agora também no relatório impresso).
            # Quando o recorte traz SOMENTE ações em ANDAMENTO, esta página
            # deixou de ser suprimida à força: agora ela apenas chega
            # DESMARCADA na janela de seleção (ver _secoes_disponiveis_
            # relatorio). A decisão passa a ser de quem gera o relatório —
            # se marcar, a página sai normalmente.
            grafico_detalhamento_financeiro = gerar_grafico_detalhamento_secretaria(df)
            if grafico_detalhamento_financeiro is not None:
                titulo_detalhamento_financeiro = (
                    f"PAC - DETALHAMENTO FINANCEIRO POR SECRETARIA - {sufixo_titulo_painel}"
                )
                elements.append(Paragraph(titulo_detalhamento_financeiro, titulo_style))
                # Mesmo princípio de sinalização do RESUMO FINANCEIRO do Painel
                # Geral: Valor Apoiado OGU, Recurso Estadual e Financiamento
                # (marcados na legenda abaixo) são estimativa orçamentária, ainda
                # não contratada. Nesta página usa-se um único "*" em vez de "**".
                elements.append(Paragraph("* Previsão Orçamentária", cabecalho_grafico_painel_style_direita))
                elements.append(Spacer(1, 10))
                elements.append(gerar_cards_resumo(len(df), _investimento_ajustado(df)))
                elements.append(Spacer(1, 14))
                elements.append(
                    Paragraph(
                        "Cada barra representa 100% do investimento daquela secretaria, "
                        "dividido entre os componentes abaixo. O texto sob a barra detalha "
                        "os valores de cada componente presente.",
                        detalhamento_texto_style,
                    )
                )
                elements.append(Spacer(1, 10))
                elements.append(desenhar_legenda_detalhamento_secretaria())
                elements.append(Spacer(1, 10))
                elements.append(grafico_detalhamento_financeiro)
                elements.append(PageBreak())

        if secao_ativa("MAPA"):
            # --- ETAPA B4: MAPA COROPLÉTICO DOS MUNICÍPIOS ---
            # Precisa do arquivo "municipios_bahia.geojson" na mesma pasta do
            # script/.exe — se não encontrar, essa página é simplesmente
            # omitida (sem travar a geração do restante do relatório).
            geojson_municipios = carregar_geojson_municipios()
            if geojson_municipios is not None:
                titulo_mapa = f"PAC - MAPA DE INVESTIMENTOS POR MUNICÍPIO - {sufixo_titulo_painel}"
                (
                    desenho_mapa,
                    bordas_classes_mapa,
                    qtd_municipios_com_dado,
                    qtd_municipios_total,
                ) = gerar_mapa_coropletico_municipios(df, geojson_municipios)
                desenho_legenda_mapa = gerar_legenda_mapa_municipios(bordas_classes_mapa)

                tabela_mapa = Table(
                    [[desenho_mapa, desenho_legenda_mapa]],
                    colWidths=[LARGURA_UTIL * 0.6, LARGURA_UTIL * 0.4],
                )
                tabela_mapa.setStyle(
                    TableStyle(
                        [
                            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                            ("ALIGN", (0, 0), (0, 0), "CENTER"),
                            ("ALIGN", (1, 0), (1, 0), "LEFT"),
                        ]
                    )
                )

                elements.append(
                    KeepTogether(
                        [
                            Paragraph(titulo_mapa, titulo_style),
                            Spacer(1, 18),
                            gerar_cards_resumo(len(df), _investimento_ajustado(df)),
                            Spacer(1, 14),
                            Paragraph(
                                "A cor de cada município varia conforme o investimento total "
                                f"(R$) no filtro atual — {qtd_municipios_com_dado} de "
                                f"{qtd_municipios_total} municípios da Bahia têm ao menos uma "
                                "ação registrada.",
                                texto_padrao_style,
                            ),
                            Spacer(1, 18),
                            tabela_mapa,
                        ]
                    )
                )
                elements.append(PageBreak())

        if secao_ativa("RESUMO_GERAL"):
            # --- ETAPA C: RESUMO EXECUTIVO GERAL ---
            # Só faz sentido mostrar essa página quando existe MAIS DE UMA gestão
            # nos dados filtrados — quando sobra só uma (ex: filtro por "GESTÃO
            # ESTADUAL"), essa página fica idêntica à primeira página da Etapa D
            # logo em seguida, então é omitida para evitar duplicidade.
            gestoes_presentes_no_df = df["GESTAO"].nunique()
            if gestoes_presentes_no_df > 1:
                if len(gestoes_presentes_titulo) == 1:
                    sufixo_titulo_resumo = gestoes_presentes_titulo[0]
                else:
                    sufixo_titulo_resumo = "GESTÃO ESTADUAL E FEDERAL"
                elements.append(
                    KeepTogether(
                        [
                            Paragraph(
                                f"PAC - {sufixo_titulo_resumo} POR STATUS",
                                titulo_style,
                            ),
                            Spacer(1, 13),
                            gerar_cards_resumo(len(df), _investimento_ajustado(df)),
                            Spacer(1, 22),
                            gerar_faixa_resumo_eixo_status(df, "GERAL"),
                        ]
                    )
                )
                elements.append(PageBreak())

        # --- ETAPA D: VISÕES SETORIAIS (ESTADUAL / FEDERAL) ---
        ordem_gestao = ["GESTÃO ESTADUAL", "GESTÃO FEDERAL"]
        gestoes = sorted(
            list(df_agrupado.groupby("GESTAO")),
            key=lambda x: ordem_gestao.index(x[0]) if x[0] in ordem_gestao else 99,
        )

        for i, (gestao, df_gestao) in enumerate(gestoes):
            gestao_str = str(gestao)
            df_contexto = df[df["GESTAO"] == gestao_str]
            total_acoes_gestao = len(df_contexto)
            total_invest_gestao = _investimento_ajustado(df_contexto)

            cards = gerar_cards_resumo(total_acoes_gestao, total_invest_gestao)

            titulo_por_status = f"PAC - {gestao_str} POR STATUS"
            titulo_por_fase = f"PAC - {gestao_str} POR FASE"

            # Marcador SEMPRE emitido, mesmo com todas as páginas desta
            # gestão desmarcadas: é ele que diz ao numerador de rodapé a
            # que gestão a página pertence, e sem ele a numeração de uma
            # gestão vazaria para a seguinte.
            elements.append(_MarcadorGestao(gestao_str))

            # Página Setorial por Status — é a "Página 01" da gestão; a
            # numeração começa aqui.
            if secao_ativa("POR_STATUS"):
                elements.append(
                    KeepTogether(
                        [
                            Paragraph(titulo_por_status, titulo_style),
                            Spacer(1, 18),
                            cards,
                            Spacer(1, 22),
                            gerar_faixa_resumo_eixo_status(df, gestao_str),
                        ]
                    )
                )
                elements.append(PageBreak())

            if secao_ativa("POR_FASE"):
                # Página Setorial por Fase
                paragrafo_titulo_fase = Paragraph(titulo_por_fase, titulo_style)
                _, altura_titulo_fase = paragrafo_titulo_fase.wrap(LARGURA_UTIL, 10000)
                # Espaço já reservado nesta página antes da tabela: título +
                # Spacer(9) + cards (cabeçalho+valor) + Spacer(13). O que sobra
                # de ALTURA_UTIL é o teto que a tabela Secretaria/Fase tem para
                # caber inteira (cabeçalho + linhas + TOTAL) em uma única página.
                # +12: preenchimento interno padrão que o Frame do SimpleDocTemplate
                # reserva em cada borda (6pt em cima, 6pt embaixo) além das margens
                # da página — não contabilizado em ALTURA_UTIL, então entra aqui
                # como margem de segurança extra para o autofit.
                altura_reservada_pagina_fase = (
                    altura_titulo_fase + 9 + CARDS_ALTURA_CABECALHO + CARDS_ALTURA_VALOR + 13 + 12
                )
                altura_disponivel_tabela_fase = ALTURA_UTIL - altura_reservada_pagina_fase
                elements.append(
                    KeepTogether(
                        [
                            paragrafo_titulo_fase,
                            Spacer(1, 9),
                            cards,
                            Spacer(1, 13),
                            gerar_tabela_secretaria_fase(
                                df, gestao_str, altura_maxima=altura_disponivel_tabela_fase
                            ),
                        ]
                    )
                )
                elements.append(PageBreak())

            if secao_ativa("DETALHAMENTO"):
                # Cabeçalho da Seção de Tabelas Analíticas
                elements.append(
                    Paragraph(
                        f"{gestao_str} - {total_acoes_gestao} ações - {moeda_sem_quebra(total_invest_gestao)}",
                        gestao_style,
                    )
                )

                # --- ETAPA E: DETALHAMENTO ANALÍTICO POR SECRETARIA ---
                df_gestao = df_gestao.assign(
                    _SEC_SORT=df_gestao["SECRETARIA_LIMPA"].apply(remover_acentos)
                )
                for _, df_eixo in df_gestao.groupby("_SEC_SORT", sort=True):
                    secretaria = str(df_eixo["SECRETARIA_LIMPA"].iloc[0])
                    df_eixo_contexto = df_contexto[df_contexto["SECRETARIA_LIMPA"] == secretaria]
                    total_invest_eixo = _investimento_ajustado(df_eixo_contexto)

                    # As colunas exibidas vêm da customização do painel, já
                    # normalizadas (dependências resolvidas, ordem do catálogo,
                    # limite aplicado) por normalizar_colunas_detalhamento.
                    # Calculado ANTES do cabeçalho porque é daqui que sai o
                    # recuo da coluna INVESTIMENTO, onde a miniatura do
                    # Detalhamento Financeiro precisa cair.
                    colunas_tab = colunas_detalhamento
                    indice_coluna = {chave: pos for pos, chave in enumerate(colunas_tab)}
                    LARGURAS_DETALHE = calcular_larguras(LARGURA_UTIL, colunas_tab)
                    largura_de = dict(zip(colunas_tab, LARGURAS_DETALHE))

                    # Miniatura só quando INVESTIMENTO está entre as colunas
                    # escolhidas: sem a coluna na tabela, não existe nada a que
                    # alinhar, e a barra viraria um gráfico solto no cabeçalho.
                    miniatura_financeira = None
                    recuo_miniatura = 0
                    if "INVESTIMENTO" in indice_coluna:
                        # A tabela analítica ocupa 97,1% da largura útil e é
                        # centralizada (hAlign padrão do Table), então ela começa
                        # deslocada — esse deslocamento entra no recuo, senão a
                        # miniatura não cai sobre a coluna.
                        #
                        # A centralização é feita dentro do FRAME, que tem 6 pt
                        # de padding de cada lado (padrão do SimpleDocTemplate),
                        # e não sobre a largura útil cheia. Ignorar esses 12 pt
                        # jogava a miniatura 6 pt à direita da coluna: pouco para
                        # notar de relance, o suficiente para o alinhamento não
                        # fechar quando se olha de perto.
                        largura_frame = LARGURA_UTIL - 2 * PADDING_FRAME_DOCUMENTO
                        deslocamento_tabela = (largura_frame - sum(LARGURAS_DETALHE)) / 2
                        recuo_miniatura = deslocamento_tabela + sum(
                            LARGURAS_DETALHE[: indice_coluna["INVESTIMENTO"]]
                        )
                        miniatura_financeira = _miniatura_detalhamento_financeiro(
                            df_eixo_contexto, largura_de["INVESTIMENTO"]
                        )

                    elements.append(
                        _cabecalho_secretaria_detalhamento(
                            f"{secretaria} - ({len(df_eixo_contexto)}) - {moeda_sem_quebra(total_invest_eixo)}",
                            indices_por_secretaria.get((gestao_str, secretaria)),
                            miniatura=miniatura_financeira,
                            recuo_miniatura=recuo_miniatura,
                        )
                    )

                    headers_analiticos = [
                        COLUNAS_DETALHAMENTO_POR_CHAVE[chave]["titulo"] for chave in colunas_tab
                    ]
                    data_tab = [
                        [
                            Paragraph(
                                f"<b>{h}</b>",
                                header_center_termo if h == "TERMO DE COMPROMISSO/ FINANCIAMENTO" else header_center,
                            )
                            for h in headers_analiticos
                        ]
                    ]

                    df_eixo = df_eixo.assign(_OBJETO_SORT=df_eixo[col_objeto].apply(remover_acentos))
                    # Hierarquia de mesclas: 1º OBJETO, 2º FASE, 3º STATUS — a
                    # ordenação segue a mesma hierarquia (FASE_ORDEM entra como
                    # critério intermediário), para que as linhas de cada nível
                    # fiquem sempre contíguas na tabela e possam ser mescladas.
                    # A ordenação NÃO muda com a customização de colunas: mesmo
                    # que STATUS esteja oculto, as linhas continuam agrupadas do
                    # mesmo jeito, e o que era um bloco continua sendo um bloco.
                    df_eixo = df_eixo.sort_values(
                        ["_OBJETO_SORT", "FASE_ORDEM", "STATUS_ORDEM"], kind="stable"
                    )

                    # Altura máxima segura (em pontos) para uma célula mesclada (SPAN): acima
                    # disso ela pode ficar mais alta que uma página inteira e travar o
                    # ReportLab com LayoutError, já que uma célula mesclada não pode ser
                    # quebrada entre páginas. A tabela continua sendo UMA só (sem cortes
                    # artificiais); o rótulo de cada nível (OBJETO/FASE/STATUS) só se
                    # repete quando o bloco realmente se aproxima do limite de altura de
                    # uma página.
                    ALTURA_MAX_BLOCO = 480

                    def gerar_ids_subgrupo_altura(seq, alturas, limite_altura):
                        ids, contador, valor_anterior, altura_no_grupo = [], 0, object(), 0
                        for val, altura in zip(seq, alturas):
                            if val != valor_anterior or altura_no_grupo + altura > limite_altura:
                                contador += 1
                                valor_anterior = val
                                altura_no_grupo = 0
                            altura_no_grupo += altura
                            ids.append(contador)
                        return ids

                    def gerar_id_nivel_alinhado_altura(seq_nivel, id_filho, alturas, limite_altura):
                        # Nível intermediário/externo da hierarquia de mesclas (ex: FASE
                        # dentro de OBJETO, ou OBJETO por cima de FASE): só quebra quando
                        # o valor deste nível mudar de verdade, ou em um limite de grupo
                        # do nível filho (nunca no meio de uma célula mesclada do nível
                        # abaixo dele).
                        #
                        # A decisão de quebrar olha a altura do PRÓXIMO bloco filho
                        # INTEIRO, e não só a da linha atual. Como o corte só pode
                        # acontecer na fronteira entre blocos filhos, decidir linha a
                        # linha deixava o bloco pai passar do limite e só quebrar
                        # depois — com colunas de texto longo (Pendências, Próximos
                        # Passos), isso produzia células mescladas mais altas que a
                        # página, que o ReportLab não consegue nem desenhar nem
                        # dividir.
                        altura_por_bloco_filho = {}
                        for filho, altura in zip(id_filho, alturas):
                            altura_por_bloco_filho[filho] = altura_por_bloco_filho.get(filho, 0) + altura

                        ids, contador, valor_anterior, altura_no_grupo = [], 0, object(), 0
                        for pos, valor in enumerate(seq_nivel):
                            novo_bloco_filho = pos == 0 or id_filho[pos] != id_filho[pos - 1]
                            estoura = (
                                novo_bloco_filho
                                and altura_no_grupo > 0
                                and altura_no_grupo + altura_por_bloco_filho[id_filho[pos]] > limite_altura
                            )
                            if valor != valor_anterior or estoura:
                                contador += 1
                                altura_no_grupo = 0
                            altura_no_grupo += alturas[pos]
                            valor_anterior = valor
                            ids.append(contador)
                        return ids

                    def gerar_ids_subgrupo_aninhado(seq_valor, seq_grupo_pai):
                        # Cria um novo id sempre que o VALOR mudar OU o grupo pai mudar —
                        # ou seja, só mescla dentro de um mesmo bloco hierárquico já
                        # comprovadamente seguro (mesmo objeto, mesma fase, mesmo status,
                        # mesma página).
                        ids, contador, valor_ant, grupo_ant = [], 0, object(), object()
                        for val, grupo in zip(seq_valor, seq_grupo_pai):
                            if val != valor_ant or grupo != grupo_ant:
                                contador += 1
                                valor_ant = val
                                grupo_ant = grupo
                            ids.append(contador)
                        return ids

                    def _combinar_fonte_recurso(fonte_recurso, fonte_financ):
                        # FONTE DE RECURSO leva, entre parênteses, a origem do
                        # financiamento (ex: "OGU CONTRAPARTIDA FINANCIAMENTO
                        # (FGTS)") sempre que o texto da fonte de recurso
                        # mencionar "FINANCIAMENTO" — se a coluna FONTE DE
                        # FINANCIAMENTO não existir ou estiver em branco pra essa
                        # linha, mostra "(A DEFINIR)" em vez de simplesmente
                        # omitir. Quando a fonte NÃO menciona financiamento (ex:
                        # só "OGU"), não tem parênteses nenhum.
                        if "FINANCIAMENTO" not in fonte_recurso.upper():
                            return fonte_recurso
                        origem = fonte_financ if (fonte_financ and fonte_financ.upper() != "NAN") else "A DEFINIR"
                        return f"{fonte_recurso} ({origem})"

                    def _texto_opcional(row, coluna):
                        # Colunas de acompanhamento podem não existir na planilha
                        # ou vir vazias/NaN — nesses casos a célula fica em branco
                        # em vez de imprimir "nan".
                        if coluna not in df_eixo.columns:
                            return ""
                        valor = row.get(coluna)
                        if valor is None or (isinstance(valor, float) and pd.isna(valor)):
                            return ""
                        texto = str(valor).strip()
                        return "" if texto.upper() == "NAN" else texto

                    # --- PASSO 1: texto de cada célula, linha a linha ---
                    # Primeiro só o CONTEÚDO, sem decidir nada de mescla. As
                    # alturas medidas aqui é que definem, no passo seguinte, onde
                    # os blocos mesclados podem ser cortados com segurança — por
                    # isso a medição precisa acontecer sobre as colunas que
                    # realmente vão aparecer, com as larguras que elas realmente
                    # terão.
                    ESTILO_COLUNA = {
                        "OBJETO": cell_nowrap,
                        "FASE": cell_nowrap,
                        "STATUS": cell_nowrap,
                        "EMISSAO_OS": cell_center,
                        "AVANCO": cell_center,
                        "MUNICIPIOS": cell_municipio,
                        "INVESTIMENTO": valor_style,
                        "FONTE": cell_center,
                        "TERMO": cell_center,
                        "PRAZO_FASE": cell_center,
                        "PENDENCIA": cell_nowrap,
                        "PROVIDENCIAS": cell_nowrap,
                        "PROXIMOS_PASSOS": cell_nowrap,
                    }

                    registros = []
                    # Referência única de "hoje" para toda a tabela: garante que
                    # todas as linhas sejam julgadas pela mesma data, mesmo se a
                    # geração do relatório atravessar a virada do dia.
                    hoje_referencia = datetime.now().date()
                    for _, row in df_eixo.iterrows():
                        obj = str(row[col_objeto])
                        fas = str(row["FASE_TEXTO"])
                        sta = str(row["STATUS_TEXTO"])
                        # Fonte, termo, prazo e avanço só sobrevivem ao
                        # agrupamento quando a coluna está sendo exibida (ver
                        # as chaves de atributo, mais acima). Leitura tolerante,
                        # portanto: fora da seleção, a coluna simplesmente não
                        # existe no dataframe agrupado.
                        tc_status = _texto_opcional(row, "SINALIZACAO_TC")

                        valor_linha_ajustado = (
                            row[col_valor_contratado]
                            + row[col_apoiado_ajustado]
                            + row[col_financiamento_ajustado]
                            + row[col_contrapartida_ajustado]
                            + row[col_complementar_ajustado]
                        )

                        fonte_financ = _texto_opcional(row, col_fonte_financiamento)

                        prazo_fase_texto = _texto_opcional(row, "PRAZO_FASE_TEXTO")
                        # Prazo vencido sai em vermelho. A cor é aplicada AQUI,
                        # antes das mesclas: assim duas linhas vizinhas com a
                        # mesma data só se juntam numa célula se estiverem no
                        # mesmo estado (as duas vencidas ou as duas em dia) —
                        # uma mescla não pode ter metade vermelha.
                        if prazo_fase_vencido(prazo_fase_texto, sta, hoje_referencia):
                            prazo_fase_texto = (
                                f"<font color='{COR_PRAZO_VENCIDO}'>{prazo_fase_texto}</font>"
                            )

                        textos = {
                            "OBJETO": obj,
                            "FASE": fas,
                            "STATUS": sta,
                            "EMISSAO_OS": _texto_opcional(row, COLUNA_EMISSAO_OS_TEXTO),
                            "AVANCO": _texto_opcional(row, "AVANCO_OBRA_TEXTO"),
                            "MUNICIPIOS": str(row[col_municipio]),
                            "INVESTIMENTO": moeda_sem_quebra(valor_linha_ajustado),
                            "FONTE": _combinar_fonte_recurso(_texto_opcional(row, col_fonte), fonte_financ),
                            "TERMO": tc_status,
                            "PRAZO_FASE": prazo_fase_texto,
                            "PENDENCIA": _texto_opcional(row, col_pendencia),
                            "PROVIDENCIAS": _texto_opcional(row, col_providencias),
                            "PROXIMOS_PASSOS": _texto_opcional(row, col_proximos_passos),
                        }

                        # Trava contra LayoutError: nenhuma célula pode deixar a
                        # linha mais alta que uma página. Só as colunas de texto
                        # livre têm risco real de estourar isso.
                        for chave in ("PENDENCIA", "PROVIDENCIAS", "PROXIMOS_PASSOS", "MUNICIPIOS"):
                            if chave in indice_coluna and textos[chave]:
                                textos[chave] = limitar_texto_para_altura(
                                    textos[chave], ESTILO_COLUNA[chave], largura_de[chave]
                                )

                        registros.append(
                            {
                                "textos": textos,
                                "obj": obj,
                                "fas": fas,
                                "sta": sta,
                                "tc_status": tc_status,
                                "eixo": str(row[col_eixo]),
                            }
                        )

                    alturas_linhas = [
                        estimar_altura_linha(
                            [
                                (reg["textos"][chave], ESTILO_COLUNA[chave], largura_de[chave])
                                for chave in colunas_tab
                            ]
                        )
                        for reg in registros
                    ]

                    # --- PASSO 2: blocos de mescla ---
                    seq_obj = [reg["obj"] for reg in registros]
                    seq_obj_fase = [(reg["obj"], reg["fas"]) for reg in registros]
                    seq_sta = [(reg["obj"], reg["fas"], reg["sta"]) for reg in registros]

                    # Nível 3 (mais interno): STATUS — bloco atômico, respeita o limite
                    # de altura da página.
                    id_sta = gerar_ids_subgrupo_altura(seq_sta, alturas_linhas, ALTURA_MAX_BLOCO)
                    # Nível 2: FASE — mescla ao longo de vários blocos de STATUS que
                    # compartilham o mesmo OBJETO + FASE, respeitando a altura da página.
                    id_fase_nivel = gerar_id_nivel_alinhado_altura(
                        seq_obj_fase, id_sta, alturas_linhas, ALTURA_MAX_BLOCO
                    )
                    # Nível 1 (mais externo): OBJETO — mescla ao longo de vários blocos
                    # de FASE que compartilham o mesmo OBJETO, respeitando a altura da
                    # página.
                    id_obj = gerar_id_nivel_alinhado_altura(
                        seq_obj, id_fase_nivel, alturas_linhas, ALTURA_MAX_BLOCO
                    )

                    # Bloco-âncora das mesclas por repetição: o nível hierárquico
                    # mais interno que estiver VISÍVEL. Se STATUS foi tirado da
                    # seleção, as colunas de repetição passam a se apoiar no bloco
                    # de FASE (e assim por diante). Sem isso, elas se apoiariam num
                    # bloco invisível e o relatório mostraria uma célula mesclada
                    # sendo interrompida sem nenhum motivo aparente na tela.
                    if "STATUS" in indice_coluna:
                        id_bloco_ancora = id_sta
                    elif "FASE" in indice_coluna:
                        id_bloco_ancora = id_fase_nivel
                    else:
                        id_bloco_ancora = id_obj

                    # Ids de mescla das colunas por repetição — só das que estão
                    # de fato na tabela.
                    ids_repeticao = {}
                    for chave in colunas_tab:
                        if not COLUNAS_DETALHAMENTO_POR_CHAVE[chave].get("mescla_repetido"):
                            continue
                        ids_repeticao[chave] = gerar_ids_subgrupo_aninhado(
                            [reg["textos"][chave] for reg in registros], id_bloco_ancora
                        )

                    # --- PASSO 3: montagem das linhas + SPANs ---
                    spans = []
                    linhas_por_bloco = {chave: {} for chave in colunas_tab}
                    id_obj_ant, id_fase_nivel_ant, id_sta_ant = None, None, None
                    ids_repeticao_ant = {chave: None for chave in ids_repeticao}
                    obj_real_ant, fase_real_ant, sta_real_ant = None, None, None
                    l_idx = 1

                    for pos, reg in enumerate(registros):
                        obj, fas, sta = reg["obj"], reg["fas"], reg["sta"]
                        textos = dict(reg["textos"])

                        qtd_obj = df_counts_objeto.get((gestao_str, reg["eixo"], obj), 1)
                        qtd_fase = df_counts_fase.get((gestao_str, reg["eixo"], obj, fas), 1)
                        qtd_sta = df_counts_status.get((gestao_str, reg["eixo"], obj, fas, sta), 1)

                        novo_bloco_obj = id_obj[pos] != id_obj_ant
                        novo_bloco_fase = id_fase_nivel[pos] != id_fase_nivel_ant
                        novo_bloco_sta = id_sta[pos] != id_sta_ant

                        mudou_valor_obj = obj != obj_real_ant
                        mudou_valor_fase = (obj, fas) != fase_real_ant
                        mudou_valor_sta = (obj, fas, sta) != sta_real_ant

                        # Primeira linha de um bloco mostra o rótulo (com a
                        # contagem); as demais ficam em branco porque estão
                        # mescladas. Quando um bloco foi cortado só por causa da
                        # altura da página, o rótulo se repete com "(cont.)".
                        if novo_bloco_obj:
                            textos["OBJETO"] = (
                                f"{obj}{formatar_contagem_opcional(qtd_obj)}" if mudou_valor_obj
                                else f"{obj} (cont.)"
                            )
                        else:
                            textos["OBJETO"] = ""

                        if novo_bloco_fase:
                            textos["FASE"] = (
                                f"{fas}{formatar_contagem_opcional(qtd_fase)}" if mudou_valor_fase
                                else f"{fas} (cont.)"
                            )
                        else:
                            textos["FASE"] = ""

                        if novo_bloco_sta:
                            textos["STATUS"] = (
                                f"{sta}{formatar_contagem_opcional(qtd_sta)}" if mudou_valor_sta
                                else f"{sta} (cont.)"
                            )
                        else:
                            textos["STATUS"] = ""

                        # Sem a coluna TERMO na seleção, tc_status vem vazio —
                        # nesse caso não faz sentido embrulhar nada em cor.
                        if reg["tc_status"]:
                            cor_tc = "darkgreen" if reg["tc_status"] == "SIM" else "red"
                            textos["TERMO"] = f"<b><font color='{cor_tc}'>{reg['tc_status']}</font></b>"

                        for chave, ids in ids_repeticao.items():
                            if ids[pos] == ids_repeticao_ant[chave]:
                                textos[chave] = ""

                        data_tab.append(
                            [
                                Paragraph(textos[chave], ESTILO_COLUNA[chave])
                                for chave in colunas_tab
                            ]
                        )

                        # Registro das linhas de cada bloco, por coluna — é daqui
                        # que saem os SPANs, sempre com o índice REAL da coluna na
                        # seleção atual (e não mais uma posição fixa).
                        if "OBJETO" in linhas_por_bloco:
                            linhas_por_bloco["OBJETO"].setdefault(id_obj[pos], []).append(l_idx)
                        if "FASE" in linhas_por_bloco:
                            linhas_por_bloco["FASE"].setdefault(id_fase_nivel[pos], []).append(l_idx)
                        if "STATUS" in linhas_por_bloco:
                            linhas_por_bloco["STATUS"].setdefault(id_sta[pos], []).append(l_idx)
                        for chave, ids in ids_repeticao.items():
                            linhas_por_bloco[chave].setdefault(ids[pos], []).append(l_idx)

                        id_obj_ant, id_fase_nivel_ant, id_sta_ant = (
                            id_obj[pos], id_fase_nivel[pos], id_sta[pos]
                        )
                        for chave, ids in ids_repeticao.items():
                            ids_repeticao_ant[chave] = ids[pos]
                        obj_real_ant, fase_real_ant, sta_real_ant = obj, (obj, fas), (obj, fas, sta)
                        l_idx += 1

                    for chave, blocos in linhas_por_bloco.items():
                        coluna = indice_coluna[chave]
                        for linhas in blocos.values():
                            if len(linhas) > 1:
                                spans.append(("SPAN", (coluna, linhas[0]), (coluna, linhas[-1])))

                    estilo_tab = [
                        ("GRID", (0, 0), (-1, -1), 0.4, colors.grey),
                        ("BACKGROUND", (0, 0), (-1, 0), COR_MARCA_TEAL_CLARA),
                        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ]
                    if "INVESTIMENTO" in indice_coluna:
                        coluna_valor = indice_coluna["INVESTIMENTO"]
                        estilo_tab.append(("ALIGN", (coluna_valor, 1), (coluna_valor, -1), "RIGHT"))

                    t_obj = Table(data_tab, colWidths=LARGURAS_DETALHE, repeatRows=1)
                    t_obj.setStyle(TableStyle(estilo_tab + spans))

                    elements.append(t_obj)
                    elements.append(Spacer(1, 15))

            if i < len(gestoes) - 1:
                elements.append(PageBreak())

        # Com seleção de páginas, dá para sobrar PageBreak no fim (quando a
        # última seção marcada não é a última do relatório) — cada um vira
        # uma folha em branco no PDF. Marcador de gestão solto no fim também
        # não tem página a que se referir.
        while elements and isinstance(elements[-1], (PageBreak, _MarcadorGestao)):
            elements.pop()
        if not elements:
            raise ValueError(
                "Nenhuma página foi selecionada para o relatório. "
                "Marque ao menos uma na janela de seleção de páginas."
            )

        # A capa tem fundo próprio; sem ela, a primeira página é uma página
        # comum e precisa do cabeçalho normal, senão sairia sem os logos e
        # sem a linha teal/gold.
        fundo_primeira_pagina = (
            desenhar_background_capa if secao_ativa("CAPA") else desenhar_cabecalho
        )
        doc.build(
            elements,
            onFirstPage=fundo_primeira_pagina,
            onLaterPages=desenhar_cabecalho,
            canvasmaker=NumeradorPaginasGestaoCanvas,
        )
        print(f"PDF GERADO COM SUCESSO: {arquivo_pdf}")
        # No modo web, quem abre o PDF é o navegador de quem está acessando
        # (o servidor devolve os bytes na resposta HTTP) — não o processo
        # Python rodando no servidor remoto.
        if not MODO_WEB and os.name == "nt":
            os.startfile(arquivo_pdf)


# =====================================================
# PUBLICAÇÃO (botão "Publicar Atualização" do painel) — SOMENTE DESKTOP.
#
# Roda comandos git de verdade (add/commit/push) na máquina onde o
# processo está rodando. Isso só é seguro porque, no modo desktop, quem
# consegue clicar no botão é literalmente a pessoa sentada na frente
# dessa máquina — ninguém mais alcança essa função.
#
# *** NUNCA registre uma rota HTTP pra nenhuma das duas funções abaixo em
# servidor_web.py. *** Exposto por HTTP, qualquer visitante do site
# conseguiria disparar um "git push" no repositório de qualquer lugar do
# mundo. Por isso elas NÃO seguem o prefixo "_api_" usado pelas funções
# compartilhadas com o servidor web (ver seção "API DO PAINEL" logo
# abaixo) — o nome diferente é de propósito, pra ficar óbvio que são de
# outra categoria.
# =====================================================

def _git_disponivel_e_repo():
    # Confere as duas condições básicas antes de tentar qualquer comando
    # git de verdade: o git precisa estar instalado nesta máquina, e esta
    # pasta precisa ser um checkout de repositório (o que nem sempre é
    # verdade — por exemplo, numa máquina que recebeu o app só pelo kit de
    # migração do MIGRACAO.md, sem clonar o repositório).
    if shutil.which("git") is None:
        return False, "git não está instalado (ou não foi encontrado) nesta máquina."
    if not os.path.isdir(os.path.join(PASTA_BASE, ".git")):
        return False, "Esta pasta não é um repositório git — não é possível publicar daqui."
    return True, None


def _git_rodar(*args):
    resultado = subprocess.run(
        # "-c core.quotepath=false": sem isso, o git escapa qualquer nome de
        # arquivo com acento como octal literal (ex.: "BALAN\303\207O"), que
        # é exatamente o tipo de texto que apareceria sem tratamento nenhum
        # no confirm() do botão "Publicar Atualização" — feio e confuso pra
        # quem só quer ver "BALANÇO".
        ["git", "-c", "core.quotepath=false", *args],
        cwd=PASTA_BASE,
        capture_output=True,
        text=True,
        encoding="utf-8",
        errors="replace",
    )
    return resultado.returncode, resultado.stdout, resultado.stderr


def _git_verificar_mudancas():
    # Primeira etapa do botão "Publicar Atualização": só olha o que mudou
    # (planilha, código, o que for) — não mexe em nada ainda. O painel usa
    # isso pra montar a lista de confirmação antes de perguntar pro usuário
    # se pode publicar.
    ok, erro = _git_disponivel_e_repo()
    if not ok:
        return {"ok": False, "erro": erro}
    codigo, saida, erro_saida = _git_rodar("status", "--porcelain")
    if codigo != 0:
        return {"ok": False, "erro": erro_saida.strip() or "Falha ao checar o status do git."}
    mudancas = [linha for linha in saida.splitlines() if linha.strip()]
    return {"ok": True, "mudancas": mudancas}


def _git_publicar_atualizacao(mensagem):
    # Segunda etapa, só chamada depois que o usuário já viu a lista de
    # _git_verificar_mudancas e confirmou: git add -A + commit + push,
    # exatamente o mesmo fluxo do 4_ATUALIZAR_BASE_E_PUBLICAR.bat, só que
    # disparado por um botão dentro do próprio painel em vez de um .bat à
    # parte.
    ok, erro = _git_disponivel_e_repo()
    if not ok:
        return {"ok": False, "erro": erro}

    mensagem = str(mensagem or "").strip() or "Atualiza planilha/código"

    codigo, _, erro_saida = _git_rodar("add", "-A")
    if codigo != 0:
        return {"ok": False, "erro": erro_saida.strip() or "Falha ao preparar as mudanças (git add)."}

    codigo, saida_commit, erro_commit = _git_rodar("commit", "-m", mensagem)
    if codigo != 0:
        # "nothing to commit" não é um erro de verdade pro usuário — só
        # significa que não havia nada pra publicar (pode acontecer se
        # alguém clicar de novo logo depois de já ter publicado).
        texto_combinado = (saida_commit or "") + (erro_commit or "")
        if "nothing to commit" in texto_combinado.lower():
            return {"ok": True, "nada_a_publicar": True}
        return {"ok": False, "erro": (erro_commit or saida_commit).strip() or "Falha ao criar o commit."}

    codigo, _, erro_push = _git_rodar("push", "origin", "master")
    if codigo != 0:
        return {
            "ok": False,
            "erro": (
                "O commit foi criado localmente, mas o envio pro GitHub falhou "
                "(confira sua internet, ou se alguém mais publicou antes de "
                "você — nesse caso, um \"git pull\" resolve):\n\n"
                + (erro_push.strip() or "erro desconhecido")
            ),
        }

    return {"ok": True}


def montar_html_painel(df_base):
    # Monta a string HTML/CSS/JS completa do painel de filtros — só isso,
    # sem abrir janela nenhuma. Usada tanto pelo modo desktop
    # (abrir_interface_filtros, logo abaixo, que pega esse HTML e abre numa
    # janela pywebview) quanto pelo servidor web (servidor_web.py, que serve
    # esse mesmo HTML direto num navegador comum, em vez de janela nativa).
    opcoes_gestao = sorted(df_base["GESTAO"].dropna().astype(str).unique().tolist())

    # EIXO: ordem alfabética insensível a acento (mesma correção usada no
    # detalhamento do relatório, evitando que "Água Para Todos" vá parar no
    # final da lista por causa do acento).
    opcoes_eixo = (
        df_base[[col_eixo, "EIXO_SORT"]]
        .drop_duplicates()
        .sort_values("EIXO_SORT")[col_eixo]
        .tolist()
    )

    # STATUS e FASE: ordem pelo número usado na própria base (STATUS_ORDEM /
    # FASE_ORDEM), igual ao critério já usado nas tabelas do relatório —
    # não é ordem alfabética.
    opcoes_status = (
        df_base[["STATUS_ORDEM", "STATUS_TEXTO"]]
        .drop_duplicates()
        .sort_values("STATUS_ORDEM")["STATUS_TEXTO"]
        .tolist()
    )
    opcoes_fase = (
        df_base[["FASE_ORDEM", "FASE_TEXTO"]]
        .drop_duplicates()
        .sort_values("FASE_ORDEM")["FASE_TEXTO"]
        .tolist()
    )

    opcoes_orgao = sorted(
        df_base["SECRETARIA_LIMPA"].dropna().astype(str).unique().tolist(), key=remover_acentos
    )
    opcoes_executor = sorted(
        df_base["EXECUTOR"].dropna().astype(str).unique().tolist(), key=remover_acentos
    )
    opcoes_objeto = sorted(
        df_base[col_objeto].dropna().astype(str).unique().tolist(), key=remover_acentos
    )
    opcoes_municipio = sorted(
        df_base[col_municipio].dropna().astype(str).unique().tolist(), key=remover_acentos
    )
    opcoes_fonte = sorted(
        df_base[col_fonte].dropna().astype(str).unique().tolist(), key=remover_acentos
    )
    # Valores vindos da coluna CLÁUSULA SUSPENSIVA da planilha (ex: "01 -
    # Retirada Total"), ordenados pelo número do início — mesmo critério já
    # usado em STATUS e FASE — e não mais em ordem alfabética.
    opcoes_clausula_suspensiva = (
        df_base[["CLAUSULA_SUSPENSIVA_ORDEM", "SITUACAO_CLAUSULA_SUSPENSIVA"]]
        .drop_duplicates()
        .sort_values("CLAUSULA_SUSPENSIVA_ORDEM")["SITUACAO_CLAUSULA_SUSPENSIVA"]
        .tolist()
    )

    blocos_config = [
        ("GESTAO", "GESTÃO", opcoes_gestao),
        ("ORGAO", "SECRETARIA/ÓRGÃO", opcoes_orgao),
        ("EXECUTOR", "EXECUTOR", opcoes_executor),
        ("EIXO", "EIXO", opcoes_eixo),
        ("OBJETO", "OBJETO", opcoes_objeto),
        ("FASE", "FASE", opcoes_fase),
        ("STATUS", "STATUS", opcoes_status),
        ("MUNICIPIO", "MUNICÍPIO", opcoes_municipio),
        ("FONTE", "FONTE DE RECURSO", opcoes_fonte),
        ("CLAUSULA_SUSPENSIVA", "CLÁUSULA SUSPENSIVA", opcoes_clausula_suspensiva),
        ("TERMO_COMPROMISSO", "TERMO DE COMPROMISSO", ["SIM", "NÃO"]),
    ]

    # Coluna do DataFrame correspondente a cada bloco de filtro — usada tanto
    # para montar o dataset enviado ao JS quanto para a lógica de
    # "disponibilidade" (opções acinzentadas quando a combinação de filtros
    # não teria nenhum resultado), agora calculada no próprio navegador.
    COLUNA_FILTRO = {
        "GESTAO": "GESTAO",
        "ORGAO": "SECRETARIA_LIMPA",
        "EXECUTOR": "EXECUTOR",
        "EIXO": col_eixo,
        "OBJETO": col_objeto,
        "FASE": "FASE_TEXTO",
        "STATUS": "STATUS_TEXTO",
        "MUNICIPIO": col_municipio,
        "FONTE": col_fonte,
        "CLAUSULA_SUSPENSIVA": "SITUACAO_CLAUSULA_SUSPENSIVA",
        "TERMO_COMPROMISSO": "SINALIZACAO_TC",
    }

    # Dataset enxuto (só as colunas usadas nos filtros, já como string, já
    # com a chave de cada bloco em vez do nome da coluna original) — enviado
    # uma única vez para o JS, que recalcula a disponibilidade de cada bloco
    # a cada clique, sem precisar chamar o Python de volta (interface fica
    # instantânea). As três colunas de data entram no mesmo dataset (como
    # texto "AAAA-MM-DD" ou None), para que os filtros de data também
    # encolham conforme os outros filtros (e entre si) — mesma lógica de
    # disponibilidade cruzada usada nos blocos de checkbox.
    def _data_para_iso(valor):
        if pd.isna(valor):
            return None
        return pd.Timestamp(valor).strftime("%Y-%m-%d")

    serie_datas_fase = pd.to_datetime(df_base[col_prazo], errors="coerce") if col_prazo in df_base.columns else None
    serie_datas_vigencia = pd.to_datetime(df_base[col_vigencia], errors="coerce") if col_vigencia in df_base.columns else None
    serie_datas_atual = pd.to_datetime(df_base[col_prazo_atual], errors="coerce") if col_prazo_atual in df_base.columns else None

    linhas_filtro = []
    for idx, row in df_base.iterrows():
        linha = {chave: str(row[coluna]) for chave, coluna in COLUNA_FILTRO.items()}
        linha["DATAS_CONCLUSAO_FASE"] = _data_para_iso(serie_datas_fase.loc[idx]) if serie_datas_fase is not None else None
        linha["DATAS_VIGENCIA"] = _data_para_iso(serie_datas_vigencia.loc[idx]) if serie_datas_vigencia is not None else None
        linha["DATAS_CONCLUSAO_ATUAL"] = _data_para_iso(serie_datas_atual.loc[idx]) if serie_datas_atual is not None else None
        linhas_filtro.append(linha)

    dados_painel = {
        "blocos": [
            {"chave": chave, "titulo": titulo, "opcoes": opcoes}
            for chave, titulo, opcoes in blocos_config
        ],
        "linhas": linhas_filtro,
        "statusPadrao": sorted(STATUS_PADRAO_SELECIONADOS),
        # Catálogo de colunas do DETALHAMENTO, para o bloco de customização
        # do painel. O JS mostra e valida em tempo real, mas quem manda de
        # verdade é normalizar_colunas_detalhamento no Python — a interface
        # é conveniência, não é a trava.
        "colunasDetalhamento": [
            {
                "chave": c["chave"],
                "titulo": c["titulo"],
                "obrigatoria": bool(c.get("obrigatoria")),
                "requer": list(c.get("requer", [])),
                "padrao": bool(c.get("padrao")),
            }
            for c in CATALOGO_COLUNAS_DETALHAMENTO
        ],
        "limiteColunasDetalhamento": LIMITE_COLUNAS_DETALHAMENTO,
        # Paleta dos gráficos do painel, já dessaturada aqui no Python. Vem
        # pronta do mesmo lugar que a do PDF (FATOR_SATURACAO_GRAFICOS), em
        # vez de repetir os códigos de cor no JavaScript — assim os dois
        # nunca saem com tons diferentes.
        "paletaGraficos": {
            "azul": dessaturar_hex("#4E92BA"),
            "vermelho": dessaturar_hex("#BB6060"),
            "verde": dessaturar_hex("#49925C"),
            "amarelo": dessaturar_hex("#BC9E2C"),
            "medidorInsatisfatorio": dessaturar_hex("#BB6060"),
            "medidorRegular": dessaturar_hex("#D9A441"),
            "medidorBom": dessaturar_hex("#9FCE9B"),
            "medidorOtimo": dessaturar_hex("#3F8F52"),
        },
        "arvoresData": {
            "DATAS_CONCLUSAO_FASE": _construir_arvore_datas(df_base, col_prazo),
            "DATAS_VIGENCIA": _construir_arvore_datas(df_base, col_vigencia),
            "DATAS_CONCLUSAO_ATUAL": _construir_arvore_datas(df_base, col_prazo_atual),
        },
    }
    # Serializado num arquivo .json separado (não mais embutido dentro do
    # HTML) — evita qualquer risco de tamanho/limite ao colocar um bloco de
    # dados potencialmente grande dentro da própria página; o JS busca esse
    # arquivo via fetch() depois que a página carrega.
    dados_painel_json = json.dumps(dados_painel, ensure_ascii=False)

    # =====================================================
    # PAINEL DE FILTROS — HTML/CSS/JS (design system AGF)
    # =====================================================
    # Reproduz o painel inteiro como uma página web local, renderizada numa
    # janela nativa via pywebview — permite aplicar TODAS as propriedades do
    # design system extraídas do app AGF (cores, raios de borda, sombras e
    # transições), coisa que o Tkinter/ttk não suporta nativamente.
    #
    # A tipografia usa a pilha "Segoe UI"/Roboto/system-ui em vez de
    # Jakarta/Roboto puras: como o relatório roda em máquinas de rede sem
    # acesso garantido à internet, não é seguro depender de fontes do Google
    # Fonts — a pilha abaixo já reproduz peso e proporção muito próximos do
    # original, usando fontes nativas do Windows.
    html_paginal = r"""<!DOCTYPE html>
<html lang="pt-BR">
<head>
<meta charset="UTF-8">
<!-- Sem isto, celular nenhum respeita a largura real da tela: o navegador
     mobile assume um "viewport" de desktop (~980px) e encolhe a página
     inteira pra caber, daí o efeito de "tudo minúsculo, precisa dar zoom".
     Com isto, 1px de CSS passa a valer 1px de tela — e é isso que faz as
     regras @media (max-width: 768px) logo abaixo (blocos de filtro em
     acordeão, grade Secretaria em 1 coluna, gráfico de prazo com rolagem)
     realmente entrarem em vigor no celular, em vez de nunca disparar. -->
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>CGAPE - BALANÇO PAC</title>
<script>
  // Aplica o tema salvo ANTES do <style> ser lido, pra não piscar o tema
  // escuro (padrão do :root) por uma fração de segundo antes de trocar pro
  // claro salvo da sessão anterior.
  (function () {
    try {
      if (localStorage.getItem("cgape-tema") === "claro") {
        document.documentElement.setAttribute("data-tema", "claro");
      }
    } catch (e) {}
  })();
</script>
<style>
  :root {
    --cor-fundo: #303030;
    --cor-card: #353B47;
    --cor-card-elevado: #41454F;
    --cor-texto-primario: #FFFFFF;
    --cor-texto-secundario: #B3BAC9;
    --cor-texto-terciario: #858B97;
    --cor-acento-mint: #B8EAE1;
    --cor-acento-teal: #72B4AE;
    --cor-acento-teal-hover: #8AC4BF;
    --cor-acento-peach: #EEB489;
    --cor-acento-gold: #E0AB45;
    /* Gold claro: mesma função que o mint tem em relação ao teal — o tom
       de texto/hover em cima do gold. Usado no bloco de colunas. */
    --cor-acento-gold-claro: #F2CE83;
    /* teal/mint usados como COR DE TEXTO (número de destaque, título de
       card marcado, hover de botão/link etc.) -- separados de
       --cor-acento-teal/mint (que continuam servindo fundo/borda/badge)
       porque no tema claro o texto precisa de um tom bem mais escuro pra
       ter contraste em cima de branco, enquanto fundo/borda ficam iguais
       nos dois temas. No escuro são idênticos aos originais. */
    --cor-acento-teal-texto: var(--cor-acento-teal);
    --cor-acento-mint-texto: var(--cor-acento-mint);
    --raio-sm: 4px;
    --raio-md: 8px;
    --raio-lg: 16px;
    --raio-xl: 24px;
    --sombra-card: 0px 5px 40px 0px rgba(9, 14, 21, 0.16);
    --transicao-rapida: 0.15s ease;
    --transicao-padrao: 0.25s ease-in-out;
    --fonte: "Segoe UI", "Roboto", system-ui, -apple-system, sans-serif;
  }

  /* --- Tema claro: mesmo design system (AGF), só troca fundo/superfície/
     texto -- acentos (teal, mint, peach, gold) e cores semânticas (fase,
     atrasada, etc., definidas mais abaixo/no JS) continuam iguais nos dois
     temas, é só a "casca" que muda. Ativado via data-tema="claro" na
     <html>, alternado pela chave de tema no topo (ver .tema-switch) e
     lembrado entre sessões (localStorage "cgape-tema"). Sem o atributo, o
     :root de cima já é o tema escuro (o original/padrão do projeto). */
  html[data-tema="claro"] {
    --cor-fundo: #F4F6F8;
    --cor-card: #FFFFFF;
    --cor-card-elevado: #EBEFF2;
    --cor-texto-primario: #1B2430;
    --cor-texto-secundario: #56606E;
    --cor-texto-terciario: #8A94A3;
    --sombra-card: 0px 5px 30px 0px rgba(27, 36, 48, 0.10);
    /* tons de teal/mint escurecidos SÓ pra uso como texto (ver comentário
       em --cor-acento-teal-texto acima) -- o teal/mint "de fundo" (botões,
       badges, bordas) continua igual ao escuro, só o texto muda. */
    --cor-acento-teal-texto: #2E6B66;
    --cor-acento-mint-texto: #1F7A73;
  }

  * { box-sizing: border-box; }

  html, body {
    margin: 0;
    padding: 0;
    background: var(--cor-fundo);
    color: var(--cor-texto-secundario);
    font-family: var(--fonte);
    font-size: 14px;
    height: 100%;
    overflow: hidden;
  }

  #app {
    display: none;
    flex-direction: column;
    position: fixed;
    inset: 0;
    z-index: 85;
    background: var(--cor-fundo);
    padding: 16px 20px;
  }

  /* --- Barra do topo --- */
  #topo {
    display: flex;
    align-items: flex-start;
    justify-content: space-between;
    gap: 16px;
    padding-bottom: 14px;
    margin-bottom: 12px;
    border-bottom: 1px solid var(--cor-card-elevado);
  }

  .topo-titulo-bloco {
    display: flex;
    flex-direction: column;
    gap: 6px;
    max-width: 760px;
  }
  .topo-titulo-textos {
    display: flex;
    flex-direction: column;
    gap: 2px;
  }
  .topo-titulo-principal {
    font-size: 16px;
    font-weight: 700;
    color: var(--cor-texto-primario);
    letter-spacing: 0.02em;
  }
  .topo-titulo-atualizacao {
    font-size: 11px;
    font-weight: 400;
    color: var(--cor-texto-secundario);
  }
  .topo-orientacao {
    margin: 0;
    font-size: 13px;
    font-weight: 600;
    color: var(--cor-texto-primario);
    line-height: 1.4;
  }

  #botoes-topo {
    display: flex;
    align-items: center;
    gap: 10px;
    flex-shrink: 0;
  }

  button {
    font-family: var(--fonte);
    cursor: pointer;
    border: none;
    outline: none;
    transition: background var(--transicao-rapida), transform var(--transicao-rapida), box-shadow var(--transicao-rapida);
  }

  .btn {
    background: var(--cor-card-elevado);
    color: var(--cor-texto-primario);
    font-size: 12px;
    font-weight: 500;
    padding: 8px 14px;
    border-radius: var(--raio-md);
  }
  .btn:hover { background: var(--cor-acento-teal-hover); color: #1A1A1A; }
  .btn:active { transform: scale(0.97); }

  .btn-acento {
    background: var(--cor-acento-teal);
    color: #16211F;
    font-size: 13px;
    font-weight: 700;
    padding: 10px 20px;
    border-radius: var(--raio-md);
    box-shadow: var(--sombra-card);
  }
  .btn-acento:hover { background: var(--cor-acento-mint); }
  .btn-acento:active { transform: scale(0.97); }

  .btn-mini {
    background: transparent;
    color: var(--cor-texto-secundario);
    border: 1px solid var(--cor-card-elevado);
    font-size: 11px;
    font-weight: 500;
    padding: 4px 10px;
    border-radius: var(--raio-sm);
  }
  .btn-mini:hover {
    border-color: var(--cor-acento-teal);
    color: var(--cor-acento-mint-texto);
  }

  /* --- Grade de blocos --- */
  #grade {
    flex: 1;
    display: grid;
    grid-template-columns: repeat(5, 1fr);
    grid-auto-rows: 1fr;
    gap: 14px;
    overflow: hidden;
    min-height: 0;
  }

  .bloco {
    background: var(--cor-card);
    border-radius: var(--raio-lg);
    box-shadow: var(--sombra-card);
    border: 1px solid rgba(114, 180, 174, 0.25);
    display: flex;
    flex-direction: column;
    min-height: 0;
    overflow: hidden;
    transition: border-color var(--transicao-padrao);
  }
  .bloco:hover { border-color: var(--cor-acento-teal); }

  /* --- Seção COLUNAS DO DETALHAMENTO ---
     Não fica mais no painel de filtros: virou uma seção da janela
     "Selecione as páginas do relatório" (ela não filtra dados nem escolhe
     páginas — escolhe o que a tabela de Detalhamento mostra). O acento em
     gold, a marca de seleção e a barra de rolagem em gold separam essa
     seção da grade de páginas logo acima. --- */
  #paginas-colunas {
    flex-shrink: 0;
    border-top: 1px solid var(--cor-card-elevado);
    padding: 12px 22px 4px 22px;
    display: flex;
    flex-direction: column;
    gap: 8px;
  }
  .paginas-colunas-topo {
    display: flex;
    align-items: center;
    justify-content: space-between;
    gap: 8px;
  }
  .paginas-colunas-titulo {
    font-size: 12.5px;
    font-weight: 700;
    letter-spacing: 0.02em;
    text-transform: uppercase;
    color: var(--cor-texto-primario);
  }
  .paginas-colunas-contador {
    font-size: 12px;
    color: var(--cor-texto-secundario);
  }
  .paginas-colunas-botoes {
    display: flex;
    flex-wrap: wrap;
    gap: 6px;
  }
  .paginas-colunas-botoes .btn-mini:hover,
  .paginas-colunas-botoes .btn-mini:focus {
    border-color: var(--cor-acento-gold);
    color: var(--cor-acento-gold-claro);
  }
  .paginas-colunas-lista {
    display: grid;
    grid-template-columns: repeat(auto-fill, minmax(210px, 1fr));
    gap: 0 14px;
    max-height: 168px;
    overflow-y: auto;
    padding: 2px 0 6px 0;
  }
  .paginas-colunas-lista::-webkit-scrollbar { width: 8px; }
  .paginas-colunas-lista::-webkit-scrollbar-track { background: var(--cor-card); border-radius: var(--raio-sm); }
  .paginas-colunas-lista::-webkit-scrollbar-thumb {
    background: var(--cor-card-elevado);
    border-radius: var(--raio-sm);
  }
  .paginas-colunas-lista::-webkit-scrollbar-thumb:hover { background: var(--cor-acento-gold); }
  #paginas-colunas label.item input[type="checkbox"]:checked {
    background: var(--cor-acento-gold);
    border-color: var(--cor-acento-gold);
  }
  /* Ao bater o limite de colunas, as não marcadas NÃO somem da lista — só
     o quadrado de seleção esmaece, deixando visível que a lista inteira
     continua ali e basta desmarcar uma para liberar outra. Mesmo
     tratamento na coluna obrigatória (sempre marcada, nunca editável). */
  #paginas-colunas label.item.col-trava,
  #paginas-colunas label.item.col-trava:hover {
    cursor: not-allowed;
    background: transparent;
  }
  #paginas-colunas label.item.col-trava input[type="checkbox"] {
    opacity: 0.3;
    cursor: not-allowed;
  }

  .bloco-titulo {
    padding: 10px 14px 8px 14px;
    font-size: 12.5px;
    font-weight: 700;
    color: var(--cor-texto-primario);
    letter-spacing: 0.02em;
    text-transform: uppercase;
    border-bottom: 1px solid var(--cor-card-elevado);
    display: flex;
    align-items: center;
    justify-content: space-between;
    gap: 8px;
  }
  .bloco-titulo .arvore-data-secao {
    flex-shrink: 0;
  }
  /* Seta de expandir/recolher do acordeão de filtros no celular — ver
     configurarAccordionFiltros() no JS e o @media (max-width: 768px) no
     fim desta folha de estilos. Some no desktop (grade normal, sem
     acordeão); só aparece dentro daquele media query. */
  .bloco-seta {
    display: none;
  }

  .bloco-botoes {
    display: flex;
    flex-wrap: wrap;
    gap: 6px;
    padding: 8px 10px;
  }

  .bloco-lista {
    flex: 1;
    overflow-y: auto;
    padding: 4px 10px 10px 10px;
  }

  /* --- Campo de busca dentro do bloco (hoje só o MUNICÍPIO usa) ---
     A lista de municípios passa de trezentos itens, e rolar até achar um
     nome é lento. O campo fica logo abaixo dos botões e vai escondendo as
     opções que não batem com o que foi digitado, sem mexer no que já está
     marcado. --- */
  .bloco-busca {
    padding: 0 10px 8px 10px;
  }
  .bloco-busca input {
    width: 100%;
    padding: 6px 9px;
    font-family: inherit;
    font-size: 11.5px;
    color: var(--cor-texto-primario);
    background: var(--cor-fundo);
    border: 1px solid var(--cor-card-elevado);
    border-radius: var(--raio-sm);
    outline: none;
    transition: border-color var(--transicao-rapida);
  }
  .bloco-busca input::placeholder {
    color: var(--cor-texto-terciario);
  }
  .bloco-busca input:focus {
    border-color: var(--cor-acento-teal);
  }
  .bloco-busca .busca-vazia {
    display: none;
    padding: 5px 2px 0 2px;
    font-size: 11px;
    color: var(--cor-texto-terciario);
  }
  .bloco-busca.sem-resultado .busca-vazia {
    display: block;
  }
  label.item.oculto-busca {
    display: none;
  }

  /* scrollbar customizada, na paleta escura */
  .bloco-lista::-webkit-scrollbar { width: 8px; }
  .bloco-lista::-webkit-scrollbar-track { background: var(--cor-card); border-radius: var(--raio-sm); }
  .bloco-lista::-webkit-scrollbar-thumb {
    background: var(--cor-card-elevado);
    border-radius: var(--raio-sm);
  }
  .bloco-lista::-webkit-scrollbar-thumb:hover { background: var(--cor-acento-teal); }

  /* --- Filtros de data em lista suspensa (dropdown) — ícone de calendário
     verde no título do bloco (FASE, STATUS e CLÁUSULA SUSPENSIVA), que
     abre um painel flutuante ao clicar, sem ocupar espaço fixo no bloco. --- */
  .arvore-data-secao { position: relative; }
  .arvore-dropdown-toggle {
    width: 26px;
    height: 26px;
    display: flex;
    align-items: center;
    justify-content: center;
    background: transparent;
    color: var(--cor-acento-teal-texto);
    border: none;
    border-radius: var(--raio-sm);
    padding: 0;
    cursor: pointer;
    transition: background var(--transicao-rapida);
    position: relative;
  }
  .arvore-dropdown-toggle:hover { background: rgba(114, 180, 174, 0.18); }
  .arvore-dropdown-toggle .contagem {
    display: none;
    position: absolute;
    top: -4px;
    right: -4px;
    align-items: center;
    justify-content: center;
    min-width: 14px;
    height: 14px;
    padding: 0 3px;
    border-radius: 999px;
    background: var(--cor-acento-teal);
    color: #16211F;
    font-size: 8.5px;
    font-weight: 700;
    font-family: var(--fonte);
  }
  .arvore-dropdown-toggle .contagem.visivel { display: flex; }
  .arvore-dropdown-painel {
    display: none;
    position: fixed;
    background: var(--cor-card-elevado);
    border: 1px solid var(--cor-acento-teal);
    border-radius: var(--raio-md);
    box-shadow: var(--sombra-card);
    padding: 8px 10px;
    width: 220px;
    max-height: 320px;
    overflow-y: auto;
    z-index: 90;
    font-size: 11px;
  }
  .arvore-dropdown-painel.aberto { display: block; }
  .arvore-dropdown-painel::-webkit-scrollbar { width: 6px; }
  .arvore-dropdown-painel::-webkit-scrollbar-thumb {
    background: var(--cor-card);
    border-radius: var(--raio-sm);
  }
  .arvore-dropdown-painel summary {
    list-style: none;
    cursor: pointer;
  }
  .arvore-dropdown-painel summary::-webkit-details-marker { display: none; }
  .arvore-dropdown-painel summary::before {
    content: "▸";
    display: inline-block;
    width: 10px;
    font-size: 9px;
    color: var(--cor-texto-terciario);
  }
  .arvore-dropdown-painel details[open] > summary::before { content: "▾"; }
  .arvore-item, .arvore-dropdown-painel summary label {
    display: flex;
    align-items: center;
    gap: 5px;
    padding: 1.5px 0;
    color: var(--cor-texto-primario);
    cursor: pointer;
    text-transform: uppercase;
  }
  .arvore-dropdown-painel details { margin-left: 2px; }
  .arvore-nivel-mes { margin-left: 14px; }
  .arvore-nivel-dia { margin-left: 28px; font-size: 10.5px; }
  .arvore-raiz { font-weight: 700; margin-bottom: 2px; }
  .arvore-dropdown-painel input[type="checkbox"] {
    appearance: none;
    -webkit-appearance: none;
    width: 12px;
    height: 12px;
    min-width: 12px;
    border-radius: 3px;
    border: 1.5px solid var(--cor-texto-terciario);
    background: transparent;
    cursor: pointer;
    position: relative;
  }
  .arvore-dropdown-painel input[type="checkbox"]:checked,
  .arvore-dropdown-painel input[type="checkbox"]:indeterminate {
    background: var(--cor-acento-teal);
    border-color: var(--cor-acento-teal);
  }
  .arvore-dropdown-painel input[type="checkbox"]:indeterminate { opacity: 0.6; }
  .arvore-dropdown-painel input[type="checkbox"]:checked::after {
    content: "";
    position: absolute;
    left: 2.5px;
    top: -0.5px;
    width: 3px;
    height: 6.5px;
    border: solid #16211F;
    border-width: 0 2px 2px 0;
    transform: rotate(40deg);
  }

  /* --- Checkbox customizado (pill) --- */
  label.item {
    display: flex;
    align-items: flex-start;
    gap: 8px;
    padding: 5px 8px;
    margin: 1px 0;
    border-radius: var(--raio-sm);
    font-size: 12px;
    color: var(--cor-texto-primario);
    cursor: pointer;
    transition: background var(--transicao-rapida);
    line-height: 1.35;
    text-transform: uppercase;
  }
  label.item:hover { background: var(--cor-card-elevado); }

  label.item input[type="checkbox"] {
    appearance: none;
    -webkit-appearance: none;
    width: 15px;
    height: 15px;
    min-width: 15px;
    margin-top: 1px;
    border-radius: 4px;
    border: 1.5px solid var(--cor-texto-terciario);
    background: transparent;
    cursor: pointer;
    position: relative;
    transition: background var(--transicao-rapida), border-color var(--transicao-rapida);
  }
  label.item input[type="checkbox"]:checked {
    background: var(--cor-acento-teal);
    border-color: var(--cor-acento-teal);
  }
  label.item input[type="checkbox"]:checked::after {
    content: "";
    position: absolute;
    left: 4px;
    top: 1px;
    width: 4px;
    height: 8px;
    border: solid #16211F;
    border-width: 0 2px 2px 0;
    transform: rotate(40deg);
  }
  label.item.indisponivel {
    display: none;
  }
  .arvore-item.indisponivel {
    display: none;
  }

  /* --- Aviso de fechamento por inatividade --- */
  #aviso-inatividade {
    display: none;
    position: fixed;
    inset: 0;
    background: rgba(0,0,0,0.55);
    align-items: center;
    justify-content: center;
    z-index: 50;
  }
  #aviso-inatividade .caixa {
    background: var(--cor-card);
    border-radius: var(--raio-lg);
    box-shadow: var(--sombra-card);
    padding: 24px 28px;
    text-align: center;
    color: var(--cor-texto-primario);
    font-size: 13px;
  }

  /* --- Modal do Controle de Qualidade --- */
  #modal-overlay {
    display: none;
    position: fixed;
    inset: 0;
    background: rgba(0,0,0,0.6);
    align-items: center;
    justify-content: center;
    z-index: 95;
  }
  #modal-qualidade {
    background: var(--cor-card);
    border-radius: var(--raio-lg);
    box-shadow: var(--sombra-card);
    border: 1px solid var(--cor-acento-teal);
    width: min(760px, 90vw);
    max-height: 80vh;
    display: flex;
    flex-direction: column;
    overflow: hidden;
  }
  #modal-qualidade .modal-titulo {
    display: flex;
    align-items: center;
    justify-content: space-between;
    gap: 12px;
    padding: 16px 22px;
    font-size: 14px;
    font-weight: 700;
    color: var(--cor-texto-primario);
    border-bottom: 1px solid var(--cor-card-elevado);
    background: var(--cor-card-elevado);
  }
  #modal-qualidade .modal-corpo {
    padding: 18px 22px;
    overflow-y: auto;
    font-size: 12.5px;
    line-height: 1.55;
    color: var(--cor-texto-primario);
  }
  #modal-qualidade .modal-corpo::-webkit-scrollbar { width: 8px; }
  #modal-qualidade .modal-corpo::-webkit-scrollbar-thumb {
    background: var(--cor-card-elevado);
    border-radius: var(--raio-sm);
  }
  .qc-cabecalho {
    margin-bottom: 12px;
  }
  .qc-grupo-titulo {
    font-weight: 700;
    color: var(--cor-acento-teal-texto);
    margin: 14px 0 6px 0;
  }
  .qc-linha {
    margin-bottom: 14px;
    line-height: 1.6;
  }
  /* Hierarquia do Controle de Qualidade: OBJETO > ALERTA > ITENS.
     O recuo é o que deixa a leitura óbvia — o objeto na margem, cada tipo
     de pendência recuado abaixo dele, e os itens daquela pendência
     recuados mais um nível. */
  .qc-objeto {
    margin-bottom: 16px;
    line-height: 1.6;
  }
  .qc-objeto-titulo {
    font-weight: 700;
  }
  .qc-objeto-contagem {
    font-weight: 400;
    color: var(--cor-texto-terciario);
  }
  .qc-alerta {
    margin: 6px 0 0 18px;
    padding-left: 10px;
    border-left: 2px solid var(--cor-card-elevado);
  }
  .qc-alerta-motivo {
    color: var(--cor-texto-secundario);
  }
  .qc-alerta-itens {
    margin-top: 2px;
  }
  .qc-item-link {
    color: var(--cor-acento-teal-texto);
    text-decoration: underline;
    cursor: pointer;
  }
  .qc-item-link:hover {
    color: var(--cor-acento-mint-texto);
  }
  .qc-instrucao {
    margin-top: 14px;
    font-weight: 700;
  }
  #modal-qualidade .modal-rodape {
    display: flex;
    align-items: center;
    justify-content: flex-end;
    gap: 10px;
    padding: 14px 22px;
    border-top: 1px solid var(--cor-card-elevado);
  }
  #modal-qualidade .modal-contagem {
    margin-right: auto;
    font-size: 12px;
    color: var(--cor-texto-secundario);
  }
  #modal-qualidade button:disabled {
    opacity: 0.4;
    cursor: not-allowed;
  }

  /* --- Ficha Cadastral de uma ação (modo tela + modo impressão) --- */
  #ficha-overlay {
    display: none;
    position: fixed;
    inset: 0;
    background: rgba(0,0,0,0.7);
    align-items: center;
    justify-content: center;
    z-index: 97;
  }
  #ficha-painel {
    width: 95vw;
    height: 92vh;
    background: var(--cor-fundo);
    border-radius: var(--raio-lg);
    box-shadow: var(--sombra-card);
    border: 1px solid var(--cor-acento-teal);
    display: flex;
    flex-direction: column;
    overflow: hidden;
  }
  .ficha-topo {
    display: flex;
    align-items: center;
    justify-content: space-between;
    gap: 16px;
    padding: 14px 24px;
    background: var(--cor-card-elevado);
    border-bottom: 1px solid var(--cor-card);
    flex-shrink: 0;
  }
  .ficha-topo-titulo {
    display: flex;
    align-items: baseline;
    gap: 12px;
    min-width: 0;
    font-family: var(--fonte-display, inherit);
    font-weight: 700;
    font-size: 17px;
    color: var(--cor-texto-primario);
    overflow: hidden;
    text-overflow: ellipsis;
    white-space: nowrap;
  }
  .ficha-topo-titulo span:last-child {
    flex-shrink: 0;
    font-size: 12px;
    font-weight: 500;
    color: var(--cor-acento-teal-texto);
  }
  .ficha-topo-acoes {
    display: flex;
    align-items: center;
    gap: 8px;
    flex-shrink: 0;
  }
  .ficha-botao-icone {
    display: inline-flex;
    align-items: center;
    justify-content: center;
    width: 34px;
    height: 34px;
    border-radius: var(--raio-md);
    border: none;
    background: var(--cor-card);
    color: var(--cor-texto-secundario);
    cursor: pointer;
    transition: var(--transicao-rapida);
  }
  .ficha-botao-icone:hover {
    background: var(--cor-acento-teal);
    color: #16211F;
  }
  #ficha-btn-fechar,
  #ficha-btn-fechar-impressao {
    font-size: 20px;
    line-height: 1;
  }
  .ficha-corpo {
    flex: 1;
    overflow-y: auto;
    padding: 22px 26px;
  }
  .ficha-secao {
    margin-bottom: 20px;
  }
  .ficha-secao-titulo {
    font-size: 11px;
    font-weight: 700;
    letter-spacing: 0.05em;
    text-transform: uppercase;
    color: var(--cor-acento-teal-texto);
    margin-bottom: 10px;
    padding-bottom: 6px;
    border-bottom: 1px solid var(--cor-card-elevado);
  }
  .ficha-secao-grid {
    display: grid;
    grid-template-columns: repeat(auto-fit, minmax(220px, 1fr));
    gap: 14px 22px;
  }
  .ficha-secao-grid.ficha-grid-largo {
    grid-template-columns: repeat(auto-fit, minmax(360px, 1fr));
  }
  .ficha-campo-rotulo {
    font-size: 10.5px;
    color: var(--cor-texto-terciario);
    text-transform: uppercase;
    letter-spacing: 0.03em;
    margin-bottom: 3px;
  }
  .ficha-campo-valor {
    font-size: 13.5px;
    color: var(--cor-texto-primario);
    line-height: 1.5;
    white-space: pre-wrap;
    word-break: break-word;
  }
  .ficha-campo-valor.ficha-campo-vazio {
    color: var(--cor-texto-terciario);
    font-style: italic;
  }
  .ficha-campo-valor a {
    color: var(--cor-acento-teal-texto);
    text-decoration: underline;
    word-break: break-all;
  }
  .ficha-campo-valor.ficha-campo-alerta {
    border: 1.5px solid #E15757;
    border-radius: var(--raio-sm);
    padding: 5px 8px;
    margin: -5px -8px;
    background: rgba(225, 87, 87, 0.08);
  }

  /* --- Modo impressão: reskin pra parecer uma folha A4 --- */
  #ficha-painel.ficha-modo-impressao {
    background: #E8E8E8;
  }
  #ficha-painel.ficha-modo-impressao .ficha-topo {
    background: #F4F4F4;
    border-bottom: 1px solid #D0D0D0;
  }
  #ficha-painel.ficha-modo-impressao .ficha-topo-titulo {
    color: #222;
  }
  #ficha-painel.ficha-modo-impressao .ficha-corpo {
    display: none;
  }
  .ficha-pagina-a4-wrap {
    display: none;
    flex: 1;
    overflow-y: auto;
    justify-content: center;
    padding: 24px;
  }
  #ficha-painel.ficha-modo-impressao .ficha-pagina-a4-wrap {
    display: flex;
  }
  #ficha-pagina-a4 {
    width: 210mm;
    min-height: 297mm;
    height: fit-content;
    background: #fff;
    color: #1a1a1a;
    box-shadow: 0 4px 24px rgba(0,0,0,0.35);
    padding: 16mm;
    box-sizing: border-box;
    font-family: Arial, Helvetica, sans-serif;
  }
  .ficha-a4-titulo {
    font-size: 20px;
    font-weight: 800;
    margin-bottom: 2px;
  }
  .ficha-a4-subtitulo {
    font-size: 11px;
    color: #666;
    margin-bottom: 14px;
  }
  .ficha-a4-secao {
    margin-bottom: 12px;
    break-inside: avoid;
  }
  .ficha-a4-secao-titulo {
    font-size: 10px;
    font-weight: 700;
    letter-spacing: 0.04em;
    text-transform: uppercase;
    color: #1565A3;
    border-bottom: 1px solid #ccc;
    padding-bottom: 3px;
    margin-bottom: 6px;
  }
  .ficha-a4-grid {
    display: grid;
    grid-template-columns: repeat(3, 1fr);
    gap: 6px 16px;
  }
  .ficha-a4-grid.ficha-a4-grid-largo {
    grid-template-columns: repeat(2, 1fr);
  }
  .ficha-a4-campo-rotulo {
    font-size: 8.5px;
    color: #777;
    text-transform: uppercase;
  }
  .ficha-a4-campo-valor {
    font-size: 10.5px;
    color: #1a1a1a;
    line-height: 1.4;
    white-space: pre-wrap;
    word-break: break-word;
  }
  .ficha-a4-campo-valor.ficha-a4-campo-alerta {
    border: 1.2px solid #D64545;
    border-radius: 3px;
    padding: 3px 5px;
    margin: -3px -5px;
    background: rgba(214, 69, 69, 0.08);
  }

  /* --- Lista de desambiguação (mais de uma ação encontrada) --- */
  /* --- Janela de seleção de páginas do relatório --- */
  #paginas-overlay {
    display: none;
    position: fixed;
    inset: 0;
    background: rgba(0,0,0,0.6);
    align-items: center;
    justify-content: center;
    z-index: 99;
  }
  #paginas-painel {
    background: var(--cor-card);
    border-radius: var(--raio-lg);
    box-shadow: var(--sombra-card);
    width: min(1100px, 92vw);
    max-height: 95vh;
    display: flex;
    flex-direction: column;
    overflow: hidden;
  }
  #paginas-painel .modal-titulo {
    padding: 13px 22px;
    font-size: 16px;
    font-weight: 700;
    border-bottom: 1px solid var(--cor-card-elevado);
  }
  .paginas-acoes {
    display: flex;
    align-items: center;
    gap: 8px;
    padding: 10px 22px 0 22px;
  }
  .paginas-contador {
    margin-left: auto;
    font-size: 12px;
    color: var(--cor-texto-secundario);
  }
  .paginas-grade {
    flex: 1 1 auto;
    min-height: 0;
    display: grid;
    grid-template-columns: repeat(auto-fill, minmax(160px, 1fr));
    gap: 12px;
    padding: 14px 22px 16px 22px;
    overflow-y: auto;
  }
  /* Cada miniatura é um botão inteiro: clicar em qualquer ponto do cartão
     marca ou desmarca a página, sem precisar acertar uma caixinha. */
  .pagina-card {
    display: flex;
    flex-direction: column;
    gap: 8px;
    padding: 10px;
    border: 2px solid var(--cor-card-elevado);
    border-radius: var(--raio-md);
    background: var(--cor-fundo);
    cursor: pointer;
    text-align: left;
    transition: border-color var(--transicao-rapida), background var(--transicao-rapida);
  }
  .pagina-card:hover { border-color: var(--cor-acento-teal); }
  .pagina-card.marcada {
    border-color: var(--cor-acento-teal);
    background: rgba(114, 180, 174, 0.10);
  }
  .pagina-card.marcada .pagina-card-titulo { color: var(--cor-acento-mint-texto); }
  /* A miniatura fica opaca quando a página está fora da seleção — a
     diferença precisa ser visível de relance numa grade de onze cartões. */
  .pagina-card svg { opacity: 0.35; transition: opacity var(--transicao-rapida); }
  .pagina-card.marcada svg { opacity: 1; }
  .pagina-card-titulo {
    font-size: 12px;
    font-weight: 700;
    color: var(--cor-texto-secundario);
    line-height: 1.25;
  }
  .pagina-card-sub {
    font-size: 11px;
    color: var(--cor-texto-terciario);
  }
  #paginas-painel .modal-rodape {
    display: flex;
    justify-content: flex-end;
    gap: 10px;
    padding: 14px 22px;
    border-top: 1px solid var(--cor-card-elevado);
  }

  #ficha-multiplos-overlay {
    display: none;
    position: fixed;
    inset: 0;
    background: rgba(0,0,0,0.6);
    align-items: center;
    justify-content: center;
    z-index: 98;
  }
  #ficha-multiplos-painel {
    background: var(--cor-card);
    border-radius: var(--raio-lg);
    box-shadow: var(--sombra-card);
    border: 1px solid var(--cor-acento-teal);
    /* Mais larga que antes: agora abriga os dois campos de busca lado a
       lado, e as descrições das ações ficam mais legíveis na lista. */
    width: min(700px, 94vw);
    max-height: 80vh;
    display: flex;
    flex-direction: column;
    overflow: hidden;
  }
  #ficha-multiplos-painel .modal-titulo {
    padding: 16px 22px;
    font-size: 14px;
    font-weight: 700;
    color: var(--cor-texto-primario);
    background: var(--cor-card-elevado);
    border-bottom: 1px solid var(--cor-card);
  }
  .ficha-multiplos-lista {
    overflow-y: auto;
    padding: 14px 18px;
  }
  .ficha-multiplos-item {
    display: flex;
    align-items: center;
    justify-content: space-between;
    gap: 10px;
    padding: 10px 12px;
    border-radius: var(--raio-md);
    background: var(--cor-card-elevado);
    margin-bottom: 6px;
    cursor: pointer;
    transition: var(--transicao-rapida);
  }
  .ficha-multiplos-item:hover {
    background: var(--cor-acento-teal);
    color: #16211F;
  }
  .ficha-multiplos-item-info { min-width: 0; }
  .ficha-multiplos-item-objeto {
    font-size: 13px;
    white-space: nowrap;
    overflow: hidden;
    text-overflow: ellipsis;
  }
  .ficha-multiplos-item-meta {
    font-size: 11px;
    opacity: 0.75;
    margin-top: 2px;
  }
  #ficha-multiplos-painel .modal-rodape {
    padding: 12px 18px;
    display: flex;
    justify-content: flex-end;
    border-top: 1px solid var(--cor-card-elevado);
  }

  /* --- Página dedicada do Mapa Mental: mesmo padrão de overlay em tela
     cheia do #app (filtros), só que por cima dele (z-index maior), já que
     o botão que abre fica dentro do próprio painel de filtros. O conteúdo
     é uma página HTML autônoma (gerada em Python, ver mapa_mental_html.py)
     carregada num <iframe> — isolado do CSS/JS deste painel de propósito,
     para os dois nunca conflitarem entre si (ids repetidos, temas etc.).
     Sem barra externa aqui: o botão de fechar mora dentro do próprio
     cabeçalho da página do iframe (mesma linha do título "... Mapa
     Mental") e avisa esta página por postMessage quando é clicado — ver
     o listener "message" logo abaixo, e o botão #fechar-mapa em
     mapa_mental_html.py. */
  #mapa-mental-overlay {
    display: none;
    position: fixed;
    inset: 0;
    z-index: 95;
    background: var(--cor-fundo);
  }
  #mapa-mental-iframe {
    width: 100%;
    height: 100%;
    border: none;
    background: #eef1f6;
  }

  /* --- Modal de Pré-visualização (dashboard) --- */
  #preview-overlay {
    display: flex;
    position: fixed;
    inset: 0;
    background: var(--cor-fundo);
    align-items: stretch;
    justify-content: stretch;
    z-index: 70;
  }
  #preview-modal {
    background: var(--cor-fundo);
    border-radius: 0;
    box-shadow: none;
    border: none;
    width: 100%;
    height: 100%;
    max-height: 100%;
    display: flex;
    flex-direction: column;
    overflow: hidden;
  }
  #preview-modal .modal-titulo {
    padding: 12px 22px;
    font-size: 15px;
    font-weight: 700;
    color: var(--cor-texto-primario);
    background: var(--cor-card-elevado);
    display: flex;
    align-items: center;
    justify-content: space-between;
  }
  .modal-titulo-textos {
    display: flex;
    flex-direction: column;
    gap: 1px;
  }
  .modal-titulo-principal {
    font-size: 16px;
    font-weight: 700;
    color: var(--cor-texto-primario);
    letter-spacing: 0.02em;
  }
  .modal-titulo-atualizacao {
    font-size: 11px;
    font-weight: 400;
    color: var(--cor-texto-secundario);
  }
  .modal-titulo-botoes {
    display: flex;
    align-items: center;
    gap: 8px;
  }
  .modal-titulo-botoes .btn-acento {
    font-size: 12px;
    padding: 7px 14px;
    box-shadow: none;
  }
  /* --- Chave de tema claro/escuro: mesmo padrão visual da chave de
     alternância do Android (trilho em pílula, bolinha que desliza, sol de
     um lado e lua do outro) — mesmo componente nas duas barras (painel de
     filtros e dash), ao lado do Acesso Rápido. Estado atual = atributo
     data-tema na <html> (ver aplicarTema() no JS); sem o atributo é
     escuro, então a bolinha começa à direita (sobre a lua). */
  .tema-switch {
    display: inline-flex;
    align-items: center;
    padding: 0;
    background: none;
    border: none;
    flex-shrink: 0;
  }
  .tema-switch-trilho {
    position: relative;
    width: 50px;
    height: 26px;
    border-radius: 999px;
    background: var(--cor-card-elevado);
    border: 1px solid rgba(114, 180, 174, 0.25);
    display: flex;
    align-items: center;
    justify-content: space-between;
    padding: 0 5px;
    transition: background var(--transicao-padrao), border-color var(--transicao-padrao);
  }
  .tema-switch-icone {
    display: flex;
    /* sem z-index de propósito: a bolinha (definida depois no HTML) tem
       que ficar por cima e encobrir por completo o ícone do lado em que
       ela está parada -- só o ícone do lado "livre" (oposto) fica
       visível na pista. */
    opacity: 0.5;
    transition: opacity var(--transicao-rapida);
  }
  .tema-switch-sol { color: var(--cor-acento-gold); }
  .tema-switch-lua { color: var(--cor-texto-secundario); }
  html:not([data-tema="claro"]) .tema-switch-lua { opacity: 1; }
  html[data-tema="claro"] .tema-switch-sol { opacity: 1; }
  .tema-switch-bolinha {
    position: absolute;
    top: 2px;
    left: 2px;
    width: 20px;
    height: 20px;
    border-radius: 50%;
    background: var(--cor-texto-secundario);
    /* sombra própria, pequena e justa -- var(--sombra-card) é pensada pra
       cards grandes (30px de blur) e numa bolinha de 20px vira um brilho
       enorme que borra o trilho inteiro */
    box-shadow: 0 1px 3px rgba(0, 0, 0, 0.35);
    transform: translateX(24px);
    transition: transform var(--transicao-padrao), background var(--transicao-padrao);
  }
  html[data-tema="claro"] .tema-switch-bolinha {
    transform: translateX(0);
    background: var(--cor-acento-gold);
  }
  .tema-switch:hover .tema-switch-trilho { border-color: var(--cor-acento-teal); }

  /* --- Acesso Rápido: botão de ícone (grade 2x2) que abre um menu com os
     atalhos MAPA MENTAL, FILTROS e PUBLICAR — mesmo componente nas duas
     barras (painel de filtros e dash), sempre no fim da fileira de
     botões. O menu (.acesso-rapido-menu) fica escondido (display:none)
     até ganhar a classe "aberto". */
  .acesso-rapido-wrap {
    position: relative;
    display: flex;
  }
  .botao-icone-topo {
    display: flex;
    align-items: center;
    justify-content: center;
    width: 32px;
    height: 32px;
    padding: 0;
    background: var(--cor-card-elevado);
    color: var(--cor-texto-primario);
    border-radius: var(--raio-sm);
    transition: background var(--transicao-rapida), color var(--transicao-rapida);
  }
  .botao-icone-topo:hover, .botao-icone-topo[aria-expanded="true"] {
    background: var(--cor-acento-teal);
    color: #1A1A1A;
  }
  .acesso-rapido-menu {
    display: none;
    position: absolute;
    top: calc(100% + 6px);
    right: 0;
    flex-direction: column;
    gap: 2px;
    min-width: 170px;
    background: var(--cor-card);
    border: 1px solid var(--cor-card-elevado);
    border-radius: var(--raio-sm);
    box-shadow: 0 8px 24px rgba(0, 0, 0, 0.35);
    padding: 6px;
    z-index: 50;
  }
  .acesso-rapido-menu.aberto {
    display: flex;
  }
  .acesso-rapido-item {
    display: flex;
    align-items: center;
    gap: 10px;
    width: 100%;
    padding: 8px 10px;
    background: transparent;
    color: var(--cor-texto-primario);
    border-radius: var(--raio-sm);
    font-size: 12px;
    font-weight: 600;
    letter-spacing: 0.02em;
    text-align: left;
    transition: background var(--transicao-rapida);
  }
  .acesso-rapido-item svg { flex-shrink: 0; }
  .acesso-rapido-item:hover { background: var(--cor-card-elevado); }
  .acesso-rapido-item:disabled { opacity: 0.5; }

  /* --- Linha de filtro rápido por Secretaria/Órgão + Executor, no
     cabeçalho do dashboard. O separador verde e as pills de Executor só
     aparecem depois que pelo menos uma Secretaria é selecionada. --- */
  .dash-filtros-rapidos {
    display: flex;
    align-items: center;
    gap: 6px;
    padding: 8px 22px;
    background: var(--cor-card);
    border-bottom: 1px solid var(--cor-card-elevado);
    overflow-x: auto;
    flex-shrink: 0;
  }
  .dash-filtros-rapidos::-webkit-scrollbar { height: 6px; }
  .dash-filtros-rapidos::-webkit-scrollbar-thumb {
    background: var(--cor-card-elevado);
    border-radius: var(--raio-sm);
  }
  .dash-secretaria-filtro {
    display: flex;
    align-items: center;
    gap: 6px;
    flex-shrink: 0;
  }
  .dash-separador-executor {
    display: none;
    width: 2px;
    align-self: stretch;
    min-height: 26px;
    background: var(--cor-acento-teal);
    border-radius: 1px;
    margin: 0 6px;
    flex-shrink: 0;
  }
  .dash-separador-executor.visivel {
    display: block;
  }

  /* --- Linha dos filtros rápidos + botão da lupa ---
     A lupa fica FORA do contêiner que rola na horizontal: os botões de
     secretaria/executor podem ser muitos e empurrariam o botão para fora
     da tela. Assim ele fica sempre visível, na mesma altura das pílulas. */
  .dash-filtros-rapidos-linha {
    display: flex;
    align-items: center;
    gap: 10px;
    padding-right: 22px;
    background: var(--cor-card);
    border-bottom: 1px solid var(--cor-card-elevado);
    flex-shrink: 0;
  }
  .dash-filtros-rapidos-linha > .dash-filtros-rapidos {
    flex: 1;
    min-width: 0;
    border-bottom: none;
  }

  /* --- Campos de busca dentro da janela da lupa ---
     Ficam junto da lista que eles filtram: digitar reduz a lista na hora,
     em vez de exigir um segundo clique para "buscar". --- */
  .ficha-multiplos-busca {
    display: flex;
    gap: 10px;
    padding: 12px 18px;
    background: var(--cor-card-elevado);
    border-bottom: 1px solid var(--cor-card);
    flex-shrink: 0;
  }
  .ficha-multiplos-busca-input {
    flex: 1;
    min-width: 0;
    padding: 8px 12px;
    border-radius: var(--raio-md);
    border: 1px solid var(--cor-card);
    background: var(--cor-card);
    color: var(--cor-texto-primario);
    font-family: inherit;
    font-size: 12.5px;
    outline: none;
    transition: var(--transicao-rapida);
  }
  .ficha-multiplos-busca-input::placeholder {
    color: var(--cor-texto-terciario);
  }
  .ficha-multiplos-busca-input:focus {
    border-color: var(--cor-acento-teal);
  }
  .ficha-multiplos-lista-vazia {
    padding: 18px 4px;
    font-size: 12.5px;
    color: var(--cor-texto-terciario);
    text-align: center;
  }
  .dash-busca-btn-lupa {
    flex-shrink: 0;
    display: inline-flex;
    align-items: center;
    justify-content: center;
    width: 36px;
    height: 36px;
    border-radius: var(--raio-md);
    border: none;
    /* Em repouso o botão é discreto: fundo escuro do card e só o desenho
       da lupa em teal — do mesmo jeito que as pílulas de secretaria ficam
       apagadas até receberem o cursor. */
    background: var(--cor-card-elevado);
    color: var(--cor-acento-teal-texto);
    cursor: pointer;
    transition: background var(--transicao-rapida), color var(--transicao-rapida);
  }
  .dash-busca-btn-lupa:hover {
    /* Sob o cursor, inverte: fundo teal e lupa escura. */
    background: var(--cor-acento-teal-hover);
    color: #1A1A1A;
  }

  /* --- Botão do Controle de Qualidade (ao lado da lupa) ---
     Mesmo tamanho/estilo da lupa, mas em pêssego, com um badge vermelho
     no padrão Android (círculo sobreposto no canto) mostrando quantas
     ações têm pendência de qualidade no recorte atual. Quando não há
     nenhuma pendência depois dos filtros, o JS esconde o botão inteiro
     (display:none). */
  .dash-qc-btn {
    position: relative;
    flex-shrink: 0;
    display: inline-flex;
    align-items: center;
    justify-content: center;
    width: 36px;
    height: 36px;
    border-radius: var(--raio-md);
    border: none;
    background: var(--cor-card-elevado);
    color: var(--cor-acento-peach);
    cursor: pointer;
    transition: background var(--transicao-rapida), color var(--transicao-rapida);
  }
  .dash-qc-btn:hover {
    background: var(--cor-acento-peach);
    color: #1A1A1A;
  }
  .dash-qc-badge {
    position: absolute;
    top: -6px;
    right: -6px;
    min-width: 18px;
    height: 18px;
    padding: 0 5px;
    border-radius: 999px;
    background: #E2574C;
    color: #FFFFFF;
    font-size: 10.5px;
    font-weight: 700;
    line-height: 18px;
    text-align: center;
    box-shadow: 0 0 0 2px var(--cor-fundo), 0 1px 3px rgba(0, 0, 0, 0.35);
    pointer-events: none;
    font-variant-numeric: tabular-nums;
  }
  /* Na linha de filtros rápidos do dashboard o fundo é o do card, então o
     anel do badge acompanha para não ficar um halo mais escuro. */
  .dash-filtros-rapidos-linha .dash-qc-badge {
    box-shadow: 0 0 0 2px var(--cor-card), 0 1px 3px rgba(0, 0, 0, 0.35);
  }
  .dash-secretaria-pill {
    flex-shrink: 0;
    white-space: nowrap;
    background: var(--cor-card-elevado);
    color: var(--cor-texto-secundario);
    border: none;
    border-radius: 999px;
    padding: 5px 12px;
    font-size: 10.5px;
    font-weight: 600;
    font-family: var(--fonte);
    cursor: pointer;
    transition: background var(--transicao-rapida), color var(--transicao-rapida);
  }
  .dash-secretaria-pill:hover {
    background: var(--cor-acento-teal-hover);
    color: #1A1A1A;
  }
  .dash-secretaria-pill.ativo {
    background: var(--cor-acento-teal);
    color: #1A1A1A;
    font-weight: 700;
  }

  #preview-nota-secretaria {
    display: none;
    background: var(--cor-card-elevado);
    border-left: 3px solid var(--cor-acento-teal);
    border-radius: var(--raio-sm);
    padding: 10px 14px;
    margin-bottom: 16px;
    font-size: 12.5px;
    color: var(--cor-texto-primario);
  }
  #preview-nota-secretaria b { color: var(--cor-acento-mint-texto); }

  #preview-corpo {
    padding: 24px 32px 40px 32px;
    overflow-y: auto;
    max-width: 1400px;
    margin: 0 auto;
    width: 100%;
    box-sizing: border-box;
  }

  /* Impressão manual (Ctrl+P do navegador) — garante que as cores do tema
     escuro sejam preservadas e que todo o conteúdo apareça, sem cortar
     pelo scroll/altura do modal. */
  @media print {
    body * { visibility: hidden; }
    #preview-modal, #preview-modal * { visibility: visible; }
    #preview-overlay {
      position: static;
      background: none;
      display: block !important;
    }
    #preview-modal {
      position: static;
      width: 100%;
      max-height: none;
      border: none;
      box-shadow: none;
    }
    #preview-corpo { overflow: visible; max-height: none; }
    .modal-titulo-botoes { display: none; }
    * {
      -webkit-print-color-adjust: exact;
      print-color-adjust: exact;
    }
  }
  #preview-corpo::-webkit-scrollbar { width: 8px; }
  #preview-corpo::-webkit-scrollbar-thumb { background: var(--cor-card-elevado); border-radius: var(--raio-sm); }

  #preview-cards {
    display: flex;
    gap: 14px;
    margin-bottom: 18px;
  }
  .preview-card-resumo {
    flex: 1;
    background: var(--cor-card);
    border-radius: var(--raio-md);
    box-shadow: var(--sombra-card);
    padding: 14px 18px;
    text-align: center;
  }
  .preview-card-resumo .rotulo {
    font-size: 11px;
    font-weight: 700;
    color: var(--cor-texto-secundario);
    text-transform: uppercase;
    letter-spacing: 0.04em;
  }
  .preview-card-resumo .valor {
    font-size: 26px;
    font-weight: 700;
    color: var(--cor-acento-mint-texto);
    margin-top: 4px;
  }

  .preview-grid {
    display: grid;
    grid-template-columns: 1fr 1fr;
    gap: 16px;
  }
  .preview-grafico-card {
    background: var(--cor-card);
    border-radius: var(--raio-lg);
    box-shadow: var(--sombra-card);
    padding: 16px 18px;
  }
  .preview-grafico-card h4 {
    margin: 0 0 14px 0;
    font-size: 12px;
    font-weight: 700;
    color: var(--cor-texto-primario);
    text-transform: uppercase;
    letter-spacing: 0.03em;
  }

  /* Gráfico de pizza via conic-gradient — sem nenhuma lib externa */
  .grafico-pizza-wrap {
    display: flex;
    align-items: center;
    gap: 24px;
    width: 100%;
  }
  .grafico-pizza {
    width: 165px;
    height: 165px;
    min-width: 165px;
    border-radius: 50%;
    box-shadow: var(--sombra-card);
  }
  .grafico-legenda {
    display: flex;
    flex-direction: column;
    gap: 9px;
    font-size: 13px;
    flex: 1;
    min-width: 0;
  }
  .grafico-legenda-item {
    display: flex;
    align-items: flex-start;
    gap: 7px;
    color: var(--cor-texto-primario);
  }
  .grafico-legenda-item-com-dado {
    cursor: pointer;
  }
  .grafico-legenda-item-com-dado:hover .grafico-legenda-bolinha {
    filter: brightness(1.15);
    transform: scale(1.15);
  }
  .grafico-legenda-bolinha {
    width: 10px;
    height: 10px;
    min-width: 10px;
    border-radius: 50%;
    flex-shrink: 0;
    margin-top: 3px;
  }
  .grafico-legenda-corpo {
    display: flex;
    flex-direction: column;
    line-height: 1.3;
  }
  .grafico-legenda-rotulo {
    font-size: 11px;
    color: var(--cor-texto-secundario);
  }
  .grafico-legenda-qtd-mini {
    font-size: 12.5px;
    font-weight: 700;
    color: var(--cor-texto-primario);
  }
  .grafico-legenda-valor {
    font-size: 14px;
    font-weight: 700;
    color: var(--cor-acento-mint-texto);
  }
  .grafico-legenda-pct {
    font-size: 10px;
    font-weight: 400;
    color: var(--cor-texto-terciario);
  }

  #tooltip-objetos {
    display: none;
    position: fixed;
    z-index: 200;
    background: var(--cor-card-elevado);
    border: 1px solid var(--cor-acento-teal);
    border-radius: var(--raio-md);
    box-shadow: var(--sombra-card);
    padding: 10px 12px;
    max-height: 280px;
    overflow-y: auto;
  }
  #tooltip-objetos.visivel { display: block; }
  #tooltip-objetos::-webkit-scrollbar { width: 6px; }
  #tooltip-objetos::-webkit-scrollbar-thumb {
    background: var(--cor-card);
    border-radius: var(--raio-sm);
  }
  .tooltip-objetos-titulo {
    font-size: 10.5px;
    font-weight: 700;
    color: var(--cor-acento-teal-texto);
    text-transform: uppercase;
    margin-bottom: 6px;
    padding-bottom: 6px;
    border-bottom: 1px solid var(--cor-card);
  }
  .tooltip-objetos-titulo-com-valor {
    display: flex;
    justify-content: space-between;
    align-items: baseline;
    gap: 10px;
  }
  .tooltip-objetos-titulo-com-valor .tooltip-objetos-valor {
    font-size: 11px;
  }
  .tooltip-objetos-item {
    display: flex;
    justify-content: space-between;
    align-items: baseline;
    gap: 10px;
    padding: 3px 0;
    font-size: 11px;
  }
  .tooltip-objetos-nome {
    color: var(--cor-texto-primario);
    flex: 1;
  }
  .tooltip-objetos-qtd {
    color: var(--cor-texto-primario);
    font-size: 12px;
    font-weight: 700;
  }
  .tooltip-objetos-valor {
    color: var(--cor-acento-mint-texto);
    font-weight: 700;
    white-space: nowrap;
  }

  /* Gráfico de barras via flex — sem nenhuma lib externa */
  .grafico-barras {
    display: flex;
    align-items: flex-end;
    gap: 8px;
    height: 150px;
    padding-top: 10px;
  }
  .grafico-barra-coluna {
    flex: 1;
    display: flex;
    flex-direction: column;
    align-items: center;
    justify-content: flex-end;
    height: 100%;
    min-width: 0;
  }
  .grafico-barra-coluna-com-dado {
    cursor: pointer;
  }
  .grafico-barra-coluna-com-dado:hover .grafico-barra {
    filter: brightness(1.15);
  }
  .grafico-barra-valor {
    font-size: 11px;
    font-weight: 700;
    color: var(--cor-texto-primario);
    margin-bottom: 4px;
  }
  .grafico-barra {
    width: 60%;
    max-width: 48px;
    /* Canto arredondado no topo, o mesmo raio do PDF (RAIO_CANTO_COLUNA).
       Não é cápsula: com o topo em semicírculo a coluna parecia mais alta
       do que é e a comparação entre trimestres ficava pior. A base fica
       reta porque encosta na linha do eixo. */
    border-radius: 4px 4px 0 0;
    background: var(--cor-acento-teal);
    transition: height var(--transicao-padrao);
  }
  .grafico-barra-rotulo {
    font-size: 10px;
    color: var(--cor-texto-secundario);
    margin-top: 6px;
    text-align: center;
    word-break: break-word;
  }

  /* Quando há muitas colunas (ex: muitos trimestres), os rótulos giram na
     diagonal em vez de quebrar em várias linhas — mesma ideia usada no
     gráfico original do PDF (ReportLab), que também inclina os rótulos do
     eixo quando o espaço fica apertado. */
  .grafico-barras.muitas-colunas {
    padding-bottom: 34px;
  }
  .grafico-barras.muitas-colunas .grafico-barra-rotulo {
    margin-top: 10px;
    white-space: nowrap;
    text-align: right;
    transform: rotate(-40deg);
    transform-origin: top right;
    width: auto;
  }

  .preview-sem-dados {
    color: var(--cor-texto-terciario);
    font-size: 12px;
    text-align: center;
    padding: 30px 0;
  }

  .preview-secao-titulo {
    margin: 26px 0 14px 0;
    padding-top: 18px;
    border-top: 1px solid var(--cor-card-elevado);
    font-size: 13px;
    font-weight: 700;
    color: var(--cor-texto-primario);
    text-transform: uppercase;
    letter-spacing: 0.03em;
    display: flex;
    align-items: center;
    justify-content: space-between;
    gap: 8px;
  }
  /* Seta de abrir/fechar seção no celular — some no desktop (seção sempre
     aberta ali). Ver .preview-secao no @media (max-width: 768px). */
  .secao-seta {
    display: none;
  }
  .preview-nav-rapida {
    display: none;
  }
  .preview-texto-explicativo {
    font-size: 12px;
    color: var(--cor-texto-secundario);
    line-height: 1.5;
    margin: -6px 0 14px 0;
  }
  .preview-legenda-fases {
    display: flex;
    flex-wrap: wrap;
    gap: 14px;
    margin: -6px 0 12px 0;
    font-size: 11px;
    color: var(--cor-texto-secundario);
  }
  .preview-legenda-fases-item {
    display: flex;
    align-items: center;
    gap: 5px;
  }
  .preview-legenda-fases-swatch {
    width: 9px;
    height: 9px;
    border-radius: 2px;
    display: inline-block;
  }

  .preview-secretaria-grid {
    display: grid;
    grid-template-columns: repeat(4, 1fr);
    gap: 12px;
  }
  .preview-gestao-subtitulo {
    font-size: 11.5px;
    font-weight: 700;
    color: var(--cor-acento-teal-texto);
    text-transform: uppercase;
    letter-spacing: 0.03em;
    margin: 16px 0 10px 0;
  }
  .preview-gestao-subtitulo:first-child {
    margin-top: 0;
  }
  .preview-mini-card {
    background: var(--cor-card);
    border-radius: var(--raio-md);
    box-shadow: var(--sombra-card);
    padding: 10px 12px;
    display: flex;
    flex-direction: column;
    position: relative;
  }
  .preview-mini-card-gauge {
    align-items: center;
  }
  .preview-mini-card-esmaecido {
    opacity: 0.55;
  }
  .preview-mini-card-esmaecido .titulo-sec-exec {
    color: var(--cor-texto-secundario) !important;
  }
  .gauge-desempenho-corpo {
    display: flex;
    justify-content: center;
  }
  .gauge-desempenho-corpo svg {
    max-width: 100%;
    height: auto;
  }
  .preview-mini-card .titulo-sec-exec {
    font-size: 12.5px;
    font-weight: 700;
    color: var(--cor-acento-mint-texto);
    margin-bottom: 8px;
    line-height: 1.3;
    padding-right: 20px;
  }
  .mini-card-icone-detalhe {
    position: absolute;
    top: 8px;
    right: 8px;
    width: 20px;
    height: 20px;
    display: flex;
    align-items: center;
    justify-content: center;
    padding: 0;
    background: transparent;
    color: var(--cor-texto-terciario);
    border: none;
    border-radius: 50%;
    cursor: help;
    transition: background var(--transicao-rapida), color var(--transicao-rapida);
  }
  .mini-card-icone-detalhe:hover {
    background: var(--cor-acento-teal);
    color: #1A1A1A;
  }
  .preview-mini-card .mini-card-corpo {
    flex: 1;
    display: flex;
    align-items: center;
  }
  .preview-mini-card .grafico-pizza {
    width: 68px;
    height: 68px;
    min-width: 68px;
  }
  .preview-mini-card .grafico-legenda {
    font-size: 11.5px;
    gap: 5px;
  }
  .preview-mini-card .grafico-legenda-bolinha {
    width: 7px;
    height: 7px;
    min-width: 7px;
  }
  .preview-mini-card .grafico-legenda-rotulo {
    font-size: 11px;
  }
  .preview-mini-card .grafico-legenda-qtd-mini {
    font-size: 12.5px;
  }
  .preview-mini-card .grafico-legenda-valor {
    font-size: 13.5px;
  }
  .preview-mini-card .grafico-legenda-pct {
    font-size: 11px;
  }

  .grafico-secretaria-legenda {
    display: flex;
    flex-wrap: wrap;
    gap: 14px;
    margin-bottom: 10px;
    font-size: 10.5px;
    color: var(--cor-texto-secundario);
  }
  .grafico-secretaria-legenda-item {
    display: flex;
    align-items: center;
    gap: 5px;
  }
  .grafico-secretaria-legenda-swatch {
    width: 9px;
    height: 9px;
    border-radius: 2px;
    display: inline-block;
  }
  .grafico-secretaria-wrap {
    background: var(--cor-card);
    border-radius: var(--raio-md);
    box-shadow: var(--sombra-card);
    padding: 14px 16px;
    display: flex;
    flex-direction: column;
    gap: 12px;
  }
  .grafico-secretaria-linha {
    display: grid;
    grid-template-columns: 200px 1fr 90px;
    align-items: center;
    gap: 12px;
    cursor: help;
    border-radius: var(--raio-sm);
    padding: 4px 6px;
    transition: background var(--transicao-rapida);
  }
  .grafico-secretaria-linha:hover {
    background: var(--cor-card-elevado);
  }
  .grafico-secretaria-rotulo {
    font-size: 12px;
    font-weight: 700;
    color: var(--cor-texto-primario);
    text-decoration: underline dotted var(--cor-texto-terciario);
    text-underline-offset: 3px;
    white-space: nowrap;
    overflow: hidden;
    text-overflow: ellipsis;
  }
  .grafico-secretaria-qtd {
    font-weight: 400;
    color: var(--cor-texto-terciario);
    font-size: 11px;
    text-decoration: none;
  }
  .grafico-secretaria-barra-trilha {
    background: var(--cor-card-elevado);
    /* Trilho em formato de cápsula, igual ao trilho do medidor. O
       overflow: hidden faz os segmentos coloridos herdarem esse recorte
       nas duas pontas. */
    border-radius: 999px;
    height: 14px;
    overflow: hidden;
  }
  .grafico-secretaria-barra-empilhada {
    width: 100%;
    height: 100%;
    display: flex;
  }
  .grafico-secretaria-segmento {
    height: 100%;
    box-sizing: border-box;
    border-right-width: 2px;
    border-right-style: solid;
    border-right-color: var(--cor-fundo);
  }
  .grafico-secretaria-segmento:last-child { border-right: none; }
  .grafico-secretaria-segmento:first-child { border-radius: 999px 0 0 999px; }
  .grafico-secretaria-segmento:last-child { border-radius: 0 999px 999px 0; }
  .grafico-secretaria-valor {
    font-size: 13px;
    font-weight: 700;
    color: var(--cor-acento-mint-texto);
    text-align: right;
    white-space: nowrap;
  }

  .preview-mapa-wrap {
    display: flex;
    gap: 20px;
    align-items: flex-start;
    background: var(--cor-card);
    border-radius: var(--raio-lg);
    box-shadow: var(--sombra-card);
    padding: 18px;
    position: relative;
  }
  .preview-mapa-wrap svg {
    flex-shrink: 0;
    max-width: 100%;
    height: auto;
  }
  .preview-mapa-svg {
    cursor: grab;
    touch-action: none;
  }
  .preview-mapa-svg.preview-mapa-arrastando {
    cursor: grabbing;
  }
  .preview-mapa-wrap polygon {
    stroke: #fff;
    stroke-width: 0.4;
  }
  .preview-mapa-poligono-com-dado {
    cursor: pointer;
    transition: var(--transicao-rapida);
  }
  .preview-mapa-poligono-com-dado:hover {
    stroke: var(--cor-acento-teal);
    stroke-width: 1.4;
  }
  .preview-mapa-controles {
    position: absolute;
    top: 28px;
    left: 28px;
    display: flex;
    flex-direction: column;
    gap: 4px;
    z-index: 5;
  }
  .preview-mapa-botao-zoom {
    width: 28px;
    height: 28px;
    border-radius: var(--raio-sm);
    border: none;
    background: var(--cor-card-elevado);
    color: var(--cor-texto-primario);
    font-size: 16px;
    font-weight: 700;
    line-height: 1;
    cursor: pointer;
    transition: var(--transicao-rapida);
  }
  .preview-mapa-botao-zoom:hover {
    background: var(--cor-acento-teal);
    color: #16211F;
  }
  .preview-mapa-botao-zoom-reset {
    font-size: 14px;
  }
  .preview-mapa-legenda {
    display: flex;
    flex-direction: column;
    gap: 8px;
    font-size: 12px;
    min-width: 180px;
  }
  .preview-mapa-legenda-item {
    display: flex;
    align-items: center;
    gap: 8px;
    color: var(--cor-texto-primario);
  }
  .preview-mapa-legenda-swatch {
    width: 18px;
    height: 18px;
    border-radius: var(--raio-sm);
    flex-shrink: 0;
  }

  /* =====================================================
     CUSTOMIZAÇÃO PARA CELULAR (retrato) — só entra em vigor com a
     <meta name="viewport"> lá no <head> (sem ela o navegador mobile nunca
     reporta uma largura estreita o bastante pra disparar este bloco).
     Três ajustes, cada um resolvendo um aperto real da tela pequena:
       1) Painel de filtros vira um acordeão de 1 coluna (em vez da grade
          fixa de 5 colunas, ilegível num celular) — ver
          configurarAccordionFiltros() no JS.
       2) Grade "Panorama por Secretaria" passa de 4 cards por linha (bem
          espremidos) para 1 por linha.
       3) Gráfico "Previsão de Conclusão da Fase" ganha rolagem horizontal
          com largura mínima por coluna, em vez de espremer ~28 trimestres
          na largura da tela.
     O card "Resumo Financeiro"/"Panorama Geral"/"Situação do Termo" (grade
     2x2 de gráficos) também passa a 1 coluna aqui: é consequência direta
     da <meta viewport> acima — sem isso, aquele 2x2 quebraria de verdade
     (pizza de 165px não cabe em meia tela de celular), em vez de só
     encolher proporcionalmente como fazia antes (sem viewport, tudo era
     desenhado numa largura de desktop e depois minimizado por igual). */
  @media (max-width: 768px) {
    #grade {
      display: flex;
      flex-direction: column;
      overflow-y: auto;
    }
    .bloco {
      flex: 0 0 auto;
    }
    .bloco-titulo {
      cursor: pointer;
      justify-content: flex-start;
    }
    .bloco-seta {
      display: inline-block;
      margin-left: auto;
      flex-shrink: 0;
      font-size: 11px;
      color: var(--cor-texto-terciario);
      transition: transform var(--transicao-rapida);
    }
    .bloco-colapsado .bloco-seta {
      transform: rotate(-90deg);
    }
    .bloco-colapsado .bloco-botoes,
    .bloco-colapsado .bloco-busca,
    .bloco-colapsado .bloco-lista {
      display: none;
    }
    .bloco-lista {
      max-height: 46vh;
      overflow-y: auto;
    }

    .preview-grid {
      grid-template-columns: 1fr;
    }

    /* 2 por linha (Panorama por Secretaria E Índice de Desempenho, que
       reaproveitam esta mesma classe de grade) — mais compacto que 1 por
       linha, aceitando ficar um pouco mais apertado. */
    .preview-secretaria-grid {
      grid-template-columns: repeat(2, 1fr);
    }

    /* --- Seções recolhíveis (Panorama por Secretaria, Índice de
       Desempenho, Detalhamento Financeiro, Mapa) — mesmo padrão de
       acordeão dos filtros, ver configurarAccordionSecoes() no JS. Todas
       começam fechadas: é o que reduz a rolagem gigante que só empilhar
       tudo em 1 coluna criava. --- */
    .preview-secao-titulo {
      cursor: pointer;
    }
    .secao-seta {
      display: inline-block;
      font-size: 12px;
      color: var(--cor-texto-terciario);
      transition: transform var(--transicao-rapida);
    }
    .preview-secao.secao-colapsada .secao-seta {
      transform: rotate(-90deg);
    }
    .preview-secao.secao-colapsada .preview-secao-corpo {
      display: none;
    }

    /* --- Nav rápida entre seções, fixa no topo da área que rola --- */
    .preview-nav-rapida {
      display: flex;
      gap: 8px;
      overflow-x: auto;
      padding: 2px 2px 12px 2px;
      margin-bottom: 4px;
      position: sticky;
      top: 0;
      background: var(--cor-fundo);
      z-index: 3;
    }
    .preview-nav-rapida button {
      flex: 0 0 auto;
      background: var(--cor-card-elevado);
      color: var(--cor-texto-primario);
      border: none;
      border-radius: var(--raio-md);
      padding: 6px 14px;
      font-size: 11.5px;
      font-weight: 600;
      white-space: nowrap;
      cursor: pointer;
    }
    .preview-nav-rapida button:active {
      background: var(--cor-acento-teal);
      color: #16211F;
    }

    #preview-grafico-prazo {
      overflow-x: auto;
      -webkit-overflow-scrolling: touch;
    }
    #preview-grafico-prazo .grafico-barras {
      width: max-content;
      min-width: 100%;
    }
    #preview-grafico-prazo .grafico-barra-coluna {
      flex: 0 0 34px;
      width: 34px;
    }
  }
</style>
</head>
<body>
<script>
  // Diagnóstico: se qualquer erro de JS acontecer (parse ou execução),
  // mostra a mensagem numa faixa vermelha no topo — e também busca o
  // próprio arquivo (agora que é carregado via url=, dá pra buscar de
  // volta) para mostrar o TRECHO EXATO da linha com problema, em vez de só
  // o número da linha.
  window.addEventListener("error", function (e) {
    var div = document.createElement("div");
    div.style.cssText = "position:fixed;top:0;left:0;right:0;background:#7A1F1F;color:#fff;padding:10px 16px;font-family:monospace;font-size:12px;z-index:9999;white-space:pre-wrap;max-height:60vh;overflow:auto;";
    var origem = e.filename ? (" (linha " + e.lineno + ", coluna " + e.colno + ")") : "";
    div.textContent = "Erro no painel: " + (e.message || e) + origem;
    document.body.appendChild(div);

    if (e.lineno) {
      fetch(location.href)
        .then(function (resp) { return resp.text(); })
        .then(function (texto) {
          var linhas = texto.split("\n");
          var idx = e.lineno - 1;
          var ini = Math.max(0, idx - 3);
          var fim = Math.min(linhas.length, idx + 4);
          var pre = document.createElement("pre");
          pre.style.cssText = "white-space:pre-wrap;color:#fff;background:#3A1414;padding:8px 12px;margin:6px 0 0 0;font-size:11px;";
          var texto_contexto = "";
          for (var i = ini; i < fim; i++) {
            texto_contexto += (i === idx ? ">>> " : "    ") + (i + 1) + ": " + linhas[i] + "\n";
          }
          pre.textContent = texto_contexto;
          div.appendChild(pre);
        })
        .catch(function () {});
    }
  });
</script>
<div id="app">
  <div id="topo">
    <div class="topo-titulo-bloco">
      <div class="topo-titulo-textos">
        <span class="topo-titulo-principal">BALANÇO PAC - BAHIA</span>
        <span class="topo-titulo-atualizacao">__ULTIMA_ATUALIZACAO__</span>
      </div>
      <p class="topo-orientacao">Selecione os filtros desejados. Caso nenhum item de um bloco seja selecionado, todos os itens desse bloco serão considerados automaticamente.</p>
    </div>
    <div id="botoes-topo">
      <button id="filtros-busca-btn-lupa" class="dash-busca-btn-lupa" title="Buscar ação e abrir a Ficha Cadastral">
        <svg viewBox="0 0 24 24" width="17" height="17" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round">
          <circle cx="11" cy="11" r="7"></circle>
          <line x1="21" y1="21" x2="16.65" y2="16.65"></line>
        </svg>
      </button>
      <button id="filtros-qc-btn" class="dash-qc-btn" style="display:none;" title="Controle de Qualidade da Base de Dados">
        <svg viewBox="0 0 24 24" width="17" height="17" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round">
          <ellipse cx="12" cy="5" rx="8" ry="3"></ellipse>
          <path d="M4 5v6c0 1.66 3.58 3 8 3s8-1.34 8-3V5"></path>
          <path d="M4 11v6c0 1.66 3.58 3 8 3 1.2 0 2.34-.1 3.36-.28"></path>
          <path d="M15 18.5l2 2 4-4.5"></path>
        </svg>
        <span class="dash-qc-badge" id="filtros-qc-badge">0</span>
      </button>
      <button class="btn" id="btn-limpar-tudo">LIMPAR FILTROS</button>
      <button class="btn" id="btn-gerencial-filtros">GERENCIAL</button>
      <button class="btn" id="btn-preview">DASHBOARD</button>
      <button class="btn-acento" id="btn-gerar">GERAR RELATÓRIO</button>
      <button type="button" class="tema-switch" id="tema-switch-filtros" role="switch" aria-checked="false" title="Alternar tema claro/escuro" aria-label="Alternar tema claro/escuro">
        <span class="tema-switch-trilho">
          <span class="tema-switch-icone tema-switch-sol">
            <svg viewBox="0 0 24 24" width="13" height="13" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round">
              <circle cx="12" cy="12" r="4"></circle>
              <line x1="12" y1="2" x2="12" y2="4"></line>
              <line x1="12" y1="20" x2="12" y2="22"></line>
              <line x1="4.2" y1="4.2" x2="5.6" y2="5.6"></line>
              <line x1="18.4" y1="18.4" x2="19.8" y2="19.8"></line>
              <line x1="2" y1="12" x2="4" y2="12"></line>
              <line x1="20" y1="12" x2="22" y2="12"></line>
              <line x1="4.2" y1="19.8" x2="5.6" y2="18.4"></line>
              <line x1="18.4" y1="5.6" x2="19.8" y2="4.2"></line>
            </svg>
          </span>
          <span class="tema-switch-icone tema-switch-lua">
            <svg viewBox="0 0 24 24" width="12" height="12" fill="currentColor">
              <path d="M20.5 14.5A8.5 8.5 0 0 1 9.5 3.5a8.5 8.5 0 1 0 11 11z"></path>
            </svg>
          </span>
          <span class="tema-switch-bolinha"></span>
        </span>
      </button>
      <div class="acesso-rapido-wrap" id="acesso-rapido-filtros-wrap">
        <button class="botao-icone-topo" id="acesso-rapido-filtros-btn" title="Acesso rápido" aria-haspopup="true" aria-expanded="false">
          <svg viewBox="0 0 24 24" width="18" height="18" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round">
            <rect x="3" y="3" width="7" height="7" rx="1.5"></rect>
            <rect x="14" y="3" width="7" height="7" rx="1.5"></rect>
            <rect x="3" y="14" width="7" height="7" rx="1.5"></rect>
            <rect x="14" y="14" width="7" height="7" rx="1.5"></rect>
          </svg>
        </button>
        <div class="acesso-rapido-menu" id="acesso-rapido-filtros-menu">
          <button class="acesso-rapido-item" id="acesso-rapido-filtros-mapa-mental" title="Mapa Mental">
            <svg viewBox="0 0 24 24" width="16" height="16" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round">
              <circle cx="18" cy="5" r="3"></circle>
              <circle cx="6" cy="12" r="3"></circle>
              <circle cx="18" cy="19" r="3"></circle>
              <line x1="8.6" y1="10.5" x2="15.4" y2="6.5"></line>
              <line x1="8.6" y1="13.5" x2="15.4" y2="17.5"></line>
            </svg>
            <span>MAPA MENTAL</span>
          </button>
          <button class="acesso-rapido-item" id="acesso-rapido-filtros-filtros" title="Filtros">
            <svg viewBox="0 0 24 24" width="16" height="16" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round">
              <line x1="4" y1="21" x2="4" y2="14"></line>
              <line x1="4" y1="10" x2="4" y2="3"></line>
              <line x1="12" y1="21" x2="12" y2="12"></line>
              <line x1="12" y1="8" x2="12" y2="3"></line>
              <line x1="20" y1="21" x2="20" y2="16"></line>
              <line x1="20" y1="12" x2="20" y2="3"></line>
              <line x1="1" y1="14" x2="7" y2="14"></line>
              <line x1="9" y1="8" x2="15" y2="8"></line>
              <line x1="17" y1="16" x2="23" y2="16"></line>
            </svg>
            <span>FILTROS</span>
          </button>
          <button class="acesso-rapido-item" id="acesso-rapido-filtros-publicar" style="display:none" title="Publicar planilha/código atualizados no site (GitHub + Render)">
            <svg viewBox="0 0 24 24" width="16" height="16" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round">
              <path d="M12 16V4"></path>
              <path d="M6 10l6-6 6 6"></path>
              <path d="M4 20h16"></path>
            </svg>
            <span>PUBLICAR</span>
          </button>
        </div>
      </div>
    </div>
  </div>
  <div id="grade"></div>
</div>

<div id="aviso-inatividade"><div class="caixa">Painel fechado por inatividade.</div></div>

<div id="modal-overlay">
  <div id="modal-qualidade">
    <div class="modal-titulo">
      <span>PAC - Controle de Qualidade da Base de Dados</span>
      <button id="modal-btn-whatsapp" class="ficha-botao-icone" title="Compartilhar no WhatsApp">
        <svg viewBox="0 0 24 24" width="18" height="18" fill="currentColor">
          <path d="M12.04 2C6.58 2 2.13 6.45 2.13 11.91c0 1.75.46 3.48 1.32 5l-1.4 5.12 5.24-1.37c1.46.8 3.11 1.22 4.75 1.22h.01c5.46 0 9.91-4.45 9.91-9.91 0-2.65-1.03-5.14-2.9-7.01A9.816 9.816 0 0 0 12.04 2zm5.8 14.16c-.24.68-1.4 1.3-1.93 1.36-.5.06-1.05.24-3.5-.73-2.93-1.16-4.83-4.14-4.98-4.33-.14-.19-1.19-1.58-1.19-3.02s.75-2.14 1.02-2.43c.26-.29.58-.36.77-.36.19 0 .39 0 .55.01.18.01.42-.07.66.5.24.58.83 2.01.9 2.16.07.15.12.32.02.51-.1.19-.15.31-.29.48-.15.17-.31.38-.44.51-.15.14-.3.3-.13.58.17.29.77 1.27 1.65 2.06 1.14 1.01 2.09 1.33 2.38 1.48.29.15.46.12.63-.07.17-.19.72-.84.92-1.13.19-.29.38-.24.63-.14.26.1 1.66.78 1.94.92.29.14.48.22.55.34.07.13.07.72-.17 1.4z"></path>
        </svg>
      </button>
    </div>
    <div class="modal-corpo" id="modal-qualidade-texto"></div>
    <div class="modal-rodape">
      <span class="modal-contagem" id="modal-contagem"></span>
      <button class="btn" id="modal-btn-cancelar">Cancelar</button>
      <button class="btn-acento" id="modal-btn-ok" disabled>OK</button>
    </div>
  </div>
</div>

<div id="ficha-overlay">
  <div id="ficha-painel">
    <div class="ficha-topo">
      <div class="ficha-topo-titulo">
        <span id="ficha-topo-objeto">—</span>
        <span id="ficha-topo-item"></span>
      </div>
      <div class="ficha-topo-acoes" id="ficha-acoes-tela">
        <button id="ficha-btn-preview-impressao" class="ficha-botao-icone" title="Visualizar impressão">
          <svg viewBox="0 0 24 24" width="18" height="18" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round">
            <polyline points="6 9 6 2 18 2 18 9"></polyline>
            <path d="M6 18H4a2 2 0 0 1-2-2v-5a2 2 0 0 1 2-2h16a2 2 0 0 1 2 2v5a2 2 0 0 1-2 2h-2"></path>
            <rect x="6" y="14" width="12" height="8"></rect>
          </svg>
        </button>
        <button id="ficha-btn-fechar" class="ficha-botao-icone" title="Fechar">&times;</button>
      </div>
      <div class="ficha-topo-acoes" id="ficha-acoes-impressao" style="display:none;">
        <button id="ficha-btn-voltar-tela" class="btn">&larr; Voltar</button>
        <button id="ficha-btn-salvar-pdf" class="ficha-botao-icone" title="Salvar em PDF">
          <svg viewBox="0 0 24 24" width="18" height="18" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round">
            <path d="M14 2H6a2 2 0 0 0-2 2v16a2 2 0 0 0 2 2h12a2 2 0 0 0 2-2V8z"></path>
            <polyline points="14 2 14 8 20 8"></polyline>
            <line x1="12" y1="18" x2="12" y2="12"></line>
            <polyline points="9 15 12 18 15 15"></polyline>
          </svg>
        </button>
        <button id="ficha-btn-fechar-impressao" class="ficha-botao-icone" title="Fechar">&times;</button>
      </div>
    </div>
    <div id="ficha-corpo" class="ficha-corpo"></div>
    <div id="ficha-pagina-a4-wrap" class="ficha-pagina-a4-wrap">
      <div id="ficha-pagina-a4" class="ficha-pagina-a4"></div>
    </div>
  </div>
</div>

<div id="ficha-multiplos-overlay">
  <div id="ficha-multiplos-painel">
    <div class="modal-titulo" id="ficha-multiplos-titulo">Mais de uma ação encontrada — escolha uma</div>
    <div class="ficha-multiplos-busca">
      <input type="text" id="busca-ficha-descricao" class="ficha-multiplos-busca-input" placeholder="Filtrar por Descrição..." autocomplete="off">
      <input type="text" id="busca-ficha-municipio" class="ficha-multiplos-busca-input" placeholder="Filtrar por Município..." autocomplete="off">
    </div>
    <div id="ficha-multiplos-lista" class="ficha-multiplos-lista"></div>
    <div class="modal-rodape">
      <button class="btn" id="ficha-multiplos-cancelar">Cancelar</button>
    </div>
  </div>
</div>

<div id="paginas-overlay">
  <div id="paginas-painel">
    <div class="modal-titulo">
      <span>Selecione as páginas do relatório</span>
    </div>
    <div class="paginas-acoes">
      <button class="btn-mini" id="paginas-marcar-tudo">Marcar tudo</button>
      <button class="btn-mini" id="paginas-limpar">Limpar</button>
      <span id="paginas-contador" class="paginas-contador"></span>
    </div>
    <div id="paginas-grade" class="paginas-grade"></div>
    <div id="paginas-colunas">
      <div class="paginas-colunas-topo">
        <span class="paginas-colunas-titulo">Colunas do detalhamento</span>
        <span id="paginas-colunas-contador" class="paginas-colunas-contador"></span>
      </div>
      <div class="paginas-colunas-botoes">
        <button class="btn-mini" id="paginas-colunas-padrao">Padrão</button>
        <button class="btn-mini" id="paginas-colunas-essencial">Só o essencial</button>
      </div>
      <div id="paginas-colunas-lista" class="paginas-colunas-lista"></div>
    </div>
    <div class="modal-rodape">
      <button class="btn" id="paginas-cancelar">Cancelar</button>
      <button class="btn-acento" id="paginas-confirmar">GERAR PDF</button>
    </div>
  </div>
</div>

<div id="mapa-mental-overlay">
  <iframe id="mapa-mental-iframe" title="Mapa Mental do BALANÇO PAC"></iframe>
</div>

<div id="preview-overlay">
  <div id="preview-modal">
    <div class="modal-titulo">
      <span class="modal-titulo-textos">
        <span class="modal-titulo-principal">BALANÇO PAC - BAHIA</span>
        <span class="modal-titulo-atualizacao">__ULTIMA_ATUALIZACAO__</span>
      </span>
      <div class="modal-titulo-botoes">
        <button class="btn" id="dash-limpar">LIMPAR FILTROS</button>
        <button class="btn" id="dash-gerencial">GERENCIAL</button>
        <button class="btn-acento" id="preview-gerar-topo">GERAR RELATÓRIO</button>
        <button type="button" class="tema-switch" id="tema-switch-dash" role="switch" aria-checked="false" title="Alternar tema claro/escuro" aria-label="Alternar tema claro/escuro">
          <span class="tema-switch-trilho">
            <span class="tema-switch-icone tema-switch-sol">
              <svg viewBox="0 0 24 24" width="13" height="13" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round">
                <circle cx="12" cy="12" r="4"></circle>
                <line x1="12" y1="2" x2="12" y2="4"></line>
                <line x1="12" y1="20" x2="12" y2="22"></line>
                <line x1="4.2" y1="4.2" x2="5.6" y2="5.6"></line>
                <line x1="18.4" y1="18.4" x2="19.8" y2="19.8"></line>
                <line x1="2" y1="12" x2="4" y2="12"></line>
                <line x1="20" y1="12" x2="22" y2="12"></line>
                <line x1="4.2" y1="19.8" x2="5.6" y2="18.4"></line>
                <line x1="18.4" y1="5.6" x2="19.8" y2="4.2"></line>
              </svg>
            </span>
            <span class="tema-switch-icone tema-switch-lua">
              <svg viewBox="0 0 24 24" width="12" height="12" fill="currentColor">
                <path d="M20.5 14.5A8.5 8.5 0 0 1 9.5 3.5a8.5 8.5 0 1 0 11 11z"></path>
              </svg>
            </span>
            <span class="tema-switch-bolinha"></span>
          </span>
        </button>
        <div class="acesso-rapido-wrap" id="acesso-rapido-dash-wrap">
          <button class="botao-icone-topo" id="acesso-rapido-dash-btn" title="Acesso rápido" aria-haspopup="true" aria-expanded="false">
            <svg viewBox="0 0 24 24" width="18" height="18" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round">
              <rect x="3" y="3" width="7" height="7" rx="1.5"></rect>
              <rect x="14" y="3" width="7" height="7" rx="1.5"></rect>
              <rect x="3" y="14" width="7" height="7" rx="1.5"></rect>
              <rect x="14" y="14" width="7" height="7" rx="1.5"></rect>
            </svg>
          </button>
          <div class="acesso-rapido-menu" id="acesso-rapido-dash-menu">
            <button class="acesso-rapido-item" id="acesso-rapido-dash-mapa-mental" title="Mapa Mental">
              <svg viewBox="0 0 24 24" width="16" height="16" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round">
                <circle cx="18" cy="5" r="3"></circle>
                <circle cx="6" cy="12" r="3"></circle>
                <circle cx="18" cy="19" r="3"></circle>
                <line x1="8.6" y1="10.5" x2="15.4" y2="6.5"></line>
                <line x1="8.6" y1="13.5" x2="15.4" y2="17.5"></line>
              </svg>
              <span>MAPA MENTAL</span>
            </button>
            <button class="acesso-rapido-item" id="acesso-rapido-dash-filtros" title="Filtros">
              <svg viewBox="0 0 24 24" width="16" height="16" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round">
                <line x1="4" y1="21" x2="4" y2="14"></line>
                <line x1="4" y1="10" x2="4" y2="3"></line>
                <line x1="12" y1="21" x2="12" y2="12"></line>
                <line x1="12" y1="8" x2="12" y2="3"></line>
                <line x1="20" y1="21" x2="20" y2="16"></line>
                <line x1="20" y1="12" x2="20" y2="3"></line>
                <line x1="1" y1="14" x2="7" y2="14"></line>
                <line x1="9" y1="8" x2="15" y2="8"></line>
                <line x1="17" y1="16" x2="23" y2="16"></line>
              </svg>
              <span>FILTROS</span>
            </button>
            <!-- Só existe no modo desktop (roda git de verdade na máquina de
                 quem clica) — nasce escondido e o JS decide se mostra, ver
                 window.PAC_MODO_WEB logo no início do <script> principal. -->
            <button class="acesso-rapido-item" id="acesso-rapido-dash-publicar" style="display:none" title="Publicar planilha/código atualizados no site (GitHub + Render)">
              <svg viewBox="0 0 24 24" width="16" height="16" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round">
                <path d="M12 16V4"></path>
                <path d="M6 10l6-6 6 6"></path>
                <path d="M4 20h16"></path>
              </svg>
              <span>PUBLICAR</span>
            </button>
          </div>
        </div>
      </div>
    </div>
    <div class="dash-filtros-rapidos-linha">
      <div id="dash-filtros-rapidos" class="dash-filtros-rapidos">
        <div id="dash-gestao-filtro" class="dash-secretaria-filtro"></div>
        <div id="dash-separador-secretaria" class="dash-separador-executor visivel"></div>
        <div id="dash-secretaria-filtro" class="dash-secretaria-filtro"></div>
        <div id="dash-separador-executor" class="dash-separador-executor"></div>
        <div id="dash-executor-filtro" class="dash-secretaria-filtro"></div>
        <div id="dash-separador-objeto" class="dash-separador-executor"></div>
        <div id="dash-objeto-filtro" class="dash-secretaria-filtro"></div>
      </div>
      <button id="dash-busca-btn-lupa" class="dash-busca-btn-lupa" title="Buscar ação e abrir a Ficha Cadastral">
        <svg viewBox="0 0 24 24" width="17" height="17" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round">
          <circle cx="11" cy="11" r="7"></circle>
          <line x1="21" y1="21" x2="16.65" y2="16.65"></line>
        </svg>
      </button>
      <button id="dash-qc-btn" class="dash-qc-btn" style="display:none;" title="Controle de Qualidade da Base de Dados">
        <svg viewBox="0 0 24 24" width="17" height="17" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round">
          <ellipse cx="12" cy="5" rx="8" ry="3"></ellipse>
          <path d="M4 5v6c0 1.66 3.58 3 8 3s8-1.34 8-3V5"></path>
          <path d="M4 11v6c0 1.66 3.58 3 8 3 1.2 0 2.34-.1 3.36-.28"></path>
          <path d="M15 18.5l2 2 4-4.5"></path>
        </svg>
        <span class="dash-qc-badge" id="dash-qc-badge">0</span>
      </button>
    </div>
    <div id="preview-corpo">
      <!-- Só aparece no celular (ver @media max-width:768px) — pula direto
           pras seções de baixo sem precisar rolar a tela toda. Também
           reabre a seção-alvo se ela estiver fechada no acordeão (ver
           configurarAccordionSecoes() no JS). -->
      <div class="preview-nav-rapida" id="preview-nav-rapida">
        <button type="button" data-alvo="secao-secretaria">Secretaria</button>
        <button type="button" data-alvo="secao-desempenho">Desempenho</button>
        <button type="button" data-alvo="secao-financeiro">Financeiro</button>
        <button type="button" data-alvo="secao-mapa">Mapa</button>
      </div>
      <div id="preview-nota-secretaria"></div>
      <div id="preview-cards"></div>
      <div class="preview-grid">
        <div class="preview-grafico-card">
          <h4>Previsão de Conclusão da Fase</h4>
          <div id="preview-grafico-prazo"></div>
        </div>
        <div class="preview-grafico-card">
          <h4>Resumo Financeiro</h4>
          <div id="preview-grafico-financeiro"></div>
        </div>
        <div class="preview-grafico-card">
          <h4>Panorama Geral das Fases</h4>
          <div id="preview-grafico-panorama"></div>
        </div>
        <div class="preview-grafico-card">
          <h4>Situação do Termo de Compromisso</h4>
          <div id="preview-grafico-termo"></div>
        </div>
      </div>

      <section class="preview-secao" id="secao-secretaria">
        <h3 class="preview-secao-titulo">Panorama por Secretaria | Executor<span class="secao-seta" aria-hidden="true">▾</span></h3>
        <div class="preview-secao-corpo">
          <div class="preview-legenda-fases">
            <span class="preview-legenda-fases-item"><span class="preview-legenda-fases-swatch" data-paleta="vermelho"></span>Captação de Recurso</span>
            <span class="preview-legenda-fases-item"><span class="preview-legenda-fases-swatch" data-paleta="amarelo"></span>Licitação</span>
            <span class="preview-legenda-fases-item"><span class="preview-legenda-fases-swatch" data-paleta="verde"></span>Execução do Objeto</span>
            <span class="preview-legenda-fases-item"><span class="preview-legenda-fases-swatch" data-paleta="azul"></span>Concluída</span>
          </div>
          <div id="preview-secretaria-container"></div>
        </div>
      </section>

      <section class="preview-secao" id="secao-desempenho">
        <h3 class="preview-secao-titulo">Índice de Desempenho por Secretaria<span class="secao-seta" aria-hidden="true">▾</span></h3>
        <div class="preview-secao-corpo">
          <div class="preview-texto-explicativo">
            O índice combina Status, Fase, Cláusula Suspensiva, o tempo entre aviso de licitação, O.S. e
            conclusão prevista, a quantidade de ações administradas e a proporção de Valor Contratado sobre o
            investimento total da própria gestão — ponderado pelo investimento de cada ação. Avaliação por
            SECRETARIA | EXECUTOR, separada por gestão, ordenada da melhor pra pior.
          </div>
          <div id="preview-desempenho-container"></div>
        </div>
      </section>

      <section class="preview-secao" id="secao-financeiro">
        <h3 class="preview-secao-titulo">Detalhamento Financeiro por Secretaria<span class="secao-seta" aria-hidden="true">▾</span></h3>
        <div class="preview-secao-corpo">
          <div id="preview-detalhamento-secretaria"></div>
        </div>
      </section>

      <section class="preview-secao" id="secao-mapa">
        <h3 class="preview-secao-titulo">Mapa de Investimentos por Município<span class="secao-seta" aria-hidden="true">▾</span></h3>
        <div class="preview-secao-corpo">
          <div id="preview-mapa"></div>
        </div>
      </section>
    </div>
  </div>
</div>

<script type="application/json" id="dados-painel">__DADOS_PAINEL__</script>
<script>
  // Os dados vêm de um bloco <script type="application/json"> (texto puro,
  // nunca interpretado como código) em vez de um literal JS embutido
  // diretamente ou de um fetch para arquivo externo — assim não há risco
  // de caracteres especiais nos dados quebrarem a sintaxe do JavaScript,
  // nem de CORS ao carregar via file://. JSON.parse lida com tudo isso de
  // forma segura, de forma síncrona.
  const DADOS = JSON.parse(document.getElementById("dados-painel").textContent);
  const NL = "\n";

  // true quando este HTML foi servido pelo servidor web (servidor_web.py);
  // false quando está rodando dentro da janela desktop (pywebview). Vem
  // pronto do Python (ver montar_html_painel) em vez do JS ter que
  // adivinhar isso por uma corrida entre o carregamento da página e a
  // injeção assíncrona de "window.pywebview" pelo WebView2.
  window.PAC_MODO_WEB = __MODO_WEB__;

  // Ponte entre o painel e o backend: no desktop, fala diretamente com
  // "window.pywebview.api" (como sempre foi); num navegador comum (sem
  // pywebview), faz a mesma chamada por HTTP contra as rotas /api/* do
  // servidor web (ver servidor_web.py) — o resto do painel não precisa
  // saber qual dos dois está em uso.
  function pacApiDesktopDisponivel() {
    return !!(window.pywebview && window.pywebview.api);
  }

  async function chamarAPI(nome, ...args) {
    if (pacApiDesktopDisponivel()) {
      return await window.pywebview.api[nome](...args);
    }
    const resposta = await fetch("/api/" + nome, {
      method: "POST",
      headers: { "Content-Type": "application/json" },
      body: JSON.stringify(args),
    });
    return await resposta.json();
  }

  // Variante para os dois botões que geram PDF (relatório completo e Ficha
  // Cadastral): no desktop, o Python já salva o arquivo direto no disco via
  // diálogo nativo e só devolve status; num navegador comum não existe
  // diálogo de "Salvar como" para o servidor abrir, então o servidor web
  // devolve os BYTES do PDF na resposta e é o próprio navegador quem
  // baixa/abre o arquivo.
  async function baixarPDF(nome, ...args) {
    if (pacApiDesktopDisponivel()) {
      return await window.pywebview.api[nome](...args);
    }
    const resposta = await fetch("/api/" + nome, {
      method: "POST",
      headers: { "Content-Type": "application/json" },
      body: JSON.stringify(args),
    });
    // O servidor devolve o PDF de verdade (Content-Type: application/pdf)
    // só quando deu certo; qualquer outro caso (recorte vazio, erro,
    // filtros inválidos) volta como JSON, no mesmo formato
    // {"ok": false, ...} de sempre — por isso checa o Content-Type em vez
    // de só o status HTTP, senão um corpo JSON de erro seria tratado como
    // se fosse o PDF.
    const tipo = resposta.headers.get("Content-Type") || "";
    if (!resposta.ok || !tipo.includes("application/pdf")) {
      let corpo = { ok: false, erro: "Falha ao gerar o PDF." };
      try {
        corpo = await resposta.json();
      } catch (e) { /* resposta sem corpo JSON — mantém o erro padrão acima */ }
      return corpo;
    }
    const disposicao = resposta.headers.get("Content-Disposition") || "";
    const nomeMatch = disposicao.match(/filename="?([^"]+)"?/);
    const nomeArquivo = nomeMatch ? nomeMatch[1] : "relatorio.pdf";
    const blob = await resposta.blob();
    const url = URL.createObjectURL(blob);
    window.open(url, "_blank");
    return { ok: true, arquivo: nomeArquivo };
  }

  const estadoSelecao = {};
  DADOS.blocos.forEach(b => { estadoSelecao[b.chave] = new Set(); });

  // Estado das três árvores de data (Conclusão da Fase, Vigência da
  // Cláusula Suspensiva, Conclusão Atual) — cada uma guarda o conjunto de
  // datas exatas marcadas ("AAAA-MM-DD"), à parte do estadoSelecao normal
  // porque são filtros de data em árvore, não listas simples de opções.
  const estadoSelecaoDatas = {
    DATAS_CONCLUSAO_FASE: new Set(),
    DATAS_VIGENCIA: new Set(),
    DATAS_CONCLUSAO_ATUAL: new Set(),
  };

  const grade = document.getElementById("grade");

  // Texto digitado no campo de busca de cada bloco que tem um. É só um
  // filtro VISUAL da lista — não entra em estadoSelecao nem no cálculo dos
  // resultados, então um município marcado continua valendo mesmo depois
  // de sumir da tela por causa da busca.
  const buscaPorBloco = {};

  // Tira acento e caixa para comparar: quem digita "sao" precisa achar
  // "SÃO FRANCISCO DO CONDE", e quem digita "amelia" precisa achar
  // "AMÉLIA RODRIGUES".
  function normalizarBusca(texto) {
    return String(texto || "")
      .normalize("NFD")
      .replace(/[\u0300-\u036f]/g, "")
      .toUpperCase()
      .trim();
  }

  // Esconde da lista os itens que não batem com a busca. Roda depois de
  // cada digitação e também no fim de renderizarBloco, porque a lista é
  // reconstruída do zero lá e perderia o estado da busca.
  function aplicarBusca(chave) {
    const lista = document.getElementById("lista-" + chave);
    const caixaBusca = document.getElementById("busca-" + chave);
    if (!lista) return;
    const termo = normalizarBusca(buscaPorBloco[chave]);
    let visiveis = 0;
    lista.querySelectorAll(".item").forEach(label => {
      const bate = !termo || normalizarBusca(label.dataset.valor).includes(termo);
      label.classList.toggle("oculto-busca", !bate);
      if (bate) visiveis++;
    });
    if (caixaBusca) {
      caixaBusca.classList.toggle("sem-resultado", termo !== "" && visiveis === 0);
    }
  }

  // Opções que a busca do bloco está deixando à mostra. Com a busca vazia,
  // são todas — é o que "Marcar tudo" e "Limpar" usam, para agirem sobre o
  // que está na tela e não sobre a lista inteira escondida atrás do filtro.
  function opcoesVisiveis(bloco) {
    const termo = normalizarBusca(buscaPorBloco[bloco.chave]);
    if (!termo) return bloco.opcoes;
    return bloco.opcoes.filter(op => normalizarBusca(op).includes(termo));
  }

  function criarBloco(bloco) {
    const el = document.createElement("div");
    el.className = "bloco";
    el.dataset.chave = bloco.chave;

    const titulo = document.createElement("div");
    titulo.className = "bloco-titulo";
    const tituloTexto = document.createElement("span");
    tituloTexto.textContent = bloco.titulo;
    titulo.appendChild(tituloTexto);
    el.appendChild(titulo);

    const botoes = document.createElement("div");
    botoes.className = "bloco-botoes";

    const btnMarcar = document.createElement("button");
    btnMarcar.className = "btn-mini";
    btnMarcar.textContent = "Marcar tudo";
    btnMarcar.onclick = () => {
      opcoesVisiveis(bloco).forEach(op => estadoSelecao[bloco.chave].add(op));
      renderizarBloco(bloco.chave);
      atualizarDisponibilidade();
    };
    botoes.appendChild(btnMarcar);

    const btnLimpar = document.createElement("button");
    btnLimpar.className = "btn-mini";
    btnLimpar.textContent = "Limpar";
    btnLimpar.onclick = () => {
      opcoesVisiveis(bloco).forEach(op => estadoSelecao[bloco.chave].delete(op));
      renderizarBloco(bloco.chave);
      atualizarDisponibilidade();
    };
    botoes.appendChild(btnLimpar);

    if (bloco.chave === "FONTE") {
      const btnFinanciamento = document.createElement("button");
      btnFinanciamento.className = "btn-mini";
      btnFinanciamento.textContent = "Financiamento";
      btnFinanciamento.onclick = () => {
        estadoSelecao.FONTE = new Set(bloco.opcoes.filter(op => op.toUpperCase().includes("FINANCIAMENTO")));
        renderizarBloco("FONTE");
        atualizarDisponibilidade();
      };
      botoes.appendChild(btnFinanciamento);
    }

    el.appendChild(botoes);

    // Campo de busca — por enquanto só o MUNICÍPIO, que é o único bloco
    // com lista longa o bastante para justificar. Filtra o que a lista
    // MOSTRA, não o que está marcado: apagar a busca traz todos de volta
    // com as marcações intactas.
    if (bloco.chave === "MUNICIPIO") {
      const caixaBusca = document.createElement("div");
      caixaBusca.className = "bloco-busca";
      caixaBusca.id = "busca-" + bloco.chave;

      const campo = document.createElement("input");
      campo.type = "search";
      campo.placeholder = "Buscar município...";
      campo.autocomplete = "off";
      campo.oninput = () => {
        buscaPorBloco[bloco.chave] = campo.value;
        aplicarBusca(bloco.chave);
      };
      caixaBusca.appendChild(campo);

      const aviso = document.createElement("div");
      aviso.className = "busca-vazia";
      aviso.textContent = "Nenhum município com esse nome.";
      caixaBusca.appendChild(aviso);

      el.appendChild(caixaBusca);
    }

    const lista = document.createElement("div");
    lista.className = "bloco-lista";
    lista.id = "lista-" + bloco.chave;
    el.appendChild(lista);

    if (bloco.chave === "FASE") {
      criarArvoreData(titulo, "Conclusão", DADOS.arvoresData.DATAS_CONCLUSAO_FASE, "DATAS_CONCLUSAO_FASE");
    }

    if (bloco.chave === "STATUS") {
      criarArvoreData(titulo, "P. de Conclusão Atual", DADOS.arvoresData.DATAS_CONCLUSAO_ATUAL, "DATAS_CONCLUSAO_ATUAL");
    }

    if (bloco.chave === "CLAUSULA_SUSPENSIVA") {
      criarArvoreData(titulo, "Vigência", DADOS.arvoresData.DATAS_VIGENCIA, "DATAS_VIGENCIA");
    }

    return el;
  }

  // Monta um filtro de data em ÁRVORE SUSPENSA (Ano > Mês > Dia, com
  // "Selecionar Tudo" no topo) — igual ao filtro de datas do Excel, só que
  // fica FECHADO por padrão (só um botão compacto com o total selecionado)
  // e abre um painel flutuante ao clicar, sem ocupar espaço fixo no bloco.
  // Marcar um nó pai marca/desmarca todos os filhos; o estado de cada nó
  // (marcado/desmarcado/parcial) é recalculado a cada mudança subindo a
  // árvore.
  const paineisDataAbertos = [];
  const resetadoresArvoreData = [];
  const registroCheckboxesDia = {
    DATAS_CONCLUSAO_FASE: [],
    DATAS_VIGENCIA: [],
    DATAS_CONCLUSAO_ATUAL: [],
  };

  function fecharTodosPaineisData(exceto) {
    paineisDataAbertos.forEach(function (p) {
      if (p !== exceto) p.classList.remove("aberto");
    });
  }
  document.addEventListener("click", function (evento) {
    var dentroDeAlgumPainel = evento.target.closest(".arvore-dropdown-painel, .arvore-dropdown-toggle");
    if (!dentroDeAlgumPainel) fecharTodosPaineisData(null);
  });

  function criarArvoreData(container, titulo, arvore, chave) {
    const secao = document.createElement("div");
    secao.className = "arvore-data-secao";

    const toggle = document.createElement("button");
    toggle.type = "button";
    toggle.className = "arvore-dropdown-toggle";
    toggle.title = titulo;
    toggle.innerHTML =
      '<svg viewBox="0 0 24 24" width="15" height="15" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round">' +
        '<rect x="3" y="4" width="18" height="18" rx="2"></rect>' +
        '<line x1="16" y1="2" x2="16" y2="6"></line>' +
        '<line x1="8" y1="2" x2="8" y2="6"></line>' +
        '<line x1="3" y1="10" x2="21" y2="10"></line>' +
      '</svg>';
    const spanContagem = document.createElement("span");
    spanContagem.className = "contagem";
    toggle.appendChild(spanContagem);
    secao.appendChild(toggle);

    const painel = document.createElement("div");
    painel.className = "arvore-dropdown-painel";
    secao.appendChild(painel);
    container.appendChild(secao);
    paineisDataAbertos.push(painel);

    function atualizarContagem() {
      const n = estadoSelecaoDatas[chave].size;
      spanContagem.textContent = n > 0 ? n : "";
      spanContagem.classList.toggle("visivel", n > 0);
    }

    toggle.addEventListener("click", function (evento) {
      evento.stopPropagation();
      const vaiAbrir = !painel.classList.contains("aberto");
      fecharTodosPaineisData(null);
      if (vaiAbrir) {
        const rect = toggle.getBoundingClientRect();
        const alturaMaxima = 320; // deve bater com max-height do CSS
        const espacoAbaixo = window.innerHeight - rect.bottom;
        const espacoAcima = rect.top;

        painel.style.top = "";
        painel.style.bottom = "";
        if (espacoAbaixo < alturaMaxima && espacoAcima > espacoAbaixo) {
          // não cabe embaixo — abre para cima do botão em vez de para baixo
          painel.style.bottom = (window.innerHeight - rect.top + 4) + "px";
          painel.style.maxHeight = Math.min(alturaMaxima, espacoAcima - 10) + "px";
        } else {
          painel.style.top = (rect.bottom + 4) + "px";
          painel.style.maxHeight = Math.min(alturaMaxima, espacoAbaixo - 10) + "px";
        }

        // não deixa vazar pela lateral direita da janela
        const larguraPainel = 220;
        const esquerdaMaxima = window.innerWidth - larguraPainel - 10;
        painel.style.left = Math.min(rect.left, Math.max(esquerdaMaxima, 0)) + "px";

        painel.classList.add("aberto");
      }
    });
    painel.addEventListener("click", function (evento) { evento.stopPropagation(); });

    const anos = Object.keys(arvore || {});
    if (anos.length === 0) {
      painel.innerHTML = '<div class="preview-sem-dados" style="padding:4px 0;font-size:10px;">Sem datas na base</div>';
      resetadoresArvoreData.push(function () { estadoSelecaoDatas[chave].clear(); });
      return;
    }

    const conjunto = estadoSelecaoDatas[chave];

    function atualizarNo(chkPai, filhos) {
      const total = filhos.length;
      let marcados = 0;
      filhos.forEach(f => { if ((f.chk || f).checked) marcados++; });
      if (marcados === 0) {
        chkPai.checked = false;
        chkPai.indeterminate = false;
      } else if (marcados === total) {
        chkPai.checked = true;
        chkPai.indeterminate = false;
      } else {
        chkPai.checked = false;
        chkPai.indeterminate = true;
      }
    }

    const chkRaiz = document.createElement("input");
    chkRaiz.type = "checkbox";
    const labelRaiz = document.createElement("label");
    labelRaiz.className = "arvore-item arvore-raiz";
    labelRaiz.appendChild(chkRaiz);
    labelRaiz.appendChild(document.createTextNode(" Selecionar Tudo"));
    painel.appendChild(labelRaiz);

    const checkboxesAno = [];

    anos.forEach(ano => {
      const detAno = document.createElement("details");
      const sumAno = document.createElement("summary");
      const labAno = document.createElement("label");
      const chkAno = document.createElement("input");
      chkAno.type = "checkbox";
      labAno.appendChild(chkAno);
      labAno.appendChild(document.createTextNode(" " + ano));
      sumAno.appendChild(labAno);
      detAno.appendChild(sumAno);

      const checkboxesMes = [];

      Object.keys(arvore[ano]).forEach(mesChave => {
        const nomeMes = mesChave.slice(3);
        const numeroMes = mesChave.slice(0, 2);
        const detMes = document.createElement("details");
        detMes.className = "arvore-nivel-mes";
        const sumMes = document.createElement("summary");
        const labMes = document.createElement("label");
        const chkMes = document.createElement("input");
        chkMes.type = "checkbox";
        labMes.appendChild(chkMes);
        labMes.appendChild(document.createTextNode(" " + nomeMes));
        sumMes.appendChild(labMes);
        detMes.appendChild(sumMes);

        const checkboxesDia = [];
        arvore[ano][mesChave].forEach(dia => {
          const dataStr = ano + "-" + numeroMes + "-" + String(dia).padStart(2, "0");
          const labDia = document.createElement("label");
          labDia.className = "arvore-item arvore-nivel-dia";
          const chkDia = document.createElement("input");
          chkDia.type = "checkbox";
          chkDia.dataset.data = dataStr;
          labDia.appendChild(chkDia);
          labDia.appendChild(document.createTextNode(" " + dia));
          detMes.appendChild(labDia);
          checkboxesDia.push(chkDia);
          registroCheckboxesDia[chave].push(chkDia);

          chkDia.addEventListener("change", () => {
            if (chkDia.checked) conjunto.add(dataStr); else conjunto.delete(dataStr);
            atualizarNo(chkMes, checkboxesDia);
            atualizarNo(chkAno, checkboxesMes);
            atualizarNo(chkRaiz, checkboxesAno);
            atualizarContagem();
            atualizarDisponibilidade();
          });
        });

        chkMes.addEventListener("change", () => {
          checkboxesDia.forEach(c => {
            c.checked = chkMes.checked;
            if (chkMes.checked) conjunto.add(c.dataset.data); else conjunto.delete(c.dataset.data);
          });
          chkMes.indeterminate = false;
          atualizarNo(chkAno, checkboxesMes);
          atualizarNo(chkRaiz, checkboxesAno);
          atualizarContagem();
          atualizarDisponibilidade();
        });

        checkboxesMes.push({ chk: chkMes, filhos: checkboxesDia });
        detAno.appendChild(detMes);
      });

      chkAno.addEventListener("change", () => {
        checkboxesMes.forEach(m => {
          m.chk.checked = chkAno.checked;
          m.chk.indeterminate = false;
          m.filhos.forEach(c => {
            c.checked = chkAno.checked;
            if (chkAno.checked) conjunto.add(c.dataset.data); else conjunto.delete(c.dataset.data);
          });
        });
        chkAno.indeterminate = false;
        atualizarNo(chkRaiz, checkboxesAno);
        atualizarContagem();
        atualizarDisponibilidade();
      });

      checkboxesAno.push({ chk: chkAno, filhosMes: checkboxesMes });
      painel.appendChild(detAno);
    });

    chkRaiz.addEventListener("change", () => {
      checkboxesAno.forEach(a => {
        a.chk.checked = chkRaiz.checked;
        a.chk.indeterminate = false;
        a.filhosMes.forEach(m => {
          m.chk.checked = chkRaiz.checked;
          m.chk.indeterminate = false;
          m.filhos.forEach(c => {
            c.checked = chkRaiz.checked;
            if (chkRaiz.checked) conjunto.add(c.dataset.data); else conjunto.delete(c.dataset.data);
          });
        });
      });
      atualizarContagem();
      atualizarDisponibilidade();
    });

    resetadoresArvoreData.push(function () {
      conjunto.clear();
      chkRaiz.checked = false;
      chkRaiz.indeterminate = false;
      checkboxesAno.forEach(a => {
        a.chk.checked = false;
        a.chk.indeterminate = false;
        a.filhosMes.forEach(m => {
          m.chk.checked = false;
          m.chk.indeterminate = false;
          m.filhos.forEach(c => { c.checked = false; });
        });
      });
      atualizarContagem();
    });
  }

  // Monta o objeto de filtros enviado à API: os blocos normais (listas de
  // opções) mais as três seleções de data em árvore — usado em todo lugar
  // que precisa dos filtros atuais (pré-visualizar, gerar relatório etc.),
  // para não repetir essa montagem em cada handler.
  function montarFiltrosAtuais() {
    var filtros = {};
    DADOS.blocos.forEach(function (b) { filtros[b.chave] = Array.from(estadoSelecao[b.chave]); });
    filtros.DATAS_CONCLUSAO_FASE = Array.from(estadoSelecaoDatas.DATAS_CONCLUSAO_FASE);
    filtros.DATAS_VIGENCIA = Array.from(estadoSelecaoDatas.DATAS_VIGENCIA);
    filtros.DATAS_CONCLUSAO_ATUAL = Array.from(estadoSelecaoDatas.DATAS_CONCLUSAO_ATUAL);
    // Customização de colunas do Detalhamento — vai junto com os filtros
    // porque é a mesma chamada que gera o relatório. Não filtra linha
    // nenhuma: só diz quais colunas a tabela mostra.
    filtros.COLUNAS_DETALHAMENTO = Array.from(estadoColunas);
    return filtros;
  }

  // --- CUSTOMIZAÇÃO DE COLUNAS do Detalhamento ---
  // Não fica mais no painel de filtros: a lista é desenhada dentro da
  // janela "Selecione as páginas do relatório" (ver mostrarJanelaPaginas),
  // já que escolher colunas é decisão de geração, não de recorte de dados.
  //
  // Três regras são aplicadas aqui, ao vivo, para o usuário não descobrir
  // o problema só depois de gerar o PDF:
  //   - colunas obrigatórias ficam marcadas e desabilitadas;
  //   - dependências de hierarquia (FASE exige OBJETO, STATUS exige
  //     OBJETO+FASE) são resolvidas automaticamente nos dois sentidos:
  //     marcar um filho puxa os pais, desmarcar um pai solta os filhos;
  //   - ao bater o limite de colunas, as não marcadas continuam na lista,
  //     só com o quadrado de seleção esmaecido (não somem).
  // O Python revalida tudo de novo antes de montar a tabela.
  const CATALOGO_COLUNAS = (DADOS.colunasDetalhamento || []);
  const LIMITE_COLUNAS = DADOS.limiteColunasDetalhamento || 9;
  const estadoColunas = new Set(
    CATALOGO_COLUNAS.filter(c => c.padrao).map(c => c.chave)
  );

  function colunaPorChave(chave) {
    return CATALOGO_COLUNAS.find(c => c.chave === chave);
  }

  function marcarColuna(chave) {
    // Marca a coluna e, em cascata, tudo de que ela depende.
    const coluna = colunaPorChave(chave);
    if (!coluna) return;
    estadoColunas.add(chave);
    (coluna.requer || []).forEach(marcarColuna);
  }

  function desmarcarColuna(chave) {
    // Desmarca a coluna e, em cascata, todas as que dependiam dela — do
    // contrário sobraria um STATUS mesclado sem a FASE que o organiza.
    const coluna = colunaPorChave(chave);
    if (!coluna || coluna.obrigatoria) return;
    estadoColunas.delete(chave);
    CATALOGO_COLUNAS.forEach(outra => {
      if ((outra.requer || []).indexOf(chave) !== -1) desmarcarColuna(outra.chave);
    });
  }

  // Quantas colunas entrariam junto se esta fosse marcada agora (ela mais
  // as dependências que ainda faltam) — é o que decide se cabe no limite.
  function custoDeMarcar(chave) {
    const pendentes = new Set();
    (function acumular(alvo) {
      if (estadoColunas.has(alvo) || pendentes.has(alvo)) return;
      pendentes.add(alvo);
      const coluna = colunaPorChave(alvo);
      (coluna && coluna.requer ? coluna.requer : []).forEach(acumular);
    })(chave);
    return pendentes.size;
  }

  function renderizarBlocoColunas() {
    const lista = document.getElementById("paginas-colunas-lista");
    if (!lista) return;
    lista.innerHTML = "";

    CATALOGO_COLUNAS.forEach(coluna => {
      const label = document.createElement("label");
      label.className = "item";
      label.dataset.valor = coluna.chave;

      const input = document.createElement("input");
      input.type = "checkbox";
      input.checked = estadoColunas.has(coluna.chave);

      const cabeLimite =
        estadoColunas.has(coluna.chave) ||
        estadoColunas.size + custoDeMarcar(coluna.chave) <= LIMITE_COLUNAS;

      if (coluna.obrigatoria) {
        input.disabled = true;
        label.classList.add("col-trava");
        label.title = "Coluna obrigatória: é ela que ancora a ordenação e as mesclas da tabela.";
      } else if (!cabeLimite) {
        input.disabled = true;
        label.classList.add("col-trava");
        label.title = "Limite de " + LIMITE_COLUNAS + " colunas atingido — desmarque outra antes.";
      } else if ((coluna.requer || []).length) {
        label.title = "Depende de: " + coluna.requer.join(", ");
      }

      input.onchange = () => {
        if (input.checked) marcarColuna(coluna.chave);
        else desmarcarColuna(coluna.chave);
        renderizarBlocoColunas();
      };

      const span = document.createElement("span");
      span.textContent = coluna.titulo;

      label.appendChild(input);
      label.appendChild(span);
      lista.appendChild(label);
    });

    const contador = document.getElementById("paginas-colunas-contador");
    if (contador) {
      contador.textContent = estadoColunas.size + "/" + LIMITE_COLUNAS;
      contador.style.color = estadoColunas.size >= LIMITE_COLUNAS ? "#C77" : "";
    }
  }

  // Atalhos "Padrão" / "Só o essencial" da seção de colunas (ligados pela
  // mostrarJanelaPaginas quando a janela abre).
  function colunasAplicarPadrao() {
    estadoColunas.clear();
    CATALOGO_COLUNAS.filter(c => c.padrao).forEach(c => estadoColunas.add(c.chave));
    renderizarBlocoColunas();
  }
  function colunasAplicarEssencial() {
    estadoColunas.clear();
    CATALOGO_COLUNAS.filter(c => c.obrigatoria).forEach(c => estadoColunas.add(c.chave));
    ["FASE", "STATUS", "MUNICIPIOS", "INVESTIMENTO"].forEach(marcarColuna);
    renderizarBlocoColunas();
  }

  // Transforma a grade de filtros num acordeão no celular (ver
  // @media (max-width: 768px) no <style>): cada ".bloco-titulo" ganha uma
  // seta e ao ser clicado alterna ".bloco-colapsado" no ".bloco" pai, que é
  // quem esconde os botões/busca/lista via CSS. Roda uma única vez, depois
  // que TODOS os blocos já estão na grade — assim a seta
  // sempre entra como o último filho de cada título, mesmo nos blocos
  // FASE/STATUS, que só ganham o ícone de calendário (arvore-data-secao)
  // dentro do título depois que criarBloco() já terminou.
  //
  // No desktop isso não faz diferença nenhuma: a seta fica sempre
  // escondida (.bloco-seta { display:none }) fora daquele media query, e
  // o clique no título só chega a colapsar algo se a tela realmente
  // estiver estreita (matchMedia checado a cada clique, não só uma vez).
  function configurarAccordionFiltros() {
    grade.querySelectorAll(".bloco-titulo").forEach(function (titulo) {
      var seta = document.createElement("span");
      seta.className = "bloco-seta";
      seta.textContent = "▾";
      seta.setAttribute("aria-hidden", "true");
      titulo.appendChild(seta);
      titulo.addEventListener("click", function () {
        if (!window.matchMedia("(max-width: 768px)").matches) return;
        titulo.parentElement.classList.toggle("bloco-colapsado");
      });
    });
    // Começa tudo recolhido no celular — só os títulos aparecem, o que é
    // o ganho de espaço que o modo acordeão existe pra dar. No desktop
    // (grade normal) essa classe fica sem efeito nenhum.
    if (window.matchMedia("(max-width: 768px)").matches) {
      grade.querySelectorAll(".bloco").forEach(function (blocoEl) {
        blocoEl.classList.add("bloco-colapsado");
      });
    }
  }

  // Mesmo padrão de acordeão acima, agora pras seções do Dashboard
  // (Panorama por Secretaria, Índice de Desempenho, Detalhamento
  // Financeiro, Mapa — ver ".preview-secao" no @media max-width:768px).
  // Diferente dos blocos de filtro (montados em JS a partir de DADOS),
  // essas seções já existem prontas no HTML da página — então isto roda
  // uma única vez, sem depender de nenhum carregamento de dados, e cobre
  // também os botões da nav rápida (#preview-nav-rapida): clicar neles
  // reabre a seção-alvo (se estiver fechada) antes de rolar até ela, pra
  // nunca "pular" pra um título com o corpo escondido.
  function configurarAccordionSecoes() {
    document.querySelectorAll(".preview-secao").forEach(function (secao) {
      var titulo = secao.querySelector(".preview-secao-titulo");
      if (!titulo) return;
      titulo.addEventListener("click", function () {
        if (!window.matchMedia("(max-width: 768px)").matches) return;
        secao.classList.toggle("secao-colapsada");
      });
      if (window.matchMedia("(max-width: 768px)").matches) {
        secao.classList.add("secao-colapsada");
      }
    });

    document.querySelectorAll(".preview-nav-rapida button").forEach(function (botao) {
      botao.addEventListener("click", function () {
        var alvo = document.getElementById(botao.dataset.alvo);
        if (!alvo) return;
        alvo.classList.remove("secao-colapsada");
        alvo.scrollIntoView({ behavior: "smooth", block: "start" });
      });
    });
  }

  function renderizarBloco(chave) {
    const bloco = DADOS.blocos.find(b => b.chave === chave);
    const lista = document.getElementById("lista-" + chave);
    // O bloco GESTÃO não tem card na grade (é controlado pelas pills do
    // topo do dashboard), mas continua no estadoSelecao — então quem chamar
    // renderizarBloco("GESTAO") simplesmente não faz nada aqui, em vez de
    // estourar um erro de elemento inexistente.
    if (!bloco || !lista) return;
    lista.innerHTML = "";
    bloco.opcoes.forEach(opcao => {
      const label = document.createElement("label");
      label.className = "item";
      label.dataset.valor = opcao;

      const input = document.createElement("input");
      input.type = "checkbox";
      input.checked = estadoSelecao[chave].has(opcao);
      input.onchange = () => {
        if (input.checked) estadoSelecao[chave].add(opcao);
        else estadoSelecao[chave].delete(opcao);
        atualizarDisponibilidade();
      };

      const span = document.createElement("span");
      span.textContent = opcao;

      label.appendChild(input);
      label.appendChild(span);
      lista.appendChild(label);
    });
    aplicarBusca(chave);
  }

  // Mesma lógica do antigo atualizar_disponibilidade() em Python: para cada
  // bloco, filtra as linhas usando a seleção atual de TODOS OS OUTROS
  // blocos, e acinzenta (desabilita) as opções que não sobram em nenhuma
  // linha dessa combinação — evita marcar um cruzamento de filtros sem
  // nenhum resultado. Roda inteiramente no navegador, então é instantâneo.
  const CHAVES_DATA = ["DATAS_CONCLUSAO_FASE", "DATAS_VIGENCIA", "DATAS_CONCLUSAO_ATUAL"];

  // Filtra DADOS.linhas usando TODOS os filtros ativos no momento — blocos
  // de checkbox normais e as três árvores de data — exceto o filtro
  // "chaveExcluida" (o próprio filtro que está sendo recalculado não pode
  // se autolimitar). Usada tanto para acinzentar opções de bloco quanto
  // para acinzentar datas indisponíveis, mantendo os dois tipos de filtro
  // se afetando mutuamente.
  function linhasFiltradasExcluindo(chaveExcluida) {
    let linhas = DADOS.linhas;
    DADOS.blocos.forEach(outroBloco => {
      if (outroBloco.chave === chaveExcluida) return;
      const selecionados = estadoSelecao[outroBloco.chave];
      if (selecionados.size > 0) {
        linhas = linhas.filter(l => selecionados.has(l[outroBloco.chave]));
      }
    });
    CHAVES_DATA.forEach(chaveData => {
      if (chaveData === chaveExcluida) return;
      const selecionadas = estadoSelecaoDatas[chaveData];
      if (selecionadas && selecionadas.size > 0) {
        linhas = linhas.filter(l => selecionadas.has(l[chaveData]));
      }
    });
    return linhas;
  }

  function atualizarDisponibilidade() {
    DADOS.blocos.forEach(blocoAlvo => {
      const linhas = linhasFiltradasExcluindo(blocoAlvo.chave);
      const disponiveis = new Set(linhas.map(l => l[blocoAlvo.chave]));

      const lista = document.getElementById("lista-" + blocoAlvo.chave);
      if (!lista) return; // bloco sem card na grade (ex: GESTÃO, que virou pills)
      lista.querySelectorAll(".item").forEach(label => {
        const valor = label.dataset.valor;
        const input = label.querySelector("input");
        if (disponiveis.has(valor)) {
          label.classList.remove("indisponivel");
          input.disabled = false;
        } else {
          label.classList.add("indisponivel");
          input.disabled = true;
        }
      });
    });
    atualizarDisponibilidadeDatas();
    agendarAtualizacaoQualidadeFiltros();
  }

  // O botão de Controle de Qualidade ao lado da lupa precisa acompanhar os
  // filtros ENQUANTO a pessoa mexe no painel — não só quando o dashboard é
  // recarregado ao fechar os filtros. Cada mudança de filtro reagenda uma
  // consulta leve ao backend (só o aviso, sem os agregados do dashboard),
  // com debounce para não disparar uma chamada por clique numa marcação em
  // lote ("Marcar tudo"/"Limpar"). Só roda com o painel de filtros à mostra:
  // do lado do dashboard quem cuida do botão é o carregarDashboard().
  var _timerQualidadeFiltros = null;
  var _qualidadeFiltrosEmCurso = 0;
  function agendarAtualizacaoQualidadeFiltros() {
    // #app é position:fixed — offsetParent é sempre null nele, mesmo aberto,
    // então quem diz se o painel está à mostra é o display computado.
    var app = document.getElementById("app");
    if (!app || getComputedStyle(app).display === "none") return;
    if (_timerQualidadeFiltros) clearTimeout(_timerQualidadeFiltros);
    _timerQualidadeFiltros = setTimeout(async function () {
      _timerQualidadeFiltros = null;
      var minhaVez = ++_qualidadeFiltrosEmCurso;
      var resultado;
      try {
        resultado = await chamarAPI("aviso_qualidade", montarFiltrosAtuais());
      } catch (erro) {
        return; // sem rede/ponte: mantém o badge como está
      }
      if (minhaVez !== _qualidadeFiltrosEmCurso) return; // resposta atrasada
      if (resultado && resultado.ok) atualizarBotaoQualidade(resultado.aviso);
    }, 350);
  }

  // Mesma lógica de acinzentamento, só que para as folhas (dias) das três
  // árvores de data — encolhe conforme os outros filtros (de bloco ou de
  // outra data) vão sendo ativados.
  function atualizarDisponibilidadeDatas() {
    CHAVES_DATA.forEach(chaveData => {
      const linhas = linhasFiltradasExcluindo(chaveData);
      const disponiveis = new Set(
        linhas.map(l => l[chaveData]).filter(v => v !== null && v !== undefined)
      );
      registroCheckboxesDia[chaveData].forEach(chk => {
        const label = chk.closest(".arvore-item");
        if (disponiveis.has(chk.dataset.data)) {
          if (label) label.classList.remove("indisponivel");
          chk.disabled = false;
        } else {
          if (label) label.classList.add("indisponivel");
          chk.disabled = true;
        }
      });
    });
  }

  if (!DADOS || !DADOS.blocos || DADOS.blocos.length === 0) {
    grade.innerHTML = '<p style="color:#fff;padding:20px;">Nenhum bloco de filtro foi carregado (DADOS vazio ou ausente).</p>';
  } else {
    // GESTÃO não tem card na grade (é controlada pelas pills do topo do
    // dashboard, ligadas ao mesmo estadoSelecao.GESTAO). Todos os demais
    // blocos — inclusive CLÁUSULA SUSPENSIVA e TERMO DE COMPROMISSO, que
    // voltaram a ocupar um slot inteiro cada — entram lado a lado na grade.
    // O slot que sobrou era do antigo bloco "Colunas do Detalhamento", que
    // se mudou para a janela "Selecione as páginas do relatório".
    DADOS.blocos.forEach(function (bloco) {
      if (bloco.chave === "GESTAO") return;
      grade.appendChild(criarBloco(bloco));
      renderizarBloco(bloco.chave);
    });
    atualizarDisponibilidade();
    montarFiltroGestaoDash();
    montarFiltroSecretariaDash();
    configurarAccordionFiltros();
  }

  // Linha de filtro rápido por Secretaria/Órgão, no cabeçalho do
  // dashboard — um botão (pill) por secretaria, marcando/desmarcando ao
  // clicar e atualizando o dashboard na hora, sem precisar abrir o painel
  // de filtros completo. Reflete e fica sincronizada com a seleção do
  // bloco SECRETARIA/ÓRGÃO normal (estadoSelecao.ORGAO) nos dois sentidos.
  function montarFiltroGestaoDash() {
    const container = document.getElementById("dash-gestao-filtro");
    if (!container) return;
    container.innerHTML = "";
    const blocoGestao = DADOS.blocos.find(b => b.chave === "GESTAO");
    if (!blocoGestao) return;
    blocoGestao.opcoes.forEach(opcao => {
      const pill = document.createElement("button");
      pill.type = "button";
      pill.className = "dash-secretaria-pill" + (estadoSelecao.GESTAO.has(opcao) ? " ativo" : "");
      pill.textContent = opcao;
      pill.onclick = async () => {
        if (estadoSelecao.GESTAO.has(opcao)) {
          estadoSelecao.GESTAO.delete(opcao);
        } else {
          estadoSelecao.GESTAO.add(opcao);
        }
        renderizarBloco("GESTAO");
        atualizarDisponibilidade();
        montarFiltroGestaoDash();
        montarFiltroSecretariaDash();
        await carregarDashboard();
      };
      container.appendChild(pill);
    });
  }

  function montarFiltroSecretariaDash() {
    const container = document.getElementById("dash-secretaria-filtro");
    if (!container) return;
    container.innerHTML = "";
    const blocoOrgao = DADOS.blocos.find(b => b.chave === "ORGAO");
    if (!blocoOrgao) return;
    // "GESTÃO FEDERAL" é um valor real da coluna SECRETARIA/ÓRGÃO na
    // planilha, mas como agora tem a pill de GESTÃO logo à esquerda desse
    // bloco (com a mesma opção "GESTÃO FEDERAL"), mostrar os dois juntos
    // fica redundante e confuso — por isso essa opção específica não
    // aparece aqui como pill de secretaria (continua funcionando
    // normalmente em qualquer outro filtro/lugar do app).
    blocoOrgao.opcoes.filter(opcao => opcao !== "GESTÃO FEDERAL").forEach(opcao => {
      const pill = document.createElement("button");
      pill.type = "button";
      pill.className = "dash-secretaria-pill" + (estadoSelecao.ORGAO.has(opcao) ? " ativo" : "");
      pill.textContent = opcao;
      pill.onclick = async () => {
        if (estadoSelecao.ORGAO.has(opcao)) {
          estadoSelecao.ORGAO.delete(opcao);
        } else {
          estadoSelecao.ORGAO.add(opcao);
        }
        renderizarBloco("ORGAO");
        atualizarDisponibilidade();
        montarFiltroSecretariaDash();
        await carregarDashboard();
      };
      container.appendChild(pill);
    });
    montarFiltroExecutorDash();
  }

  // Pills de Executor — só aparecem (com o separador verde antes delas)
  // depois que pelo menos uma Secretaria/Órgão é selecionada, e mostram só
  // os executores que de fato existem dentro da secretaria já escolhida
  // (mesmo cálculo de disponibilidade usado pra acinzentar opções no
  // painel de filtros, só que aqui decide o que aparece ou não).
  function montarFiltroExecutorDash() {
    const separador = document.getElementById("dash-separador-executor");
    const container = document.getElementById("dash-executor-filtro");
    if (!container || !separador) return;
    container.innerHTML = "";

    // Normalmente só mostra Executor depois de escolher uma Secretaria —
    // mas como a secretaria "GESTÃO FEDERAL" fica escondida (ver
    // montarFiltroSecretariaDash), selecionar essa Gestão sozinha já deve
    // levar direto pro bloco de Executor, sem precisar de uma secretaria
    // explícita (que nem apareceria pra escolher).
    var gestaoFederalAtiva = estadoSelecao.GESTAO.has("GESTÃO FEDERAL");
    if (estadoSelecao.ORGAO.size === 0 && !gestaoFederalAtiva) {
      separador.classList.remove("visivel");
      container.style.display = "none";
      montarFiltroObjetoDash();
      return;
    }

    const blocoExecutor = DADOS.blocos.find(b => b.chave === "EXECUTOR");
    if (!blocoExecutor) {
      separador.classList.remove("visivel");
      container.style.display = "none";
      montarFiltroObjetoDash();
      return;
    }

    const linhasDisponiveis = linhasFiltradasExcluindo("EXECUTOR");
    // Se nenhuma secretaria foi escolhida (o bloco só está visível aqui
    // por causa da Gestão Federal — ver acima), restringe às linhas de
    // GESTÃO FEDERAL especificamente, mesmo que Gestão Estadual também
    // esteja marcada — senão os executores estaduais vazariam pra cá sem
    // nenhuma secretaria estadual ter sido escolhida.
    const linhasParaExecutor =
      estadoSelecao.ORGAO.size === 0
        ? linhasDisponiveis.filter(l => l.GESTAO === "GESTÃO FEDERAL")
        : linhasDisponiveis;
    const disponiveis = new Set(linhasParaExecutor.map(l => l.EXECUTOR));
    const opcoesVisiveis = blocoExecutor.opcoes.filter(op => disponiveis.has(op));

    if (opcoesVisiveis.length === 0) {
      separador.classList.remove("visivel");
      container.style.display = "none";
      montarFiltroObjetoDash();
      return;
    }

    separador.classList.add("visivel");
    container.style.display = "flex";
    opcoesVisiveis.forEach(opcao => {
      const pill = document.createElement("button");
      pill.type = "button";
      pill.className = "dash-secretaria-pill" + (estadoSelecao.EXECUTOR.has(opcao) ? " ativo" : "");
      pill.textContent = opcao;
      pill.onclick = async () => {
        if (estadoSelecao.EXECUTOR.has(opcao)) {
          estadoSelecao.EXECUTOR.delete(opcao);
        } else {
          estadoSelecao.EXECUTOR.add(opcao);
        }
        renderizarBloco("EXECUTOR");
        atualizarDisponibilidade();
        montarFiltroExecutorDash();
        await carregarDashboard();
      };
      container.appendChild(pill);
    });
    montarFiltroObjetoDash();
  }

  // Pills de OBJETO — à direita do Executor. Aparecem quando o Executor
  // está "definido": ou porque a pessoa selecionou um (ou mais) Executor
  // explicitamente, ou porque, mesmo sem seleção, só existe UM Executor
  // disponível dentro do que já foi filtrado até aqui (nesse caso não
  // precisa clicar em nada — já é inequívoco qual Executor está em jogo).
  function montarFiltroObjetoDash() {
    const separador = document.getElementById("dash-separador-objeto");
    const container = document.getElementById("dash-objeto-filtro");
    if (!container || !separador) return;
    container.innerHTML = "";

    const blocoExecutor = DADOS.blocos.find(b => b.chave === "EXECUTOR");
    const blocoObjeto = DADOS.blocos.find(b => b.chave === "OBJETO");
    if (!blocoExecutor || !blocoObjeto) {
      separador.classList.remove("visivel");
      container.style.display = "none";
      return;
    }

    const linhasDisponiveisExecutor = linhasFiltradasExcluindo("EXECUTOR");
    const executoresDisponiveis = new Set(linhasDisponiveisExecutor.map(l => l.EXECUTOR));
    const opcoesExecutorVisiveis = blocoExecutor.opcoes.filter(op => executoresDisponiveis.has(op));

    var executorDefinido = estadoSelecao.EXECUTOR.size > 0 || opcoesExecutorVisiveis.length === 1;
    if (!executorDefinido) {
      separador.classList.remove("visivel");
      container.style.display = "none";
      return;
    }

    const linhasDisponiveisObjeto = linhasFiltradasExcluindo("OBJETO");
    const objetosDisponiveis = new Set(linhasDisponiveisObjeto.map(l => l.OBJETO));
    const opcoesObjetoVisiveis = blocoObjeto.opcoes.filter(op => objetosDisponiveis.has(op));

    if (opcoesObjetoVisiveis.length === 0) {
      separador.classList.remove("visivel");
      container.style.display = "none";
      return;
    }

    separador.classList.add("visivel");
    container.style.display = "flex";
    opcoesObjetoVisiveis.forEach(opcao => {
      const pill = document.createElement("button");
      pill.type = "button";
      pill.className = "dash-secretaria-pill" + (estadoSelecao.OBJETO.has(opcao) ? " ativo" : "");
      pill.textContent = opcao;
      pill.onclick = async () => {
        if (estadoSelecao.OBJETO.has(opcao)) {
          estadoSelecao.OBJETO.delete(opcao);
        } else {
          estadoSelecao.OBJETO.add(opcao);
        }
        renderizarBloco("OBJETO");
        atualizarDisponibilidade();
        montarFiltroObjetoDash();
        await carregarDashboard();
      };
      container.appendChild(pill);
    });
  }

  function limparTodosFiltros() {
    // Zera a busca dos blocos antes de redesenhar, senão a lista voltaria
    // com todos desmarcados mas ainda escondida atrás do último termo
    // digitado.
    Object.keys(buscaPorBloco).forEach(chave => {
      buscaPorBloco[chave] = "";
      var caixa = document.getElementById("busca-" + chave);
      var campo = caixa ? caixa.querySelector("input") : null;
      if (campo) campo.value = "";
    });
    DADOS.blocos.forEach(b => {
      estadoSelecao[b.chave].clear();
      renderizarBloco(b.chave);
    });
    resetadoresArvoreData.forEach(resetar => resetar());
    fecharTodosPaineisData(null);
    atualizarDisponibilidade();
    montarFiltroGestaoDash();
    montarFiltroSecretariaDash();
    // Os campos de busca agora vivem na janela da lupa; limpar os filtros
    // do painel também os zera, para a próxima busca começar do recorte novo.
    var campoDescricao = document.getElementById("busca-ficha-descricao");
    var campoMunicipio = document.getElementById("busca-ficha-municipio");
    if (campoDescricao) campoDescricao.value = "";
    if (campoMunicipio) campoMunicipio.value = "";
  }
  document.getElementById("btn-limpar-tudo").onclick = limparTodosFiltros;

  document.getElementById("btn-gerencial-filtros").onclick = function () {
    var blocoStatus = DADOS.blocos.find(b => b.chave === "STATUS");
    estadoSelecao.STATUS = new Set(DADOS.statusPadrao.filter(op => blocoStatus.opcoes.includes(op)));
    renderizarBloco("STATUS");
    atualizarDisponibilidade();
  };

  // --- Paleta usada nos gráficos, mesma dos gráficos do PDF ---
  // Os tons vêm prontos do Python (já dessaturados). Os valores à direita
  // do "||" são só uma rede de segurança caso o painel seja aberto com um
  // arquivo de dados antigo, sem a paleta.
  const PALETA_GRAFICOS = DADOS.paletaGraficos || {};
  const COR_AZUL = PALETA_GRAFICOS.azul || "#4E92BA";
  const COR_VERMELHO = PALETA_GRAFICOS.vermelho || "#BB6060";
  const COR_VERDE = PALETA_GRAFICOS.verde || "#49925C";
  const COR_AMARELO = PALETA_GRAFICOS.amarelo || "#BC9E2C";

  // Pinta os quadradinhos de legenda que ficam fixos no HTML.
  document.querySelectorAll("[data-paleta]").forEach(function (elemento) {
    const cor = PALETA_GRAFICOS[elemento.dataset.paleta];
    if (cor) elemento.style.background = cor;
  });
  const CORES_FASE = {
    "CAPTAÇÃO DE RECURSO": COR_VERMELHO,
    "LICITAÇÃO": COR_AMARELO,
    "EXECUÇÃO DO OBJETO": COR_VERDE,
    "CONCLUÍDA": COR_AZUL,
  };
  const CORES_FINANCEIRO = {
    "VALOR CONTRATADO": COR_VERDE,
    "OGU": COR_AZUL,
    "FINANCIAMENTO": COR_VERMELHO,
    "RECURSO ESTADUAL": COR_AMARELO,
  };

  function formatarMiBi(valor) {
    var texto;
    if (valor >= 1000000000) {
      texto = "R$ " + (valor / 1000000000).toFixed(1) + " Bi";
    } else if (valor >= 1000000) {
      texto = "R$ " + (valor / 1000000).toFixed(1) + " Mi";
    } else {
      texto = "R$ " + valor.toLocaleString("pt-BR", { maximumFractionDigits: 0 });
    }
    return texto;
  }

  // Gráfico de pizza feito só com CSS (conic-gradient) — sem lib externa.
  // Aceita um container (elemento) direto, para poder ser reaproveitado
  // tanto nos gráficos do Painel Geral quanto nos mini-gráficos da grade
  // de Secretaria | Executor.
  function criarGraficoPizzaEm(container, itens, corPorRotulo, opcoes) {
    opcoes = opcoes || {};
    container.innerHTML = "";
    if (!itens || itens.length === 0) {
      container.innerHTML = '<div class="preview-sem-dados">Sem investimento no filtro atual</div>';
      return;
    }
    var total = itens.reduce(function (soma, item) { return soma + item.valor; }, 0);
    var wrap = document.createElement("div");
    wrap.className = "grafico-pizza-wrap";

    var pizza = document.createElement("div");
    pizza.className = "grafico-pizza";
    if (opcoes.tamanho) {
      pizza.style.width = opcoes.tamanho + "px";
      pizza.style.height = opcoes.tamanho + "px";
      pizza.style.minWidth = opcoes.tamanho + "px";
    }
    var partes = [];
    var acumulado = 0;
    // Uma fina folga entre fatias, na MESMA cor de fundo do card (não mais
    // branca) — dá a impressão de um "vazado" entre as fatias, em vez de
    // uma linha divisória branca chamando atenção.
    var meiaFolga = 0.8;
    itens.forEach(function (item, indice) {
      var cor = corPorRotulo[item.rotulo] || "#888";
      var inicio = (acumulado / total) * 360;
      acumulado += item.valor;
      var fim = (acumulado / total) * 360;
      var inicioCor = indice === 0 ? inicio : inicio + meiaFolga;
      var fimCor = indice === itens.length - 1 ? fim : fim - meiaFolga;
      partes.push(cor + " " + inicioCor.toFixed(2) + "deg " + fimCor.toFixed(2) + "deg");
      if (indice < itens.length - 1) {
        partes.push("var(--cor-card) " + fimCor.toFixed(2) + "deg " + (fim + meiaFolga).toFixed(2) + "deg");
      }
    });
    pizza.style.background = "conic-gradient(" + partes.join(", ") + ")";
    wrap.appendChild(pizza);

    if (!opcoes.semLegenda) {
      var legenda = document.createElement("div");
      legenda.className = "grafico-legenda";
      itens.forEach(function (item) {
        var cor = corPorRotulo[item.rotulo] || "#888";
        var linha = document.createElement("div");
        linha.className = "grafico-legenda-item";
        var pct = ((item.valor / total) * 100).toFixed(1);
        var sufixoQtd = item.qtd !== undefined ? ' <span class="grafico-legenda-qtd-mini">(' + item.qtd + ')</span>' : "";
        var rotuloHtml = opcoes.semRotulo
          ? '<span class="grafico-legenda-rotulo">' + sufixoQtd.trim() + '</span>'
          : '<span class="grafico-legenda-rotulo">' + item.rotulo + sufixoQtd + '</span>';
        linha.innerHTML =
          '<span class="grafico-legenda-bolinha" style="background:' + cor + ';"></span>' +
          '<span class="grafico-legenda-corpo">' +
            '<span class="grafico-legenda-valor">' + formatarMiBi(item.valor) + ' <span class="grafico-legenda-pct">(' + pct + '%)</span></span>' +
            rotuloHtml +
          '</span>';
        if (item.objetos && item.objetos.length > 0) {
          linha.classList.add("grafico-legenda-item-com-dado");
          ativarTooltipPrazo(linha, item);
        }
        legenda.appendChild(linha);
      });
      wrap.appendChild(legenda);
    }
    container.appendChild(wrap);
  }

  // --- Tooltip com o detalhamento por OBJETO de uma fase, usado nos
  // mini-gráficos de Secretaria | Executor. Um único elemento flutuante é
  // reaproveitado para todos os itens (criado uma vez, reposicionado a
  // cada hover), em vez de um por gráfico.
  var elementoTooltipObjetos = null;
  function obterTooltipObjetos() {
    if (!elementoTooltipObjetos) {
      elementoTooltipObjetos = document.createElement("div");
      elementoTooltipObjetos.id = "tooltip-objetos";
      elementoTooltipObjetos.className = "tooltip-objetos";
      elementoTooltipObjetos.addEventListener("mouseleave", function () {
        elementoTooltipObjetos.classList.remove("visivel");
      });
      document.body.appendChild(elementoTooltipObjetos);
    }
    return elementoTooltipObjetos;
  }

  function construirConteudoTooltipCard(item) {
    var corpo = "";
    item.fases.forEach(function (fase) {
      if (!fase.objetos || fase.objetos.length === 0) return;
      corpo += '<div class="tooltip-objetos-titulo">' + fase.rotulo + '</div>';
      fase.objetos.forEach(function (o) {
        corpo +=
          '<div class="tooltip-objetos-item">' +
            '<span class="tooltip-objetos-nome">' + o.objeto + ' <span class="tooltip-objetos-qtd">(' + o.qtd + ')</span></span>' +
            '<span class="tooltip-objetos-valor">' + formatarMiBi(o.valor) + '</span>' +
          '</div>';
      });
    });
    return corpo;
  }

  function ativarTooltipCard(elementoGatilho, item) {
    elementoGatilho.addEventListener("mouseenter", function () {
      var tooltip = obterTooltipObjetos();
      tooltip.innerHTML = construirConteudoTooltipCard(item);
      tooltip.classList.add("visivel");
      posicionarTooltipObjetos(elementoGatilho);
    });
    elementoGatilho.addEventListener("mouseleave", function () {
      var tooltip = obterTooltipObjetos();
      // Só esconde se o mouse não foi para dentro do próprio tooltip (dá
      // pra rolar uma lista de objetos longa sem o tooltip sumir).
      setTimeout(function () {
        if (!tooltip.matches(":hover") && !elementoGatilho.matches(":hover")) {
          tooltip.classList.remove("visivel");
        }
      }, 50);
    });
  }

  function posicionarTooltipObjetos(elementoReferencia) {
    var tooltip = obterTooltipObjetos();
    var rect = elementoReferencia.getBoundingClientRect();
    var larguraTooltip = 260;
    var espacoAbaixo = window.innerHeight - rect.bottom;

    tooltip.style.width = larguraTooltip + "px";
    var esquerda = Math.min(rect.left, Math.max(window.innerWidth - larguraTooltip - 10, 0));
    tooltip.style.left = esquerda + "px";

    if (espacoAbaixo < 160 && rect.top > espacoAbaixo) {
      tooltip.style.top = "";
      tooltip.style.bottom = (window.innerHeight - rect.top + 6) + "px";
    } else {
      tooltip.style.bottom = "";
      tooltip.style.top = (rect.bottom + 6) + "px";
    }
  }

  // --- Tooltip do mapa: nome do município + tabela OBJETO | INVESTIMENTO
  // TOTAL, ao passar o mouse em cima do polígono. Mesmo elemento flutuante
  // compartilhado usado nos mini-cards e nas barras de secretaria.
  function construirConteudoTooltipMunicipio(item) {
    var corpo =
      '<div class="tooltip-objetos-titulo tooltip-objetos-titulo-com-valor">' +
        '<span>' + item.municipio + ' <span class="tooltip-objetos-qtd">(' + item.qtd + ')</span></span>' +
        '<span class="tooltip-objetos-valor">' + formatarMiBi(item.valor) + '</span>' +
      '</div>';
    item.objetos.forEach(function (o) {
      corpo +=
        '<div class="tooltip-objetos-item">' +
          '<span class="tooltip-objetos-nome">' + o.objeto + '</span>' +
          '<span class="tooltip-objetos-valor">' + formatarMiBi(o.valor) + '</span>' +
        '</div>';
    });
    return corpo;
  }

  function ativarTooltipMunicipio(elementoGatilho, item) {
    elementoGatilho.addEventListener("mouseenter", function () {
      var tooltip = obterTooltipObjetos();
      tooltip.innerHTML = construirConteudoTooltipMunicipio(item);
      tooltip.classList.add("visivel");
      posicionarTooltipObjetos(elementoGatilho);
    });
    elementoGatilho.addEventListener("mousemove", function () {
      posicionarTooltipObjetos(elementoGatilho);
    });
    elementoGatilho.addEventListener("mouseleave", function () {
      var tooltip = obterTooltipObjetos();
      setTimeout(function () {
        if (!tooltip.matches(":hover") && !elementoGatilho.matches(":hover")) {
          tooltip.classList.remove("visivel");
        }
      }, 50);
    });
  }

  function criarGraficoPizza(idAlvo, itens, corPorRotulo) {
    criarGraficoPizzaEm(document.getElementById(idAlvo), itens, corPorRotulo);
  }

  // Gráfico de barras feito só com CSS (flexbox) — sem lib externa.
  function construirConteudoTooltipPrazo(item) {
    if (!item.objetos || item.objetos.length === 0) return "";
    var valorTotal = item.objetos.reduce(function (soma, o) { return soma + o.valor; }, 0);
    var corpo =
      '<div class="tooltip-objetos-titulo tooltip-objetos-titulo-com-valor">' +
        '<span>' + item.rotulo + ' <span class="tooltip-objetos-qtd">(' + item.qtd + ')</span></span>' +
        '<span class="tooltip-objetos-valor">' + formatarMiBi(valorTotal) + '</span>' +
      '</div>';
    item.objetos.forEach(function (o) {
      corpo +=
        '<div class="tooltip-objetos-item">' +
          '<span class="tooltip-objetos-nome">' + o.objeto + (o.qtd > 1 ? ' <span class="tooltip-objetos-qtd">(' + o.qtd + ')</span>' : '') + '</span>' +
          '<span class="tooltip-objetos-valor">' + formatarMiBi(o.valor) + '</span>' +
        '</div>';
    });
    return corpo;
  }

  function ativarTooltipPrazo(elementoGatilho, item) {
    elementoGatilho.addEventListener("mouseenter", function () {
      var tooltip = obterTooltipObjetos();
      tooltip.innerHTML = construirConteudoTooltipPrazo(item);
      tooltip.classList.add("visivel");
      posicionarTooltipObjetos(elementoGatilho);
    });
    elementoGatilho.addEventListener("mouseleave", function () {
      var tooltip = obterTooltipObjetos();
      setTimeout(function () {
        if (!tooltip.matches(":hover") && !elementoGatilho.matches(":hover")) {
          tooltip.classList.remove("visivel");
        }
      }, 50);
    });
  }

  function criarGraficoBarras(idAlvo, itens) {
    var container = document.getElementById(idAlvo);
    container.innerHTML = "";
    if (!itens || itens.length === 0) {
      container.innerHTML = '<div class="preview-sem-dados">Sem dados no filtro atual</div>';
      return;
    }
    var maiorQtd = Math.max.apply(null, itens.map(function (i) { return i.qtd; }));
    var muitasColunas = itens.length > 8;
    var grafico = document.createElement("div");
    grafico.className = "grafico-barras" + (muitasColunas ? " muitas-colunas" : "");
    itens.forEach(function (item) {
      var coluna = document.createElement("div");
      coluna.className = "grafico-barra-coluna";
      var alturaPct = maiorQtd > 0 ? (item.qtd / maiorQtd) * 100 : 0;
      var cor = item.rotulo === "A definir" ? COR_VERMELHO : COR_AZUL;
      coluna.innerHTML =
        '<div class="grafico-barra-valor">' + item.qtd + '</div>' +
        '<div class="grafico-barra" style="height:' + alturaPct + '%;background:' + cor + ';"></div>' +
        '<div class="grafico-barra-rotulo">' + item.rotulo + '</div>';
      if (item.objetos && item.objetos.length > 0) {
        coluna.classList.add("grafico-barra-coluna-com-dado");
        ativarTooltipPrazo(coluna, item);
      }
      grafico.appendChild(coluna);
    });
    container.appendChild(grafico);
  }

  // Grade de mini-gráficos por combinação SECRETARIA | EXECUTOR — mesma
  // ideia da página "Panorama por Secretaria | Executor" do PDF.
  // --- Gráfico de medidor (gauge/velocímetro) em SVG puro — usado na
  // grade de Índice de Desempenho por Secretaria. Semicírculo com 4
  // faixas fixas de cor (Insatisfatório/Regular/Bom/Ótimo) e um ponteiro
  // apontando pro índice — mesmo desenho usado no PDF (gerar_medidor_
  // desempenho), só que aqui em SVG em vez de ReportLab.
  var CORES_CATEGORIA_DESEMPENHO_JS = {
    "Insatisfatório": PALETA_GRAFICOS.medidorInsatisfatorio || "#BB6060",
    "Regular": PALETA_GRAFICOS.medidorRegular || "#D9A441",
    "Bom": PALETA_GRAFICOS.medidorBom || "#9FCE9B",
    "Ótimo": PALETA_GRAFICOS.medidorOtimo || "#3F8F52",
  };

  // Geometria do medidor — os mesmos números do PDF (ver
  // gerar_medidor_desempenho): anel aberto de 250°, começando às 215° e
  // terminando às -35°, com o índice 50 caindo no topo.
  var ANGULO_INICIAL_GAUGE = 215;
  var VARREDURA_GAUGE = 250;

  function _pontoGauge(cx, cy, r, anguloGraus) {
    var rad = (anguloGraus * Math.PI) / 180;
    // y invertido: em SVG o eixo cresce para baixo.
    return { x: cx + r * Math.cos(rad), y: cy - r * Math.sin(rad) };
  }

  function _arcoGauge(cx, cy, r, anguloIni, anguloFim, espessura, cor) {
    // Arco desenhado como traço com ponta arredondada (stroke-linecap), em
    // vez do setor preenchido de antes — é o que dá as pontas redondas sem
    // precisar de círculos extras em cada extremidade.
    var p1 = _pontoGauge(cx, cy, r, anguloIni);
    var p2 = _pontoGauge(cx, cy, r, anguloFim);
    var largeArc = Math.abs(anguloFim - anguloIni) > 180 ? 1 : 0;
    var varredura = anguloFim < anguloIni ? 1 : 0;
    return '<path d="M ' + p1.x + ' ' + p1.y + ' A ' + r + ' ' + r + ' 0 ' +
      largeArc + ' ' + varredura + ' ' + p2.x + ' ' + p2.y +
      '" fill="none" stroke="' + cor + '" stroke-width="' + espessura +
      '" stroke-linecap="round"></path>';
  }

  function criarGaugeDesempenho(indice, esmaecido, categoria) {
    var largura = 200, altura = 150;
    var espessura = 18;
    // Espaço reservado embaixo para a categoria, na mesma proporção do PDF.
    var espacoCategoria = altura * 0.22;
    var raio = Math.min((largura - espessura) / 2, (altura - espacoCategoria - espessura) / 1.574);
    var cx = largura / 2;
    // cy medido de cima para baixo, ao contrário do PDF.
    var cy = altura - (espacoCategoria + espessura / 2 + 0.574 * raio);
    var valor = Math.max(0, Math.min(100, indice));
    var anguloValor = ANGULO_INICIAL_GAUGE - (valor / 100) * VARREDURA_GAUGE;
    var cor = esmaecido ? "#B0B0B0" : (CORES_CATEGORIA_DESEMPENHO_JS[categoria] || "#8C8C8C");
    var corTrilho = "var(--cor-card-elevado)";
    var svg = '<svg width="100%" viewBox="0 0 ' + largura + ' ' + altura + '">';
    svg += _arcoGauge(cx, cy, raio, ANGULO_INICIAL_GAUGE, ANGULO_INICIAL_GAUGE - VARREDURA_GAUGE, espessura, corTrilho);
    if (raio * ((ANGULO_INICIAL_GAUGE - anguloValor) * Math.PI / 180) > espessura * 1.5) {
      svg += _arcoGauge(cx, cy, raio, ANGULO_INICIAL_GAUGE, anguloValor, espessura, cor);
    }
    var marcador = _pontoGauge(cx, cy, raio, anguloValor);
    svg += '<circle cx="' + marcador.x + '" cy="' + marcador.y + '" r="' + (espessura * 0.86) +
      '" fill="var(--cor-card)"></circle>';
    svg += '<circle cx="' + marcador.x + '" cy="' + marcador.y + '" r="' + (espessura * 0.58) +
      '" fill="' + cor + '"></circle>';
    var corNumero = esmaecido ? "#9A9A9A" : "var(--cor-texto-primario)";
    var fonteNumero = Math.min(raio * 0.85, largura * 0.34);
    svg += '<text x="' + cx + '" y="' + (cy + fonteNumero * 0.36) + '" text-anchor="middle" font-size="' +
      fonteNumero + '" font-weight="700" fill="' + corNumero + '">' + Math.round(indice) + '</text>';

    // A categoria vai DENTRO do SVG, e não num elemento HTML embaixo do
    // gráfico. O SVG escala junto com a largura do card, então só assim a
    // proporção entre a nota e a categoria fica travada em qualquer
    // tamanho de tela — com a categoria em pixels fixos, ela encolhia em
    // relação ao número conforme o card crescia.
    //
    // A largura do texto é estimada (0,46 por caractere e por ponto de
    // fonte, aferido na Helvetica negrito) porque o SVG ainda não foi
    // medido pelo navegador nesse momento. A estimativa erra para mais, o
    // que é o lado seguro: no pior caso a fonte fica um pouco menor que o
    // necessário, nunca transbordando.
    var fonteCategoria = Math.min(
      fonteNumero * 0.38,
      (largura * 0.94) / Math.max(1, categoria.length * 0.46)
    );
    svg += '<text x="' + cx + '" y="' + (altura - espacoCategoria * 0.30) +
      '" text-anchor="middle" font-size="' + fonteCategoria + '" font-weight="700" fill="' +
      cor + '">' + categoria + '</text>';
    return svg + '</svg>';
  }

  function montarIndiceDesempenho(grupos) {
    var container = document.getElementById("preview-desempenho-container");
    if (!container) return;
    container.innerHTML = "";
    if (!grupos || grupos.length === 0) {
      container.innerHTML = '<div class="preview-sem-dados">Nenhuma combinação Secretaria/Executor com dados suficientes no filtro atual</div>';
      return;
    }
    grupos.forEach(function (grupo) {
      var subtitulo = document.createElement("div");
      subtitulo.className = "preview-gestao-subtitulo";
      subtitulo.textContent = grupo.gestao;
      container.appendChild(subtitulo);

      var grid = document.createElement("div");
      grid.className = "preview-secretaria-grid";
      grupo.itens.forEach(function (item, indiceOrdem) {
        var card = document.createElement("div");
        card.className = "preview-mini-card preview-mini-card-gauge" + (item.esmaecido ? " preview-mini-card-esmaecido" : "");
        if (item.esmaecido) {
          card.title = "Essa combinação Secretaria | Executor não aparece no restante do relatório com os filtros atuais — o índice continua calculado com a base completa.";
        }
        var titulo = document.createElement("div");
        titulo.className = "titulo-sec-exec";
        titulo.textContent = (indiceOrdem + 1) + "º — " + item.rotulo;
        card.appendChild(titulo);

        var corpoGauge = document.createElement("div");
        corpoGauge.className = "gauge-desempenho-corpo";
        corpoGauge.innerHTML = criarGaugeDesempenho(item.indice, item.esmaecido, item.categoria);
        card.appendChild(corpoGauge);

        grid.appendChild(card);
      });
      container.appendChild(grid);
    });
  }

  function montarPanoramaSecretaria(grupos) {
    var container = document.getElementById("preview-secretaria-container");
    container.innerHTML = "";
    if (!grupos || grupos.length === 0) {
      container.innerHTML = '<div class="preview-sem-dados">Nenhuma combinação Secretaria/Executor no filtro atual</div>';
      return;
    }
    grupos.forEach(function (grupo) {
      var subtitulo = document.createElement("div");
      subtitulo.className = "preview-gestao-subtitulo";
      subtitulo.textContent = grupo.gestao;
      container.appendChild(subtitulo);

      var grid = document.createElement("div");
      grid.className = "preview-secretaria-grid";

      grupo.itens.forEach(function (item) {
        var card = document.createElement("div");
        card.className = "preview-mini-card";
        var titulo = document.createElement("div");
        titulo.className = "titulo-sec-exec";
        titulo.textContent = item.secretaria + " | " + item.executor;
        card.appendChild(titulo);

        var temObjetos = item.fases.some(function (f) { return f.objetos && f.objetos.length > 0; });
        if (temObjetos) {
          var iconeDetalhe = document.createElement("button");
          iconeDetalhe.type = "button";
          iconeDetalhe.className = "mini-card-icone-detalhe";
          iconeDetalhe.title = "Ver objetos por fase";
          iconeDetalhe.innerHTML =
            '<svg viewBox="0 0 24 24" width="13" height="13" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round">' +
              '<circle cx="12" cy="12" r="9"></circle>' +
              '<line x1="12" y1="11" x2="12" y2="16.5"></line>' +
              '<circle cx="12" cy="8" r="0.5" fill="currentColor" stroke="none"></circle>' +
            '</svg>';
          // Guarda os dados do card no próprio HTML (atributo), não só na
          // função JS em memória — assim, quando o painel "Compartilhar"
          // copia esse ícone pra um arquivo isolado, o dado vai junto (um
          // clone de elemento preserva atributos, mas não os eventos JS
          // presos via addEventListener).
          iconeDetalhe.setAttribute("data-item", JSON.stringify(item));
          ativarTooltipCard(iconeDetalhe, item);
          card.appendChild(iconeDetalhe);
        }

        var graficoDiv = document.createElement("div");
        graficoDiv.className = "mini-card-corpo";
        card.appendChild(graficoDiv);
        criarGraficoPizzaEm(graficoDiv, item.fases, CORES_FASE, { tamanho: 68, semRotulo: true });

        grid.appendChild(card);
      });

      container.appendChild(grid);
    });
  }

  // Mapa coroplético via SVG puro — os polígonos já vêm projetados e
  // coloridos do Python (mesma lógica usada para gerar o Drawing do PDF),
  // o JS só desenha.
  // Handlers de arrasto que a última montagem do mapa deixou registrados na
  // janela — guardados para poderem ser removidos na montagem seguinte.
  var _arrastoMapaRegistrado = null;

  function montarMapa(mapaDados) {
    var container = document.getElementById("preview-mapa");
    container.innerHTML = "";
    if (!mapaDados) {
      container.innerHTML = '<div class="preview-sem-dados">Arquivo do mapa (municipios_bahia.geojson) não encontrado.</div>';
      return;
    }

    var wrap = document.createElement("div");
    wrap.className = "preview-mapa-wrap";

    var svgNS = "http://www.w3.org/2000/svg";
    var svg = document.createElementNS(svgNS, "svg");
    // Estado do zoom/pan — guardado à parte do viewBox original, que fica
    // reservado como o "zoom 100%" pro botão de resetar.
    var viewBoxOriginal = { x: 0, y: 0, w: mapaDados.largura, h: mapaDados.altura };
    var viewBoxAtual = { x: 0, y: 0, w: mapaDados.largura, h: mapaDados.altura };
    function aplicarViewBox() {
      svg.setAttribute(
        "viewBox",
        viewBoxAtual.x.toFixed(2) + " " + viewBoxAtual.y.toFixed(2) + " " +
          viewBoxAtual.w.toFixed(2) + " " + viewBoxAtual.h.toFixed(2)
      );
    }
    svg.setAttribute("width", mapaDados.largura);
    svg.setAttribute("height", mapaDados.altura);
    svg.classList.add("preview-mapa-svg");
    aplicarViewBox();

    mapaDados.poligonos.forEach(function (poly) {
      var el = document.createElementNS(svgNS, "polygon");
      el.setAttribute("points", poly.pontos);
      el.setAttribute("fill", poly.cor);
      if (poly.objetos) {
        el.classList.add("preview-mapa-poligono-com-dado");
        ativarTooltipMunicipio(el, poly);
      }
      svg.appendChild(el);
    });
    wrap.appendChild(svg);

    // --- Zoom com a roda do mouse, centralizado no cursor ---
    var ZOOM_MIN = 1, ZOOM_MAX = 12;
    var zoomAtual = 1;
    svg.addEventListener("wheel", function (ev) {
      ev.preventDefault();
      var fator = ev.deltaY < 0 ? 1.15 : 1 / 1.15;
      var novoZoom = Math.min(ZOOM_MAX, Math.max(ZOOM_MIN, zoomAtual * fator));
      fator = novoZoom / zoomAtual;
      zoomAtual = novoZoom;

      var rect = svg.getBoundingClientRect();
      // Ponto do cursor, convertido de coordenadas de tela pra coordenadas
      // do viewBox atual — é em torno DELE que o zoom cresce/encolhe.
      var pxRel = (ev.clientX - rect.left) / rect.width;
      var pyRel = (ev.clientY - rect.top) / rect.height;
      var pontoX = viewBoxAtual.x + pxRel * viewBoxAtual.w;
      var pontoY = viewBoxAtual.y + pyRel * viewBoxAtual.h;

      var novaLargura = viewBoxOriginal.w / zoomAtual;
      var novaAltura = viewBoxOriginal.h / zoomAtual;
      viewBoxAtual = {
        x: pontoX - pxRel * novaLargura,
        y: pontoY - pyRel * novaAltura,
        w: novaLargura,
        h: novaAltura,
      };
      aplicarViewBox();
    }, { passive: false });

    // --- Pan arrastando o mouse (só ativa depois de já ter dado zoom) ---
    var arrastando = false, ultimoX = 0, ultimoY = 0;
    svg.addEventListener("mousedown", function (ev) {
      if (zoomAtual <= 1) return;
      arrastando = true;
      ultimoX = ev.clientX;
      ultimoY = ev.clientY;
      svg.classList.add("preview-mapa-arrastando");
    });
    // Os dois listeners abaixo ficam na JANELA (e não no SVG), porque o
    // arrasto precisa continuar funcionando mesmo com o ponteiro saindo do
    // mapa. Só que montarMapa roda de novo a cada recarga do dashboard —
    // ou seja, a cada mudança de filtro — e sem remover os anteriores eles
    // se acumulavam: depois de vinte filtros, todo movimento do mouse
    // disparava vinte handlers, cada um mexendo no viewBox de um SVG
    // antigo que já nem está mais na tela. A tela continuava certa (só o
    // mapa atual aparece), mas o painel ia ficando pesado ao longo do uso.
    function moverArrastandoMapa(ev) {
      if (!arrastando) return;
      var rect = svg.getBoundingClientRect();
      var dx = (ev.clientX - ultimoX) * (viewBoxAtual.w / rect.width);
      var dy = (ev.clientY - ultimoY) * (viewBoxAtual.h / rect.height);
      viewBoxAtual.x -= dx;
      viewBoxAtual.y -= dy;
      ultimoX = ev.clientX;
      ultimoY = ev.clientY;
      aplicarViewBox();
    }
    function soltarArrastoMapa() {
      arrastando = false;
      svg.classList.remove("preview-mapa-arrastando");
    }
    if (_arrastoMapaRegistrado) {
      window.removeEventListener("mousemove", _arrastoMapaRegistrado.mover);
      window.removeEventListener("mouseup", _arrastoMapaRegistrado.soltar);
    }
    window.addEventListener("mousemove", moverArrastandoMapa);
    window.addEventListener("mouseup", soltarArrastoMapa);
    _arrastoMapaRegistrado = { mover: moverArrastandoMapa, soltar: soltarArrastoMapa };

    // --- Controles de zoom (botões, pra quem não tem roda de mouse) ---
    function zoomCentralizado(fator) {
      var novoZoom = Math.min(ZOOM_MAX, Math.max(ZOOM_MIN, zoomAtual * fator));
      fator = novoZoom / zoomAtual;
      zoomAtual = novoZoom;
      var centroX = viewBoxAtual.x + viewBoxAtual.w / 2;
      var centroY = viewBoxAtual.y + viewBoxAtual.h / 2;
      var novaLargura = viewBoxOriginal.w / zoomAtual;
      var novaAltura = viewBoxOriginal.h / zoomAtual;
      viewBoxAtual = {
        x: centroX - novaLargura / 2,
        y: centroY - novaAltura / 2,
        w: novaLargura,
        h: novaAltura,
      };
      aplicarViewBox();
    }
    var controles = document.createElement("div");
    controles.className = "preview-mapa-controles";
    var btnMais = document.createElement("button");
    btnMais.type = "button";
    btnMais.className = "preview-mapa-botao-zoom";
    btnMais.textContent = "+";
    btnMais.title = "Aumentar zoom";
    btnMais.addEventListener("click", function () { zoomCentralizado(1.4); });
    var btnMenos = document.createElement("button");
    btnMenos.type = "button";
    btnMenos.className = "preview-mapa-botao-zoom";
    btnMenos.textContent = "\u2212";
    btnMenos.title = "Diminuir zoom";
    btnMenos.addEventListener("click", function () { zoomCentralizado(1 / 1.4); });
    var btnReset = document.createElement("button");
    btnReset.type = "button";
    btnReset.className = "preview-mapa-botao-zoom preview-mapa-botao-zoom-reset";
    btnReset.textContent = "\u21BA";
    btnReset.title = "Redefinir zoom";
    btnReset.addEventListener("click", function () {
      zoomAtual = 1;
      viewBoxAtual = { x: viewBoxOriginal.x, y: viewBoxOriginal.y, w: viewBoxOriginal.w, h: viewBoxOriginal.h };
      aplicarViewBox();
    });
    controles.appendChild(btnMais);
    controles.appendChild(btnMenos);
    controles.appendChild(btnReset);
    wrap.appendChild(controles);

    var legenda = document.createElement("div");
    legenda.className = "preview-mapa-legenda";
    var itensLegenda = mapaDados.legenda.slice().reverse();
    itensLegenda.forEach(function (fx) {
      var linha = document.createElement("div");
      linha.className = "preview-mapa-legenda-item";
      linha.innerHTML =
        '<span class="preview-mapa-legenda-swatch" style="background:' + fx.cor + ';"></span>' +
        '<span>' + fx.de + ' a ' + fx.ate + '</span>';
      legenda.appendChild(linha);
    });
    var linhaSemDado = document.createElement("div");
    linhaSemDado.className = "preview-mapa-legenda-item";
    linhaSemDado.innerHTML =
      '<span class="preview-mapa-legenda-swatch" style="background:' + mapaDados.cor_sem_dado + ';"></span>' +
      '<span>Sem ação no filtro atual</span>';
    legenda.appendChild(linhaSemDado);
    var linhaResumo = document.createElement("div");
    linhaResumo.style.cssText = "margin-top:8px;color:var(--cor-texto-secundario);font-size:11px;";
    linhaResumo.textContent = mapaDados.qtd_municipios_com_dado + " de " + mapaDados.qtd_municipios_total + " municípios com ao menos uma ação.";
    legenda.appendChild(linhaResumo);
    wrap.appendChild(legenda);

    container.appendChild(wrap);
  }

  function montarPreVisualizacao(dados) {
    var nota = document.getElementById("preview-nota-secretaria");
    if (dados.secretaria_unica) {
      nota.style.display = "block";
      nota.innerHTML =
        "Filtro atual restrito a uma única Secretaria/Órgão e Executor: " +
        "<b>" + dados.secretaria_unica.secretaria + "</b> | <b>" + dados.secretaria_unica.executor + "</b>";
    } else {
      nota.style.display = "none";
      nota.innerHTML = "";
    }

    var cards = document.getElementById("preview-cards");
    cards.innerHTML =
      '<div class="preview-card-resumo"><div class="rotulo">QTD Total</div><div class="valor">' + dados.qtd + '</div></div>' +
      '<div class="preview-card-resumo"><div class="rotulo">Investimento</div><div class="valor">' + formatarMiBi(dados.investimento) + '</div></div>';

    criarGraficoBarras("preview-grafico-prazo", dados.prazo);
    criarGraficoPizza("preview-grafico-financeiro", dados.financeiro, CORES_FINANCEIRO);
    criarGraficoPizza("preview-grafico-panorama", dados.panorama, CORES_FASE);
    criarGraficoPizza("preview-grafico-termo", dados.termo, {
      "Termo Assinado": COR_VERDE,
      "Termo Não Assinado": COR_VERMELHO,
    });
    montarPanoramaSecretaria(dados.secretaria);
    montarIndiceDesempenho(dados.desempenho);
    montarDetalhamentoSecretaria(dados.detalhamentoSecretaria);
    montarMapa(dados.mapa);
  }

  // Monta a tabela de detalhamento financeiro por Fonte de Recurso — uma
  // linha por fonte (CONTRAPARTIDA, OGU, FINANCIAMENTO etc.), com
  // quantidade, investimento, valor contratado, recurso estadual e
  // financiamento, ordenada da maior pra menor investimento.
  function montarDetalhamentoSecretaria(lista) {
    var container = document.getElementById("preview-detalhamento-secretaria");
    container.innerHTML = "";
    if (!lista || lista.length === 0) {
      container.innerHTML = '<div class="preview-sem-dados">Sem dados no filtro atual</div>';
      return;
    }

    var legenda = document.createElement("div");
    legenda.className = "grafico-secretaria-legenda";
    legenda.innerHTML =
      '<span class="grafico-secretaria-legenda-item"><span class="grafico-secretaria-legenda-swatch" style="background:' + COR_VERDE + ';"></span>Valor Contratado</span>' +
      '<span class="grafico-secretaria-legenda-item"><span class="grafico-secretaria-legenda-swatch" style="background:' + COR_AZUL + ';"></span>Valor Apoiado OGU</span>' +
      '<span class="grafico-secretaria-legenda-item"><span class="grafico-secretaria-legenda-swatch" style="background:' + COR_AMARELO + ';"></span>Recurso Estadual</span>' +
      '<span class="grafico-secretaria-legenda-item"><span class="grafico-secretaria-legenda-swatch" style="background:' + COR_VERMELHO + ';"></span>Financiamento</span>';
    container.appendChild(legenda);

    var wrap = document.createElement("div");
    wrap.className = "grafico-secretaria-wrap";
    lista.forEach(function (item) {
      var linha = document.createElement("div");
      linha.className = "grafico-secretaria-linha";
      var base = item.investimento > 0 ? item.investimento : 1;
      var segmentos = [
        { valor: item.valorContratado, cor: COR_VERDE },
        { valor: item.valorApoiadoOgu, cor: COR_AZUL },
        { valor: item.recursoEstadual, cor: COR_AMARELO },
        { valor: item.financiamento, cor: COR_VERMELHO },
      ];
      var segmentosHtml = segmentos
        .filter(function (s) { return s.valor > 0; })
        .map(function (s) {
          var pctSegmento = (s.valor / base) * 100;
          return '<div class="grafico-secretaria-segmento" style="width:' + pctSegmento + '%;background:' + s.cor + ';"></div>';
        })
        .join("");
      linha.innerHTML =
        '<div class="grafico-secretaria-rotulo">' + item.secretaria + ' <span class="grafico-secretaria-qtd">(' + item.qtd + ')</span></div>' +
        '<div class="grafico-secretaria-barra-trilha"><div class="grafico-secretaria-barra-empilhada">' + segmentosHtml + '</div></div>' +
        '<div class="grafico-secretaria-valor">' + formatarMiBi(item.investimento) + '</div>';
      // Guarda o detalhamento no próprio HTML (mesmo padrão do ícone dos
      // mini-cards) — assim o tooltip sobrevive ao arquivo "Compartilhar".
      linha.setAttribute("data-secretaria", JSON.stringify(item));
      ativarTooltipSecretaria(linha, item);
      wrap.appendChild(linha);
    });
    container.appendChild(wrap);
  }

  function ativarTooltipSecretaria(linha, item) {
    linha.addEventListener("mouseenter", function () {
      var tooltip = obterTooltipObjetos();
      tooltip.innerHTML =
        '<div class="tooltip-objetos-titulo">' + item.secretaria + '</div>' +
        '<div class="tooltip-objetos-item"><span class="tooltip-objetos-nome">Valor Contratado</span><span class="tooltip-objetos-valor">' + formatarMiBi(item.valorContratado) + '</span></div>' +
        '<div class="tooltip-objetos-item"><span class="tooltip-objetos-nome">Valor Apoiado OGU</span><span class="tooltip-objetos-valor">' + formatarMiBi(item.valorApoiadoOgu) + '</span></div>' +
        '<div class="tooltip-objetos-item"><span class="tooltip-objetos-nome">Recurso Estadual</span><span class="tooltip-objetos-valor">' + formatarMiBi(item.recursoEstadual) + '</span></div>' +
        '<div class="tooltip-objetos-item"><span class="tooltip-objetos-nome">Financiamento</span><span class="tooltip-objetos-valor">' + formatarMiBi(item.financiamento) + '</span></div>';
      tooltip.classList.add("visivel");
      posicionarTooltipObjetos(linha);
    });
    linha.addEventListener("mouseleave", function () {
      var tooltip = obterTooltipObjetos();
      setTimeout(function () {
        if (!tooltip.matches(":hover") && !linha.matches(":hover")) {
          tooltip.classList.remove("visivel");
        }
      }, 50);
    });
  }

  // Botão "Publicar Atualização" — só existe no modo desktop (roda git de
  // verdade na máquina de quem clica; não tem equivalente no site, por
  // segurança, ver os comentários em _git_publicar_atualizacao no lado
  // Python). Fica escondido e sem ação nenhuma quando o painel está rodando
  // no servidor web. Existe um botão dentro de cada menu de Acesso Rápido
  // (painel de filtros e dash) — os dois chamam a mesma ação.
  var botoesPublicar = [
    document.getElementById("acesso-rapido-filtros-publicar"),
    document.getElementById("acesso-rapido-dash-publicar"),
  ];
  if (!window.PAC_MODO_WEB) {
    botoesPublicar.forEach(function (btnPublicar) {
      btnPublicar.style.display = "";
      btnPublicar.onclick = async function () {
        botoesPublicar.forEach(function (b) { b.disabled = true; });

        var verificacao;
        try {
          verificacao = await window.pywebview.api.verificar_mudancas_git();
        } catch (erro) {
          alert("Não foi possível verificar o que mudou:" + NL + NL + erro);
          botoesPublicar.forEach(function (b) { b.disabled = false; });
          return;
        }
        if (!verificacao || verificacao.ok === false) {
          alert(verificacao ? verificacao.erro : "Erro desconhecido ao verificar o que mudou.");
          botoesPublicar.forEach(function (b) { b.disabled = false; });
          return;
        }
        if (!verificacao.mudancas || verificacao.mudancas.length === 0) {
          alert("Nada para publicar — a base local já está igual ao site.");
          botoesPublicar.forEach(function (b) { b.disabled = false; });
          return;
        }

        var confirmar = window.confirm(
          "Isto vai publicar as mudanças abaixo no repositório PÚBLICO do GitHub, " +
          "atualizando o link do painel web em alguns minutos:" + NL + NL +
          verificacao.mudancas.join(NL) + NL + NL + "Confirma?"
        );
        if (!confirmar) {
          botoesPublicar.forEach(function (b) { b.disabled = false; });
          return;
        }

        var mensagem = window.prompt("Descreva rapidamente o que mudou (opcional):", "") || "";

        var resultadoPublicacao;
        try {
          resultadoPublicacao = await window.pywebview.api.publicar_atualizacao_git(mensagem);
        } catch (erro) {
          alert("Ocorreu um erro ao publicar:" + NL + NL + erro);
          botoesPublicar.forEach(function (b) { b.disabled = false; });
          return;
        }
        botoesPublicar.forEach(function (b) { b.disabled = false; });

        if (resultadoPublicacao && resultadoPublicacao.ok) {
          if (resultadoPublicacao.nada_a_publicar) {
            alert("Nada para publicar — a base local já está igual ao site.");
          } else {
            alert("Publicado com sucesso! O link do painel web vai se atualizar sozinho em alguns minutos.");
          }
        } else {
          alert("Não foi possível publicar:" + NL + NL + (resultadoPublicacao ? resultadoPublicacao.erro : "erro desconhecido"));
        }
      };
    });
  }

  var btnPreview = document.getElementById("btn-preview");

  // Busca os dados atualizados e (re)desenha o dashboard — chamada tanto no
  // carregamento inicial da página (o dashboard agora É a tela inicial)
  // quanto sempre que os filtros mudam através de um dos atalhos no topo
  // do próprio dashboard (FILTROS/GERENCIAL/LIMPAR), ou ao fechar o painel
  // de filtros voltando pro dashboard.
  async function carregarDashboard() {
    // A espera pelo evento "pywebviewready" só faz sentido no modo desktop
    // (window.pywebview é injetado de forma assíncrona pelo WebView2 pouco
    // depois do carregamento da página). No modo web (window.PAC_MODO_WEB)
    // esse objeto nunca vai existir — nem precisa, porque chamarAPI() já
    // sabe falar direto com o servidor por HTTP.
    if (!window.PAC_MODO_WEB && !window.pywebview) {
      await new Promise(function (resolve) {
        var jaResolveu = false;
        var finalizar = function () { if (!jaResolveu) { jaResolveu = true; resolve(); } };
        window.addEventListener("pywebviewready", finalizar, { once: true });
        setTimeout(finalizar, 5000);
      });
      if (!window.pywebview) {
        alert("O painel ainda não terminou de carregar — tente novamente em instantes.");
        return;
      }
    }

    var filtros = montarFiltrosAtuais();
    var cards = document.getElementById("preview-cards");
    cards.innerHTML = '<div class="preview-card-resumo"><div class="rotulo">Carregando...</div><div class="valor">—</div></div>';

    var resultado;
    try {
      resultado = await chamarAPI("pre_visualizar", filtros);
    } catch (erro) {
      alert("Ocorreu um erro ao carregar o dashboard:" + NL + NL + erro);
      return;
    }

    if (!resultado || resultado.ok === false) {
      atualizarBotaoQualidade(null);
      if (resultado && resultado.vazio) {
        alert("Nenhum registro encontrado para os filtros selecionados.");
      } else {
        alert("Ocorreu um erro ao carregar o dashboard:" + NL + NL + (resultado ? resultado.erro : "erro desconhecido"));
      }
      return;
    }

    montarPreVisualizacao(resultado.dados);
    atualizarBotaoQualidade(resultado.aviso);
    document.getElementById("preview-overlay").style.display = "flex";
  }

  // Botão "DASHBOARD" DENTRO do painel de filtros: agora fecha o painel
  // (que passou a ser um overlay por cima do dashboard) e atualiza o
  // dashboard por trás, para refletir os filtros que acabaram de mudar.
  btnPreview.onclick = async function () {
    document.getElementById("app").style.display = "none";
    montarFiltroGestaoDash();
    montarFiltroSecretariaDash();
    await carregarDashboard();
  };

  // Botão "MAPA MENTAL" (dentro do menu de Acesso Rápido, nas duas barras):
  // pede ao backend a página autônoma do mapa mental (mesmo recorte dos
  // filtros atuais) e mostra numa página dedicada, DENTRO do mesmo ambiente
  // (um <iframe> em tela cheia por cima do painel — ver #mapa-mental-overlay),
  // em vez de abrir aba/janela separada. Um window.open com Blob de HTML
  // chegou a ser usado aqui, mas no modo desktop (WebView2/pywebview) isso
  // faz o Windows tratar o link como um arquivo para "abrir com outro app"
  // em vez de exibir a página — por isso o iframe embutido, que nunca sai
  // da janela do programa.
  var botoesMapaMental = [
    document.getElementById("acesso-rapido-filtros-mapa-mental"),
    document.getElementById("acesso-rapido-dash-mapa-mental"),
  ];
  var mapaMentalOverlay = document.getElementById("mapa-mental-overlay");
  var mapaMentalIframe = document.getElementById("mapa-mental-iframe");

  botoesMapaMental.forEach(function (btnMapaMental) {
    btnMapaMental.onclick = async function () {
      if (!window.PAC_MODO_WEB && !window.pywebview) {
        await new Promise(function (resolve) {
          var jaResolveu = false;
          var finalizar = function () { if (!jaResolveu) { jaResolveu = true; resolve(); } };
          window.addEventListener("pywebviewready", finalizar, { once: true });
          setTimeout(finalizar, 5000);
        });
        if (!window.pywebview) {
          alert("O painel ainda não terminou de carregar — tente novamente em instantes.");
          return;
        }
      }

      var filtros = montarFiltrosAtuais();
      // o mapa mental é um documento à parte (iframe srcdoc) sem acesso ao
      // localStorage desta página — manda o tema atual junto pra ele nascer
      // já na cor certa (ver meta["tema"] em _api_mapa_mental)
      filtros.tema = obterTemaAtual();
      botoesMapaMental.forEach(function (b) { b.disabled = true; });
      var resultado;
      try {
        resultado = await chamarAPI("mapa_mental", filtros);
      } catch (erro) {
        alert("Ocorreu um erro ao abrir o mapa mental:" + NL + NL + erro);
        botoesMapaMental.forEach(function (b) { b.disabled = false; });
        return;
      }
      botoesMapaMental.forEach(function (b) { b.disabled = false; });

      if (!resultado || resultado.ok === false) {
        if (resultado && resultado.vazio) {
          alert("Nenhum registro encontrado para os filtros selecionados.");
        } else {
          alert("Ocorreu um erro ao abrir o mapa mental:" + NL + NL + (resultado ? resultado.erro : "erro desconhecido"));
        }
        return;
      }

      mapaMentalIframe.srcdoc = resultado.html;
      mapaMentalOverlay.style.display = "flex";
    };
  });

  // Botão de fechar mora dentro do próprio iframe (mesma linha do título
  // "... Mapa Mental", ver #fechar-mapa em mapa_mental_html.py) — como
  // ele não pode chamar direto o overlay da página pai (documentos
  // isolados), avisa por postMessage.
  window.addEventListener("message", function (ev) {
    if (ev.data && ev.data.tipo === "mapa-mental-fechar") {
      mapaMentalOverlay.style.display = "none";
      mapaMentalIframe.srcdoc = "";
    }
  });

  // Botão "FILTROS" (dentro do menu de Acesso Rápido, nas duas barras):
  // abre o painel de filtros maximizado, por cima do dashboard. No próprio
  // painel de filtros não tem efeito visível (já está aberto) — existe ali
  // só pra manter o mesmo menu nas duas telas.
  [
    document.getElementById("acesso-rapido-filtros-filtros"),
    document.getElementById("acesso-rapido-dash-filtros"),
  ].forEach(function (btnFiltros) {
    btnFiltros.onclick = function () {
      document.getElementById("app").style.display = "flex";
    };
  });

  // Menu de Acesso Rápido: o botão de ícone (grade 2x2) alterna a exibição
  // do menu com MAPA MENTAL/FILTROS/PUBLICAR; clicar num item do menu ou
  // fora dele fecha de volta. Mesmo componente nas duas barras.
  function configurarAcessoRapido(idWrap) {
    var wrap = document.getElementById(idWrap);
    var botao = wrap.querySelector(".botao-icone-topo");
    var menu = wrap.querySelector(".acesso-rapido-menu");
    function fechar() {
      menu.classList.remove("aberto");
      botao.setAttribute("aria-expanded", "false");
    }
    botao.onclick = function (e) {
      e.stopPropagation();
      var vaiAbrir = !menu.classList.contains("aberto");
      document.querySelectorAll(".acesso-rapido-menu.aberto").forEach(function (m) {
        m.classList.remove("aberto");
      });
      document.querySelectorAll(".botao-icone-topo[aria-expanded=\"true\"]").forEach(function (b) {
        b.setAttribute("aria-expanded", "false");
      });
      if (vaiAbrir) {
        menu.classList.add("aberto");
        botao.setAttribute("aria-expanded", "true");
      }
    };
    menu.addEventListener("click", function (e) {
      if (e.target.closest(".acesso-rapido-item")) fechar();
    });
  }
  configurarAcessoRapido("acesso-rapido-filtros-wrap");
  configurarAcessoRapido("acesso-rapido-dash-wrap");
  document.addEventListener("click", function (e) {
    document.querySelectorAll(".acesso-rapido-wrap").forEach(function (wrap) {
      if (!wrap.contains(e.target)) {
        wrap.querySelector(".acesso-rapido-menu").classList.remove("aberto");
        wrap.querySelector(".botao-icone-topo").setAttribute("aria-expanded", "false");
      }
    });
  });

  // Chave de tema claro/escuro: alterna o atributo data-tema na <html>
  // (é ele que troca as variáveis --cor-* do :root, ver <style> no topo
  // deste HTML), lembra a escolha entre sessões (localStorage) e mantém as
  // duas chaves (barra de filtros e barra do dash) sincronizadas entre si.
  // O tema atual também vai junto quando o Mapa Mental é aberto (ver
  // filtros.tema logo abaixo, no botão MAPA MENTAL), já que ele roda num
  // <iframe> à parte e não teria como puxar isso do localStorage sozinho.
  function obterTemaAtual() {
    return document.documentElement.getAttribute("data-tema") === "claro" ? "claro" : "escuro";
  }
  function aplicarTema(tema) {
    if (tema === "claro") {
      document.documentElement.setAttribute("data-tema", "claro");
    } else {
      document.documentElement.removeAttribute("data-tema");
    }
    document.querySelectorAll(".tema-switch").forEach(function (chave) {
      chave.setAttribute("aria-checked", tema === "claro" ? "true" : "false");
    });
    try { localStorage.setItem("cgape-tema", tema); } catch (e) {}
  }
  document.querySelectorAll(".tema-switch").forEach(function (chave) {
    chave.onclick = function () {
      aplicarTema(obterTemaAtual() === "claro" ? "escuro" : "claro");
    };
  });
  aplicarTema(obterTemaAtual()); // sincroniza aria-checked das duas chaves com o que o <script> do <head> já aplicou

  // Botão "GERENCIAL" no topo do dashboard: mesma seleção rápida de STATUS
  // do botão "Gerencial" de dentro do painel de filtros, só que acionável
  // direto do dashboard, sem precisar abrir os filtros — e já atualiza o
  // dashboard com o resultado.
  document.getElementById("dash-gerencial").onclick = async function () {
    var blocoStatus = DADOS.blocos.find(b => b.chave === "STATUS");
    estadoSelecao.STATUS = new Set(DADOS.statusPadrao.filter(op => blocoStatus.opcoes.includes(op)));
    renderizarBloco("STATUS");
    atualizarDisponibilidade();
    await carregarDashboard();
  };

  // Botão "LIMPAR TODOS OS FILTROS" no topo do dashboard: mesma função do
  // botão de dentro do painel de filtros, acionável direto do dashboard —
  // e já atualiza o dashboard com o resultado.
  document.getElementById("dash-limpar").onclick = async function () {
    limparTodosFiltros();
    await carregarDashboard();
  };

  const btnGerar = document.getElementById("btn-gerar");
  const previewGerarTopo = document.getElementById("preview-gerar-topo");
  const botoesGerar = [btnGerar, previewGerarTopo].filter(Boolean);
  const textoOriginalBtnGerar = btnGerar.textContent;

  function travarBotaoGerar(texto) {
    botoesGerar.forEach(function (b) {
      b.disabled = true;
      b.textContent = texto;
    });
  }
  function destravarBotaoGerar() {
    botoesGerar.forEach(function (b) {
      b.disabled = false;
      b.textContent = textoOriginalBtnGerar;
    });
  }

  // Modal do Controle de Qualidade: mesmo comportamento de antes (botao OK
  // fica bloqueado por 10s com contagem regressiva, forcando a leitura do
  // aviso antes de prosseguir), so que como um modal HTML em vez de uma
  // janela Tkinter separada.
  function montarHtmlAvisoQualidade(aviso, somenteVisualizar) {
    var html = '<div class="qc-cabecalho">' + escaparHtmlFicha(aviso.cabecalho).replace(/\n/g, "<br>") + '</div>';
    aviso.grupos.forEach(function (grupo) {
      html += '<div class="qc-grupo-titulo">— ' + escaparHtmlFicha(grupo.gestao) + ' —</div>';
      grupo.linhas.forEach(function (linha) {
        // Um bloco por OBJETO; dentro dele, um subtópico por tipo de
        // pendência, com os itens que têm aquela pendência específica.
        var qtdItens = linha.itens.length;
        var resumoItens = qtdItens === 1 ? "1 item" : qtdItens + " itens";
        html += '<div class="qc-objeto">';
        html += '<div class="qc-objeto-titulo">• ' + escaparHtmlFicha(linha.objeto)
             + ' <span class="qc-objeto-contagem">(' + resumoItens + ')</span></div>';
        (linha.alertas || []).forEach(function (alerta) {
          var itensHtml;
          if (alerta.itens.length > 0) {
            var rotuloItem = alerta.itens.length > 1 ? "Itens" : "Item";
            var linksItens = alerta.itens.map(function (it) {
              return '<a href="#" class="qc-item-link" data-item="' + escaparHtmlFicha(it) + '" title="Abrir ficha cadastral">' + escaparHtmlFicha(it) + '</a>';
            }).join(", ");
            itensHtml = rotuloItem + " " + linksItens;
          } else {
            itensHtml = "Item não informado";
          }
          html += '<div class="qc-alerta">'
               + '<div class="qc-alerta-motivo">' + escaparHtmlFicha(alerta.motivo) + '</div>'
               + '<div class="qc-alerta-itens">' + itensHtml + '</div>'
               + '</div>';
        });
        html += '</div>';
      });
      if (grupo.qtd_restante > 0) {
        var textoResto = grupo.qtd_restante === 1 ? "objeto" : "objetos";
        html += '<div class="qc-linha">... e mais ' + grupo.qtd_restante + ' ' + textoResto + ' com prazo vencido.</div>';
      }
    });
    if (somenteVisualizar) {
      html += '<div class="qc-instrucao">Clique num item para abrir a Ficha Cadastral e corrigir a pendência na planilha.</div>';
    } else {
      html += '<div class="qc-instrucao">Clique em OK para gerar o relatório normalmente, ou em Cancelar para interromper a geração.</div>';
    }
    return html;
  }

  var _avisoQualidadeAtual = null;

  function montarTextoCompartilhamentoQualidade(aviso) {
    var linhas = ["*PAC - Controle de Qualidade da Base de Dados*", "", aviso.cabecalho];
    aviso.grupos.forEach(function (grupo) {
      linhas.push("");
      linhas.push("— " + grupo.gestao + " —");
      grupo.linhas.forEach(function (linha) {
        var qtdItens = linha.itens.length;
        linhas.push("• *" + linha.objeto + "* (" + (qtdItens === 1 ? "1 item" : qtdItens + " itens") + ")");
        (linha.alertas || []).forEach(function (alerta) {
          var itensTexto = alerta.itens.length > 0
            ? (alerta.itens.length > 1 ? "Itens " : "Item ") + alerta.itens.join(", ")
            : "Item não informado";
          linhas.push("   - " + alerta.motivo + ": " + itensTexto);
        });
        linhas.push("");
      });
      if (grupo.qtd_restante > 0) {
        var textoResto = grupo.qtd_restante === 1 ? "objeto" : "objetos";
        linhas.push("... e mais " + grupo.qtd_restante + " " + textoResto + " com prazo vencido.");
        linhas.push("");
      }
    });
    return linhas.join("\n");
  }

  async function compartilharQualidadeWhatsapp() {
    if (!_avisoQualidadeAtual) return;
    var texto = montarTextoCompartilhamentoQualidade(_avisoQualidadeAtual);
    if (pacApiDesktopDisponivel()) {
      await window.pywebview.api.compartilhar_whatsapp(texto);
    } else {
      // Num navegador comum não precisa passar pelo servidor — o próprio
      // JS já sabe montar o link do WhatsApp Web.
      window.open("https://wa.me/?text=" + encodeURIComponent(texto), "_blank");
    }
  }
  document.getElementById("modal-btn-whatsapp").onclick = compartilharQualidadeWhatsapp;

  // --- Miniaturas das páginas ---
  // São esquemas desenhados em SVG, não imagens do PDF: renderizar o PDF de
  // verdade exigiria uma biblioteca a mais (e o relatório completo levaria
  // segundos para ser gerado só para produzir as miniaturas). O esquema
  // mostra o formato da página — onde ficam título, cards, tabelas e
  // gráficos — que é o que permite reconhecê-la na lista.
  var MINIATURAS_PAGINA = {
    capa:
      '<rect x="6" y="6" width="148" height="96" rx="3" fill="#2E6B66"/>' +
      '<rect x="18" y="30" width="70" height="7" rx="2" fill="#FFFFFF"/>' +
      '<rect x="18" y="43" width="96" height="10" rx="2" fill="#FFFFFF"/>' +
      '<rect x="18" y="60" width="52" height="5" rx="2" fill="#E0AB45"/>',
    texto:
      '<rect x="18" y="16" width="60" height="7" rx="2" fill="#2E6B66"/>' +
      linhasMini(18, 32, 124, 6, 7),
    graficos:
      '<rect x="14" y="12" width="52" height="6" rx="2" fill="#2E6B66"/>' +
      '<rect x="14" y="26" width="60" height="34" rx="3" fill="#9FCE9B"/>' +
      '<circle cx="118" cy="43" r="17" fill="#4E92BA"/>' +
      colunasMini(14, 96, 5, 12, 26),
    tabela:
      '<rect x="14" y="12" width="52" height="6" rx="2" fill="#2E6B66"/>' +
      '<rect x="14" y="26" width="132" height="8" rx="2" fill="#D7EAE7"/>' +
      linhasMini(14, 40, 132, 6, 7),
    texto_tabelas:
      '<rect x="14" y="12" width="52" height="6" rx="2" fill="#2E6B66"/>' +
      linhasMini(14, 24, 132, 4, 5) +
      '<rect x="14" y="52" width="76" height="7" rx="2" fill="#D7EAE7"/>' +
      linhasMini(14, 62, 76, 4, 6) +
      medidoresMini(104, 56, 2),
    medidores:
      '<rect x="14" y="12" width="52" height="6" rx="2" fill="#2E6B66"/>' +
      medidoresMini(18, 34, 4) +
      medidoresMini(18, 68, 4),
    barras:
      '<rect x="14" y="12" width="52" height="6" rx="2" fill="#2E6B66"/>' +
      barrasMini(14, 28, 132, 6, 8),
    mapa:
      '<rect x="14" y="12" width="52" height="6" rx="2" fill="#2E6B66"/>' +
      '<path d="M40 30 L104 26 L128 48 L112 84 L56 88 L28 62 Z" fill="#9FCE9B" ' +
      'stroke="#49925C" stroke-width="1.5"/>' +
      '<path d="M72 28 L74 86 M32 56 L126 52" stroke="#FFFFFF" stroke-width="1"/>',
    cards_tabelas:
      '<rect x="14" y="12" width="52" height="6" rx="2" fill="#2E6B66"/>' +
      '<rect x="14" y="24" width="40" height="16" rx="3" fill="#2E6B66"/>' +
      '<rect x="58" y="24" width="40" height="16" rx="3" fill="#2E6B66"/>' +
      '<rect x="14" y="48" width="64" height="7" rx="2" fill="#D7EAE7"/>' +
      linhasMini(14, 58, 64, 4, 6) +
      '<rect x="86" y="48" width="60" height="7" rx="2" fill="#D7EAE7"/>' +
      linhasMini(86, 58, 60, 4, 6),
    cards_tabela:
      '<rect x="14" y="12" width="52" height="6" rx="2" fill="#2E6B66"/>' +
      '<rect x="14" y="24" width="40" height="16" rx="3" fill="#2E6B66"/>' +
      '<rect x="58" y="24" width="40" height="16" rx="3" fill="#2E6B66"/>' +
      '<rect x="14" y="48" width="132" height="8" rx="2" fill="#D7EAE7"/>' +
      linhasMini(14, 60, 132, 4, 6),
    detalhamento:
      '<rect x="14" y="10" width="46" height="6" rx="2" fill="#2E6B66"/>' +
      '<rect x="14" y="22" width="132" height="7" rx="2" fill="#D7EAE7"/>' +
      linhasMini(14, 33, 132, 3, 5) +
      '<rect x="14" y="56" width="46" height="6" rx="2" fill="#2E6B66"/>' +
      '<rect x="14" y="66" width="132" height="7" rx="2" fill="#D7EAE7"/>' +
      linhasMini(14, 77, 132, 3, 3)
  };

  function linhasMini(x, y, largura, altura, quantidade) {
    var out = "";
    for (var i = 0; i < quantidade; i++) {
      // A última linha sai mais curta, como um parágrafo de verdade.
      var w = i === quantidade - 1 ? largura * 0.6 : largura;
      out += '<rect x="' + x + '" y="' + (y + i * (altura + 3)) + '" width="' + w +
        '" height="' + altura + '" rx="1.5" fill="#B8C4C2"/>';
    }
    return out;
  }

  function colunasMini(x, base, largura, quantidade, alturaMax) {
    var out = "";
    var alturas = [0.5, 0.85, 0.35, 1, 0.6, 0.75, 0.45, 0.9, 0.55, 0.3, 0.7, 0.4];
    for (var i = 0; i < quantidade; i++) {
      var h = alturaMax * alturas[i % alturas.length];
      out += '<rect x="' + (x + i * (largura + 4)) + '" y="' + (base - h) +
        '" width="' + largura + '" height="' + h + '" rx="2" fill="#4E92BA"/>';
    }
    return out;
  }

  function barrasMini(x, y, largura, altura, quantidade) {
    var out = "";
    var fatias = [0.9, 0.7, 0.55, 0.85, 0.4, 0.65, 0.5, 0.75];
    for (var i = 0; i < quantidade; i++) {
      var w = largura * fatias[i % fatias.length];
      out += '<rect x="' + x + '" y="' + (y + i * (altura + 4)) + '" width="' + w +
        '" height="' + altura + '" rx="' + (altura / 2) + '" fill="#49925C"/>';
    }
    return out;
  }

  function medidoresMini(x, y, quantidade) {
    var out = "";
    var voltas = [0.75, 0.45, 0.9, 0.3];
    for (var i = 0; i < quantidade; i++) {
      var cx = x + i * 34 + 13;
      out += '<path d="M' + (cx - 12) + ' ' + (y + 20) + ' A 12 12 0 1 1 ' +
        (cx + 12) + ' ' + (y + 20) + '" fill="none" stroke="#D8D8D8" stroke-width="4" ' +
        'stroke-linecap="round"/>';
      out += '<circle cx="' + (cx + 12 * Math.cos(Math.PI * (1 - voltas[i % voltas.length]))) +
        '" cy="' + (y + 20 - 12 * Math.sin(Math.PI * (1 - voltas[i % voltas.length]))) +
        '" r="3.5" fill="#3F8F52"/>';
    }
    return out;
  }

  function mostrarJanelaPaginas(filtros) {
    // Resolve com a lista de chaves escolhidas, ou com null quando a pessoa
    // cancela (aí a geração inteira é abortada, sem abrir o "Salvar como").
    return new Promise(async function (resolve) {
      var overlay = document.getElementById("paginas-overlay");
      var grade = document.getElementById("paginas-grade");
      var contador = document.getElementById("paginas-contador");

      var lista;
      try {
        lista = await chamarAPI("listar_paginas_relatorio", filtros);
      } catch (erro) {
        alert("Não foi possível montar a lista de páginas:" + NL + NL + erro);
        resolve(null);
        return;
      }
      if (!lista || lista.ok === false || !lista.secoes || !lista.secoes.length) {
        // Sem lista utilizável, segue com o relatório completo em vez de
        // travar a geração por causa da janela de seleção. Resolve com []
        // (não null) porque null é reservado para "usuário cancelou" — o
        // backend trata None/[] do mesmo jeito (relatório completo).
        resolve([]);
        return;
      }

      // O padrão vem do backend: quase tudo marcado, mas o Detalhamento
      // Financeiro chega desmarcado quando o recorte só tem ANDAMENTO.
      var marcadas = {};
      lista.secoes.forEach(function (sec) {
        marcadas[sec.chave] = sec.marcada !== false;
      });

      function atualizarContador() {
        var n = lista.secoes.filter(function (s) { return marcadas[s.chave]; }).length;
        contador.textContent = n + " de " + lista.secoes.length + " selecionadas";
        document.getElementById("paginas-confirmar").disabled = n === 0;
      }

      function desenhar() {
        grade.innerHTML = "";
        lista.secoes.forEach(function (sec) {
          var card = document.createElement("div");
          card.className = "pagina-card" + (marcadas[sec.chave] ? " marcada" : "");
          card.innerHTML =
            '<svg viewBox="0 0 160 108" width="100%">' +
            '<rect x="0" y="0" width="160" height="108" rx="4" fill="#FFFFFF"/>' +
            (MINIATURAS_PAGINA[sec.layout] || "") +
            "</svg>" +
            '<div class="pagina-card-titulo">' + escaparHtmlFicha(sec.titulo) + "</div>" +
            '<div class="pagina-card-sub">' +
            (sec.paginas === null
              ? "várias páginas"
              : sec.paginas > 1
              ? sec.paginas + " páginas"
              : "1 página") +
            "</div>";
          card.onclick = function () {
            marcadas[sec.chave] = !marcadas[sec.chave];
            card.classList.toggle("marcada", marcadas[sec.chave]);
            atualizarContador();
          };
          grade.appendChild(card);
        });
        atualizarContador();
      }

      function fechar() {
        overlay.style.display = "none";
        document.getElementById("paginas-confirmar").onclick = null;
        document.getElementById("paginas-cancelar").onclick = null;
        document.getElementById("paginas-marcar-tudo").onclick = null;
        document.getElementById("paginas-limpar").onclick = null;
        document.getElementById("paginas-colunas-padrao").onclick = null;
        document.getElementById("paginas-colunas-essencial").onclick = null;
      }

      // Seção "Colunas do Detalhamento" — mesma seleção (estadoColunas) que
      // antes vivia no painel de filtros. renderizarBlocoColunas() já lê e
      // escreve direto em estadoColunas; aqui é só (re)desenhá-la com a
      // janela aberta e ligar os dois atalhos. O que ficar marcado ao
      // confirmar entra no relatório porque executarGeracaoRelatorio
      // remonta os filtros depois que esta janela fecha.
      renderizarBlocoColunas();
      document.getElementById("paginas-colunas-padrao").onclick = colunasAplicarPadrao;
      document.getElementById("paginas-colunas-essencial").onclick = colunasAplicarEssencial;

      document.getElementById("paginas-marcar-tudo").onclick = function () {
        lista.secoes.forEach(function (s) { marcadas[s.chave] = true; });
        desenhar();
      };
      document.getElementById("paginas-limpar").onclick = function () {
        lista.secoes.forEach(function (s) { marcadas[s.chave] = false; });
        desenhar();
      };
      document.getElementById("paginas-cancelar").onclick = function () {
        fechar();
        resolve(null);
      };
      document.getElementById("paginas-confirmar").onclick = function () {
        var escolhidas = lista.secoes
          .filter(function (s) { return marcadas[s.chave]; })
          .map(function (s) { return s.chave; });
        if (!escolhidas.length) return;
        fechar();
        resolve(escolhidas);
      };

      desenhar();
      overlay.style.display = "flex";
    });
  }

  // somenteVisualizar: aberto pelo botão de Controle de Qualidade ao lado
  // da lupa — é só consulta, então sem contagem regressiva, sem "Cancelar"
  // e o OK vira "Fechar". Sem esse sinal mantém o comportamento do gate da
  // geração do relatório (OK travado por 10s, OK/Cancelar decidem seguir).
  function mostrarModalQualidade(aviso, somenteVisualizar) {
    return new Promise(function (resolve) {
      _avisoQualidadeAtual = aviso;
      var overlay = document.getElementById("modal-overlay");
      var corpo = document.getElementById("modal-qualidade-texto");
      var contagem = document.getElementById("modal-contagem");
      var btnOk = document.getElementById("modal-btn-ok");
      var btnCancelar = document.getElementById("modal-btn-cancelar");

      corpo.innerHTML = montarHtmlAvisoQualidade(aviso, somenteVisualizar);
      corpo.querySelectorAll(".qc-item-link").forEach(function (link) {
        link.addEventListener("click", async function (ev) {
          ev.preventDefault();
          var item = link.getAttribute("data-item");
          var resultado = await chamarAPI("buscar_ficha_por_item", item);
          if (resultado && resultado.ok) {
            abrirFicha(resultado.dados);
          } else {
            alert((resultado && resultado.erro) || "Não foi possível abrir essa ação.");
          }
        });
      });
      overlay.style.display = "flex";

      var intervalo = null;
      if (somenteVisualizar) {
        contagem.textContent = "";
        btnOk.disabled = false;
        btnOk.textContent = "Fechar";
        btnCancelar.style.display = "none";
      } else {
        btnOk.disabled = true;
        btnOk.textContent = "OK";
        btnCancelar.style.display = "";
        var restante = 10;
        contagem.textContent = "Aguarde " + restante + "s...";
        intervalo = setInterval(function () {
          restante -= 1;
          if (restante <= 0) {
            clearInterval(intervalo);
            contagem.textContent = "";
            btnOk.disabled = false;
          } else {
            contagem.textContent = "Aguarde " + restante + "s...";
          }
        }, 1000);
      }

      function finalizar(confirmado) {
        if (intervalo) clearInterval(intervalo);
        overlay.style.display = "none";
        btnOk.onclick = null;
        btnCancelar.onclick = null;
        btnOk.textContent = "OK";
        btnCancelar.style.display = "";
        resolve(confirmado);
      }

      btnOk.onclick = function () { finalizar(true); };
      btnCancelar.onclick = function () { finalizar(false); };
    });
  }

  // Botão do Controle de Qualidade ao lado da lupa (nas duas barras).
  // Acende com um badge vermelho quando o recorte atual tem ações com
  // pendência de qualidade; some por inteiro quando não tem nenhuma.
  // O aviso é o mesmo objeto que _api_pre_visualizar devolve, guardado em
  // _avisoQualidadeAtual (também usado pelo compartilhamento no WhatsApp).
  function atualizarBotaoQualidade(aviso) {
    _avisoQualidadeAtual = aviso || null;
    var total = (aviso && aviso.total_itens) ? aviso.total_itens : 0;
    document.querySelectorAll(".dash-qc-btn").forEach(function (btn) {
      if (total > 0) {
        btn.style.display = "inline-flex";
        var badge = btn.querySelector(".dash-qc-badge");
        if (badge) badge.textContent = total > 99 ? "99+" : String(total);
        btn.title = "Controle de Qualidade — " + total
          + (total === 1 ? " ação com pendência na base" : " ações com pendências na base");
      } else {
        btn.style.display = "none";
      }
    });
  }

  function abrirModalQualidadeVisualizacao() {
    if (!_avisoQualidadeAtual) return;
    mostrarModalQualidade(_avisoQualidadeAtual, true);
  }
  document.getElementById("dash-qc-btn").onclick = abrirModalQualidadeVisualizacao;
  document.getElementById("filtros-qc-btn").onclick = abrirModalQualidadeVisualizacao;

  // =====================================================
  // Ficha Cadastral de uma ação — aberta pela lupa ao lado dos campos de
  // busca (ITEM/Objeto/Descrição). Não filtra o dashboard: é só uma busca
  // pontual pra abrir os dados completos de UMA ação.
  // =====================================================
  function escaparHtmlFicha(texto) {
    var div = document.createElement("div");
    div.textContent = texto == null ? "" : String(texto);
    return div.innerHTML;
  }

  function definicaoSecoesFicha() {
    return [
      // Sem o campo "Item": o número já está no cabeçalho da ficha.
      { titulo: "Identificação", largo: true, campos: [
        { rotulo: "Objeto", chave: "objeto" },
        { rotulo: "Descrição", chave: "descricao" },
      ]},
      { titulo: "Execução", campos: [
        { rotulo: "Secretaria/Órgão", chave: "secretaria" },
        { rotulo: "Órgão Executor", chave: "executor" },
        { rotulo: "Gestão", chave: "gestao" },
        { rotulo: "Eixo", chave: "eixo" },
        { rotulo: "Município", chave: "municipio" },
        { rotulo: "Fonte de Recurso", chave: "fonte" },
      ]},
      { titulo: "Situação", campos: [
        { rotulo: "Fase", chave: "fase" },
        { rotulo: "Status", chave: "status" },
        { rotulo: "Cláusula Suspensiva", chave: "clausula_suspensiva" },
        { rotulo: "Motivo da Cláusula Suspensiva", chave: "motivo_clausula_suspensiva" },
      ]},
      { titulo: "Prazos", campos: [
        { rotulo: "Vigência", chave: "vigencia" },
        { rotulo: "Previsão de Conclusão Atual", chave: "prazo_atual" },
        { rotulo: "Prazo de Conclusão da Fase", chave: "prazo_fase" },
        { rotulo: "Avanço da Obra", chave: "avanco" },
      ]},
      { titulo: "Licitação", campos: [
        { rotulo: "Aviso de Licitação", chave: "aviso_licitacao" },
        { rotulo: "Abertura de Licitação", chave: "abertura_licitacao" },
        { rotulo: "Emissão de O.S.", chave: "emissao_os" },
      ]},
      { titulo: "Financeiro", campos: [
        { rotulo: "Valor Contratado", chave: "valor_contratado" },
        { rotulo: "Financiamento", chave: "financiamento" },
        { rotulo: "Apoiado (OGU)", chave: "apoiado" },
        { rotulo: "Contrapartida", chave: "contrapartida" },
        { rotulo: "Complementar", chave: "complementar" },
        { rotulo: "Investimento Total", chave: "investimento_total" },
      ]},
      { titulo: "Acompanhamento", largo: true, campos: [
        { rotulo: "Pendências / Tarefa", chave: "pendencia" },
        { rotulo: "Providências (Datas)", chave: "providencias" },
        { rotulo: "Prazo da Pendência / Tarefa", chave: "prazo_pendencia" },
        { rotulo: "Próximos Passos", chave: "proximos_passos" },
      ]},
      { titulo: "Links", largo: true, campos: [
        { rotulo: "Link Monitora", chave: "link_monitora", link: true },
        { rotulo: "Link Localização", chave: "link_localizacao", link: true },
      ]},
    ];
  }

  function valorFichaHtml(valorBruto, ehLink) {
    var valor = (valorBruto == null) ? "" : String(valorBruto).trim();
    if (!valor) return '<span class="ficha-campo-vazio">—</span>';
    if (ehLink) {
      // O Python já normaliza o endereço (acrescenta o https:// quando
      // falta), mas o "www." é aceito aqui também como cinto de segurança:
      // um campo de link que aparece como texto morto é justamente o tipo
      // de defeito que passa despercebido.
      var destino = /^https?:\/\//i.test(valor) ? valor
                  : (/^www\./i.test(valor) ? "https://" + valor : "");
      if (destino) {
        return '<a href="' + escaparHtmlFicha(destino) + '" target="_blank" rel="noopener">'
             + escaparHtmlFicha(valor) + '</a>';
      }
    }
    return escaparHtmlFicha(valor);
  }

  function renderizarFichaTela(dados) {
    var corpo = document.getElementById("ficha-corpo");
    corpo.innerHTML = "";
    var camposAlerta = dados._campos_alerta || [];
    definicaoSecoesFicha().forEach(function (secao) {
      var secaoEl = document.createElement("div");
      secaoEl.className = "ficha-secao";
      var gridClasse = "ficha-secao-grid" + (secao.largo ? " ficha-grid-largo" : "");
      var camposHtml = secao.campos.map(function (c) {
        var classeAlerta = camposAlerta.indexOf(c.chave) !== -1 ? " ficha-campo-alerta" : "";
        return (
          '<div>' +
            '<div class="ficha-campo-rotulo">' + c.rotulo + '</div>' +
            '<div class="ficha-campo-valor' + classeAlerta + '">' + valorFichaHtml(dados[c.chave], c.link) + '</div>' +
          '</div>'
        );
      }).join("");
      secaoEl.innerHTML =
        '<div class="ficha-secao-titulo">' + secao.titulo + '</div>' +
        '<div class="' + gridClasse + '">' + camposHtml + '</div>';
      corpo.appendChild(secaoEl);
    });
  }

  function renderizarFichaA4(dados) {
    var pagina = document.getElementById("ficha-pagina-a4");
    var camposAlerta = dados._campos_alerta || [];
    var html =
      '<div class="ficha-a4-titulo">' + escaparHtmlFicha(dados.objeto || "Ficha Cadastral") + '</div>' +
      '<div class="ficha-a4-subtitulo">PAC - Ficha Cadastral da Ação — Item ' + escaparHtmlFicha(dados.item) + '</div>';
    definicaoSecoesFicha().forEach(function (secao) {
      var gridClasse = "ficha-a4-grid" + (secao.largo ? " ficha-a4-grid-largo" : "");
      var camposHtml = secao.campos.map(function (c) {
        var classeAlerta = camposAlerta.indexOf(c.chave) !== -1 ? " ficha-a4-campo-alerta" : "";
        return (
          '<div>' +
            '<div class="ficha-a4-campo-rotulo">' + c.rotulo + '</div>' +
            '<div class="ficha-a4-campo-valor' + classeAlerta + '">' + valorFichaHtml(dados[c.chave], c.link) + '</div>' +
          '</div>'
        );
      }).join("");
      html +=
        '<div class="ficha-a4-secao">' +
          '<div class="ficha-a4-secao-titulo">' + secao.titulo + '</div>' +
          '<div class="' + gridClasse + '">' + camposHtml + '</div>' +
        '</div>';
    });
    pagina.innerHTML = html;
  }

  function sairModoImpressaoFicha() {
    document.getElementById("ficha-painel").classList.remove("ficha-modo-impressao");
    document.getElementById("ficha-acoes-tela").style.display = "flex";
    document.getElementById("ficha-acoes-impressao").style.display = "none";
  }

  var _fichaDadosAtuais = null;
  // Lembra se a ficha foi aberta a partir da janela de busca. Quem chega
  // pela busca costuma estar percorrendo vários itens da lista, então
  // fechar a ficha devolve a lista em vez de largar a pessoa no dashboard
  // e obrigá-la a refazer a busca a cada ação consultada. Quem chega pelo
  // link do Controle de Qualidade continua voltando direto ao dashboard.
  var _fichaVeioDaBusca = false;

  function abrirFicha(dados, veioDaBusca) {
    _fichaDadosAtuais = dados;
    _fichaVeioDaBusca = !!veioDaBusca;
    document.getElementById("ficha-topo-objeto").textContent = dados.objeto || "(sem objeto)";
    document.getElementById("ficha-topo-item").textContent = dados.item ? ("Item " + dados.item) : "";
    renderizarFichaTela(dados);
    renderizarFichaA4(dados);
    sairModoImpressaoFicha();
    document.getElementById("ficha-overlay").style.display = "flex";
  }

  function fecharFicha() {
    document.getElementById("ficha-overlay").style.display = "none";
    if (_fichaVeioDaBusca) {
      // A janela de busca reaparece com a lista e os filtros como estavam
      // — não é uma busca nova, é a mesma de onde a ficha saiu.
      document.getElementById("ficha-multiplos-overlay").style.display = "flex";
      _fichaVeioDaBusca = false;
    }
  }

  // --- Busca de ação (janela da lupa) ---
  // Os campos de filtro moram DENTRO da janela, junto da lista que eles
  // filtram: digitar reduz a lista na hora, sem precisar de um segundo
  // clique. Antes eles ficavam no dashboard e serviam só para alimentar um
  // dropdown de sugestões, que era uma segunda lista, paralela a esta.
  var _debounceBuscaFicha = null;
  var _buscaFichaEmCurso = 0;

  function campoBuscaMunicipio() { return document.getElementById("busca-ficha-municipio"); }
  function campoBuscaDescricao() { return document.getElementById("busca-ficha-descricao"); }

  async function atualizarListaBuscaFicha() {
    // Cada digitação dispara uma consulta; como elas voltam fora de ordem,
    // só a mais recente pode escrever na tela (o contador abaixo descarta
    // respostas atrasadas de buscas já superadas).
    var minhaVez = ++_buscaFichaEmCurso;
    var municipio = campoBuscaMunicipio().value;
    var descricao = campoBuscaDescricao().value;
    var filtros = montarFiltrosAtuais();
    var resultado = await chamarAPI("buscar_ficha_acao", municipio, descricao, filtros, true);
    if (minhaVez !== _buscaFichaEmCurso) return;
    if (!resultado || !resultado.ok) {
      renderizarListaBuscaFicha([], 0, 0, (resultado && resultado.erro) || "Não foi possível buscar.");
      return;
    }
    renderizarListaBuscaFicha(resultado.itens, resultado.total, resultado.limite, "");
  }

  function agendarAtualizacaoListaBusca() {
    clearTimeout(_debounceBuscaFicha);
    _debounceBuscaFicha = setTimeout(atualizarListaBuscaFicha, 250);
  }

  function renderizarListaBuscaFicha(itens, total, limite, mensagemVazio) {
    var lista = document.getElementById("ficha-multiplos-lista");
    var titulo = document.getElementById("ficha-multiplos-titulo");
    itens = itens || [];
    var qtd = total || itens.length;

    if (titulo) {
      if (!itens.length) {
        titulo.textContent = "Nenhuma ação encontrada";
      } else if (limite && qtd > limite) {
        titulo.textContent = "Mostrando " + limite + " de " + qtd
          + " ações — refine a busca para ver as demais";
      } else {
        titulo.textContent = qtd + (qtd === 1 ? " ação encontrada" : " ações encontradas");
      }
    }

    lista.innerHTML = "";
    if (!itens.length) {
      var vazio = document.createElement("div");
      vazio.className = "ficha-multiplos-lista-vazia";
      vazio.textContent = mensagemVazio || "Nenhuma ação com esses critérios no recorte atual.";
      lista.appendChild(vazio);
      return;
    }

    itens.forEach(function (item) {
      var el = document.createElement("div");
      el.className = "ficha-multiplos-item";
      var metaTexto = escaparHtmlFicha(item.secretaria || "Sem secretaria");
      if (item.descricao) {
        metaTexto += '<br>' + escaparHtmlFicha(item.descricao);
      }
      el.innerHTML =
        '<div class="ficha-multiplos-item-info">' +
          '<div class="ficha-multiplos-item-objeto">#' + escaparHtmlFicha(item.item) + ' — ' + escaparHtmlFicha(item.objeto) + '</div>' +
          '<div class="ficha-multiplos-item-meta">' + metaTexto + '</div>' +
        '</div>';
      el.addEventListener("click", function () { abrirFichaDaBusca(item.item); });
      lista.appendChild(el);
    });
  }

  async function abrirFichaDaBusca(item) {
    document.getElementById("ficha-multiplos-overlay").style.display = "none";
    var resultado = await chamarAPI("buscar_ficha_por_item", item);
    if (resultado && resultado.ok) {
      abrirFicha(resultado.dados, true);
    } else {
      // Falhou ao abrir: devolve a janela de busca, senão a pessoa fica
      // sem lista e sem ficha depois do alerta.
      document.getElementById("ficha-multiplos-overlay").style.display = "flex";
      alert((resultado && resultado.erro) || "Não foi possível abrir essa ação.");
    }
  }

  function abrirJanelaBuscaFicha() {
    // A janela abre já com a lista do recorte atual (campos limpos), que é
    // o caso de quem quer navegar em vez de procurar um nome específico.
    campoBuscaMunicipio().value = "";
    campoBuscaDescricao().value = "";
    renderizarListaBuscaFicha([], 0, 0, "Carregando ações do recorte atual...");
    document.getElementById("ficha-multiplos-overlay").style.display = "flex";
    // Foco no primeiro campo da linha — Descrição, que é por onde a busca
    // costuma começar.
    campoBuscaDescricao().focus();
    atualizarListaBuscaFicha();
  }

  document.getElementById("dash-busca-btn-lupa").onclick = abrirJanelaBuscaFicha;
  // Mesma lupa no painel de filtros: a busca por ação é útil nos dois
  // lugares, e quem está montando o recorte não precisa passar pelo
  // dashboard só para consultar uma ficha. A janela de busca já respeita
  // os filtros marcados no momento (montarFiltrosAtuais), então o
  // resultado é o mesmo recorte nos dois pontos de entrada.
  document.getElementById("filtros-busca-btn-lupa").onclick = abrirJanelaBuscaFicha;
  [campoBuscaMunicipio(), campoBuscaDescricao()].forEach(function (campo) {
    campo.addEventListener("input", agendarAtualizacaoListaBusca);
    campo.addEventListener("keydown", function (ev) {
      if (ev.key === "Escape") {
        document.getElementById("ficha-multiplos-overlay").style.display = "none";
      }
      if (ev.key === "Enter") {
        // Enter abre direto quando a busca já se resumiu a uma única ação.
        var itens = document.querySelectorAll("#ficha-multiplos-lista .ficha-multiplos-item");
        if (itens.length === 1) itens[0].click();
      }
    });
  });

  document.getElementById("ficha-btn-fechar").onclick = fecharFicha;
  document.getElementById("ficha-multiplos-cancelar").onclick = function () {
    document.getElementById("ficha-multiplos-overlay").style.display = "none";
    _fichaVeioDaBusca = false;
  };
  document.getElementById("ficha-btn-preview-impressao").onclick = function () {
    document.getElementById("ficha-painel").classList.add("ficha-modo-impressao");
    document.getElementById("ficha-acoes-tela").style.display = "none";
    document.getElementById("ficha-acoes-impressao").style.display = "flex";
    document.getElementById("ficha-corpo").scrollTop = 0;
  };
  document.getElementById("ficha-btn-voltar-tela").onclick = sairModoImpressaoFicha;
  document.getElementById("ficha-btn-fechar-impressao").onclick = fecharFicha;
  async function salvarFichaComoPdf() {
    if (!_fichaDadosAtuais) return;
    // No desktop, o Python salva o PDF direto no disco escolhido pelo
    // diálogo nativo; num navegador comum, baixarPDF() já abre o PDF numa
    // nova aba a partir dos bytes devolvidos pelo servidor — por isso a
    // mensagem fala em "gerado", não "salvo em disco", e serve pros dois
    // casos.
    var resultado = await baixarPDF("exportar_ficha_pdf", _fichaDadosAtuais.item);
    if (!resultado || resultado.cancelado) return;
    if (resultado.ok) {
      alert("PDF gerado com sucesso: " + resultado.arquivo);
    } else {
      alert(resultado.erro || "Não foi possível gerar o PDF.");
    }
  }

  document.getElementById("ficha-btn-salvar-pdf").onclick = salvarFichaComoPdf;

  async function executarGeracaoRelatorio() {
    var filtros = montarFiltrosAtuais();

    travarBotaoGerar("Verificando...");
    var inicio;
    try {
      inicio = await chamarAPI("iniciar_geracao", filtros);
    } catch (erro) {
      alert("Ocorreu um erro ao verificar os filtros:" + NL + NL + erro);
      destravarBotaoGerar();
      return;
    }

    if (!inicio || inicio.ok === false) {
      if (inicio && inicio.vazio) {
        alert("Nenhum registro encontrado para os filtros selecionados. O relatorio nao sera gerado.");
      } else {
        alert("Ocorreu um erro ao verificar os filtros:" + NL + NL + (inicio ? inicio.erro : "erro desconhecido"));
      }
      destravarBotaoGerar();
      return;
    }

    if (inicio.aviso) {
      destravarBotaoGerar();
      var prosseguir = await mostrarModalQualidade(inicio.aviso);
      if (!prosseguir) return;
    }

    // Janela de seleção de páginas — entre a checagem de qualidade e o
    // diálogo de salvar. Sai daqui com a lista de seções escolhidas.
    var secoesEscolhidas = await mostrarJanelaPaginas(filtros);
    if (secoesEscolhidas === null) {
      destravarBotaoGerar();
      return;
    }
    // A janela de páginas é também onde as colunas do Detalhamento são
    // escolhidas agora — remonta os filtros para levar a seleção final.
    filtros = montarFiltrosAtuais();

    travarBotaoGerar("Gerando...");
    var resultado;
    try {
      resultado = await baixarPDF("escolher_local_e_gerar", filtros, secoesEscolhidas);
    } catch (erro) {
      alert("Ocorreu um erro ao gerar o relatorio:" + NL + NL + erro);
      destravarBotaoGerar();
      return;
    }
    destravarBotaoGerar();

    if (!resultado || resultado.ok === false) {
      if (resultado && resultado.cancelado) return;
      if (resultado && resultado.vazio) {
        alert("Nenhum registro encontrado para os filtros selecionados. O relatorio nao sera gerado.");
        return;
      }
      alert("Ocorreu um erro ao gerar o relatorio:" + NL + NL + (resultado ? resultado.erro : "erro desconhecido"));
      return;
    }

    if (resultado.fallback) {
      alert(
        "Nao foi possivel acessar a pasta de rede configurada para salvar o PDF "
        + "(unidade X: pode nao estar mapeada nesta maquina)." + NL + NL
        + "O relatorio foi salvo em:" + NL + resultado.arquivo
      );
    }
  }

  botoesGerar.forEach(function (b) { b.onclick = executarGeracaoRelatorio; });

  // --- Fechamento automático por inatividade (5 minutos) ---
  const TEMPO_INATIVIDADE_MS = 15 * 60 * 1000;
  let temporizadorInatividade = null;

  function reiniciarTemporizadorInatividade() {
    if (temporizadorInatividade) clearTimeout(temporizadorInatividade);
    temporizadorInatividade = setTimeout(() => {
      // Enquanto a Ficha Cadastral (ou a pré-visualização de impressão
      // dela) está aberta, a pessoa pode ficar um bom tempo só lendo, sem
      // mexer no mouse/teclado — isso não pode fechar o programa. Só
      // adia o fechamento, tentando de novo mais tarde.
      var fichaAberta = document.getElementById("ficha-overlay").style.display === "flex";
      if (fichaAberta) {
        reiniciarTemporizadorInatividade();
        return;
      }
      document.getElementById("aviso-inatividade").style.display = "flex";
      // Fechar a própria janela só existe no modo desktop — no navegador
      // não há janela nativa pra fechar, o aviso na tela já é suficiente.
      if (pacApiDesktopDisponivel()) {
        window.pywebview.api.fechar_por_inatividade();
      }
    }, TEMPO_INATIVIDADE_MS);
  }
  ["mousemove", "keydown", "click", "wheel"].forEach(evento => {
    document.addEventListener(evento, reiniciarTemporizadorInatividade);
  });
  reiniciarTemporizadorInatividade();

  configurarAccordionSecoes();

  // O dashboard agora é a tela inicial — carrega os dados assim que a
  // página abre, sem esperar nenhum clique. No modo desktop, o pywebview
  // injeta o objeto "window.pywebview" de forma assíncrona (evento
  // "pywebviewready"), então é preciso esperar esse evento explicitamente
  // antes de chamar a API, ou "window.pywebview" pode ainda não existir no
  // instante certo. No modo web esse objeto nunca existe — carrega direto,
  // sem esperar nada (chamarAPI já fala com o servidor por HTTP).
  if (window.PAC_MODO_WEB || window.pywebview) {
    carregarDashboard();
  } else {
    window.addEventListener("pywebviewready", function () {
      carregarDashboard();
    });
  }
</script>
</body>
</html>
"""

    html_paginal = html_paginal.replace("__DADOS_PAINEL__", dados_painel_json)
    texto_ultima_atualizacao_painel = (
        f"Base de dados atualizada em {ultima_atualizacao.strftime('%d/%m/%Y às %Hh%Mmin')}"
    )
    html_paginal = html_paginal.replace("__ULTIMA_ATUALIZACAO__", texto_ultima_atualizacao_painel)
    # Diz pro JS do painel se ele está rodando dentro da janela desktop
    # (pywebview) ou servido por um navegador comum (servidor web) — em vez
    # de o JS ter que adivinhar isso por uma corrida entre o carregamento da
    # página e a injeção assíncrona de "window.pywebview" pelo WebView2.
    html_paginal = html_paginal.replace("__MODO_WEB__", "true" if MODO_WEB else "false")

    return html_paginal


def _data_iso_mapa_mental(valor):
    if valor is None or (isinstance(valor, float) and pd.isna(valor)):
        return None
    try:
        ts = pd.Timestamp(valor)
    except (TypeError, ValueError):
        return None
    if pd.isna(ts):
        return None
    return ts.date().isoformat()


def _construir_arvore_mapa_mental(df):
    # Secretaria/Órgão > Executor > Objeto > Ação (item) — sem o nível de
    # Eixo (o Objeto já separa bem as ações dentro de cada Executor, e o
    # Eixo só acrescentava mais um clique sem separar muita coisa). O nó do
    # Objeto não expande mais dentro do próprio mapa: clicar nele abre a
    # lista das ações no painel lateral (ver abrirListaAcoes no JS), e
    # clicar numa ação da lista abre a Ficha Cadastral completa (ver
    # abrirPainel) — os nós do tipo "item" continuam existindo nos dados
    # (para contagem, busca e a lista lateral), só não são mais desenhados
    # como cartão dentro da árvore. Cada folha reaproveita EXATAMENTE os
    # mesmos dados já usados na Ficha Cadastral (_montar_dados_ficha_acao)
    # — uma única fonte de verdade para o que aparece nos dois lugares — e
    # só acrescenta um "id" (garantidamente único, pelo índice da linha) e
    # a data em ISO da Previsão de Conclusão Atual (para o JS calcular
    # dias restantes sem precisar reinterpretar o texto já formatado).
    raiz = {"nome": "BALANÇO PAC - BAHIA", "tipo": "raiz", "filhos": {}}

    for idx, row in df.iterrows():
        orgao = str(row.get("SECRETARIA_LIMPA") or "").strip() or "Sem órgão definido"

        dados = _montar_dados_ficha_acao(row)
        dados["id"] = str(idx)
        dados["prazo_atual_iso"] = _data_iso_mapa_mental(row.get(col_prazo_atual))

        executor = dados["executor"] or "Sem executor definido"
        objeto = dados["objeto"] or "Sem objeto"

        n_org = raiz["filhos"].setdefault(orgao, {"nome": orgao, "tipo": "orgao", "filhos": {}})
        n_exec = n_org["filhos"].setdefault(executor, {"nome": executor, "tipo": "executor", "filhos": {}})
        n_obj = n_exec["filhos"].setdefault(objeto, {"nome": objeto, "tipo": "objeto", "filhos": {}})

        n_obj["filhos"][f"item-{idx}"] = {
            "nome": dados["descricao"] or dados["objeto"],
            "tipo": "item",
            "dados": dados,
            "filhos": {},
        }

    def para_lista(no):
        no = dict(no)
        no["filhos"] = [para_lista(f) for f in no["filhos"].values()]
        return no

    return para_lista(raiz)


# =====================================================
# API DO PAINEL — funções de módulo compartilhadas pela ponte pywebview
# (classe APIFiltros, usada no modo desktop) e pelas rotas do servidor web
# (servidor_web.py). Cada uma recebe os mesmos argumentos que o JS do
# painel já envia e devolve um dict JSON-serializável, sempre no formato
# {"ok": True, ...} ou {"ok": False, "erro"/"vazio"/"cancelado": ...} —
# mesmo contrato que o pywebview já usava, para o JS do painel não precisar
# saber se está falando com uma janela nativa ou com um servidor HTTP.
# =====================================================

def _api_mapa_mental(filtros):
    # Monta a página HTML autônoma do mapa mental (ver mapa_mental_html.py)
    # a partir do MESMO recorte definido pelos filtros ativos no painel no
    # momento do clique — igual ao botão DASHBOARD. Devolve o HTML pronto
    # como texto (em vez de já abrir alguma coisa): quem chamou (JS do
    # botão "MAPA MENTAL") é quem decide como abrir — sempre numa aba/
    # janela nova, tanto no desktop quanto no navegador.
    try:
        df = _filtrar_dataframe(filtros)
        if df.empty:
            return {"ok": False, "vazio": True}
        arvore = _construir_arvore_mapa_mental(df)
        meta = {
            "titulo": "BALANÇO PAC - BAHIA",
            "titulo_aba": "Mapa Mental — BALANÇO PAC",
            "subtitulo": "Secretarias/Órgãos, Executores e Objetos com ações do PAC",
            "total": int(len(df)),
            "gerado_em": datetime.now().strftime("%d/%m/%Y %H:%M"),
            # tema atual do painel (chave claro/escuro no topo) — o mapa
            # mental é um <iframe srcdoc> isolado, sem acesso ao
            # localStorage da página principal, então recebe pronto aqui
            # (ver filtros.tema no JS do botão MAPA MENTAL).
            "tema": filtros.get("tema"),
        }
        return {"ok": True, "html": montar_html_mapa_mental(arvore, meta)}
    except Exception as erro:
        import traceback
        traceback.print_exc()
        return {"ok": False, "erro": str(erro)}


def _api_pre_visualizar(filtros):
    # Filtra os dados e devolve os números já agregados (sem gerar nenhum
    # PDF) para o JS desenhar os mesmos 4 gráficos do Painel Geral em CSS —
    # um preview rápido de como o relatório vai sair, sem precisar gerar o
    # arquivo.
    try:
        df = _filtrar_dataframe(filtros)
        if df.empty:
            return {"ok": False, "vazio": True}
        # Vai junto o aviso do Controle de Qualidade do MESMO recorte, para
        # o painel acender (ou esconder) o botão de qualidade ao lado da
        # lupa sem uma segunda chamada.
        return {
            "ok": True,
            "dados": _dados_pre_visualizacao(df),
            "aviso": _montar_aviso_qualidade(df),
        }
    except Exception as erro:
        import traceback
        traceback.print_exc()
        return {"ok": False, "erro": str(erro)}


def _api_iniciar_geracao(filtros):
    # Primeira etapa: filtra os dados e verifica o Controle de Qualidade.
    # Não abre nenhum diálogo — só devolve os dados para o JS decidir o que
    # mostrar.
    try:
        df = _filtrar_dataframe(filtros)
        if df.empty:
            return {"ok": False, "vazio": True}
        aviso = _montar_aviso_qualidade(df)
        return {"ok": True, "aviso": aviso}
    except Exception as erro:
        import traceback
        traceback.print_exc()
        return {"ok": False, "erro": str(erro)}


def _api_aviso_qualidade(filtros):
    # Só o aviso do Controle de Qualidade do recorte atual — sem os
    # agregados do dashboard nem geração de PDF. O painel de filtros chama
    # isto a cada mudança de filtro (com debounce) para o botão de qualidade
    # ao lado da lupa acender, esconder ou atualizar o badge na hora, sem
    # precisar fechar os filtros e recarregar o dashboard. Recorte vazio não
    # é erro aqui: só significa "nenhuma pendência" (aviso = None).
    try:
        df = _filtrar_dataframe(filtros)
        if df.empty:
            return {"ok": True, "aviso": None}
        return {"ok": True, "aviso": _montar_aviso_qualidade(df)}
    except Exception as erro:
        import traceback
        traceback.print_exc()
        return {"ok": False, "erro": str(erro)}


def _api_listar_paginas_relatorio(filtros):
    # Alimenta a janela de seleção de páginas: devolve só as seções que
    # este recorte realmente produziria. Barato — não gera PDF nenhum, só
    # inspeciona o dataframe filtrado.
    try:
        df = _filtrar_dataframe(filtros)
        if df.empty:
            return {"ok": False, "vazio": True}
        return {"ok": True, "secoes": _secoes_disponiveis_relatorio(df)}
    except Exception as erro:
        import traceback
        traceback.print_exc()
        return {"ok": False, "erro": str(erro)}


def _api_buscar_ficha_acao(municipio, descricao, filtros, sempre_lista=False):
    # Busca por MUNICÍPIO (parcial) e/ou DESCRIÇÃO (parcial), combinados
    # com E — quando os dois estão preenchidos, os dois precisam bater.
    # Sempre dentro do recorte definido pelos filtros ativos no momento
    # (secretaria/executor selecionados etc.) — os botões acima do campo de
    # busca influenciam o resultado.
    try:
        df_busca = _filtrar_dataframe(filtros) if filtros else df_original
        municipio_t = str(municipio or "").strip()
        descricao_t = str(descricao or "").strip()
        # Sem nada digitado a busca NÃO é recusada: vale o recorte atual dos
        # filtros, e a lupa vira uma forma de navegar pelas ações que já
        # estão na tela.

        if municipio_t and col_municipio in df_busca.columns:
            municipio_norm = remover_acentos(municipio_t).upper()
            municipio_serie = df_busca[col_municipio].astype(str).apply(
                lambda v: remover_acentos(v.upper())
            )
            df_busca = df_busca[municipio_serie.str.contains(municipio_norm, regex=False)]
        elif municipio_t:
            df_busca = df_busca.iloc[0:0]
        if descricao_t and col_descricao in df_busca.columns:
            descricao_norm = remover_acentos(descricao_t).upper()
            descricao_serie = df_busca[col_descricao].astype(str).apply(
                lambda v: remover_acentos(v.upper())
            )
            df_busca = df_busca[descricao_serie.str.contains(descricao_norm, regex=False)]

        # sempre_lista: a janela da lupa filtra a lista enquanto a pessoa
        # digita, então ela nunca quer um alerta nem quer que a ficha abra
        # sozinha no meio da digitação — quer a lista, mesmo que vazia ou
        # com um item só. Quem chama sem esse sinal mantém o comportamento
        # antigo.
        if df_busca.empty:
            if sempre_lista:
                return {"ok": True, "multiplos": True, "itens": [], "total": 0,
                        "limite": LIMITE_LISTA_BUSCA_FICHA}
            if not municipio_t and not descricao_t:
                return {"ok": False, "erro": "Nenhuma ação no recorte atual — reveja os filtros do painel."}
            return {"ok": False, "erro": "Nenhuma ação encontrada com esses critérios."}

        acoes_unicas = df_busca.drop_duplicates(subset=[col_item])
        if sempre_lista or len(acoes_unicas) > 1:
            itens = [
                {
                    "item": normalizar_item(r[col_item]),
                    "objeto": str(r[col_objeto]),
                    "descricao": _texto_campo_ficha(r, col_descricao),
                    "secretaria": str(r.get(col_orgao, "")),
                }
                for _, r in acoes_unicas.head(LIMITE_LISTA_BUSCA_FICHA).iterrows()
            ]
            # "total" é quantas ações existem no recorte; "limite", quantas
            # cabem na lista. A tela usa os dois para avisar que a lista
            # está cortada — sem isso, uma busca vazia num recorte grande
            # pareceria devolver só 30 ações.
            return {
                "ok": True,
                "multiplos": True,
                "itens": itens,
                "total": int(len(acoes_unicas)),
                "limite": LIMITE_LISTA_BUSCA_FICHA,
            }

        return {"ok": True, "dados": _montar_dados_ficha_acao(df_busca.iloc[0])}
    except Exception as erro:
        import traceback
        traceback.print_exc()
        return {"ok": False, "erro": str(erro)}


def _api_buscar_ficha_por_item(item):
    # Usado quando a busca acima encontra mais de uma ação e o usuário
    # escolhe qual delas abrir.
    try:
        item_norm = remover_acentos(normalizar_item(item)).upper()
        item_serie = df_original[col_item].apply(lambda v: remover_acentos(normalizar_item(v)).upper())
        linhas = df_original[item_serie == item_norm]
        if linhas.empty:
            return {"ok": False, "erro": "Ação não encontrada."}
        return {"ok": True, "dados": _montar_dados_ficha_acao(linhas.iloc[0])}
    except Exception as erro:
        import traceback
        traceback.print_exc()
        return {"ok": False, "erro": str(erro)}


def _localizar_acao_por_item(item):
    # Helper comum a exportar_ficha_pdf (desktop) e à rota web
    # equivalente: acha a linha da planilha correspondente a um ITEM e
    # devolve os dados já montados para a Ficha Cadastral, ou None se não
    # existir.
    item_norm = remover_acentos(normalizar_item(item)).upper()
    item_serie = df_original[col_item].apply(lambda v: remover_acentos(normalizar_item(v)).upper())
    linhas = df_original[item_serie == item_norm]
    if linhas.empty:
        return None
    return _montar_dados_ficha_acao(linhas.iloc[0])


def _api_gerar_pdf_bytes(filtros, secoes=None):
    # Equivalente web de APIFiltros.escolher_local_e_gerar: em vez de abrir
    # o diálogo nativo "Salvar como" do pywebview (que não existe num
    # navegador comum), gera o PDF num arquivo temporário do próprio
    # servidor, lê os bytes de volta e apaga o arquivo — quem baixa/abre o
    # PDF de fato é o navegador de quem está acessando, a partir da
    # resposta HTTP (ver servidor_web.py).
    df = _filtrar_dataframe(filtros)
    if df.empty:
        return {"ok": False, "vazio": True}

    # tempfile.mkstemp gera um nome garantidamente único de forma atômica
    # (ao contrário de montar o nome à mão com pid/id de objeto) — importa
    # aqui porque o servidor web pode atender vários pedidos de PDF ao mesmo
    # tempo, cada um numa requisição/thread diferente.
    descritor, caminho_temp = tempfile.mkstemp(suffix=".pdf", prefix="pac_relatorio_web_")
    os.close(descritor)
    try:
        _gerar_pdf(df, caminho_temp, filtros.get("COLUNAS_DETALHAMENTO"), secoes)
        with open(caminho_temp, "rb") as f:
            conteudo_pdf = f.read()
        return {"ok": True, "nome_arquivo": nome_arquivo_pdf, "conteudo": conteudo_pdf}
    finally:
        try:
            os.remove(caminho_temp)
        except Exception:
            pass


def _api_exportar_ficha_pdf_bytes(item):
    # Equivalente web de APIFiltros.exportar_ficha_pdf — mesma lógica do
    # gerar_pdf_bytes acima, mas para o PDF de uma única ação.
    dados = _localizar_acao_por_item(item)
    if dados is None:
        return {"ok": False, "erro": "Ação não encontrada."}

    nome_sugerido = f"Ficha_Item_{normalizar_item(item)}.pdf"
    descritor, caminho_temp = tempfile.mkstemp(suffix=".pdf", prefix="pac_ficha_web_")
    os.close(descritor)
    try:
        gerar_pdf_ficha_acao(dados, caminho_temp)
        with open(caminho_temp, "rb") as f:
            conteudo_pdf = f.read()
        return {"ok": True, "nome_arquivo": nome_sugerido, "conteudo": conteudo_pdf}
    finally:
        try:
            os.remove(caminho_temp)
        except Exception:
            pass


def abrir_interface_filtros(df_base):
    # Ponto de entrada exclusivo do modo desktop: pega o HTML pronto de
    # montar_html_painel e abre numa janela nativa (pywebview/WebView2). O
    # servidor web NÃO chama esta função — ele usa montar_html_painel
    # diretamente e serve o resultado como resposta HTTP normal (ver
    # servidor_web.py).
    html_paginal = montar_html_painel(df_base)

    class APIFiltros:
        # Ponte entre o JS do painel e as funções Python já existentes.
        # NÃO usa mais nenhum diálogo do Tkinter (messagebox/filedialog/
        # Toplevel): misturar o loop de mensagens do Tcl/Tk com o loop do
        # pywebview, chamado de uma thread de fundo, causava a janela em
        # branco primeiro e depois o travamento "Gerando...". O diálogo de
        # salvar agora usa o recurso nativo do próprio pywebview
        # (window.create_file_dialog), e o aviso de Controle de Qualidade
        # virou um modal em HTML, exibido dentro do próprio painel.

        def pre_visualizar(self, filtros):
            return _api_pre_visualizar(filtros)

        def mapa_mental(self, filtros):
            return _api_mapa_mental(filtros)

        def iniciar_geracao(self, filtros):
            return _api_iniciar_geracao(filtros)

        def aviso_qualidade(self, filtros):
            return _api_aviso_qualidade(filtros)

        def listar_paginas_relatorio(self, filtros):
            return _api_listar_paginas_relatorio(filtros)

        def escolher_local_e_gerar(self, filtros, secoes=None):
            # Segunda etapa: reaplica os mesmos filtros (barato — é só
            # filtrar o DataFrame já carregado em memória), pede o local de
            # salvamento pelo diálogo nativo do pywebview e gera o PDF.
            try:
                df = _filtrar_dataframe(filtros)
                if df.empty:
                    return {"ok": False, "vazio": True}

                pasta_inicial = (
                    PASTA_DOWNLOADS_PADRAO if os.path.isdir(PASTA_DOWNLOADS_PADRAO) else PASTA_BASE
                )
                resultado_dialogo = janela.create_file_dialog(
                    webview.SAVE_DIALOG,
                    directory=pasta_inicial,
                    save_filename=nome_arquivo_pdf,
                    file_types=("Arquivo PDF (*.pdf)",),
                )
                if not resultado_dialogo:
                    # Usuário cancelou (ou fechou) o diálogo "Salvar como".
                    return {"ok": False, "cancelado": True}

                caminho_escolhido = (
                    resultado_dialogo[0]
                    if isinstance(resultado_dialogo, (list, tuple))
                    else resultado_dialogo
                )
                if not caminho_escolhido.lower().endswith(".pdf"):
                    caminho_escolhido += ".pdf"

                arquivo_pdf = caminho_escolhido
                fallback_usado = False
                try:
                    os.makedirs(os.path.dirname(arquivo_pdf), exist_ok=True)
                except Exception:
                    arquivo_pdf = caminho_recurso(nome_arquivo_pdf)
                    fallback_usado = True

                _gerar_pdf(df, arquivo_pdf, filtros.get("COLUNAS_DETALHAMENTO"), secoes)
                return {"ok": True, "arquivo": arquivo_pdf, "fallback": fallback_usado}
            except Exception as erro:
                import traceback
                traceback.print_exc()
                return {"ok": False, "erro": str(erro)}

        def buscar_ficha_acao(self, municipio, descricao, filtros, sempre_lista=False):
            return _api_buscar_ficha_acao(municipio, descricao, filtros, sempre_lista)

        def buscar_ficha_por_item(self, item):
            return _api_buscar_ficha_por_item(item)

        def exportar_ficha_pdf(self, item):
            # Botão "Salvar em PDF" da Ficha Cadastral — gera um PDF de uma
            # página A4 só com essa ação, e pede o local de salvamento pelo
            # diálogo nativo do pywebview (mesmo padrão do relatório
            # completo). Exclusivo do modo desktop: o equivalente web é
            # _api_exportar_ficha_pdf_bytes, que devolve os bytes do PDF em
            # vez de salvar direto no disco (não há diálogo nativo num
            # navegador comum).
            try:
                dados = _localizar_acao_por_item(item)
                if dados is None:
                    return {"ok": False, "erro": "Ação não encontrada."}

                pasta_inicial = (
                    PASTA_DOWNLOADS_PADRAO if os.path.isdir(PASTA_DOWNLOADS_PADRAO) else PASTA_BASE
                )
                nome_sugerido = f"Ficha_Item_{normalizar_item(item)}.pdf"
                resultado_dialogo = janela.create_file_dialog(
                    webview.SAVE_DIALOG,
                    directory=pasta_inicial,
                    save_filename=nome_sugerido,
                    file_types=("Arquivo PDF (*.pdf)",),
                )
                if not resultado_dialogo:
                    return {"ok": False, "cancelado": True}

                caminho_escolhido = (
                    resultado_dialogo[0]
                    if isinstance(resultado_dialogo, (list, tuple))
                    else resultado_dialogo
                )
                if not caminho_escolhido.lower().endswith(".pdf"):
                    caminho_escolhido += ".pdf"

                gerar_pdf_ficha_acao(dados, caminho_escolhido)
                return {"ok": True, "arquivo": caminho_escolhido}
            except Exception as erro:
                import traceback
                traceback.print_exc()
                return {"ok": False, "erro": str(erro)}

        def compartilhar_whatsapp(self, texto):
            # Botão de compartilhar do painel de Controle de Qualidade —
            # abre o WhatsApp Web (ou o aplicativo do WhatsApp, se
            # instalado e configurado como padrão pra esses links) no
            # navegador padrão do usuário, já com a mensagem preenchida.
            # new=0 pede pro navegador reaproveitar uma janela já aberta em
            # vez de abrir uma nova sempre — o Python não tem como saber
            # com certeza se JÁ existe uma aba do WhatsApp Web aberta
            # especificamente (isso depende do navegador de cada um), mas
            # essa opção reduz bastante a chance de abrir janela/aba nova
            # à toa quando o navegador já está aberto.
            try:
                import urllib.parse

                url = "https://wa.me/?text=" + urllib.parse.quote(str(texto or ""))
                webbrowser.open(url, new=0, autoraise=True)
                return {"ok": True}
            except Exception as erro:
                import traceback
                traceback.print_exc()
                return {"ok": False, "erro": str(erro)}

        def fechar_por_inatividade(self):
            if janela is not None:
                janela.destroy()

        def verificar_mudancas_git(self):
            return _git_verificar_mudancas()

        def publicar_atualizacao_git(self, mensagem):
            return _git_publicar_atualizacao(mensagem)

    # Escreve o HTML num arquivo temporário e abre via caminho de arquivo
    # (url=) em vez de passar a string inteira direto (html=). Passar uma
    # string HTML muito grande para o WebView2 via html= tem limitações de
    # tamanho/escaping que podem cortar ou corromper o conteúdo de forma
    # imprevisível (o sintoma era um SyntaxError sempre no mesmo tipo de
    # lugar, mesmo com o JS validado como correto) — carregando de um
    # arquivo real, o navegador lê exatamente como está gravado no disco,
    # sem esse tipo de limite.
    arquivo_html_temp = os.path.join(
        tempfile.gettempdir(), f"pac_filtros_{os.getpid()}.html"
    )
    with open(arquivo_html_temp, "w", encoding="utf-8") as f:
        f.write(html_paginal)

    janela = webview.create_window(
        "CGAPE - BALANÇO PAC",
        url=arquivo_html_temp,
        js_api=APIFiltros(),
        width=1600,
        height=900,
        maximized=True,
        background_color="#303030",
    )
    # No Windows, o pywebview usa o WebView2 (motor do Edge) por padrão,
    # já presente de fábrica no Windows 10/11 — não exige instalar nada
    # além do pacote "pywebview" (pip install pywebview) para empacotar
    # com o PyInstaller.
    webview.start()

    # Limpeza do arquivo temporário ao fechar o painel.
    try:
        os.remove(arquivo_html_temp)
    except Exception:
        pass


if __name__ == "__main__":
    # Só abre a janela desktop quando o arquivo é executado diretamente
    # (duplo clique no .exe, ou "python CGAPE - BALANÇO PAC.py"). Quando
    # este arquivo é IMPORTADO — caso do servidor web, em servidor_web.py —
    # todo o carregamento/ETL da planilha acima ainda roda normalmente
    # (dados prontos em df_original), mas a janela nativa não abre sozinha.
    abrir_interface_filtros(df_original)
